import sys
import sqlite3
import re
import json
import hashlib
import requests
import markdown
import gzip
import uuid
from datetime import datetime, timedelta
import math
import matplotlib
matplotlib.use('Qt5Agg')  # 设置matplotlib使用Qt5后端
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.dates as mdates
# 尝试使用系统字体，避免斜体问题
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'SimSun']  # 设置中文字体优先级
plt.rcParams['font.size'] = 8  # 设置更小的默认字体大小
plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题
plt.rcParams['figure.titlesize'] = 9   # 图表标题字体大小
plt.rcParams['axes.titlesize'] = 8    # 轴标题字体大小
plt.rcParams['axes.labelsize'] = 7     # 轴标签字体大小
plt.rcParams['xtick.labelsize'] = 6   # X轴刻度标签字体大小
plt.rcParams['ytick.labelsize'] = 6   # Y轴刻度标签字体大小
plt.rcParams['legend.fontsize'] = 6   # 图例字体大小

# 强制设置字体为非斜体
plt.rcParams['font.style'] = 'normal'

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QGroupBox, QLabel, QComboBox, QLineEdit, QCheckBox, QPushButton, QTableWidget,
    QTableWidgetItem, QMessageBox, QFileDialog, QInputDialog, QHeaderView, QAbstractItemView,
    QFrame, QStatusBar, QDateEdit, QDialog, QDialogButtonBox, QFormLayout, QShortcut, QAction, QMenu,
    QColorDialog, QListWidget, QListWidgetItem, QItemDelegate, QFontDialog, QSpinBox, QSlider, QSplitter,
    QSizePolicy, QProgressDialog, QTextEdit, QSystemTrayIcon, QTabWidget
)
from PyQt5.QtCore import Qt, QDate, pyqtSignal, QTimer, QRect, QPoint, QPropertyAnimation, QObject, Q_ARG, QSignalBlocker
from PyQt5.QtGui import QColor, QKeySequence, QClipboard, QFont, QPalette, QIcon
from PyQt5.uic import loadUi

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import stat
import subprocess  # 用于启动更新程序
import shutil  # 用于文件操作
import tempfile  # 用于创建临时目录
import threading  # 用于后台下载
import time

# 导入帮助对话框模块
from help_dialog import HelpDialog

# ==================== 软件版本配置 ====================
# 【重要】每次发布新版本时，必须修改这里的版本号！
# 版本号格式：主版本.次版本.修订号
CURRENT_VERSION = "2.3.1"

# GitHub仓库配置
# 【重要】请修改为你的GitHub用户名和仓库名
GITHUB_OWNER = "244727418"  # 你的GitHub用户名
GITHUB_REPO = "shouhou"  # 你的仓库名

# GitHub Releases API地址
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/releases/latest"

# 是否启用自动更新检查
ENABLE_AUTO_UPDATE = True

# 更新检查间隔（秒），启动后延迟多久检查更新
UPDATE_CHECK_DELAY = 3

def get_resource_path(relative_path):
    """获取资源文件的绝对路径，支持打包后的exe文件"""
    # 方法1：首先尝试PyInstaller的临时目录
    try:
        # PyInstaller创建临时文件夹，将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
        full_path = os.path.join(base_path, relative_path)
        if os.path.exists(full_path):
            return full_path
    except Exception:
        pass
    
    # 方法2：尝试当前目录
    base_path = os.path.abspath(".")
    full_path = os.path.join(base_path, relative_path)
    if os.path.exists(full_path):
        return full_path
    
    # 方法3：尝试exe文件所在目录（打包后）
    if getattr(sys, 'frozen', False):
        # 如果是打包后的exe文件
        base_path = os.path.dirname(sys.executable)
        full_path = os.path.join(base_path, relative_path)
        if os.path.exists(full_path):
            return full_path
    
    # 方法4：如果以上都失败，返回相对路径（让PyQt5尝试处理）
    return relative_path


def get_colormap_colors(colormap, count):
    """生成均匀分布的颜色序列，避免额外的直接数值库依赖。"""
    if count <= 0:
        return []
    if count == 1:
        return [colormap(0.5)]
    return [colormap(index / (count - 1)) for index in range(count)]





# 自定义多选下拉框组件（基于复选框）
class MultiSelectComboBox(QWidget):
    itemsChanged = pyqtSignal()  # 定义信号
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected_items = set()
        self.items = []
        self.init_ui()

    def init_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # 下拉按钮
        self.dropdown_btn = QPushButton("选择退款原因 ▼")
        self.dropdown_btn.setMinimumHeight(32)
        self.dropdown_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.dropdown_btn.setStyleSheet("""
            QPushButton { 
                border: 1px solid #ccc; 
                border-radius: 8px; 
                padding: 6px 10px; 
                text-align: left; 
                background-color: white;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #f0f0f0;
            }
        """)
        self.dropdown_btn.clicked.connect(self.toggle_dropdown)
        layout.addWidget(self.dropdown_btn)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        # 下拉窗口
        self.dropdown_widget = QWidget()
        self.dropdown_widget.setWindowFlags(Qt.Popup)
        self.dropdown_widget.setFixedSize(300, 200)
        self.dropdown_widget.setStyleSheet("""
            QWidget {
                border: 1px solid #ccc;
                border-radius: 3px;
                background-color: white;
            }
        """)
        # 当下拉窗口失去焦点时自动关闭并触发刷新
        self.dropdown_widget.focusOutEvent = self.dropdown_focus_out
        
        dropdown_layout = QVBoxLayout(self.dropdown_widget)
        
        # 搜索框
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("搜索退款原因...")
        self.search_edit.textChanged.connect(self.filter_items)
        dropdown_layout.addWidget(self.search_edit)
        
        # 全选/清空按钮
        button_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("全选")
        self.select_all_btn.clicked.connect(self.select_all)
        self.clear_btn = QPushButton("清空")
        self.clear_btn.clicked.connect(self.clear_selection)
        button_layout.addWidget(self.select_all_btn)
        button_layout.addWidget(self.clear_btn)
        dropdown_layout.addLayout(button_layout)
        
        # 选项列表（使用QListWidget + 复选框）
        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QListWidget.NoSelection)  # 禁用选择，使用复选框
        dropdown_layout.addWidget(self.list_widget)
        
        self.dropdown_widget.hide()

    def addItems(self, items):
        """添加选项"""
        self.items = items
        self.update_list_widget()

    def update_list_widget(self):
        """更新列表控件"""
        self.list_widget.clear()
        
        for item in self.items:
            list_item = QListWidgetItem(item)
            list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
            list_item.setCheckState(Qt.Unchecked)
            self.list_widget.addItem(list_item)
        
        # 连接复选框状态变化信号
        self.list_widget.itemChanged.connect(self.on_item_changed)

    def filter_items(self, text):
        """过滤选项"""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def checkedItems(self):
        """获取选中的项目"""
        checked = []
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item.checkState() == Qt.Checked:
                checked.append(item.text())
        return checked

    def clearChecked(self):
        """清空选择"""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setCheckState(Qt.Unchecked)

    def select_all(self):
        """全选"""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if not item.isHidden():
                item.setCheckState(Qt.Checked)
        
        # 同步更新 selected_items 属性
        self.selected_items = set(self.checkedItems())

    def clear_selection(self):
        """清空选择"""
        self.clearChecked()
        
        # 同步更新 selected_items 属性
        self.selected_items = set()

    def on_item_changed(self, item):
        """复选框状态变化处理"""
        # 同步更新 selected_items 属性
        self.selected_items = set(self.checkedItems())
        
        # 立即触发变化信号，实现实时刷新
        self.itemsChanged.emit()
        
        # 更新按钮显示
        self.update_display()

    def toggle_dropdown(self):
        """切换下拉列表显示"""
        if self.dropdown_widget.isVisible():
            self.dropdown_widget.hide()
            # 当下拉窗口关闭时，触发变化信号（确保实时刷新）
            self.itemsChanged.emit()
        else:
            # 显示在下拉按钮下方
            pos = self.dropdown_btn.mapToGlobal(QPoint(0, self.dropdown_btn.height()))
            self.dropdown_widget.move(pos)
            self.dropdown_widget.show()
            self.search_edit.setFocus()

    def update_display(self):
        """更新按钮显示"""
        selected = self.checkedItems()
        if selected:
            # 显示已选项数量
            if len(selected) == 1:
                self.dropdown_btn.setText(f"{selected[0]} ▼")
            else:
                self.dropdown_btn.setText(f"已选{len(selected)}项 ▼")
        else:
            self.dropdown_btn.setText("选择退款原因 ▼")
        
        # 触发变化信号
        self.itemsChanged.emit()

    def setMaximumWidth(self, width):
        """设置最大宽度"""
        self.dropdown_btn.setMaximumWidth(width)
        QWidget.setMaximumWidth(self, width)
    
    def dropdown_focus_out(self, event):
        """当下拉窗口失去焦点时关闭并触发刷新"""
        # 调用父类的焦点失去事件处理
        QWidget.focusOutEvent(self.dropdown_widget, event)
        
        # 延迟关闭下拉窗口，避免立即关闭导致的问题
        QTimer.singleShot(100, self.close_dropdown_and_refresh)
    
    def close_dropdown_and_refresh(self):
        """关闭下拉窗口并触发刷新"""
        if self.dropdown_widget.isVisible():
            self.dropdown_widget.hide()
            # 触发变化信号，确保实时刷新
            self.itemsChanged.emit()


# ---------------------------- 添加店铺对话框 --------------------------------
class AddStoreDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        loadUi(get_resource_path("dialog_add_store.ui"), self)
        self.setup_connections()

    def setup_connections(self):
        """设置信号连接"""
        self.add_btn.clicked.connect(self.add_store)
        self.cancel_btn.clicked.connect(self.reject)

    def add_store(self):
        """添加店铺"""
        store_name = self.store_name_edit.text().strip()
        if not store_name:
            QMessageBox.warning(self, "输入错误", "请输入店铺名称")
            return
        
        # 这里可以添加保存到数据库的逻辑
        self.accept()

    def get_store_name(self):
        """获取店铺名称"""
        return self.store_name_edit.text().strip()

# ---------------------------- 店铺基本信息设置对话框 --------------------------------
class StoreWeeklyHistoryDialog(QDialog):
    """店铺周数据历史记录对话框。"""
    def __init__(self, db, store_id, store_name, parent=None):
        super().__init__(parent)
        self.db = db
        self.store_id = store_id
        self.store_name = store_name
        self.selected_record = None

        self.setWindowTitle(f"历史记录 - {store_name}")
        self.resize(760, 420)

        layout = QVBoxLayout(self)
        title_label = QLabel(f"店铺：{store_name}")
        title_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 14px; font-weight: bold;")
        layout.addWidget(title_label)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["周范围", "上周单量", "上周销售额", "退款预算", "保存时间"])
        for col in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(col)
            if header_item:
                header_item.setTextAlignment(Qt.AlignCenter)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.doubleClicked.connect(self.apply_selected_record)
        layout.addWidget(self.table)

        button_layout = QHBoxLayout()
        self.apply_btn = QPushButton("应用到设置窗口")
        self.delete_btn = QPushButton("删除记录")
        self.close_btn = QPushButton("关闭")
        self.apply_btn.clicked.connect(self.apply_selected_record)
        self.delete_btn.clicked.connect(self.delete_selected_record)
        self.close_btn.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(self.apply_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addWidget(self.close_btn)
        layout.addLayout(button_layout)

        self.load_history()

    @staticmethod
    def _format_week_range_for_display(start_date_text, end_date_text):
        start_date = QDate.fromString(str(start_date_text or ""), "yyyy-MM-dd")
        end_date = QDate.fromString(str(end_date_text or ""), "yyyy-MM-dd")
        if start_date.isValid() and end_date.isValid():
            return f"{start_date.month()}月{start_date.day()}日 至 {end_date.month()}月{end_date.day()}日"
        return f"{start_date_text} 至 {end_date_text}"

    def load_history(self):
        """加载当前店铺的周数据历史。"""
        self.history_records = self.db.get_store_weekly_settings_history(self.store_id)
        self.table.setRowCount(len(self.history_records))

        for row, record in enumerate(self.history_records):
            week_range = self._format_week_range_for_display(record['week_start_date'], record['week_end_date'])
            values = [
                week_range,
                str(record['weekly_orders']),
                f"{record['weekly_sales']:.2f}",
                f"{record['refund_budget']:.2f}",
                record.get('updated_at') or record.get('created_at') or "",
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setData(Qt.UserRole, record['id'])
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, col, item)

        if self.history_records:
            self.table.selectRow(0)

    def _get_selected_record(self):
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.information(self, "提示", "请先选择一条历史记录")
            return None

        row = selected_rows[0].row()
        if row < 0 or row >= len(self.history_records):
            QMessageBox.information(self, "提示", "选择的历史记录无效")
            return None

        return self.history_records[row]

    def apply_selected_record(self):
        """确认选中的记录，并交由店铺设置窗口回填。"""
        record = self._get_selected_record()
        if not record:
            return

        self.selected_record = record
        self.accept()

    def delete_selected_record(self):
        """删除选中的历史记录。"""
        record = self._get_selected_record()
        if not record:
            return

        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定删除 {record['week_start_date']} 至 {record['week_end_date']} 的历史记录吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        deleted = self.db.delete_store_weekly_settings_history(record['id'], self.store_id)
        if deleted:
            QMessageBox.information(self, "删除完成", "历史记录已删除")
            self.load_history()
        else:
            QMessageBox.warning(self, "删除失败", "未能删除选中的历史记录")


class StoreSpecOrdersDialog(QDialog):
    """识别并保存店铺每周规格单量。"""
    def __init__(self, db, store_id, store_name, week_start_date, week_end_date, parent=None):
        super().__init__(parent)
        self.db = db
        self.store_id = store_id
        self.store_name = store_name
        self.week_start_date = week_start_date
        self.week_end_date = week_end_date
        self.parsed_items = []

        self.setWindowTitle(f"识别规格单量 - {store_name}")
        self.resize(680, 560)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        title_label = QLabel(f"店铺：{store_name}    周期：{week_start_date} 至 {week_end_date}")
        title_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 14px; font-weight: bold;")
        layout.addWidget(title_label)

        tip_label = QLabel("粘贴格式：左边规格文字，右边单数；规格编码只取左侧开头数字，无法识别数字编码的行会忽略。")
        tip_label.setStyleSheet("color: #6c757d; font-size: 12px;")
        layout.addWidget(tip_label)

        self.input_edit = QTextEdit()
        self.input_edit.setPlaceholderText("例如：\n605\t719\n705\t330\n805礼盒\t69")
        self.input_edit.setMinimumHeight(160)
        layout.addWidget(self.input_edit)

        action_layout = QHBoxLayout()
        self.parse_btn = QPushButton("识别预览")
        self.save_btn = QPushButton("保存识别结果")
        self.history_btn = QPushButton("历史记录")
        self.close_btn = QPushButton("关闭")
        StoreSettingsDialog._apply_history_button_style(self.parse_btn)
        StoreSettingsDialog._apply_history_button_style(self.save_btn)
        StoreSettingsDialog._apply_history_button_style(self.history_btn)
        self.close_btn.setMinimumSize(80, 35)
        self.parse_btn.clicked.connect(self.parse_input)
        self.save_btn.clicked.connect(self.save_result)
        self.history_btn.clicked.connect(self.open_history_records)
        self.close_btn.clicked.connect(self.reject)
        action_layout.addStretch()
        action_layout.addWidget(self.parse_btn)
        action_layout.addWidget(self.save_btn)
        action_layout.addWidget(self.history_btn)
        action_layout.addWidget(self.close_btn)
        layout.addLayout(action_layout)

        self.summary_label = QLabel("尚未识别")
        self.summary_label.setStyleSheet("color: #374151; font-size: 12px;")
        layout.addWidget(self.summary_label)

        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["规格编码", "单数"])
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table, 1)

        existing_items = self.db.get_store_weekly_spec_orders_by_week(store_id, week_start_date)
        if existing_items:
            self.input_edit.setPlainText("\n".join(
                f"{item['spec_code']}\t{item['order_count']}" for item in existing_items
            ))
            self.parse_input(show_message=False)

    @staticmethod
    def parse_spec_orders_text(text):
        merged = {}
        invalid_lines = []

        for line_no, raw_line in enumerate((text or "").splitlines(), start=1):
            line = raw_line.strip()
            if not line:
                continue

            match = re.match(r"^(.+?)\s+(\d+)\s*$", line)
            if not match:
                invalid_lines.append((line_no, raw_line))
                continue

            spec_text = match.group(1).strip()
            order_count = int(match.group(2))
            spec_match = re.match(r"^(\d+)", spec_text)
            if not spec_match:
                invalid_lines.append((line_no, raw_line))
                continue

            spec_code = spec_match.group(1)
            merged[spec_code] = merged.get(spec_code, 0) + order_count

        items = [
            {"spec_code": spec_code, "order_count": merged[spec_code]}
            for spec_code in sorted(merged.keys(), key=lambda value: (-merged[value], value))
        ]
        return items, invalid_lines

    def parse_input(self, show_message=True):
        items, invalid_lines = self.parse_spec_orders_text(self.input_edit.toPlainText())
        self.parsed_items = items
        self.table.setRowCount(len(items))

        for row, item in enumerate(items):
            spec_item = QTableWidgetItem(item["spec_code"])
            count_item = QTableWidgetItem(str(item["order_count"]))
            spec_item.setTextAlignment(Qt.AlignCenter)
            count_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 0, spec_item)
            self.table.setItem(row, 1, count_item)

        total_orders = sum(item["order_count"] for item in items)
        invalid_text = f"，未识别 {len(invalid_lines)} 行" if invalid_lines else ""
        self.summary_label.setText(f"识别到 {len(items)} 个规格，合计 {total_orders} 单{invalid_text}")

        if invalid_lines and show_message:
            preview = "\n".join(f"第{line_no}行：{raw_line}" for line_no, raw_line in invalid_lines[:8])
            if len(invalid_lines) > 8:
                preview += f"\n... 还有 {len(invalid_lines) - 8} 行"
            QMessageBox.warning(self, "存在未识别行", f"以下行未按“规格编码 单数”格式识别：\n{preview}")

        if not items and show_message:
            QMessageBox.information(self, "提示", "没有识别到有效的规格单量数据")

        return items, invalid_lines

    def save_result(self):
        items, invalid_lines = self.parse_input(show_message=False)
        if not items:
            QMessageBox.information(self, "提示", "没有可保存的规格单量数据")
            return

        if invalid_lines:
            reply = QMessageBox.question(
                self,
                "存在未识别行",
                f"有 {len(invalid_lines)} 行未识别，是否只保存已识别的 {len(items)} 个规格？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

        existing_items = self.db.get_store_weekly_spec_orders_by_week(self.store_id, self.week_start_date)
        if existing_items:
            reply = QMessageBox.question(
                self,
                "覆盖规格单量",
                f"{self.store_name} 已存在 {self.week_start_date} 至 {self.week_end_date} 的规格单量记录，是否覆盖？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            if reply != QMessageBox.Yes:
                return

        saved_count = self.db.save_store_weekly_spec_orders(
            self.store_id,
            self.store_name,
            self.week_start_date,
            self.week_end_date,
            items
        )
        QMessageBox.information(
            self,
            "保存完成",
            f"已保存 {len(items)} 个规格，合计 {sum(item['order_count'] for item in items)} 单"
        )
        self.accept()

    def open_history_records(self):
        """打开规格单量历史记录窗口。"""
        dialog = StoreSpecOrdersHistoryDialog(self.db, self.store_id, self.store_name, self)
        if dialog.exec_() == QDialog.Accepted and dialog.selected_items:
            self.apply_history_items(dialog.selected_summary, dialog.selected_items)

    def apply_history_items(self, summary, items):
        """把历史规格单量回填到识别窗口。"""
        self.input_edit.setPlainText("\n".join(
            f"{item['spec_code']}\t{item['order_count']}" for item in items
        ))
        week_start = summary.get('week_start_date', '') if summary else ''
        week_end = summary.get('week_end_date', '') if summary else ''
        if week_start and week_end:
            self.week_start_date = week_start
            self.week_end_date = week_end
            self.setWindowTitle(f"识别规格单量 - {self.store_name}")
        self.parse_input(show_message=False)
        QMessageBox.information(self, "应用完成", "历史规格单量已填入识别窗口，可确认后重新保存")


class StoreSpecOrdersHistoryDialog(QDialog):
    """按自然周查看店铺规格单量历史。"""
    def __init__(self, db, store_id, store_name, parent=None):
        super().__init__(parent)
        self.db = db
        self.store_id = store_id
        self.store_name = store_name
        self.week_summaries = []
        self.current_items = []
        self.selected_summary = None
        self.selected_items = []

        self.setWindowTitle(f"规格单量历史记录 - {store_name}")
        self.resize(760, 560)

        layout = QVBoxLayout(self)
        title_label = QLabel(f"店铺：{store_name}")
        title_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 14px; font-weight: bold;")
        layout.addWidget(title_label)

        self.week_table = QTableWidget(0, 4)
        self.week_table.setHorizontalHeaderLabels(["周范围", "规格数", "总单数", "保存时间"])
        self.week_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.week_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.week_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.week_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.week_table.verticalHeader().setVisible(False)
        self.week_table.currentCellChanged.connect(self.on_week_changed)
        self.week_table.doubleClicked.connect(self.apply_selected_week)
        layout.addWidget(self.week_table)

        detail_label = QLabel("规格明细")
        detail_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px; font-weight: bold;")
        layout.addWidget(detail_label)

        self.detail_table = QTableWidget(0, 2)
        self.detail_table.setHorizontalHeaderLabels(["规格编码", "单数"])
        self.detail_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.detail_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.detail_table.verticalHeader().setVisible(False)
        layout.addWidget(self.detail_table, 1)

        button_layout = QHBoxLayout()
        self.apply_btn = QPushButton("应用到识别窗口")
        self.delete_btn = QPushButton("删除本周记录")
        self.close_btn = QPushButton("关闭")
        self.apply_btn.clicked.connect(self.apply_selected_week)
        self.delete_btn.clicked.connect(self.delete_selected_week)
        self.close_btn.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(self.apply_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addWidget(self.close_btn)
        layout.addLayout(button_layout)

        self.load_history()

    @staticmethod
    def _format_week_range_for_display(start_date_text, end_date_text):
        return StoreWeeklyHistoryDialog._format_week_range_for_display(start_date_text, end_date_text)

    def load_history(self):
        """加载按周汇总的规格单量历史。"""
        self.week_summaries = self.db.get_store_weekly_spec_orders_history_summary(self.store_id)
        self.week_table.setRowCount(len(self.week_summaries))

        for row, summary in enumerate(self.week_summaries):
            week_range = self._format_week_range_for_display(
                summary.get('week_start_date', ''),
                summary.get('week_end_date', '')
            )
            values = [
                week_range,
                str(summary.get('spec_count', 0)),
                str(summary.get('total_orders', 0)),
                summary.get('updated_at') or summary.get('created_at') or "",
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                self.week_table.setItem(row, col, item)

        if self.week_summaries:
            self.week_table.selectRow(0)
            self.load_week_detail(0)
        else:
            self.detail_table.setRowCount(0)

    def on_week_changed(self, current_row, current_column, previous_row, previous_column):
        self.load_week_detail(current_row)

    def load_week_detail(self, row):
        """加载选中周的规格明细。"""
        if row < 0 or row >= len(self.week_summaries):
            self.current_items = []
            self.detail_table.setRowCount(0)
            return

        summary = self.week_summaries[row]
        self.current_items = self.db.get_store_weekly_spec_orders_by_week(
            self.store_id,
            summary.get('week_start_date', '')
        )
        self.detail_table.setRowCount(len(self.current_items))
        for detail_row, item in enumerate(self.current_items):
            spec_item = QTableWidgetItem(str(item.get('spec_code', '')))
            count_item = QTableWidgetItem(str(item.get('order_count', 0)))
            spec_item.setTextAlignment(Qt.AlignCenter)
            count_item.setTextAlignment(Qt.AlignCenter)
            self.detail_table.setItem(detail_row, 0, spec_item)
            self.detail_table.setItem(detail_row, 1, count_item)

    def _get_selected_summary(self):
        selected_rows = self.week_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.information(self, "提示", "请先选择一个周期")
            return None

        row = selected_rows[0].row()
        if row < 0 or row >= len(self.week_summaries):
            QMessageBox.information(self, "提示", "选择的周期无效")
            return None

        return self.week_summaries[row]

    def apply_selected_week(self):
        """选择一周历史记录并回填到识别窗口。"""
        summary = self._get_selected_summary()
        if not summary:
            return

        items = self.db.get_store_weekly_spec_orders_by_week(
            self.store_id,
            summary.get('week_start_date', '')
        )
        if not items:
            QMessageBox.information(self, "提示", "当前周期没有规格明细")
            return

        self.selected_summary = summary
        self.selected_items = items
        self.accept()

    def delete_selected_week(self):
        """删除选中自然周的规格单量历史。"""
        summary = self._get_selected_summary()
        if not summary:
            return

        week_start = summary.get('week_start_date', '')
        week_end = summary.get('week_end_date', '')
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定删除 {week_start} 至 {week_end} 的规格单量历史记录吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        deleted = self.db.delete_store_weekly_spec_orders_by_week(self.store_id, week_start)
        if deleted:
            QMessageBox.information(self, "删除完成", "规格单量历史记录已删除")
            self.load_history()
        else:
            QMessageBox.warning(self, "删除失败", "未能删除选中的规格单量历史记录")


class StoreSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        loadUi(get_resource_path("dialog_store_settings.ui"), self)
        self.resize(760, max(self.height(), 430))
        self._setup_week_history_controls()
        self.setup_connections()

    def setup_connections(self):
        """设置信号连接"""
        self.save_btn.clicked.connect(self.save_settings)
        self.cancel_btn.clicked.connect(self.reject)
        self.refund_budget_amount_edit.textChanged.connect(self.on_amount_changed)
        self.refund_budget_percent_edit.textChanged.connect(self.on_percent_changed)
        self.history_week_start_edit.dateChanged.connect(self.on_history_week_changed)
        self.save_history_btn.clicked.connect(self.save_current_data_to_history)
        self.history_records_btn.clicked.connect(self.open_history_records)
        self.spec_orders_btn.clicked.connect(self.open_spec_orders_recognizer)
        if hasattr(self.parent, 'search_store_combo') and self.parent.search_store_combo:
            self.parent.search_store_combo.currentTextChanged.connect(self._update_current_store_scope_label)
        if hasattr(self.parent, 'store_combo') and self.parent.store_combo:
            self.parent.store_combo.currentTextChanged.connect(self._update_current_store_scope_label)

    def _setup_week_history_controls(self):
        """在UI文件加载后动态添加周数据历史控件。"""
        self.history_week_start_edit = QDateEdit()
        self.history_week_start_edit.setCalendarPopup(True)
        self.history_week_start_edit.setDisplayFormat("yyyy-MM-dd")
        self.history_week_start_edit.setMinimumDate(QDate(2000, 1, 1))
        self.history_week_start_edit.setMaximumDate(QDate(2100, 12, 31))
        self.history_week_end_label = QLabel()
        self.history_week_tip_label = QLabel("选择该周任意日期，保存时按自然周一至周日记录")
        self.history_week_tip_label.setStyleSheet("color: #6c757d; font-size: 12px;")

        week_layout = QHBoxLayout()
        week_layout.setSpacing(8)
        week_layout.addWidget(self.history_week_start_edit)
        week_layout.addWidget(self.history_week_end_label)
        week_layout.addWidget(self.history_week_tip_label)

        form_layout = self.findChild(QFormLayout, "formLayout")
        if form_layout:
            form_layout.addRow("数据所属周：", week_layout)
            self.current_store_scope_label = QLabel()
            self.current_store_scope_label.setWordWrap(True)
            self.current_store_scope_label.setStyleSheet("color: #1F2937; font-size: 12px; font-weight: bold;")
            form_layout.addRow("当前范围：", self.current_store_scope_label)

        self.save_history_btn = QPushButton("保存到历史记录")
        self.history_records_btn = QPushButton("历史记录")
        self.spec_orders_btn = QPushButton("识别规格单量")
        self._apply_history_button_style(self.save_history_btn)
        self._apply_history_button_style(self.history_records_btn)
        self._apply_history_button_style(self.spec_orders_btn)

        button_layout = self.findChild(QHBoxLayout, "horizontalLayout_buttons")
        if button_layout:
            insert_index = max(0, button_layout.count() - 2)
            button_layout.insertWidget(insert_index, self.save_history_btn)
            button_layout.insertWidget(insert_index + 1, self.history_records_btn)
            button_layout.insertWidget(insert_index + 2, self.spec_orders_btn)

        self.history_week_start_edit.setDate(self._default_previous_week_start())
        self.on_history_week_changed(self.history_week_start_edit.date())
        self._update_history_buttons_state()
        self._update_current_store_scope_label()

    @staticmethod
    def _default_previous_week_start():
        today = QDate.currentDate()
        current_week_monday = today.addDays(1 - today.dayOfWeek())
        return current_week_monday.addDays(-7)

    @staticmethod
    def _normalize_week_start(date):
        return date.addDays(1 - date.dayOfWeek())

    @staticmethod
    def _apply_history_button_style(button):
        button.setCursor(Qt.PointingHandCursor)
        button.setMinimumSize(108, 35)
        button.setStyleSheet("""
            QPushButton {
                font-family: 'Microsoft YaHei';
                font-size: 13px;
                font-weight: bold;
                color: white;
                background-color: #2563EB;
                border: 1px solid #1D4ED8;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #1D4ED8;
                border-color: #1E40AF;
            }
            QPushButton:pressed {
                background-color: #1E40AF;
                border-color: #1E3A8A;
                padding-top: 7px;
                padding-bottom: 5px;
            }
            QPushButton:disabled {
                color: #F3F4F6;
                background-color: #9CA3AF;
                border-color: #6B7280;
            }
        """)

    def on_history_week_changed(self, date):
        week_start = self._normalize_week_start(date)
        if week_start != date:
            blocker = QSignalBlocker(self.history_week_start_edit)
            self.history_week_start_edit.setDate(week_start)
            del blocker
        week_end = week_start.addDays(6)
        self.history_week_end_label.setText(f"至 {week_end.toString('yyyy-MM-dd')}")

    def _get_current_store_id_name(self):
        search_store = ""
        if hasattr(self.parent, 'search_store_combo') and self.parent.search_store_combo:
            search_store = self.parent.search_store_combo.currentText().strip()

        if search_store:
            if search_store == "全部":
                return None, search_store, "all"
            store_id = self.parent.db.get_store_id_by_name(search_store)
            if store_id:
                return store_id, search_store, "search"
            return None, search_store, "search"

        input_store = ""
        if hasattr(self.parent, 'store_combo') and self.parent.store_combo:
            input_store = self.parent.store_combo.currentText().strip()

        invalid_names = ("", "全部", "请先添加店铺")
        if input_store not in invalid_names:
            store_id = self.parent.db.get_store_id_by_name(input_store)
            if store_id:
                return store_id, input_store, "input"

        return None, input_store, None

    def _get_current_store_scope_text(self):
        store_id, store_name, store_source = self._get_current_store_id_name()
        if store_id and store_source == "search":
            return f"搜索筛选区店铺 - {store_name}"
        if store_id and store_source == "input":
            return f"信息录入区店铺 - {store_name}"
        if store_source == "all":
            return "全部店铺（请先在搜索筛选区选择具体店铺后再保存/识别规格单量）"
        display_name = store_name or "未选择店铺"
        return f"{display_name}（请先选择具体店铺后再保存/识别规格单量）"

    def _update_current_store_scope_label(self, *args):
        if hasattr(self, 'current_store_scope_label') and self.current_store_scope_label:
            self.current_store_scope_label.setText(self._get_current_store_scope_text())

    def _update_history_buttons_state(self):
        self.save_history_btn.setEnabled(True)
        self.history_records_btn.setEnabled(True)
        self.spec_orders_btn.setEnabled(True)
        self.save_history_btn.setToolTip("保存当前填写的周单量、销售额和退款预算到历史记录")
        self.history_records_btn.setToolTip("查看历史周数据，并可回填到当前设置窗口")
        self.spec_orders_btn.setToolTip("粘贴识别每个规格编码本周卖出多少单，并保存到历史记录")

    def _read_form_settings(self):
        """读取当前窗口录入值，沿用保存设置的数字校验规则。"""
        weekly_orders = int(self.daily_orders_edit.text()) if self.daily_orders_edit.text() else 0
        weekly_sales = float(self.daily_sales_edit.text()) if self.daily_sales_edit.text() else 0.0

        if self.refund_budget_amount_edit.text():
            refund_budget = float(self.refund_budget_amount_edit.text())
        elif self.refund_budget_percent_edit.text():
            percent = float(self.refund_budget_percent_edit.text())
            refund_budget = (percent / 100) * self._weekly_to_daily_avg(weekly_sales)
        else:
            refund_budget = 0.0

        return weekly_orders, weekly_sales, refund_budget

    def save_current_data_to_history(self):
        """把当前窗口填写的周数据保存为店铺历史记录。"""
        store_id, store_name, store_source = self._get_current_store_id_name()
        if not store_id:
            QMessageBox.information(self, "提示", "当前范围是全部店铺，请先在搜索筛选区选择具体店铺，再保存历史记录")
            self._update_current_store_scope_label()
            return

        try:
            weekly_orders, weekly_sales, refund_budget = self._read_form_settings()
        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的数字")
            return

        week_start = self._normalize_week_start(self.history_week_start_edit.date())
        week_end = week_start.addDays(6)
        week_start_text = week_start.toString("yyyy-MM-dd")
        week_end_text = week_end.toString("yyyy-MM-dd")

        existing = self.parent.db.get_store_weekly_settings_history_by_week(store_id, week_start_text)
        if existing:
            reply = QMessageBox.question(
                self,
                "覆盖历史记录",
                f"{store_name} 已存在 {week_start_text} 至 {week_end_text} 的历史记录，是否覆盖？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            if reply != QMessageBox.Yes:
                return

        self.parent.db.save_store_weekly_settings_history(
            store_id,
            store_name,
            week_start_text,
            week_end_text,
            weekly_orders,
            weekly_sales,
            refund_budget
        )
        source_text = "信息录入区" if store_source == "input" else "搜索筛选区"
        QMessageBox.information(
            self,
            "保存完成",
            f"已按{source_text}店铺【{store_name}】保存 {week_start_text} 至 {week_end_text} 的历史记录"
        )

    def open_history_records(self):
        """打开店铺周数据历史记录窗口。"""
        store_id, store_name, store_source = self._get_current_store_id_name()
        if not store_id:
            QMessageBox.information(self, "提示", "当前范围是全部店铺，请先在搜索筛选区选择具体店铺，再查看历史记录")
            self._update_current_store_scope_label()
            return

        dialog = StoreWeeklyHistoryDialog(self.parent.db, store_id, store_name, self)
        if dialog.exec_() == QDialog.Accepted and dialog.selected_record:
            self.apply_history_record(dialog.selected_record)

    def open_spec_orders_recognizer(self):
        """打开规格单量识别窗口。"""
        store_id, store_name, store_source = self._get_current_store_id_name()
        if not store_id:
            QMessageBox.information(self, "提示", "当前范围是全部店铺，请先在搜索筛选区选择具体店铺，再识别规格单量")
            self._update_current_store_scope_label()
            return

        week_start = self._normalize_week_start(self.history_week_start_edit.date())
        week_end = week_start.addDays(6)
        dialog = StoreSpecOrdersDialog(
            self.parent.db,
            store_id,
            store_name,
            week_start.toString("yyyy-MM-dd"),
            week_end.toString("yyyy-MM-dd"),
            self
        )
        dialog.exec_()

    def apply_history_record(self, record):
        """把历史记录回填到当前设置窗口，不直接保存到数据库。"""
        self.daily_orders_edit.setText(str(record.get('weekly_orders', 0)))
        self.daily_sales_edit.setText(str(record.get('weekly_sales', 0.0)))
        self.refund_budget_amount_edit.setText(str(record.get('refund_budget', 0.0)))
        daily_avg_sales = self._weekly_to_daily_avg(record.get('weekly_sales', 0.0))
        if daily_avg_sales > 0:
            percent = (self._safe_float(record.get('refund_budget', 0.0)) / daily_avg_sales) * 100
            self.refund_budget_percent_edit.setText(f"{percent:.2f}")
        else:
            self.refund_budget_percent_edit.clear()
        week_start = QDate.fromString(record.get('week_start_date', ''), "yyyy-MM-dd")
        if week_start.isValid():
            self.history_week_start_edit.setDate(week_start)
        QMessageBox.information(self, "应用完成", "历史记录已填入设置窗口，请确认后点击“保存设置”生效")

    @staticmethod
    def _weekly_to_daily_avg(value):
        """将用户录入的7天总值换算为日均值。"""
        return value / 7 if value else 0.0

    @staticmethod
    def _safe_float(value):
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def on_amount_changed(self, text):
        """金额输入变化时自动计算百分比"""
        if text and self.daily_sales_edit.text():
            try:
                amount = float(text)
                sales = self._weekly_to_daily_avg(float(self.daily_sales_edit.text()))
                if sales > 0:
                    percent = (amount / sales) * 100
                    # 临时断开信号避免循环
                    self.refund_budget_percent_edit.textChanged.disconnect(self.on_percent_changed)
                    self.refund_budget_percent_edit.setText(f"{percent:.2f}")
                    self.refund_budget_percent_edit.textChanged.connect(self.on_percent_changed)
            except ValueError:
                pass

    def on_percent_changed(self, text):
        """百分比输入变化时自动计算金额"""
        if text and self.daily_sales_edit.text():
            try:
                percent = float(text)
                sales = self._weekly_to_daily_avg(float(self.daily_sales_edit.text()))
                amount = (percent / 100) * sales
                # 临时断开信号避免循环
                self.refund_budget_amount_edit.textChanged.disconnect(self.on_amount_changed)
                self.refund_budget_amount_edit.setText(f"{amount:.2f}")
                self.refund_budget_amount_edit.textChanged.connect(self.on_amount_changed)
            except ValueError:
                pass

    def save_settings(self):
        """保存设置"""
        try:
            # 保持存储字段不变，但录入含义改为7天总值
            daily_orders, daily_sales, refund_budget = self._read_form_settings()
            
            # 保存到主窗口
            self.parent.store_settings = {
                'daily_orders': daily_orders,
                'daily_sales': daily_sales,
                'refund_budget': refund_budget
            }
            
            # 保存到数据库
            current_store = self.parent.search_store_combo.currentText()
            
            if current_store and current_store != "全部":
                # 保存到当前店铺
                stores = self.parent.db.get_stores()
                store_id = None
                for sid, sname in stores:
                    if sname == current_store:
                        store_id = sid
                        break
                
                if store_id:
                    # 更新店铺设置到数据库
                    self.parent.db.update_store_settings(store_id, daily_orders, daily_sales, refund_budget)
            else:
                # 选择"全部"店铺时，保存到全局设置
                self.parent.db.save_global_settings(daily_orders, daily_sales, refund_budget)
            
            # 更新显示
            self.parent.update_store_stats_display()
            
            self.accept()
            
        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的数字")

    def load_settings(self, settings):
        """加载现有设置"""
        # 优先从数据库加载设置
        current_store = self.parent.search_store_combo.currentText()
        if current_store and current_store != "全部":
            # 获取店铺ID
            stores = self.parent.db.get_stores()
            store_id = None
            for sid, sname in stores:
                if sname == current_store:
                    store_id = sid
                    break
            
            if store_id:
                # 从数据库加载设置
                db_settings = self.parent.db.get_store_settings(store_id)
                if db_settings:
                    settings = db_settings
        
        if settings:
            self.daily_orders_edit.setText(str(settings.get('daily_orders', 0)))
            self.daily_sales_edit.setText(str(settings.get('daily_sales', 0.0)))
            
            refund_budget = settings.get('refund_budget', 0.0)
            self.refund_budget_amount_edit.setText(str(refund_budget))
            
            # 按周销售额录入、按日均销售额计算退款预算百分比
            daily_avg_sales = self._weekly_to_daily_avg(settings.get('daily_sales', 0.0))
            if daily_avg_sales > 0:
                percent = (refund_budget / daily_avg_sales) * 100
                self.refund_budget_percent_edit.setText(f"{percent:.2f}")


class ColumnMappingDialog(QDialog):
    """Excel列映射确认对话框。"""
    def __init__(self, headers, column_configs, initial_mapping, required_fields, parent=None):
        super().__init__(parent)
        self.headers = headers
        self.column_configs = column_configs
        self.required_fields = required_fields
        self.combos = {}

        self.setWindowTitle("确认导入列映射")
        self.resize(620, 520)

        layout = QVBoxLayout(self)

        tip_label = QLabel(
            "已自动识别 Excel 表头，请确认或手动修改列映射。\n"
            "必填字段不能为空，且同一列不能重复映射到多个字段。"
        )
        tip_label.setWordWrap(True)
        layout.addWidget(tip_label)

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignRight)
        options = ["-"] + headers
        required_set = set(required_fields)

        for config in column_configs:
            target_name = config['target']
            combo = QComboBox()
            combo.addItems(options)
            combo.setMinimumWidth(260)

            initial_header = initial_mapping.get(target_name, "")
            if initial_header in headers:
                combo.setCurrentText(initial_header)
            else:
                combo.setCurrentText("-")

            self.combos[target_name] = combo
            label_text = f"{target_name} *" if target_name in required_set else target_name
            form_layout.addRow(label_text, combo)

        layout.addLayout(form_layout)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def validate_and_accept(self):
        mapping = self.get_mapping()

        missing_fields = [field for field in self.required_fields if field not in mapping]
        if missing_fields:
            QMessageBox.warning(self, "缺少必填列", f"请为以下字段选择对应列：{', '.join(missing_fields)}")
            return

        selected_headers = list(mapping.values())
        duplicate_headers = sorted({header for header in selected_headers if selected_headers.count(header) > 1})
        if duplicate_headers:
            QMessageBox.warning(
                self,
                "列映射重复",
                f"以下 Excel 列被重复映射，请调整后再继续：{', '.join(duplicate_headers)}"
            )
            return

        self.accept()

    def get_mapping(self):
        mapping = {}
        for target_name, combo in self.combos.items():
            selected = combo.currentText().strip()
            if selected and selected != "-":
                mapping[target_name] = selected
        return mapping


# ---------------------------- 气泡提示组件 --------------------------------
class BubbleMessage(QWidget):
    def __init__(self, message, parent=None):
        super().__init__(parent)
        self.message = message
        self.parent = parent
        self.init_ui()
        self.setup_animation()
    
    def init_ui(self):
        """初始化气泡界面"""
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.ToolTip)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setStyleSheet("""
            QWidget {
                background-color: rgba(52, 152, 219, 0.9);
                border-radius: 15px;
                padding: 12px 18px;
                color: white;
                font-size: 14px;
                font-weight: bold;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        self.label = QLabel(self.message)
        self.label.setStyleSheet("color: white; font-size: 14px; font-weight: bold;")
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)
        
        # 设置气泡大小
        self.setFixedSize(300, 80)
    
    def setup_animation(self):
        """设置淡入淡出动画"""
        # 淡入动画
        self.fade_in = QPropertyAnimation(self, b"windowOpacity")
        self.fade_in.setDuration(300)  # 300毫秒淡入
        self.fade_in.setStartValue(0.0)
        self.fade_in.setEndValue(1.0)
        
        # 淡出动画
        self.fade_out = QPropertyAnimation(self, b"windowOpacity")
        self.fade_out.setDuration(300)  # 300毫秒淡出
        self.fade_out.setStartValue(1.0)
        self.fade_out.setEndValue(0.0)
        
        # 连接动画
        self.fade_in.finished.connect(self.start_fade_out)
        self.fade_out.finished.connect(self.close)
    
    def start_fade_out(self):
        """开始淡出动画"""
        QTimer.singleShot(1000, self.fade_out.start)  # 显示1秒后开始淡出
    
    def show_bubble(self):
        """显示气泡"""
        # 定位到父窗口中心
        if self.parent:
            parent_rect = self.parent.geometry()
            x = parent_rect.center().x() - self.width() // 2
            y = parent_rect.center().y() - self.height() // 2
            self.move(x, y)
        
        self.show()
        self.fade_in.start()


# ==================== 自动更新模块 ====================

class RejectSelectionDialog(QDialog):
    """
    驳回选择对话框
    用于选择第一轮驳回、第二轮驳回或驳回成功
    """
    def __init__(self, current_round=0, parent=None):
        super().__init__(parent)
        self.current_round = current_round
        self.selected_option = None
        self.init_ui()
    
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("驳回流程选择")
        self.setFixedSize(380, 380)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f6fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # 标题
        title_label = QLabel("请选择驳回流程")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 说明文字
        desc_label = QLabel("驳回后将开始30分钟倒计时，\n时间到后会提醒您继续操作。")
        desc_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        desc_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(desc_label)
        
        layout.addSpacing(20)
        
        # 第一轮驳回按钮
        self.first_round_btn = QPushButton("第一轮驳回")
        self.first_round_btn.setMinimumHeight(18)
        self.first_round_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 12px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.first_round_btn.clicked.connect(lambda: self.select_option("first"))
        layout.addWidget(self.first_round_btn)
        
        # 第二轮驳回按钮
        self.second_round_btn = QPushButton("第二轮驳回")
        self.second_round_btn.setMinimumHeight(18)
        self.second_round_btn.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 12px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d68910;
            }
        """)
        self.second_round_btn.clicked.connect(lambda: self.select_option("second"))
        layout.addWidget(self.second_round_btn)
        
        # 驳回成功按钮
        self.success_btn = QPushButton("驳回成功")
        self.success_btn.setMinimumHeight(18)
        self.success_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 12px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        self.success_btn.clicked.connect(lambda: self.select_option("success"))
        layout.addWidget(self.success_btn)
        
        # 根据当前轮次禁用按钮
        if self.current_round == 0:
            # 还没开始，禁用第二轮和成功
            self.second_round_btn.setEnabled(False)
            self.success_btn.setEnabled(False)
        elif self.current_round == 1:
            # 第一轮进行中，禁用第一轮
            self.first_round_btn.setEnabled(False)
        elif self.current_round == 2:
            # 第二轮进行中，禁用第一轮和第二轮
            self.first_round_btn.setEnabled(False)
            self.second_round_btn.setEnabled(False)
        
        # 取消按钮
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        layout.addWidget(cancel_btn)
        
        self.setLayout(layout)
    
    def select_option(self, option):
        """选择选项"""
        self.selected_option = option
        self.accept()
    
    def get_selected_option(self):
        """获取选择的选项"""
        return self.selected_option


class RejectSkipDialog(QDialog):
    """
    跳过等待确认对话框
    """
    def __init__(self, order_no, current_round, parent=None):
        super().__init__(parent)
        self.order_no = order_no
        self.current_round = current_round
        self.init_ui()
    
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("跳过等待")
        self.setFixedSize(350, 200)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f6fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        title_label = QLabel("跳过等待确认")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #e74c3c;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 说明文字
        round_text = "第一轮" if self.current_round == 1 else "第二轮"
        desc_label = QLabel(f"订单号: {self.order_no}\n当前处于{round_text}驳回等待中\n\n是否跳过等待时间？")
        desc_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        desc_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(desc_label)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        
        skip_btn = QPushButton("跳过等待")
        skip_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        skip_btn.clicked.connect(self.accept)
        
        cancel_btn = QPushButton("继续等待")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(skip_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)


class RejectSuccessDialog(QDialog):
    """
    驳回成功设置对话框
    """
    def __init__(self, order_no, store_name, parent=None):
        super().__init__(parent)
        self.order_no = order_no
        self.store_name = store_name
        self.remind_48h = False
        self.init_ui()
    
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("驳回成功")
        self.setFixedSize(400, 250)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f6fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QCheckBox {
                font-size: 14px;
                color: #2c3e50;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        title_label = QLabel("🎉 驳回成功")
        title_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #27ae60;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 信息
        info_label = QLabel(f"店铺: {self.store_name}\n订单号: {self.order_no}")
        info_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        info_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(info_label)
        
        # 48小时提醒复选框
        self.remind_checkbox = QCheckBox("48小时后提醒我")
        self.remind_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                color: #2c3e50;
                spacing: 8px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
            }
        """)
        self.remind_checkbox.setChecked(True)
        layout.addWidget(self.remind_checkbox)
        
        # 说明
        desc_label = QLabel("勾选后，系统将在48小时后弹出提醒")
        desc_label.setStyleSheet("color: #95a5a6; font-size: 11px;")
        desc_label.setIndent(25)
        layout.addWidget(desc_label)
        
        layout.addSpacing(10)
        
        # 确定按钮
        confirm_btn = QPushButton("确定")
        confirm_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 12px 30px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        confirm_btn.clicked.connect(self.on_confirm)
        layout.addWidget(confirm_btn, alignment=Qt.AlignCenter)
        
        self.setLayout(layout)
    
    def on_confirm(self):
        """确认按钮点击"""
        self.remind_48h = self.remind_checkbox.isChecked()
        self.accept()
    
    def should_remind_48h(self):
        """是否设置48小时提醒"""
        return self.remind_48h


class RejectCountdownFinishedDialog(QDialog):
    """
    驳回倒计时结束对话框
    显示订单信息（可复制），并提供继续下一轮驳回的选项
    """
    def __init__(self, order_no, store_name, round_text, parent=None):
        super().__init__(parent)
        self.order_no = order_no
        self.store_name = store_name
        self.round_text = round_text
        self.init_ui()
    
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("⏰ 驳回时间到")
        self.setFixedSize(420, 300)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f6fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
            }
            QPushButton {
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QLineEdit {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 8px;
                font-size: 13px;
                selection-background-color: #3498db;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        title_label = QLabel(f"⏰ {self.round_text}驳回时间到！")
        title_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #e74c3c;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 店铺信息
        store_label = QLabel(f"店铺: {self.store_name}")
        store_label.setStyleSheet("color: #7f8c8d; font-size: 13px;")
        store_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(store_label)
        
        # 订单号（可复制）
        order_label = QLabel("订单号（可复制）:")
        order_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        layout.addWidget(order_label)
        
        self.order_edit = QLineEdit(self.order_no)
        self.order_edit.setReadOnly(True)
        self.order_edit.setAlignment(Qt.AlignCenter)
        self.order_edit.setStyleSheet("""
            QLineEdit {
                background-color: #ecf0f1;
                color: #2c3e50;
                font-weight: bold;
                padding: 10px;
            }
        """)
        # 自动选中所有文本，方便复制
        self.order_edit.selectAll()
        layout.addWidget(self.order_edit)
        
        layout.addSpacing(10)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        # 稍后处理按钮
        later_btn = QPushButton("稍后处理")
        later_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        later_btn.clicked.connect(self.reject)
        btn_layout.addWidget(later_btn)
        
        btn_layout.addStretch()
        
        # 继续按钮（根据轮次显示不同文字）
        if self.round_text == "第一轮":
            next_btn_text = "开始第二轮驳回"
            next_btn_color = "#3498db"
            next_btn_hover = "#2980b9"
        else:
            next_btn_text = "驳回成功"
            next_btn_color = "#27ae60"
            next_btn_hover = "#229954"
        
        next_btn = QPushButton(next_btn_text)
        next_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {next_btn_color};
                color: white;
            }}
            QPushButton:hover {{
                background-color: {next_btn_hover};
            }}
        """)
        next_btn.clicked.connect(self.accept)
        btn_layout.addWidget(next_btn)
        
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)


class Reminder48hDialog(QDialog):
    """
    48小时提醒对话框
    显示订单信息（可复制）
    """
    def __init__(self, order_no, store_name, parent=None):
        super().__init__(parent)
        self.order_no = order_no
        self.store_name = store_name
        self.init_ui()
    
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("⏰ 48小时提醒")
        self.setFixedSize(400, 280)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f6fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 30px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QLineEdit {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 8px;
                font-size: 13px;
                selection-background-color: #3498db;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        title_label = QLabel("⏰ 48小时提醒")
        title_label.setStyleSheet("font-size: 22px; font-weight: bold; color: #e74c3c;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 说明文字
        desc_label = QLabel("该订单驳回成功已满48小时，请注意跟进！")
        desc_label.setStyleSheet("color: #7f8c8d; font-size: 13px;")
        desc_label.setAlignment(Qt.AlignCenter)
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)
        
        layout.addSpacing(10)
        
        # 店铺信息
        store_layout = QHBoxLayout()
        store_label_title = QLabel("店铺:")
        store_label_title.setStyleSheet("font-weight: bold;")
        store_label_value = QLabel(self.store_name)
        store_label_value.setStyleSheet("color: #2c3e50;")
        store_layout.addWidget(store_label_title)
        store_layout.addWidget(store_label_value)
        store_layout.addStretch()
        layout.addLayout(store_layout)
        
        # 订单号（可复制）
        order_label = QLabel("订单号（可复制）:")
        order_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        layout.addWidget(order_label)
        
        self.order_edit = QLineEdit(self.order_no)
        self.order_edit.setReadOnly(True)
        self.order_edit.setAlignment(Qt.AlignCenter)
        self.order_edit.setStyleSheet("""
            QLineEdit {
                background-color: #ecf0f1;
                color: #2c3e50;
                font-weight: bold;
                padding: 10px;
                font-size: 14px;
            }
        """)
        # 自动选中所有文本，方便复制
        self.order_edit.selectAll()
        layout.addWidget(self.order_edit)
        
        layout.addSpacing(15)
        
        # 确定按钮
        confirm_btn = QPushButton("确定")
        confirm_btn.clicked.connect(self.accept)
        layout.addWidget(confirm_btn, alignment=Qt.AlignCenter)
        
        self.setLayout(layout)


class RejectSuccessActionsDialog(QDialog):
    """
    驳回成功后操作对话框
    提供平台介入退款等选项
    """
    def __init__(self, order_no, store_name, parent=None):
        super().__init__(parent)
        self.order_no = order_no
        self.store_name = store_name
        self.init_ui()
    
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("驳回成功 - 后续操作")
        self.setFixedSize(400, 250)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f6fa;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
            }
            QPushButton {
                border: none;
                border-radius: 5px;
                padding: 12px 25px;
                font-size: 14px;
                font-weight: bold;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        title_label = QLabel("✅ 驳回成功")
        title_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #27ae60;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 信息
        info_label = QLabel(f"店铺: {self.store_name}\n订单号: {self.order_no}")
        info_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        info_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(info_label)
        
        layout.addSpacing(10)
        
        # 说明
        desc_label = QLabel("如果平台介入退款，请点击下方按钮标记")
        desc_label.setStyleSheet("color: #e74c3c; font-size: 12px;")
        desc_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(desc_label)
        
        layout.addSpacing(10)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        close_btn.clicked.connect(self.reject)
        btn_layout.addWidget(close_btn)
        
        btn_layout.addStretch()
        
        # 平台介入退款按钮
        platform_btn = QPushButton("平台介入退款")
        platform_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        platform_btn.clicked.connect(self.accept)
        btn_layout.addWidget(platform_btn)
        
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)


class UpdateDialog(QDialog):
    """
    更新对话框
    显示新版本信息和更新按钮
    """
    def __init__(self, current_version, new_version, release_notes, download_url, parent=None):
        super().__init__(parent)
        self.current_version = current_version
        self.new_version = new_version
        self.release_notes = release_notes
        self.download_url = download_url
        self.init_ui()
    
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("发现新版本")
        self.setFixedSize(500, 400)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f6fa;
            }
            QLabel {
                color: #2c3e50;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
            QTextEdit {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 10px;
                font-size: 12px;
            }
            QProgressBar {
                border: 1px solid #ddd;
                border-radius: 5px;
                text-align: center;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 5px;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        title_label = QLabel("🎉 发现新版本！")
        title_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #e74c3c;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 版本信息
        version_layout = QHBoxLayout()
        current_label = QLabel(f"当前版本: {self.current_version}")
        current_label.setStyleSheet("color: #7f8c8d;")
        new_label = QLabel(f"最新版本: {self.new_version}")
        new_label.setStyleSheet("color: #27ae60; font-weight: bold;")
        version_layout.addWidget(current_label)
        version_layout.addStretch()
        version_layout.addWidget(new_label)
        layout.addLayout(version_layout)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("background-color: #ddd;")
        layout.addWidget(line)
        
        # 更新内容标签
        notes_label = QLabel("更新内容:")
        notes_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(notes_label)
        
        # 更新内容文本框
        self.notes_text = QTextEdit()
        self.notes_text.setReadOnly(True)
        self.notes_text.setText(self.release_notes if self.release_notes else "暂无更新说明")
        self.notes_text.setMinimumHeight(120)
        layout.addWidget(self.notes_text)
        
        # 进度条（初始隐藏）
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # 状态标签
        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet("color: #3498db; font-size: 12px;")
        layout.addWidget(self.status_label)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        
        self.later_btn = QPushButton("稍后更新")
        self.later_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.later_btn.clicked.connect(self.reject)
        
        self.update_btn = QPushButton("立即更新")
        self.update_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        self.update_btn.clicked.connect(self.start_update)
        
        button_layout.addWidget(self.later_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.update_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
    
    def start_update(self):
        """开始更新"""
        self.update_btn.setEnabled(False)
        self.later_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("正在下载新版本...")
        
        # 在后台线程中下载
        self.download_thread = threading.Thread(target=self.download_update)
        self.download_thread.daemon = True
        self.download_thread.start()
    
    def download_update(self):
        """下载更新文件"""
        try:
            # 获取当前exe路径
            current_exe = sys.executable
            
            # 创建临时目录
            temp_dir = tempfile.mkdtemp()
            new_exe_path = os.path.join(temp_dir, "售后登记表_new.exe")
            
            # 下载文件
            response = requests.get(self.download_url, stream=True, timeout=300)
            total_size = int(response.headers.get('content-length', 0))
            
            downloaded = 0
            chunk_size = 8192
            
            with open(new_exe_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=chunk_size):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total_size > 0:
                            progress = int((downloaded / total_size) * 100)
                            # 使用信号更新UI
                            from PyQt5.QtCore import QMetaObject, Qt, Q_ARG
                            QMetaObject.invokeMethod(
                                self.progress_bar,
                                "setValue",
                                Qt.QueuedConnection,
                                Q_ARG(int, progress)
                            )
            
            # 下载完成，创建更新脚本
            self.create_updater_script(current_exe, new_exe_path)
            
            # 更新UI
            from PyQt5.QtCore import QMetaObject, Qt
            QMetaObject.invokeMethod(
                self.status_label,
                "setText",
                Qt.QueuedConnection,
                Q_ARG(str, "下载完成！即将安装更新...")
            )
            
            # 延迟后启动更新脚本
            QTimer.singleShot(1500, self.launch_updater)
            
        except Exception as e:
            from PyQt5.QtCore import QMetaObject, Qt
            QMetaObject.invokeMethod(
                self.status_label,
                "setText",
                Qt.QueuedConnection,
                Q_ARG(str, f"下载失败: {str(e)}")
            )
            QMetaObject.invokeMethod(
                self.update_btn,
                "setEnabled",
                Qt.QueuedConnection,
                Q_ARG(bool, True)
            )
            QMetaObject.invokeMethod(
                self.later_btn,
                "setEnabled",
                Qt.QueuedConnection,
                Q_ARG(bool, True)
            )
    
    def create_updater_script(self, old_exe, new_exe):
        """创建更新脚本"""
        # 更新脚本路径
        updater_path = os.path.join(tempfile.gettempdir(), "update_script.bat")
        
        # 创建批处理脚本
        script_content = f"""@echo off
chcp 65001 >nul
echo 正在安装更新...
timeout /t 2 /nobreak >nul

:: 等待原程序退出
:wait_loop
tasklist | findstr "{os.path.basename(old_exe)}" >nul
if errorlevel 1 goto continue
timeout /t 1 /nobreak >nul
goto wait_loop

:continue
:: 替换文件
copy /Y "{new_exe}" "{old_exe}"
if errorlevel 1 (
    echo 更新失败，请手动替换文件
    pause
    exit /b 1
)

:: 删除临时文件
del "{new_exe}"

:: 启动新版本
echo 更新完成，正在启动新版本...
start "" "{old_exe}"

:: 删除自己
del "%~f0"
"""
        
        with open(updater_path, 'w', encoding='utf-8') as f:
            f.write(script_content)
        
        self.updater_path = updater_path
    
    def launch_updater(self):
        """启动更新程序"""
        try:
            # 启动更新脚本
            subprocess.Popen(
                self.updater_path,
                shell=True,
                creationflags=subprocess.CREATE_NEW_CONSOLE
            )
            
            # 接受对话框并退出程序
            self.accept()
            
            # 退出当前程序
            QApplication.instance().quit()
            
        except Exception as e:
            QMessageBox.critical(self, "更新错误", f"启动更新失败: {str(e)}")
            self.update_btn.setEnabled(True)
            self.later_btn.setEnabled(True)


# ==================== 驳回流程管理类 ====================

class RejectProcessManager(QObject):
    """
    驳回流程管理器
    管理订单的驳回流程：第一轮驳回、第二轮驳回、驳回成功
    """
    # 信号定义
    countdown_updated = pyqtSignal(str, int, str)  # 订单号, 剩余秒数, 当前轮次
    countdown_finished = pyqtSignal(str, str)  # 订单号, 当前轮次
    reminder_48h_triggered = pyqtSignal(str, str)  # 订单号, 店铺名称
    
    def __init__(self, db=None):
        super().__init__()
        # 数据库引用
        self.db = db
        # 存储进行中的驳回流程
        # 格式: {order_no: {'round': 1/2, 'end_time': datetime, 'timer': QTimer, 'store_name': str}}
        self.active_processes = {}
        # 存储48小时提醒
        # 格式: {order_no: {'end_time': datetime, 'timer': QTimer, 'store_name': str}}
        self.reminder_48h = {}
    
    def start_first_round(self, order_no, store_name):
        """开始第一轮驳回"""
        # 如果已有流程，先停止
        if order_no in self.active_processes:
            self.stop_process(order_no)
        
        # 创建倒计时，30分钟 = 1800秒
        end_time = datetime.now() + timedelta(seconds=1800)
        timer = QTimer()
        timer.timeout.connect(lambda: self._update_countdown(order_no))
        timer.start(1000)  # 每秒更新一次
        
        self.active_processes[order_no] = {
            'round': 1,
            'end_time': end_time,
            'timer': timer,
            'store_name': store_name,
            'total_seconds': 1800
        }
        
        # 保存到数据库
        if self.db:
            self.db.save_reject_countdown(order_no, store_name, 1, end_time)
        
        return True
    
    def start_second_round(self, order_no, store_name):
        """开始第二轮驳回"""
        # 停止第一轮
        if order_no in self.active_processes:
            self.stop_process(order_no)
        
        # 创建倒计时，30分钟 = 1800秒
        end_time = datetime.now() + timedelta(seconds=1800)
        timer = QTimer()
        timer.timeout.connect(lambda: self._update_countdown(order_no))
        timer.start(1000)
        
        self.active_processes[order_no] = {
            'round': 2,
            'end_time': end_time,
            'timer': timer,
            'store_name': store_name,
            'total_seconds': 1800
        }
        
        # 保存到数据库
        if self.db:
            self.db.save_reject_countdown(order_no, store_name, 2, end_time)
        
        return True
    
    def skip_wait(self, order_no):
        """跳过当前等待"""
        if order_no in self.active_processes:
            process = self.active_processes[order_no]
            process['end_time'] = datetime.now()  # 将结束时间设为现在
            self._update_countdown(order_no)  # 立即更新一次
            return process['round']
        return None
    
    def stop_process(self, order_no):
        """停止指定订单的驳回流程"""
        if order_no in self.active_processes:
            process = self.active_processes[order_no]
            if process['timer']:
                process['timer'].stop()
            del self.active_processes[order_no]
            
            # 从数据库删除
            if self.db:
                self.db.delete_reject_countdown(order_no)
    
    def _update_countdown(self, order_no):
        """更新倒计时"""
        if order_no not in self.active_processes:
            return
        
        process = self.active_processes[order_no]
        now = datetime.now()
        remaining = (process['end_time'] - now).total_seconds()
        
        if remaining <= 0:
            # 倒计时结束
            round_num = process['round']
            store_name = process['store_name']
            self.stop_process(order_no)
            self.countdown_finished.emit(order_no, f"第{round_num}轮")
        else:
            # 发射更新信号
            self.countdown_updated.emit(order_no, int(remaining), f"第{process['round']}轮")
    
    def get_remaining_time(self, order_no):
        """获取剩余时间（秒）"""
        if order_no not in self.active_processes:
            return None
        
        process = self.active_processes[order_no]
        remaining = (process['end_time'] - datetime.now()).total_seconds()
        return max(0, int(remaining))
    
    def get_process_info(self, order_no):
        """获取流程信息"""
        if order_no not in self.active_processes:
            return None
        
        process = self.active_processes[order_no]
        remaining = self.get_remaining_time(order_no)
        return {
            'round': process['round'],
            'remaining': remaining,
            'store_name': process['store_name']
        }
    
    def set_48h_reminder(self, order_no, store_name):
        """设置48小时提醒"""
        # 如果已有提醒，先停止
        if order_no in self.reminder_48h:
            self.stop_48h_reminder(order_no)
        
        # 创建48小时倒计时
        end_time = datetime.now() + timedelta(hours=48)
        timer = QTimer()
        timer.setSingleShot(True)  # 只执行一次
        timer.timeout.connect(lambda: self._on_48h_reminder(order_no))
        timer.start(48 * 60 * 60 * 1000)  # 48小时 = 172800000毫秒
        
        self.reminder_48h[order_no] = {
            'end_time': end_time,
            'timer': timer,
            'store_name': store_name
        }
        
        print(f"[DEBUG] 订单 {order_no} 48小时提醒已设置，结束时间: {end_time}")
        return True
    
    def stop_48h_reminder(self, order_no):
        """停止48小时提醒"""
        if order_no in self.reminder_48h:
            reminder = self.reminder_48h[order_no]
            if reminder['timer']:
                reminder['timer'].stop()
            del self.reminder_48h[order_no]
            print(f"[DEBUG] 订单 {order_no} 48小时提醒已取消")
    
    def _on_48h_reminder(self, order_no):
        """48小时提醒触发"""
        if order_no in self.reminder_48h:
            reminder = self.reminder_48h[order_no]
            store_name = reminder['store_name']
            del self.reminder_48h[order_no]
            
            # 发射信号
            self.reminder_48h_triggered.emit(order_no, store_name)
            print(f"[DEBUG] 订单 {order_no} 48小时提醒触发")
    
    def get_48h_reminder_remaining(self, order_no):
        """获取48小时提醒的剩余时间（秒）"""
        if order_no not in self.reminder_48h:
            return None
        
        reminder = self.reminder_48h[order_no]
        remaining = (reminder['end_time'] - datetime.now()).total_seconds()
        return max(0, int(remaining))
    
    def has_48h_reminder(self, order_no):
        """检查是否有48小时提醒"""
        return order_no in self.reminder_48h
    
    def get_48h_reminder_info(self, order_no):
        """获取48小时提醒的详细信息"""
        if order_no not in self.reminder_48h:
            return None
        
        reminder = self.reminder_48h[order_no]
        remaining_seconds = (reminder['end_time'] - datetime.now()).total_seconds()
        remaining_seconds = max(0, int(remaining_seconds))
        
        # 计算剩余小时数（向上取整）
        remaining_hours = (remaining_seconds + 3599) // 3600  # 加3599是为了向上取整
        
        return {
            'end_time': reminder['end_time'],
            'remaining_seconds': remaining_seconds,
            'remaining_hours': remaining_hours,
            'store_name': reminder['store_name']
        }
            
    def clear_all(self):
        """清除所有流程"""
        # 停止所有活动流程
        for order_no in list(self.active_processes.keys()):
            self.stop_process(order_no)
        
        # 停止所有48小时提醒
        for order_no in list(self.reminder_48h.keys()):
            self.stop_48h_reminder(order_no)

    def restore_countdowns_from_db(self):
        """从数据库恢复倒计时状态（软件启动时调用）"""
        if not self.db:
            return
        
        # 获取所有活动的倒计时
        active_countdowns = self.db.get_all_active_reject_countdowns()
        
        for countdown in active_countdowns:
            order_no = countdown['order_no']
            store_name = countdown['store_name']
            current_round = countdown['current_round']
            end_time = countdown['end_time']
            
            # 计算剩余秒数
            remaining_seconds = (end_time - datetime.now()).total_seconds()
            
            if remaining_seconds <= 0:
                # 已经过期，删除记录
                self.db.delete_reject_countdown(order_no)
                continue
            
            # 创建倒计时
            timer = QTimer()
            timer.timeout.connect(lambda on=order_no: self._update_countdown(on))
            timer.start(1000)
            
            self.active_processes[order_no] = {
                'round': current_round,
                'end_time': end_time,
                'timer': timer,
                'store_name': store_name,
                'total_seconds': int(remaining_seconds)
            }
            
            # 发射信号更新UI
            self.countdown_updated.emit(order_no, int(remaining_seconds), f"第{current_round}轮")


class UpdateChecker(QObject):
    """
    更新检查器
    负责检查GitHub上的最新版本
    """
    update_available = pyqtSignal(dict)  # 发现更新时发射信号
    check_finished = pyqtSignal()  # 检查完成时发射信号
    
    def __init__(self):
        super().__init__()
        self.latest_version = None
        self.download_url = None
        self.release_notes = None
    
    def check_for_updates(self):
        """检查更新（在后台线程中运行）"""
        thread = threading.Thread(target=self._check_update_thread)
        thread.daemon = True
        thread.start()
    
    def _check_update_thread(self):
        """后台检查更新的线程"""
        try:
            # 发送GitHub API请求
            headers = {
                'Accept': 'application/vnd.github.v3+json',
                'User-Agent': 'RefundManager-UpdateChecker'
            }
            
            response = requests.get(GITHUB_API_URL, headers=headers, timeout=10)
            
            if response.status_code == 200:
                release_data = response.json()
                
                # 获取最新版本号（去掉v前缀）
                latest_version = release_data.get('tag_name', '').lstrip('v')
                
                if not latest_version:
                    self.check_finished.emit()
                    return
                
                # 比较版本号
                if self._compare_versions(latest_version, CURRENT_VERSION) > 0:
                    # 有新版本
                    self.latest_version = latest_version
                    self.release_notes = release_data.get('body', '暂无更新说明')
                    
                    # 查找exe文件的下载链接
                    assets = release_data.get('assets', [])
                    self.download_url = None
                    
                    for asset in assets:
                        name = asset.get('name', '')
                        if name.endswith('.exe') and '售后登记表' in name:
                            self.download_url = asset.get('browser_download_url')
                            break
                    
                    # 如果没找到特定文件，使用第一个exe文件
                    if not self.download_url:
                        for asset in assets:
                            if asset.get('name', '').endswith('.exe'):
                                self.download_url = asset.get('browser_download_url')
                                break
                    
                    if self.download_url:
                        from PyQt5.QtCore import QMetaObject, Qt
                        QMetaObject.invokeMethod(
                            self,
                            "_emit_update_available",
                            Qt.QueuedConnection
                        )
                    else:
                        self.check_finished.emit()
                else:
                    self.check_finished.emit()
            else:
                self.check_finished.emit()
                
        except Exception as e:
            print(f"检查更新时出错: {e}")
            self.check_finished.emit()
    
    def _emit_update_available(self):
        """发射更新可用信号（在主线程中调用）"""
        self.update_available.emit({
            'version': self.latest_version,
            'notes': self.release_notes,
            'url': self.download_url
        })
    
    def _compare_versions(self, version1, version2):
        """
        比较两个版本号
        返回: 1表示v1>v2, 0表示相等, -1表示v1<v2
        """
        try:
            v1_parts = [int(x) for x in version1.split('.')]
            v2_parts = [int(x) for x in version2.split('.')]
            
            # 补齐版本号位数
            while len(v1_parts) < len(v2_parts):
                v1_parts.append(0)
            while len(v2_parts) < len(v1_parts):
                v2_parts.append(0)
            
            for i in range(len(v1_parts)):
                if v1_parts[i] > v2_parts[i]:
                    return 1
                elif v1_parts[i] < v2_parts[i]:
                    return -1
            
            return 0
        except:
            return 0


# ---------------------------- 自定义表格委托类 --------------------------------
class CustomItemDelegate(QItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent

    def createEditor(self, parent, option, index):
        """创建编辑器时检查编辑权限"""
        # 补偿金额列（第6列）检查打款补偿状态
        if index.column() == 6:  # 补偿金额列
            # 获取当前行的记录信息
            row = index.row()
            record_id = self.parent.get_record_id_from_row(row)
            if record_id:
                record = self.parent.db.get_record_by_id(record_id)
                if record and not record['compensate']:
                    # 如果没有勾选打款补偿，不允许编辑
                    return None
        
        # 其他列正常创建编辑器
        return super().createEditor(parent, option, index)
    
    def setEditorData(self, editor, index):
        """设置编辑器数据，在编辑时保持选中状态"""
        # 在开始编辑时，确保当前行保持选中状态
        if self.parent and hasattr(self.parent, 'table'):
            # 获取当前行
            row = index.row()
            # 确保该行被选中
            self.parent.table.setCurrentCell(row, index.column())
        
        # 调用父类方法设置编辑器数据
        super().setEditorData(editor, index)
    
    def setModelData(self, editor, model, index):
        """设置模型数据，在编辑完成后保持选中状态"""
        # 调用父类方法设置模型数据
        super().setModelData(editor, model, index)
        
        # 在编辑完成后，确保当前行保持选中状态
        if self.parent and hasattr(self.parent, 'table'):
            # 获取当前行
            row = index.row()
            # 确保该行被选中
            self.parent.table.setCurrentCell(row, index.column())

# ---------------------------- 数据库操作类 ---------------------------------
SYNC_TABLES = {
    'stores': {
        'key': 'id',
        'columns': ['store_name', 'color', 'estimated_orders', 'daily_orders', 'daily_sales', 'refund_budget'],
        'unique': ['store_name'],
    },
    'refund_records': {
        'key': 'id',
        'columns': [
            'store_id', 'order_no', 'spec_name', 'spec_code', 'reason', 'real_refund_reason',
            'real_refund_reason_detail', 'real_refund_reason_updated_at', 'real_refund_reason_note_hash',
            'quality_refund_reason', 'quality_not_cancelled_reason', 'quality_refund_reason_detail',
            'quality_refund_reason_updated_at', 'quality_refund_reason_note_hash', 'refund_amount',
            'cancel', 'compensate', 'comp_amount', 'reject', 'reject_result', 'notes',
            'order_status', 'after_sale_status', 'refund_apply_time', 'refund_agree_time', 'record_date'
        ],
        'store_ref': 'store_id',
        'unique': ['store_id', 'order_no'],
    },
    'global_settings': {
        'key': 'id',
        'columns': ['setting_key', 'setting_value'],
        'unique': ['setting_key'],
    },
    'window_settings': {
        'key': 'id',
        'columns': ['setting_key', 'setting_value'],
        'unique': ['setting_key'],
    },
    'ai_summary_history': {
        'key': 'id',
        'columns': ['filter_summary', 'snapshot_json', 'created_at'],
    },
    'real_refund_reason_categories': {
        'key': 'id',
        'columns': ['category_name', 'keywords_text', 'status', 'sort_order', 'created_at', 'updated_at'],
        'unique': ['category_name'],
    },
    'quality_not_cancelled_reason_categories': {
        'key': 'id',
        'columns': ['category_name', 'status', 'sort_order', 'created_at', 'updated_at'],
        'unique': ['category_name'],
    },
    'reject_countdown': {
        'key': 'id',
        'columns': ['order_no', 'store_name', 'current_round', 'end_time', 'created_at'],
        'unique': ['order_no'],
    },
    'store_weekly_settings_history': {
        'key': 'id',
        'columns': [
            'store_id', 'store_name_snapshot', 'week_start_date', 'week_end_date',
            'weekly_orders', 'weekly_sales', 'refund_budget', 'created_at', 'updated_at'
        ],
        'store_ref': 'store_id',
        'unique': ['store_id', 'week_start_date'],
    },
    'store_weekly_spec_orders': {
        'key': 'id',
        'columns': [
            'store_id', 'store_name_snapshot', 'week_start_date', 'week_end_date',
            'spec_code', 'order_count', 'created_at', 'updated_at'
        ],
        'store_ref': 'store_id',
        'unique': ['store_id', 'week_start_date', 'spec_code'],
    },
}


def utc_now_text():
    return datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%fZ')


class CloudSyncError(Exception):
    pass


class CloudSyncService:
    MANIFEST_NAME = 'manifest.json'
    REFUND_RECORDS_INDEX_NAME = 'index/refund_records.json.gz'
    REFUND_RECORDS_SNAPSHOT_NAME = 'snapshot/refund_records.json.gz'

    def __init__(self, db, config):
        self.db = db
        self.config = config or {}
        self.prefix = self._normalize_prefix(self.config.get('prefix') or 'shouhou-sync/')
        self.client = self._create_client()

    @staticmethod
    def _normalize_prefix(prefix):
        prefix = str(prefix or 'shouhou-sync/').strip().strip('/')
        return f"{prefix}/" if prefix else ''

    def _create_client(self):
        try:
            from qcloud_cos import CosConfig, CosS3Client
        except Exception as exc:
            raise CloudSyncError("缺少腾讯云 COS SDK，请先安装 cos-python-sdk-v5。") from exc

        for field in ('secret_id', 'secret_key', 'bucket', 'region'):
            if not str(self.config.get(field) or '').strip():
                raise CloudSyncError(f"请先填写 {field}")

        cos_config = CosConfig(
            Region=self.config['region'].strip(),
            SecretId=self.config['secret_id'].strip(),
            SecretKey=self.config['secret_key'].strip(),
            Scheme='https',
        )
        return CosS3Client(cos_config)

    def _key(self, name):
        return f"{self.prefix}{name}"

    def test_connection(self):
        self.client.head_bucket(Bucket=self.config['bucket'].strip())
        return True

    def _load_manifest(self):
        try:
            response = self.client.get_object(
                Bucket=self.config['bucket'].strip(),
                Key=self._key(self.MANIFEST_NAME),
            )
            data = response['Body'].get_raw_stream().read()
            return json.loads(data.decode('utf-8'))
        except Exception as exc:
            status_code = getattr(exc, 'get_status_code', lambda: '')()
            error_code = getattr(exc, 'get_error_code', lambda: '')()
            if str(status_code) == '404' or error_code in ('NoSuchKey', 'NoSuchResource'):
                return {'version': 1, 'changes': [], 'updated_at': None}
            raise

    def _save_manifest(self, manifest):
        manifest['version'] = 1
        manifest['updated_at'] = utc_now_text()
        body = json.dumps(manifest, ensure_ascii=False, indent=2).encode('utf-8')
        self.client.put_object(Bucket=self.config['bucket'].strip(), Key=self._key(self.MANIFEST_NAME), Body=body)

    def _load_refund_records_index(self):
        try:
            response = self.client.get_object(
                Bucket=self.config['bucket'].strip(),
                Key=self._key(self.REFUND_RECORDS_INDEX_NAME),
            )
            compressed = response['Body'].get_raw_stream().read()
            return json.loads(gzip.decompress(compressed).decode('utf-8'))
        except Exception as exc:
            status_code = getattr(exc, 'get_status_code', lambda: '')()
            error_code = getattr(exc, 'get_error_code', lambda: '')()
            if str(status_code) == '404' or error_code in ('NoSuchKey', 'NoSuchResource'):
                return None
            raise

    def _load_refund_records_snapshot(self):
        try:
            response = self.client.get_object(
                Bucket=self.config['bucket'].strip(),
                Key=self._key(self.REFUND_RECORDS_SNAPSHOT_NAME),
            )
            compressed = response['Body'].get_raw_stream().read()
            return json.loads(gzip.decompress(compressed).decode('utf-8'))
        except Exception as exc:
            status_code = getattr(exc, 'get_status_code', lambda: '')()
            error_code = getattr(exc, 'get_error_code', lambda: '')()
            if str(status_code) == '404' or error_code in ('NoSuchKey', 'NoSuchResource'):
                return None
            raise

    def _save_refund_records_index(self):
        index = self.db.export_refund_records_cloud_index()
        body = gzip.compress(json.dumps(index, ensure_ascii=False).encode('utf-8'))
        self.client.put_object(
            Bucket=self.config['bucket'].strip(),
            Key=self._key(self.REFUND_RECORDS_INDEX_NAME),
            Body=body,
        )
        return index

    def _save_refund_records_snapshot(self):
        snapshot = self.db.export_refund_records_cloud_snapshot()
        body = gzip.compress(json.dumps(snapshot, ensure_ascii=False).encode('utf-8'))
        self.client.put_object(
            Bucket=self.config['bucket'].strip(),
            Key=self._key(self.REFUND_RECORDS_SNAPSHOT_NAME),
            Body=body,
        )
        return snapshot

    def upload_incremental(self):
        self.db.conn.commit()
        dedupe_count = self.db.dedupe_refund_records_by_store_order()
        config = self.db.load_cloud_sync_config()
        since = config.get('last_upload_at') or ''
        package = self.db.export_sync_changes(since)
        row_count = sum(len(items) for items in package['tables'].values())
        change_name = None

        manifest = self._load_manifest()
        if row_count > 0:
            timestamp = utc_now_text().replace(':', '').replace('.', '')
            change_name = f"changes/{timestamp}-{package['device_id']}.json.gz"
            body = gzip.compress(json.dumps(package, ensure_ascii=False).encode('utf-8'))
            self.client.put_object(Bucket=self.config['bucket'].strip(), Key=self._key(change_name), Body=body)

            changes = manifest.setdefault('changes', [])
            changes.append({
                'key': change_name,
                'device_id': package['device_id'],
                'created_at': package['created_at'],
                'row_count': row_count,
            })
            changes.sort(key=lambda item: item.get('created_at') or '')
            manifest['latest_change_at'] = package['created_at']

        index = self._save_refund_records_index()
        snapshot = self._save_refund_records_snapshot()
        manifest['refund_records_index_at'] = index.get('created_at')
        manifest['refund_records_snapshot_at'] = snapshot.get('created_at')
        manifest['refund_records_index_count'] = len(index.get('records', []))
        manifest['refund_records_snapshot_count'] = len(snapshot.get('records', []))
        self._save_manifest(manifest)
        if row_count > 0:
            self.db.update_cloud_sync_state(last_upload_at=package['created_at'])
        return {
            'uploaded': row_count,
            'key': change_name,
            'deduped': dedupe_count,
            'index_count': len(index.get('records', [])),
            'snapshot_count': len(snapshot.get('records', [])),
        }

    def download_incremental(self):
        manifest = self._load_manifest()
        applied = set(self.db.get_applied_cloud_change_keys())
        current_device_id = self.db.get_cloud_device_id()
        downloaded = 0
        applied_count = 0
        latest_time = None

        for change in sorted(manifest.get('changes', []), key=lambda item: item.get('created_at') or ''):
            key = change.get('key')
            if not key or key in applied:
                continue
            if change.get('device_id') == current_device_id:
                self.db.mark_cloud_change_applied(key, change.get('created_at') or utc_now_text(), 0)
                continue

            response = self.client.get_object(Bucket=self.config['bucket'].strip(), Key=self._key(key))
            compressed = response['Body'].get_raw_stream().read()
            package = json.loads(gzip.decompress(compressed).decode('utf-8'))
            applied_rows = self.db.apply_sync_package(package)
            self.db.mark_cloud_change_applied(key, change.get('created_at') or package.get('created_at') or utc_now_text(), applied_rows)
            downloaded += applied_rows
            applied_count += 1
            latest_time = change.get('created_at') or latest_time

        if latest_time:
            self.db.update_cloud_sync_state(last_download_at=latest_time)
        cloud_index = self._load_refund_records_index()
        cloud_snapshot = self._load_refund_records_snapshot()
        snapshot_restored = self.db.apply_refund_records_cloud_snapshot(cloud_snapshot) if cloud_snapshot else 0
        cloud_missing_deleted = (
            self.db.delete_refund_records_missing_from_cloud_index(cloud_index)
            if cloud_index and cloud_snapshot
            else 0
        )
        dedupe_count = self.db.dedupe_refund_records_by_store_order(cloud_index)
        return {
            'downloaded': downloaded,
            'packages': applied_count,
            'snapshot_restored': snapshot_restored,
            'cloud_missing_deleted': cloud_missing_deleted,
            'deduped': dedupe_count,
            'index_found': cloud_index is not None,
            'snapshot_found': cloud_snapshot is not None,
        }


class Database:
    def __init__(self, db_file='refund_data.db'):
        # 使用用户本地的数据文件，不打包进exe
        self.db_file = db_file
        self.conn = None
        self._ensure_database_file_writable()
        self.init_db()

    def _ensure_database_file_writable(self):
        """确保数据库文件和所在目录可写，避免启动迁移时报只读数据库。"""
        db_path = os.path.abspath(self.db_file)
        db_dir = os.path.dirname(db_path) or os.getcwd()

        if not os.access(db_dir, os.W_OK):
            raise RuntimeError(f"数据库目录不可写：{db_dir}")

        if not os.path.exists(db_path):
            return

        mode = os.stat(db_path).st_mode
        if mode & stat.S_IWRITE:
            return

        try:
            os.chmod(db_path, mode | stat.S_IWRITE)
        except OSError as e:
            raise RuntimeError(f"数据库文件为只读且无法自动解除：{db_path}") from e

    def init_db(self):
        """初始化数据库，创建表"""
        self.conn = sqlite3.connect(self.db_file)
        # 启用外键约束（SQLite默认不启用）
        self.conn.execute("PRAGMA foreign_keys = ON")
        cursor = self.conn.cursor()
        
        # 检查表是否存在，如果存在则添加缺失的列
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='stores'")
        if cursor.fetchone():
            # 表已存在，检查并添加缺失的列
            self._add_missing_columns()
        else:
            # 表不存在，创建新表
            cursor.execute('''
                CREATE TABLE stores (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    store_name TEXT UNIQUE NOT NULL,
                    color TEXT DEFAULT NULL,
                    estimated_orders INTEGER DEFAULT 0,
                    daily_orders INTEGER DEFAULT 0,
                    daily_sales REAL DEFAULT 0.0,
                    refund_budget REAL DEFAULT 0.0
                )
            ''')
        
        # 性能优化：自动修复缺失的表（确保exe运行时表一定存在）
        self._auto_fix_missing_tables()
        
        # 创建 refund_records 表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS refund_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                store_id INTEGER NOT NULL,
                order_no TEXT NOT NULL,
                spec_name TEXT DEFAULT '',
                spec_code TEXT DEFAULT '',
                reason TEXT NOT NULL,
                real_refund_reason TEXT DEFAULT '',
                real_refund_reason_detail TEXT DEFAULT '',
                real_refund_reason_updated_at TEXT DEFAULT '',
                real_refund_reason_note_hash TEXT DEFAULT '',
                quality_refund_reason TEXT DEFAULT '',
                quality_not_cancelled_reason TEXT DEFAULT '',
                quality_refund_reason_detail TEXT DEFAULT '',
                quality_refund_reason_updated_at TEXT DEFAULT '',
                quality_refund_reason_note_hash TEXT DEFAULT '',
                refund_amount REAL NOT NULL,
                cancel INTEGER DEFAULT 0,
                compensate INTEGER DEFAULT 0,
                comp_amount REAL DEFAULT 0,
                reject INTEGER DEFAULT 0,  -- 是否驳回：0=否，1=是
                reject_result TEXT DEFAULT '',  -- 驳回结果：成功、失败
                notes TEXT DEFAULT '',  -- 备注信息
                order_status TEXT DEFAULT '',
                after_sale_status TEXT DEFAULT '',
                refund_apply_time TEXT DEFAULT '',
                refund_agree_time TEXT DEFAULT '',
                record_date TEXT DEFAULT '',
                FOREIGN KEY (store_id) REFERENCES stores (id) ON DELETE CASCADE
            )
        ''')
        # 添加索引
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_order_no ON refund_records (order_no)')
        
        # 创建 API 配置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS api_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                api_url TEXT DEFAULT 'https://api.deepseek.com/v1/chat/completions',
                api_key TEXT DEFAULT '',
                model TEXT DEFAULT 'deepseek-chat',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_record_date ON refund_records (record_date)')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ai_summary_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filter_summary TEXT DEFAULT '',
                snapshot_json TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        self._create_store_weekly_settings_history_table(cursor)
        self._create_store_weekly_spec_orders_table(cursor)

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS real_refund_reason_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_name TEXT UNIQUE NOT NULL,
                keywords_text TEXT DEFAULT '',
                status TEXT DEFAULT 'ACTIVE',
                sort_order INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS quality_not_cancelled_reason_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_name TEXT UNIQUE NOT NULL,
                status TEXT DEFAULT 'ACTIVE',
                sort_order INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 创建驳回倒计时状态表（用于软件重启后恢复倒计时）
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS reject_countdown (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT UNIQUE NOT NULL,
                store_name TEXT NOT NULL,
                current_round INTEGER NOT NULL,  -- 1=第一轮, 2=第二轮
                end_time TEXT NOT NULL,  -- 倒计时结束时间 ISO格式
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_reject_countdown_order ON reject_countdown (order_no)')
        
        self.conn.commit()
        self._init_sync_metadata()
        self.cleanup_empty_date_records()

    def cleanup_empty_date_records(self):
        """清理登记日期为空的隐藏订单，避免界面不显示但导入仍判重。"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                DELETE FROM refund_records
                WHERE IFNULL(TRIM(record_date), '') = ''
            """)
            deleted_count = cursor.rowcount
            self.conn.commit()

            if deleted_count > 0:
                print(f"✅ 自动清理空登记日期记录 {deleted_count} 条")

            return deleted_count
        except Exception as e:
            print(f"自动清理空登记日期记录失败: {e}")
            return 0
    
    def _add_missing_columns(self):
        """添加缺失的列到现有表"""
        cursor = self.conn.cursor()
        
        # 检查daily_orders列是否存在
        cursor.execute("PRAGMA table_info(stores)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'daily_orders' not in columns:
            cursor.execute("ALTER TABLE stores ADD COLUMN daily_orders INTEGER DEFAULT 0")
        
        if 'daily_sales' not in columns:
            cursor.execute("ALTER TABLE stores ADD COLUMN daily_sales REAL DEFAULT 0.0")
        
        if 'refund_budget' not in columns:
            cursor.execute("ALTER TABLE stores ADD COLUMN refund_budget REAL DEFAULT 0.0")

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='refund_records'")
        if cursor.fetchone():
            cursor.execute("PRAGMA table_info(refund_records)")
            refund_columns = [column[1] for column in cursor.fetchall()]

            if 'order_status' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN order_status TEXT DEFAULT ''")

            if 'after_sale_status' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN after_sale_status TEXT DEFAULT ''")

            if 'spec_name' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN spec_name TEXT DEFAULT ''")

            if 'spec_code' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN spec_code TEXT DEFAULT ''")

            if 'refund_apply_time' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN refund_apply_time TEXT DEFAULT ''")

            if 'refund_agree_time' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN refund_agree_time TEXT DEFAULT ''")

            if 'real_refund_reason' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN real_refund_reason TEXT DEFAULT ''")

            if 'real_refund_reason_detail' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN real_refund_reason_detail TEXT DEFAULT ''")

            if 'real_refund_reason_updated_at' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN real_refund_reason_updated_at TEXT DEFAULT ''")

            if 'real_refund_reason_note_hash' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN real_refund_reason_note_hash TEXT DEFAULT ''")

            if 'quality_refund_reason' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN quality_refund_reason TEXT DEFAULT ''")

            if 'quality_not_cancelled_reason' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN quality_not_cancelled_reason TEXT DEFAULT ''")

            if 'quality_refund_reason_detail' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN quality_refund_reason_detail TEXT DEFAULT ''")

            if 'quality_refund_reason_updated_at' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN quality_refund_reason_updated_at TEXT DEFAULT ''")

            if 'quality_refund_reason_note_hash' not in refund_columns:
                cursor.execute("ALTER TABLE refund_records ADD COLUMN quality_refund_reason_note_hash TEXT DEFAULT ''")

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='real_refund_reason_categories'")
        if cursor.fetchone():
            cursor.execute("PRAGMA table_info(real_refund_reason_categories)")
            category_columns = [column[1] for column in cursor.fetchall()]
            if 'keywords_text' not in category_columns:
                cursor.execute("ALTER TABLE real_refund_reason_categories ADD COLUMN keywords_text TEXT DEFAULT ''")
        
        # 创建全局设置表（用于存储"全部店铺"的设置）
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS global_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setting_key TEXT UNIQUE NOT NULL,
                setting_value TEXT
            )
        ''')
        
        # 创建窗口设置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS window_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setting_key TEXT UNIQUE NOT NULL,
                setting_value TEXT
            )
        ''')

        self._create_store_weekly_settings_history_table(cursor)
        self._create_store_weekly_spec_orders_table(cursor)
        
        self.conn.commit()

    def _auto_fix_missing_tables(self):
        """自动修复缺失的表（确保exe运行时表一定存在）"""
        cursor = self.conn.cursor()
        
        # 检查 global_settings 表是否存在
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='global_settings'")
        if not cursor.fetchone():
            # 创建 global_settings 表
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS global_settings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    setting_key TEXT UNIQUE NOT NULL,
                    setting_value TEXT
                )
            ''')
            print("✅ 自动修复：global_settings 表已创建")
        
        # 检查 window_settings 表是否存在
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='window_settings'")
        if not cursor.fetchone():
            # 创建 window_settings 表
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS window_settings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    setting_key TEXT UNIQUE NOT NULL,
                    setting_value TEXT
                )
            ''')
            print("✅ 自动修复：window_settings 表已创建")

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ai_summary_history'")
        if not cursor.fetchone():
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS ai_summary_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    filter_summary TEXT DEFAULT '',
                    snapshot_json TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            print("✅ 自动修复：ai_summary_history 表已创建")

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='store_weekly_settings_history'")
        if not cursor.fetchone():
            self._create_store_weekly_settings_history_table(cursor)
            print("✅ 自动修复：store_weekly_settings_history 表已创建")

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='store_weekly_spec_orders'")
        if not cursor.fetchone():
            self._create_store_weekly_spec_orders_table(cursor)
            print("✅ 自动修复：store_weekly_spec_orders 表已创建")

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='real_refund_reason_categories'")
        if not cursor.fetchone():
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS real_refund_reason_categories (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    category_name TEXT UNIQUE NOT NULL,
                    keywords_text TEXT DEFAULT '',
                    status TEXT DEFAULT 'ACTIVE',
                    sort_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            print("✅ 自动修复：real_refund_reason_categories 表已创建")

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='quality_not_cancelled_reason_categories'")
        if not cursor.fetchone():
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS quality_not_cancelled_reason_categories (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    category_name TEXT UNIQUE NOT NULL,
                    status TEXT DEFAULT 'ACTIVE',
                    sort_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            print("✅ 自动修复：quality_not_cancelled_reason_categories 表已创建")
        
        # 检查 refund_records 表是否存在
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='refund_records'")
        if not cursor.fetchone():
            # 创建 refund_records 表
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS refund_records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    store_id INTEGER NOT NULL,
                    order_no TEXT NOT NULL,
                    spec_name TEXT DEFAULT '',
                    spec_code TEXT DEFAULT '',
                    reason TEXT NOT NULL,
                    real_refund_reason TEXT DEFAULT '',
                    real_refund_reason_detail TEXT DEFAULT '',
                    real_refund_reason_updated_at TEXT DEFAULT '',
                    real_refund_reason_note_hash TEXT DEFAULT '',
                    quality_refund_reason TEXT DEFAULT '',
                    quality_not_cancelled_reason TEXT DEFAULT '',
                    quality_refund_reason_detail TEXT DEFAULT '',
                    quality_refund_reason_updated_at TEXT DEFAULT '',
                    quality_refund_reason_note_hash TEXT DEFAULT '',
                    refund_amount REAL NOT NULL,
                    cancel INTEGER DEFAULT 0,
                    compensate INTEGER DEFAULT 0,
                    comp_amount REAL DEFAULT 0,
                reject INTEGER DEFAULT 0,
                reject_result TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                order_status TEXT DEFAULT '',
                after_sale_status TEXT DEFAULT '',
                refund_apply_time TEXT DEFAULT '',
                refund_agree_time TEXT DEFAULT '',
                record_date TEXT DEFAULT '',
                FOREIGN KEY (store_id) REFERENCES stores (id) ON DELETE CASCADE
            )
            ''')
            print("✅ 自动修复：refund_records 表已创建")
        
        self.conn.commit()

    def _create_store_weekly_settings_history_table(self, cursor):
        """创建店铺周数据历史记录表。"""
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS store_weekly_settings_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                store_id INTEGER NOT NULL,
                store_name_snapshot TEXT NOT NULL,
                week_start_date TEXT NOT NULL,
                week_end_date TEXT NOT NULL,
                weekly_orders INTEGER DEFAULT 0,
                weekly_sales REAL DEFAULT 0.0,
                refund_budget REAL DEFAULT 0.0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(store_id, week_start_date),
                FOREIGN KEY (store_id) REFERENCES stores (id) ON DELETE CASCADE
            )
        ''')
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_store_weekly_history_store_week
            ON store_weekly_settings_history (store_id, week_start_date DESC)
        ''')

    def _create_store_weekly_spec_orders_table(self, cursor):
        """创建店铺周规格单量历史表。"""
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS store_weekly_spec_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                store_id INTEGER NOT NULL,
                store_name_snapshot TEXT NOT NULL,
                week_start_date TEXT NOT NULL,
                week_end_date TEXT NOT NULL,
                spec_code TEXT NOT NULL,
                order_count INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(store_id, week_start_date, spec_code),
                FOREIGN KEY (store_id) REFERENCES stores (id) ON DELETE CASCADE
            )
        ''')
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_store_weekly_spec_orders_store_week
            ON store_weekly_spec_orders (store_id, week_start_date DESC)
        ''')

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()

    def _init_sync_metadata(self):
        """初始化云同步所需的元数据列、状态表和触发器。"""
        cursor = self.conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cloud_sync_state (
                setting_key TEXT PRIMARY KEY,
                setting_value TEXT
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cloud_sync_applied_changes (
                change_key TEXT PRIMARY KEY,
                change_created_at TEXT DEFAULT '',
                applied_at TEXT DEFAULT '',
                row_count INTEGER DEFAULT 0
            )
        ''')

        for table in SYNC_TABLES:
            if not self._table_exists(cursor, table):
                continue
            self._ensure_sync_columns(cursor, table)
            self._backfill_sync_columns(cursor, table)
            self._create_sync_triggers(cursor, table)

        if not self._get_sync_state(cursor, 'device_id'):
            cursor.execute(
                'INSERT OR REPLACE INTO cloud_sync_state (setting_key, setting_value) VALUES (?, ?)',
                ('device_id', str(uuid.uuid4()))
            )
        self.conn.commit()

    def _table_exists(self, cursor, table):
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
        return cursor.fetchone() is not None

    def _get_table_columns(self, cursor, table):
        cursor.execute(f'PRAGMA table_info({table})')
        return [row[1] for row in cursor.fetchall()]

    def _ensure_sync_columns(self, cursor, table):
        columns = self._get_table_columns(cursor, table)
        if 'sync_id' not in columns:
            cursor.execute(f'ALTER TABLE {table} ADD COLUMN sync_id TEXT')
        if 'updated_at' not in columns:
            cursor.execute(f'ALTER TABLE {table} ADD COLUMN updated_at TEXT')
        if 'deleted_at' not in columns:
            cursor.execute(f'ALTER TABLE {table} ADD COLUMN deleted_at TEXT')
        if 'sync_deleted' not in columns:
            cursor.execute(f'ALTER TABLE {table} ADD COLUMN sync_deleted INTEGER DEFAULT 0')
        cursor.execute(f'CREATE UNIQUE INDEX IF NOT EXISTS idx_{table}_sync_id ON {table} (sync_id)')
        cursor.execute(f'CREATE INDEX IF NOT EXISTS idx_{table}_sync_updated_at ON {table} (updated_at)')

    def _backfill_sync_columns(self, cursor, table):
        key = SYNC_TABLES[table]['key']
        rows = cursor.execute(f"SELECT {key} FROM {table} WHERE sync_id IS NULL OR sync_id = ''").fetchall()
        now = utc_now_text()
        for (row_id,) in rows:
            cursor.execute(
                f"UPDATE {table} SET sync_id=?, updated_at=COALESCE(NULLIF(updated_at, ''), ?), sync_deleted=COALESCE(sync_deleted, 0) WHERE {key}=?",
                (str(uuid.uuid4()), now, row_id)
            )
        cursor.execute(f"UPDATE {table} SET updated_at=? WHERE updated_at IS NULL OR updated_at = ''", (now,))
        cursor.execute(f"UPDATE {table} SET sync_deleted=0 WHERE sync_deleted IS NULL")

    def _create_sync_triggers(self, cursor, table):
        key = SYNC_TABLES[table]['key']
        cursor.execute(f'''
            CREATE TRIGGER IF NOT EXISTS trg_{table}_sync_insert
            AFTER INSERT ON {table}
            FOR EACH ROW
            WHEN NEW.sync_id IS NULL OR NEW.sync_id = '' OR NEW.updated_at IS NULL OR NEW.updated_at = ''
            BEGIN
                UPDATE {table}
                SET sync_id = CASE WHEN NEW.sync_id IS NULL OR NEW.sync_id = '' THEN lower(hex(randomblob(16))) ELSE NEW.sync_id END,
                    updated_at = CASE WHEN NEW.updated_at IS NULL OR NEW.updated_at = '' THEN strftime('%Y-%m-%dT%H:%M:%fZ', 'now') ELSE NEW.updated_at END,
                    sync_deleted = COALESCE(NEW.sync_deleted, 0)
                WHERE {key} = NEW.{key};
            END
        ''')
        cursor.execute(f'''
            CREATE TRIGGER IF NOT EXISTS trg_{table}_sync_update
            AFTER UPDATE ON {table}
            FOR EACH ROW
            WHEN NEW.updated_at = OLD.updated_at
            BEGIN
                UPDATE {table}
                SET updated_at = strftime('%Y-%m-%dT%H:%M:%fZ', 'now')
                WHERE {key} = NEW.{key};
            END
        ''')

    def _get_sync_state(self, cursor, key):
        row = cursor.execute('SELECT setting_value FROM cloud_sync_state WHERE setting_key=?', (key,)).fetchone()
        return row[0] if row else ''

    def get_cloud_device_id(self):
        cursor = self.conn.cursor()
        device_id = self._get_sync_state(cursor, 'device_id')
        if not device_id:
            device_id = str(uuid.uuid4())
            cursor.execute(
                'INSERT OR REPLACE INTO cloud_sync_state (setting_key, setting_value) VALUES (?, ?)',
                ('device_id', device_id)
            )
            self.conn.commit()
        return device_id

    def load_cloud_sync_config(self):
        cursor = self.conn.cursor()
        rows = cursor.execute('SELECT setting_key, setting_value FROM cloud_sync_state').fetchall()
        data = {key: value for key, value in rows}
        return {
            'device_id': data.get('device_id') or self.get_cloud_device_id(),
            'secret_id': data.get('secret_id', ''),
            'secret_key': data.get('secret_key', ''),
            'bucket': data.get('bucket', ''),
            'region': data.get('region', ''),
            'prefix': data.get('prefix', 'shouhou-sync/'),
            'last_upload_at': data.get('last_upload_at', ''),
            'last_download_at': data.get('last_download_at', ''),
        }

    def save_cloud_sync_config(self, config):
        cursor = self.conn.cursor()
        for key in ('secret_id', 'secret_key', 'bucket', 'region', 'prefix'):
            cursor.execute(
                'INSERT OR REPLACE INTO cloud_sync_state (setting_key, setting_value) VALUES (?, ?)',
                (key, str(config.get(key) or '').strip())
            )
        if not self._get_sync_state(cursor, 'device_id'):
            cursor.execute(
                'INSERT OR REPLACE INTO cloud_sync_state (setting_key, setting_value) VALUES (?, ?)',
                ('device_id', str(uuid.uuid4()))
            )
        self.conn.commit()

    def update_cloud_sync_state(self, **kwargs):
        cursor = self.conn.cursor()
        for key, value in kwargs.items():
            cursor.execute(
                'INSERT OR REPLACE INTO cloud_sync_state (setting_key, setting_value) VALUES (?, ?)',
                (key, str(value or ''))
            )
        self.conn.commit()

    def get_applied_cloud_change_keys(self):
        cursor = self.conn.cursor()
        return [row[0] for row in cursor.execute('SELECT change_key FROM cloud_sync_applied_changes').fetchall()]

    def mark_cloud_change_applied(self, change_key, change_created_at, row_count):
        cursor = self.conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO cloud_sync_applied_changes
            (change_key, change_created_at, applied_at, row_count)
            VALUES (?, ?, ?, ?)
        ''', (change_key, change_created_at or '', utc_now_text(), int(row_count or 0)))
        self.conn.commit()

    def export_refund_records_cloud_index(self):
        """导出当前有效订单索引，用于下载时按云端优先清理本地订单。"""
        cursor = self.conn.cursor()
        rows = cursor.execute('''
            SELECT s.sync_id, r.order_no, r.sync_id, r.updated_at
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE IFNULL(r.sync_deleted, 0)=0
              AND IFNULL(s.sync_deleted, 0)=0
              AND IFNULL(TRIM(r.order_no), '') != ''
              AND IFNULL(TRIM(s.sync_id), '') != ''
            ORDER BY s.sync_id ASC, r.order_no ASC
        ''').fetchall()
        return {
            'version': 1,
            'device_id': self.get_cloud_device_id(),
            'created_at': utc_now_text(),
            'records': [
                {
                    'store_sync_id': row[0],
                    'order_no': row[1],
                    'sync_id': row[2],
                    'updated_at': row[3] or '',
                }
                for row in rows
            ],
        }

    def export_refund_records_cloud_snapshot(self):
        """导出当前有效订单快照。下载时即使没有新增量包，也可恢复本地缺失订单。"""
        cursor = self.conn.cursor()
        snapshot = {
            'version': 1,
            'device_id': self.get_cloud_device_id(),
            'created_at': utc_now_text(),
            'stores': [],
            'records': [],
        }

        store_meta = SYNC_TABLES['stores']
        store_columns = [store_meta['key'], 'sync_id', 'updated_at', 'deleted_at', 'sync_deleted'] + store_meta['columns']
        store_columns = [column for column in store_columns if column in set(self._get_table_columns(cursor, 'stores'))]
        store_rows = cursor.execute(f'''
            SELECT {', '.join('s.' + column for column in store_columns)}
            FROM stores s
            WHERE IFNULL(s.sync_deleted, 0)=0
            ORDER BY s.store_name ASC
        ''').fetchall()
        snapshot['stores'] = [dict(zip(store_columns, row)) for row in store_rows]

        record_meta = SYNC_TABLES['refund_records']
        record_columns = [record_meta['key'], 'sync_id', 'updated_at', 'deleted_at', 'sync_deleted'] + record_meta['columns']
        record_columns = [column for column in record_columns if column in set(self._get_table_columns(cursor, 'refund_records'))]
        record_rows = cursor.execute(f'''
            SELECT {', '.join('r.' + column for column in record_columns)}, s.sync_id AS _store_sync_id
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE IFNULL(r.sync_deleted, 0)=0
              AND IFNULL(s.sync_deleted, 0)=0
              AND IFNULL(TRIM(r.order_no), '') != ''
            ORDER BY s.sync_id ASC, r.order_no ASC
        ''').fetchall()
        for row in record_rows:
            item = dict(zip(record_columns + ['_store_sync_id'], row))
            item['sync_deleted'] = 0
            item['deleted_at'] = ''
            snapshot['records'].append(item)
        return snapshot

    def apply_refund_records_cloud_snapshot(self, snapshot):
        """用云端当前订单快照恢复本地缺失或被软删除的订单。"""
        if not snapshot:
            return 0

        cursor = self.conn.cursor()
        applied_records = 0
        cursor.execute('PRAGMA foreign_keys = OFF')
        try:
            for store in snapshot.get('stores', []):
                store = dict(store)
                store['sync_deleted'] = 0
                store['deleted_at'] = ''
                self._apply_sync_row(cursor, 'stores', store)

            for record in snapshot.get('records', []):
                record = dict(record)
                record['sync_deleted'] = 0
                record['deleted_at'] = ''
                if not self._refund_snapshot_record_needs_apply(cursor, record):
                    continue
                if self._apply_sync_row(cursor, 'refund_records', record):
                    applied_records += 1
            self.conn.commit()
        finally:
            cursor.execute('PRAGMA foreign_keys = ON')
        return applied_records

    def _refund_snapshot_record_needs_apply(self, cursor, record):
        sync_id = str(record.get('sync_id') or '').strip()
        store_sync_id = str(record.get('_store_sync_id') or '').strip()
        order_no = str(record.get('order_no') or '').strip()
        incoming_updated_at = str(record.get('updated_at') or '').strip()

        local_row = None
        if sync_id:
            local_row = cursor.execute(
                '''
                SELECT id, sync_id, updated_at, sync_deleted
                FROM refund_records
                WHERE sync_id=?
                ORDER BY IFNULL(sync_deleted, 0) ASC, updated_at DESC, id DESC
                LIMIT 1
                ''',
                (sync_id,)
            ).fetchone()

        if not local_row and store_sync_id and order_no:
            store_row = cursor.execute('SELECT id FROM stores WHERE sync_id=?', (store_sync_id,)).fetchone()
            if store_row:
                local_row = cursor.execute(
                    '''
                    SELECT id, sync_id, updated_at, sync_deleted
                    FROM refund_records
                    WHERE store_id=? AND order_no=?
                    ORDER BY IFNULL(sync_deleted, 0) ASC, updated_at DESC, id DESC
                    LIMIT 1
                    ''',
                    (store_row[0], order_no)
                ).fetchone()

        if not local_row:
            return True
        if int(local_row[3] or 0) != 0:
            return True
        if sync_id and str(local_row[1] or '') != sync_id:
            return True
        if incoming_updated_at and str(local_row[2] or '') != incoming_updated_at:
            return True
        return False

    @staticmethod
    def _cloud_refund_index_keys(cloud_index):
        keys = set()
        sync_ids = set()
        for item in (cloud_index or {}).get('records', []):
            store_sync_id = str(item.get('store_sync_id') or '').strip()
            order_no = str(item.get('order_no') or '').strip()
            sync_id = str(item.get('sync_id') or '').strip()
            if store_sync_id and order_no:
                keys.add((store_sync_id, order_no))
            if sync_id:
                sync_ids.add(sync_id)
        return keys, sync_ids

    def delete_refund_records_missing_from_cloud_index(self, cloud_index):
        """按云端有效订单索引软删除本地多余订单。cloud_index 为 None 时不处理。"""
        if cloud_index is None:
            return 0

        cloud_keys, _ = self._cloud_refund_index_keys(cloud_index)
        cursor = self.conn.cursor()
        rows = cursor.execute('''
            SELECT r.id, s.sync_id, r.order_no
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE IFNULL(r.sync_deleted, 0)=0
              AND IFNULL(s.sync_deleted, 0)=0
        ''').fetchall()
        now = utc_now_text()
        deleted = 0
        for record_id, store_sync_id, order_no in rows:
            key = (str(store_sync_id or '').strip(), str(order_no or '').strip())
            if key not in cloud_keys:
                cursor.execute(
                    '''
                    UPDATE refund_records
                    SET sync_deleted=1, deleted_at=?, updated_at=?
                    WHERE id=? AND IFNULL(sync_deleted, 0)=0
                    ''',
                    (now, now, record_id)
                )
                deleted += cursor.rowcount
        self.conn.commit()
        return deleted

    def dedupe_refund_records_by_store_order(self, cloud_index=None):
        """同一店铺同一订单号只保留一条有效记录，其余软删除。"""
        _, cloud_sync_ids = self._cloud_refund_index_keys(cloud_index)
        cursor = self.conn.cursor()
        groups = cursor.execute('''
            SELECT r.store_id, r.order_no
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE IFNULL(r.sync_deleted, 0)=0
              AND IFNULL(s.sync_deleted, 0)=0
              AND IFNULL(TRIM(r.order_no), '') != ''
            GROUP BY r.store_id, r.order_no
            HAVING COUNT(*) > 1
        ''').fetchall()

        now = utc_now_text()
        deleted = 0
        for store_id, order_no in groups:
            rows = cursor.execute('''
                SELECT id, sync_id, updated_at
                FROM refund_records
                WHERE store_id=? AND order_no=? AND IFNULL(sync_deleted, 0)=0
                ORDER BY updated_at DESC, id DESC
            ''', (store_id, order_no)).fetchall()
            if not rows:
                continue

            keep_id = None
            for record_id, sync_id, _ in rows:
                if sync_id in cloud_sync_ids:
                    keep_id = record_id
                    break
            if keep_id is None:
                keep_id = rows[0][0]

            for record_id, _, _ in rows:
                if record_id == keep_id:
                    continue
                cursor.execute(
                    '''
                    UPDATE refund_records
                    SET sync_deleted=1, deleted_at=?, updated_at=?
                    WHERE id=? AND IFNULL(sync_deleted, 0)=0
                    ''',
                    (now, now, record_id)
                )
                deleted += cursor.rowcount

        self.conn.commit()
        return deleted

    def export_sync_changes(self, since):
        cursor = self.conn.cursor()
        package = {
            'version': 1,
            'device_id': self.get_cloud_device_id(),
            'created_at': utc_now_text(),
            'tables': {},
        }
        for table, meta in SYNC_TABLES.items():
            if not self._table_exists(cursor, table):
                continue
            columns = [meta['key'], 'sync_id', 'updated_at', 'deleted_at', 'sync_deleted'] + meta['columns']
            available = set(self._get_table_columns(cursor, table))
            columns = [column for column in columns if column in available]
            where = "WHERE updated_at > ?" if since else ""
            params = (since,) if since else ()
            rows = cursor.execute(
                f"SELECT {', '.join(columns)} FROM {table} {where} ORDER BY updated_at ASC",
                params
            ).fetchall()
            items = []
            for row in rows:
                item = dict(zip(columns, row))
                if meta.get('store_ref') and item.get(meta['store_ref']):
                    store_sync = cursor.execute(
                        'SELECT sync_id FROM stores WHERE id=?',
                        (item[meta['store_ref']],)
                    ).fetchone()
                    item['_store_sync_id'] = store_sync[0] if store_sync else ''
                items.append(item)
            package['tables'][table] = items
        return package

    def apply_sync_package(self, package):
        cursor = self.conn.cursor()
        applied = 0
        cursor.execute('PRAGMA foreign_keys = OFF')
        try:
            for table in SYNC_TABLES:
                for row in package.get('tables', {}).get(table, []):
                    if self._apply_sync_row(cursor, table, row):
                        applied += 1
            self.conn.commit()
        finally:
            cursor.execute('PRAGMA foreign_keys = ON')
        return applied

    def _apply_sync_row(self, cursor, table, row):
        meta = SYNC_TABLES[table]
        sync_id = str(row.get('sync_id') or '').strip()
        if not sync_id:
            return False

        local_id = self._find_local_sync_row(cursor, table, meta, row)
        incoming_deleted = int(row.get('sync_deleted') or 0)
        incoming_updated_at = row.get('updated_at') or utc_now_text()
        deleted_at = row.get('deleted_at') or (incoming_updated_at if incoming_deleted else '')
        values = {}

        for column in meta['columns']:
            if column not in row:
                continue
            values[column] = row.get(column)

        if meta.get('store_ref'):
            store_sync_id = row.get('_store_sync_id') or ''
            mapped_store_id = self._get_or_create_store_by_sync_id(cursor, store_sync_id)
            if mapped_store_id:
                values[meta['store_ref']] = mapped_store_id
            elif not incoming_deleted:
                return False

        values.update({
            'sync_id': sync_id,
            'updated_at': incoming_updated_at,
            'deleted_at': deleted_at,
            'sync_deleted': incoming_deleted,
        })

        if local_id:
            set_clause = ', '.join([f'{column}=?' for column in values])
            cursor.execute(
                f"UPDATE {table} SET {set_clause} WHERE {meta['key']}=?",
                [values[column] for column in values] + [local_id]
            )
        else:
            if incoming_deleted:
                return False
            columns = list(values.keys())
            placeholders = ', '.join(['?'] * len(columns))
            cursor.execute(
                f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})",
                [values[column] for column in columns]
            )
        return True

    def _find_local_sync_row(self, cursor, table, meta, row):
        found = cursor.execute(f"SELECT {meta['key']} FROM {table} WHERE sync_id=?", (row.get('sync_id'),)).fetchone()
        if found:
            return found[0]
        unique = meta.get('unique') or []
        if not unique:
            return None
        clauses = []
        params = []
        for column in unique:
            value = row.get(column)
            if meta.get('store_ref') == column:
                value = self._get_or_create_store_by_sync_id(cursor, row.get('_store_sync_id') or '')
            if value is None:
                return None
            clauses.append(f"{column}=?")
            params.append(value)
        order_clause = f" ORDER BY IFNULL(sync_deleted, 0) ASC, updated_at DESC, {meta['key']} DESC"
        found = cursor.execute(
            f"SELECT {meta['key']} FROM {table} WHERE {' AND '.join(clauses)}{order_clause}",
            params
        ).fetchone()
        return found[0] if found else None

    def _get_or_create_store_by_sync_id(self, cursor, store_sync_id):
        if not store_sync_id:
            return None
        row = cursor.execute('SELECT id FROM stores WHERE sync_id=?', (store_sync_id,)).fetchone()
        if row:
            return row[0]
        return None

    def _soft_delete(self, table, where_clause, params):
        cursor = self.conn.cursor()
        now = utc_now_text()
        cursor.execute(
            f"UPDATE {table} SET sync_deleted=1, deleted_at=?, updated_at=? WHERE {where_clause} AND IFNULL(sync_deleted, 0)=0",
            (now, now, *params)
        )
        self.conn.commit()
        return cursor.rowcount

    def save_ai_summary_history(self, filter_summary, snapshot):
        """保存AI总结历史快照。"""
        cursor = self.conn.cursor()
        cursor.execute(
            '''
            INSERT INTO ai_summary_history (filter_summary, snapshot_json)
            VALUES (?, ?)
            ''',
            (filter_summary, json.dumps(snapshot, ensure_ascii=False))
        )
        self.conn.commit()
        return cursor.lastrowid

    def get_ai_summary_history_list(self, limit=50):
        """获取AI总结历史列表。"""
        cursor = self.conn.cursor()
        cursor.execute(
            '''
            SELECT id, filter_summary, created_at
            FROM ai_summary_history
            WHERE IFNULL(sync_deleted, 0)=0
            ORDER BY id DESC
            LIMIT ?
            ''',
            (limit,)
        )
        return [
            {"id": row[0], "filter_summary": row[1], "created_at": row[2]}
            for row in cursor.fetchall()
        ]

    def get_ai_summary_history(self, history_id):
        """按ID获取AI总结历史详情。"""
        cursor = self.conn.cursor()
        cursor.execute(
            '''
            SELECT id, filter_summary, snapshot_json, created_at
            FROM ai_summary_history
            WHERE id=? AND IFNULL(sync_deleted, 0)=0
            ''',
            (history_id,)
        )
        row = cursor.fetchone()
        if not row:
            return None

        snapshot = {}
        try:
            snapshot = json.loads(row[2] or "{}")
        except Exception:
            snapshot = {}

        return {
            "id": row[0],
            "filter_summary": row[1],
            "snapshot": snapshot,
            "created_at": row[3]
        }

    def delete_ai_summary_history(self, history_ids):
        """删除选中的AI总结历史记录。"""
        ids = [int(item) for item in history_ids or [] if item]
        if not ids:
            return 0
        placeholders = ",".join("?" for _ in ids)
        return self._soft_delete('ai_summary_history', f"id IN ({placeholders})", tuple(ids))

    def get_real_refund_reason_categories(self, active_only=True):
        """获取真实退款原因分类列表。"""
        cursor = self.conn.cursor()
        query = '''
            SELECT id, category_name, keywords_text, status, sort_order, created_at, updated_at
            FROM real_refund_reason_categories
            WHERE IFNULL(sync_deleted, 0)=0
        '''
        if active_only:
            query += " AND status = 'ACTIVE'"
        query += " ORDER BY sort_order ASC, id ASC"
        cursor.execute(query)
        return [
            {
                "id": row[0],
                "category_name": row[1],
                "keywords_text": row[2] or "",
                "status": row[3],
                "sort_order": row[4],
                "created_at": row[5],
                "updated_at": row[6],
            }
            for row in cursor.fetchall()
        ]

    def save_real_refund_reason_categories(self, categories):
        """保存真实退款原因分类。支持字符串列表和配置字典列表。"""
        cursor = self.conn.cursor()
        cleaned = []
        seen = set()
        for index, item in enumerate(categories or []):
            if isinstance(item, dict):
                text = str(item.get("category_name") or item.get("name") or "").strip()
                keywords_text = str(item.get("keywords_text") or "").strip()
                status = str(item.get("status") or "ACTIVE").strip() or "ACTIVE"
                sort_order = int(item.get("sort_order", index) or 0)
            else:
                text = str(item or "").strip()
                keywords_text = ""
                status = "ACTIVE"
                sort_order = index

            if not text or text in seen:
                continue
            seen.add(text)
            cleaned.append({
                "category_name": text,
                "keywords_text": keywords_text,
                "status": status,
                "sort_order": sort_order,
            })

        for item in cleaned:
            cursor.execute(
                '''
                INSERT INTO real_refund_reason_categories (category_name, keywords_text, status, sort_order)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(category_name) DO UPDATE SET
                    keywords_text=excluded.keywords_text,
                    status=excluded.status,
                    sort_order=excluded.sort_order,
                    updated_at=CURRENT_TIMESTAMP
                ''',
                (item["category_name"], item["keywords_text"], item["status"], item["sort_order"])
            )
        self.conn.commit()
        return cleaned

    def replace_real_refund_reason_categories(self, categories):
        """用当前管理窗口里的分类配置替换分类表，删除用户已移除的分类。"""
        cleaned = self.save_real_refund_reason_categories(categories)
        cursor = self.conn.cursor()
        keep_names = [item["category_name"] for item in cleaned]
        if keep_names:
            placeholders = ",".join("?" for _ in keep_names)
            self._soft_delete('real_refund_reason_categories', f"category_name NOT IN ({placeholders})", tuple(keep_names))
        else:
            self._soft_delete('real_refund_reason_categories', "1=1", ())
        return cleaned

    def ensure_default_real_refund_reason_categories(self, default_categories):
        """首次使用时写入默认分类；用户删除后的分类不再自动补回。"""
        existing = self.get_real_refund_reason_categories(active_only=False)
        if not existing:
            self.save_real_refund_reason_categories(default_categories)

    def update_real_refund_reason(self, record_id, category, detail="", note_hash="", updated_at=""):
        """更新单条记录的真实退款原因归因结果。"""
        cursor = self.conn.cursor()
        cursor.execute(
            '''
            UPDATE refund_records SET
                real_refund_reason = ?,
                real_refund_reason_detail = ?,
                real_refund_reason_note_hash = ?,
                real_refund_reason_updated_at = ?
            WHERE id = ?
            ''',
            (
                str(category or "").strip(),
                str(detail or "").strip(),
                str(note_hash or "").strip(),
                str(updated_at or "").strip(),
                record_id,
            )
        )
        self.conn.commit()
        return cursor.rowcount > 0

    def get_stores(self):
        """获取所有店铺，返回列表 [(id, name), ...]"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT id, store_name FROM stores WHERE IFNULL(sync_deleted, 0)=0 ORDER BY store_name')
        return cursor.fetchall()

    def add_store(self, name):
        """添加店铺，返回新ID，如果已存在返回None"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('INSERT INTO stores (store_name) VALUES (?)', (name,))
            self.conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            cursor = self.conn.cursor()
            now = utc_now_text()
            cursor.execute(
                "UPDATE stores SET sync_deleted=0, deleted_at='', updated_at=? WHERE store_name=? AND IFNULL(sync_deleted, 0)=1",
                (now, name)
            )
            self.conn.commit()
            if cursor.rowcount > 0:
                row = cursor.execute('SELECT id FROM stores WHERE store_name=?', (name,)).fetchone()
                return row[0] if row else None
            return None

    def set_store_color(self, store_name, color):
        """设置店铺颜色"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE stores SET color = ? WHERE store_name = ?', (color, store_name))
        self.conn.commit()
        return cursor.rowcount > 0

    def get_store_color(self, store_name):
        """获取店铺颜色"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT color FROM stores WHERE store_name = ?', (store_name,))
        result = cursor.fetchone()
        return result[0] if result and result[0] else None

    def clear_store_color(self, store_name):
        """清除店铺颜色"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE stores SET color = NULL WHERE store_name = ?', (store_name,))
        self.conn.commit()
        return cursor.rowcount > 0

    def set_estimated_orders(self, store_name, estimated_orders):
        """设置店铺预估订单量"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE stores SET estimated_orders = ? WHERE store_name = ?', (estimated_orders, store_name))
        self.conn.commit()
        return cursor.rowcount > 0

    def get_estimated_orders(self, store_name):
        """获取店铺预估订单量"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT estimated_orders FROM stores WHERE store_name = ?', (store_name,))
        result = cursor.fetchone()
        return result[0] if result else 0

    def update_store_settings(self, store_id, daily_orders, daily_sales, refund_budget):
        """更新店铺设置"""
        cursor = self.conn.cursor()
        cursor.execute('''
            UPDATE stores SET 
                daily_orders = ?, daily_sales = ?, refund_budget = ?
            WHERE id = ?
        ''', (daily_orders, daily_sales, refund_budget, store_id))
        self.conn.commit()
        return cursor.rowcount > 0

    def get_store_settings(self, store_id):
        """获取店铺设置"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT daily_orders, daily_sales, refund_budget
            FROM stores WHERE id = ?
              AND IFNULL(sync_deleted, 0)=0
        ''', (store_id,))
        result = cursor.fetchone()
        if result:
            return {
                'daily_orders': result[0],
                'daily_sales': result[1],
                'refund_budget': result[2]
            }
        return None

    def save_store_weekly_settings_history(
        self,
        store_id,
        store_name,
        week_start_date,
        week_end_date,
        weekly_orders,
        weekly_sales,
        refund_budget
    ):
        """保存或覆盖店铺周数据历史记录。"""
        cursor = self.conn.cursor()
        cursor.execute('''
            INSERT INTO store_weekly_settings_history (
                store_id, store_name_snapshot, week_start_date, week_end_date,
                weekly_orders, weekly_sales, refund_budget
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(store_id, week_start_date) DO UPDATE SET
                store_name_snapshot = excluded.store_name_snapshot,
                week_end_date = excluded.week_end_date,
                weekly_orders = excluded.weekly_orders,
                weekly_sales = excluded.weekly_sales,
                refund_budget = excluded.refund_budget,
                updated_at = CURRENT_TIMESTAMP
        ''', (
            store_id,
            store_name,
            week_start_date,
            week_end_date,
            weekly_orders,
            weekly_sales,
            refund_budget
        ))
        self.conn.commit()
        return cursor.lastrowid

    def get_store_weekly_settings_history(self, store_id, limit=200):
        """获取指定店铺的周数据历史记录。"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT id, store_id, store_name_snapshot, week_start_date, week_end_date,
                   weekly_orders, weekly_sales, refund_budget, created_at, updated_at
            FROM store_weekly_settings_history
            WHERE store_id = ? AND IFNULL(sync_deleted, 0)=0
            ORDER BY week_start_date DESC, id DESC
            LIMIT ?
        ''', (store_id, limit))
        return [self._row_to_store_weekly_history(row) for row in cursor.fetchall()]

    def get_store_weekly_settings_history_by_week(self, store_id, week_start_date):
        """按店铺和周开始日期获取一条历史记录。"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT id, store_id, store_name_snapshot, week_start_date, week_end_date,
                   weekly_orders, weekly_sales, refund_budget, created_at, updated_at
            FROM store_weekly_settings_history
            WHERE store_id = ? AND week_start_date = ? AND IFNULL(sync_deleted, 0)=0
        ''', (store_id, week_start_date))
        row = cursor.fetchone()
        return self._row_to_store_weekly_history(row) if row else None

    def delete_store_weekly_settings_history(self, history_id, store_id):
        """删除指定店铺的一条周数据历史记录。"""
        return self._soft_delete(
            'store_weekly_settings_history',
            'id = ? AND store_id = ?',
            (history_id, store_id)
        ) > 0

    @staticmethod
    def _row_to_store_weekly_history(row):
        return {
            'id': row[0],
            'store_id': row[1],
            'store_name_snapshot': row[2],
            'week_start_date': row[3],
            'week_end_date': row[4],
            'weekly_orders': row[5] or 0,
            'weekly_sales': row[6] or 0.0,
            'refund_budget': row[7] or 0.0,
            'created_at': row[8],
            'updated_at': row[9],
        }

    def save_store_weekly_spec_orders(
        self,
        store_id,
        store_name,
        week_start_date,
        week_end_date,
        spec_orders
    ):
        """覆盖保存某店铺某自然周的规格单量明细。"""
        merged_items = {}
        for item in spec_orders or []:
            spec_code = str(item.get('spec_code') or '').strip()
            if not spec_code:
                continue
            try:
                order_count = int(item.get('order_count') or 0)
            except (TypeError, ValueError):
                continue
            if order_count < 0:
                continue
            merged_items[spec_code] = merged_items.get(spec_code, 0) + order_count

        cleaned_items = sorted(
            merged_items.items(),
            key=lambda item: (-item[1], item[0])
        )

        cursor = self.conn.cursor()
        self._soft_delete('store_weekly_spec_orders', 'store_id = ? AND week_start_date = ?', (store_id, week_start_date))

        for spec_code, order_count in cleaned_items:
            cursor.execute('''
                INSERT INTO store_weekly_spec_orders (
                    store_id, store_name_snapshot, week_start_date, week_end_date,
                    spec_code, order_count
                )
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(store_id, week_start_date, spec_code) DO UPDATE SET
                    store_name_snapshot=excluded.store_name_snapshot,
                    week_end_date=excluded.week_end_date,
                    order_count=excluded.order_count,
                    sync_deleted=0,
                    deleted_at='',
                    updated_at=CURRENT_TIMESTAMP
            ''', (
                store_id,
                store_name,
                week_start_date,
                week_end_date,
                spec_code,
                order_count
            ))

        self.conn.commit()
        return len(cleaned_items)

    def get_store_weekly_spec_orders_by_week(self, store_id, week_start_date):
        """获取某店铺某自然周的规格单量明细。"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT id, store_id, store_name_snapshot, week_start_date, week_end_date,
                   spec_code, order_count, created_at, updated_at
            FROM store_weekly_spec_orders
            WHERE store_id = ? AND week_start_date = ? AND IFNULL(sync_deleted, 0)=0
            ORDER BY order_count DESC, spec_code ASC
        ''', (store_id, week_start_date))
        return [self._row_to_store_weekly_spec_order(row) for row in cursor.fetchall()]

    def get_store_weekly_spec_orders_history_summary(self, store_id, limit=200):
        """按自然周获取某店铺规格单量历史汇总。"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT store_id, MAX(store_name_snapshot), week_start_date, MAX(week_end_date),
                   COUNT(*) AS spec_count, SUM(order_count) AS total_orders,
                   MIN(created_at) AS created_at, MAX(updated_at) AS updated_at
            FROM store_weekly_spec_orders
            WHERE store_id = ? AND IFNULL(sync_deleted, 0)=0
            GROUP BY store_id, week_start_date
            ORDER BY week_start_date DESC
            LIMIT ?
        ''', (store_id, limit))
        return [
            {
                'store_id': row[0],
                'store_name_snapshot': row[1] or '',
                'week_start_date': row[2],
                'week_end_date': row[3],
                'spec_count': row[4] or 0,
                'total_orders': row[5] or 0,
                'created_at': row[6],
                'updated_at': row[7],
            }
            for row in cursor.fetchall()
        ]

    def get_store_weekly_spec_orders_history(self, store_id, limit=1000):
        """获取某店铺已保存的规格单量历史明细。"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT id, store_id, store_name_snapshot, week_start_date, week_end_date,
                   spec_code, order_count, created_at, updated_at
            FROM store_weekly_spec_orders
            WHERE store_id = ? AND IFNULL(sync_deleted, 0)=0
            ORDER BY week_start_date DESC, order_count DESC, spec_code ASC
            LIMIT ?
        ''', (store_id, limit))
        return [self._row_to_store_weekly_spec_order(row) for row in cursor.fetchall()]

    def delete_store_weekly_spec_orders_by_week(self, store_id, week_start_date):
        """删除某店铺某自然周的规格单量明细。"""
        return self._soft_delete(
            'store_weekly_spec_orders',
            'store_id = ? AND week_start_date = ?',
            (store_id, week_start_date)
        ) > 0

    @staticmethod
    def _row_to_store_weekly_spec_order(row):
        return {
            'id': row[0],
            'store_id': row[1],
            'store_name_snapshot': row[2],
            'week_start_date': row[3],
            'week_end_date': row[4],
            'spec_code': row[5] or '',
            'order_count': row[6] or 0,
            'created_at': row[7],
            'updated_at': row[8],
        }

    def delete_store(self, store_id):
        """删除店铺及其相关数据（由于外键约束，相关记录会自动删除）"""
        try:
            self._soft_delete('refund_records', 'store_id = ?', (store_id,))
            self._soft_delete('store_weekly_settings_history', 'store_id = ?', (store_id,))
            self._soft_delete('store_weekly_spec_orders', 'store_id = ?', (store_id,))
            self._soft_delete('stores', 'id = ?', (store_id,))
            return True
        except Exception as e:
            print(f"删除店铺失败: {e}")
            return False

    def update_store_name(self, store_id, new_name):
        """修改店铺名称"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('UPDATE stores SET store_name = ? WHERE id = ?', (new_name, store_id))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            # 店铺名称已存在
            return False
        except Exception as e:
            print(f"修改店铺名称失败: {e}")
            return False

    def save_global_settings(self, daily_orders, daily_sales, refund_budget):
        """保存全局设置（全部店铺）"""
        cursor = self.conn.cursor()
        
        # 保存设置到全局设置表
        settings = {
            'daily_orders': daily_orders,
            'daily_sales': daily_sales,
            'refund_budget': refund_budget
        }
        
        for key, value in settings.items():
            cursor.execute('''
                INSERT OR REPLACE INTO global_settings (setting_key, setting_value)
                VALUES (?, ?)
            ''', (key, str(value)))
        
        self.conn.commit()

    def get_global_settings(self):
        """获取全局设置（全部店铺）"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT setting_key, setting_value FROM global_settings')
        results = cursor.fetchall()
        
        if not results:
            return {'daily_orders': 0, 'daily_sales': 0.0, 'refund_budget': 0.0}
        
        settings = {}
        for key, value in results:
            if key == 'daily_orders':
                settings[key] = int(value) if value else 0
            elif key in ['daily_sales', 'refund_budget']:
                settings[key] = float(value) if value else 0.0
            else:
                settings[key] = value
        
        return settings

    def save_window_settings(self, settings):
        """保存窗口设置到数据库"""
        cursor = self.conn.cursor()
        
        for key, value in settings.items():
            cursor.execute('''
                INSERT OR REPLACE INTO window_settings (setting_key, setting_value)
                VALUES (?, ?)
            ''', (key, str(value)))
        
        self.conn.commit()

    def load_window_settings(self):
        """从数据库加载窗口设置"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT setting_key, setting_value FROM window_settings')
        results = cursor.fetchall()
        
        if not results:
            return None
        
        settings = {}
        for key, value in results:
            # 解析不同类型的设置值
            if key in ['window_size', 'main_splitter', 'top_splitter', 'bottom_splitter']:
                # 列表类型设置（如分割器比例）
                try:
                    settings[key] = eval(value)  # 使用eval将字符串转换为列表
                except:
                    settings[key] = []
            elif value.isdigit():
                settings[key] = int(value)
            elif value.replace('.', '', 1).isdigit():
                settings[key] = float(value)
            else:
                settings[key] = value
        
        return settings

    def get_store_refund_stats(self, store_name):
        """获取店铺退款统计（排除撤销订单）"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT COUNT(*), SUM(refund_amount), SUM(comp_amount)
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE s.store_name = ? AND r.cancel = 0
              AND IFNULL(r.sync_deleted, 0)=0 AND IFNULL(s.sync_deleted, 0)=0
        ''', (store_name,))
        result = cursor.fetchone()
        if result and result[0] is not None:
            return {
                'refund_count': result[0],
                'total_refund': result[1] or 0.0,
                'total_comp': result[2] or 0.0
            }
        return {'refund_count': 0, 'total_refund': 0.0, 'total_comp': 0.0}

    def get_all_records(self):
        """获取所有退款记录"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT r.id, r.order_no, r.spec_name, r.spec_code, r.reason,
                   r.real_refund_reason, r.real_refund_reason_detail, r.real_refund_reason_updated_at, r.real_refund_reason_note_hash,
                   r.quality_refund_reason, r.quality_not_cancelled_reason, r.quality_refund_reason_detail, r.quality_refund_reason_updated_at, r.quality_refund_reason_note_hash,
                   r.refund_amount, r.cancel, r.compensate, r.comp_amount,
                   r.order_status, r.after_sale_status, r.refund_apply_time, r.refund_agree_time, r.record_date, s.store_name, r.store_id
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE IFNULL(r.sync_deleted, 0)=0 AND IFNULL(s.sync_deleted, 0)=0
            ORDER BY r.record_date DESC, r.id DESC
        ''')
        records = []
        for row in cursor.fetchall():
            records.append({
                'id': row[0], 'order_no': row[1], 'spec_name': row[2] or '', 'spec_code': row[3] or '',
                'reason': row[4], 'real_refund_reason': row[5] or '', 'real_refund_reason_detail': row[6] or '',
                'real_refund_reason_updated_at': row[7] or '', 'real_refund_reason_note_hash': row[8] or '',
                'quality_refund_reason': row[9] or '', 'quality_not_cancelled_reason': row[10] or '',
                'quality_refund_reason_detail': row[11] or '', 'quality_refund_reason_updated_at': row[12] or '',
                'quality_refund_reason_note_hash': row[13] or '',
                'refund_amount': row[14], 'cancel': bool(row[15]), 'compensate': bool(row[16]), 'comp_amount': row[17],
                'order_status': row[18], 'after_sale_status': row[19],
                'refund_apply_time': row[20] or '', 'refund_agree_time': row[21] or '',
                'record_date': row[22], 'store_name': row[23], 'store_id': row[24]
            })
        return records

    def get_records_missing_spec_code_with_name(self):
        """获取已录入规格名称但规格编码为空的记录。"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT id, spec_name, spec_code
            FROM refund_records
            WHERE IFNULL(TRIM(spec_name), '') != ''
              AND IFNULL(TRIM(spec_code), '') IN ('', '-')
              AND IFNULL(sync_deleted, 0)=0
            ORDER BY id
        ''')
        return [
            {
                'id': row[0],
                'spec_name': row[1] or '',
                'spec_code': row[2] or '',
            }
            for row in cursor.fetchall()
        ]

    def get_records_with_spec_name(self):
        """获取所有已录入规格名称的记录，用于手动重新识别规格编码。"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT id, spec_name, spec_code
            FROM refund_records
            WHERE IFNULL(TRIM(spec_name), '') != ''
              AND IFNULL(sync_deleted, 0)=0
            ORDER BY id
        ''')
        return [
            {
                'id': row[0],
                'spec_name': row[1] or '',
                'spec_code': row[2] or '',
            }
            for row in cursor.fetchall()
        ]

    def get_total_record_count(self):
        """获取数据库中的总记录数"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM refund_records WHERE IFNULL(sync_deleted, 0)=0')
        result = cursor.fetchone()
        if result and isinstance(result, (tuple, list)) and len(result) > 0:
            return result[0] if isinstance(result[0], int) else int(result[0])
        return 0

    def cleanup_orphan_records(self):
        """清理没有对应店铺的孤儿记录"""
        try:
            cursor = self.conn.cursor()
            # 删除没有对应店铺的记录
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE store_id NOT IN (SELECT id FROM stores)
            ''')
            deleted_count = cursor.rowcount
            self.conn.commit()
            return deleted_count
        except Exception as e:
            print(f"清理孤儿记录失败: {e}")
            return 0

    def debug_database_records(self):
        """调试功能：查看数据库中的所有记录"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('''
                SELECT r.id, r.order_no, r.store_id, s.store_name, r.reason, r.record_date
                FROM refund_records r
                LEFT JOIN stores s ON r.store_id = s.id
                ORDER BY r.id
            ''')
            records = cursor.fetchall()
            
            result = []
            for record in records:
                result.append({
                    'id': record[0],
                    'order_no': record[1],
                    'store_id': record[2],
                    'store_name': record[3] if record[3] else '无对应店铺',
                    'reason': record[4],
                    'record_date': record[5]
                })
            
            return result
        except Exception as e:
            print(f"调试查询失败: {e}")
            return []

    def force_global_sync(self):
        """强制全局同步：彻底清理所有不一致数据"""
        try:
            cursor = self.conn.cursor()
            
            # 第一步：清理孤儿记录
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE store_id NOT IN (SELECT id FROM stores)
            ''')
            orphan_count = cursor.rowcount
            
            # 第二步：清理重复记录（保留最新的）
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE id NOT IN (
                    SELECT MAX(id) 
                    FROM refund_records 
                    GROUP BY order_no, store_id
                )
            ''')
            duplicate_count = cursor.rowcount
            
            # 第三步：清理无效数据（订单号为空或店铺ID为0）
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE order_no = '' OR order_no IS NULL OR store_id = 0
            ''')
            invalid_count = cursor.rowcount
            
            # 第四步：清理所有隐藏的不一致数据（终极清理）
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE id NOT IN (
                    SELECT r.id 
                    FROM refund_records r
                    JOIN stores s ON r.store_id = s.id
                )
            ''')
            hidden_count = cursor.rowcount
            
            self.conn.commit()
            
            return {
                'orphan_count': orphan_count,
                'duplicate_count': duplicate_count,
                'invalid_count': invalid_count,
                'hidden_count': hidden_count,
                'total_cleaned': orphan_count + duplicate_count + invalid_count + hidden_count
            }
        except Exception as e:
            print(f"强制同步失败: {e}")
            return {'orphan_count': 0, 'duplicate_count': 0, 'invalid_count': 0, 'hidden_count': 0, 'total_cleaned': 0}

    def get_filtered_record_count(self, order_no='', reason='全部', cancel='全部',
                                 reject='全部', start_date=None, end_date=None, store_name='全部'):
        """根据筛选条件获取记录数"""
        cursor = self.conn.cursor()
        query = 'SELECT COUNT(*) FROM refund_records r JOIN stores s ON r.store_id = s.id WHERE IFNULL(r.sync_deleted, 0)=0 AND IFNULL(s.sync_deleted, 0)=0'
        params = []
        
        if order_no:
            query += ' AND r.order_no LIKE ?'
            params.append(f'%{order_no}%')
        
        if reason != '全部':
            query += ' AND r.reason = ?'
            params.append(reason)
        
        if cancel != '全部':
            if cancel == '是':
                query += ' AND r.cancel = 1'
            elif cancel == '否':
                query += ' AND r.cancel = 0'
        
        if reject != '全部':
            if reject == '是':
                query += ' AND r.reject = 1'
            elif reject == '否':
                query += ' AND r.reject = 0'
        
        if start_date:
            query += ' AND r.record_date >= ?'
            params.append(start_date)
        
        if end_date:
            query += ' AND r.record_date <= ?'
            params.append(end_date)
        
        if store_name != '全部':
            query += ' AND s.store_name = ?'
            params.append(store_name)
        
        cursor.execute(query, params)
        result = cursor.fetchone()
        if result and isinstance(result, (tuple, list)) and len(result) > 0:
            return result[0] if isinstance(result[0], int) else int(result[0])
        return 0

    def add_record(self, store_id, order_no, reason, refund_amount, cancel, compensate, comp_amount, reject, reject_result, notes, record_date, order_status='', after_sale_status='', spec_name='', spec_code='', refund_apply_time='', refund_agree_time=''):
        """添加退款记录"""
        cursor = self.conn.cursor()
        cursor.execute('''
            INSERT INTO refund_records 
            (store_id, order_no, spec_name, spec_code, reason, refund_amount, cancel, compensate, comp_amount, reject, reject_result, notes, order_status, after_sale_status, refund_apply_time, refund_agree_time, record_date,
             real_refund_reason, real_refund_reason_detail, real_refund_reason_updated_at, real_refund_reason_note_hash,
             quality_refund_reason, quality_not_cancelled_reason, quality_refund_reason_detail, quality_refund_reason_updated_at, quality_refund_reason_note_hash)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', '', '', '', '', '', '', '', '')
        ''', (store_id, order_no, spec_name, spec_code, reason, refund_amount, 1 if cancel else 0, 1 if compensate else 0, comp_amount, 1 if reject else 0, reject_result, notes, order_status, after_sale_status, refund_apply_time, refund_agree_time, record_date))
        self.conn.commit()
        return cursor.lastrowid

    def update_record(self, record_id, store_id, order_no, reason, refund_amount, cancel, compensate, comp_amount, reject, reject_result, notes, record_date, order_status='', after_sale_status='', spec_name='', spec_code='', refund_apply_time=None, refund_agree_time=None):
        """更新退款记录"""
        cursor = self.conn.cursor()
        existing_note_row = cursor.execute(
            'SELECT notes FROM refund_records WHERE id=?',
            (record_id,)
        ).fetchone()
        notes_changed = bool(existing_note_row and str(existing_note_row[0] or '') != str(notes or ''))
        real_reason = ''
        real_reason_detail = ''
        real_reason_updated_at = ''
        real_reason_note_hash = ''
        quality_reason = ''
        quality_not_cancelled_reason = ''
        quality_reason_detail = ''
        quality_reason_updated_at = ''
        quality_reason_note_hash = ''
        if not notes_changed:
            real_reason_row = cursor.execute(
                '''
                SELECT real_refund_reason, real_refund_reason_detail, real_refund_reason_updated_at, real_refund_reason_note_hash,
                       quality_refund_reason, quality_not_cancelled_reason, quality_refund_reason_detail,
                       quality_refund_reason_updated_at, quality_refund_reason_note_hash
                FROM refund_records WHERE id=?
                ''',
                (record_id,)
            ).fetchone()
            if real_reason_row:
                real_reason, real_reason_detail, real_reason_updated_at, real_reason_note_hash = real_reason_row[:4]
                quality_reason, quality_not_cancelled_reason, quality_reason_detail, quality_reason_updated_at, quality_reason_note_hash = real_reason_row[4:9]
        if refund_apply_time is None or refund_agree_time is None:
            time_row = cursor.execute(
                'SELECT refund_apply_time, refund_agree_time FROM refund_records WHERE id=?',
                (record_id,)
            ).fetchone()
            if refund_apply_time is None:
                refund_apply_time = time_row[0] if time_row else ''
            if refund_agree_time is None:
                refund_agree_time = time_row[1] if time_row else ''
        cursor.execute('''
            UPDATE refund_records SET
                store_id=?, order_no=?, spec_name=?, spec_code=?, reason=?, refund_amount=?,
                cancel=?, compensate=?, comp_amount=?, reject=?, reject_result=?, notes=?, order_status=?, after_sale_status=?, refund_apply_time=?, refund_agree_time=?, record_date=?,
                real_refund_reason=?, real_refund_reason_detail=?, real_refund_reason_updated_at=?, real_refund_reason_note_hash=?,
                quality_refund_reason=?, quality_not_cancelled_reason=?, quality_refund_reason_detail=?, quality_refund_reason_updated_at=?, quality_refund_reason_note_hash=?
            WHERE id=?
        ''', (store_id, order_no, spec_name, spec_code, reason, refund_amount, 1 if cancel else 0, 1 if compensate else 0, comp_amount, 1 if reject else 0, reject_result, notes, order_status, after_sale_status, refund_apply_time, refund_agree_time, record_date, real_reason, real_reason_detail, real_reason_updated_at, real_reason_note_hash, quality_reason, quality_not_cancelled_reason, quality_reason_detail, quality_reason_updated_at, quality_reason_note_hash, record_id))
        self.conn.commit()

    def update_refund_amount(self, record_id, refund_amount):
        """更新退款金额"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE refund_records SET refund_amount=? WHERE id=?', (refund_amount, record_id))
        self.conn.commit()
        return cursor.rowcount > 0

    def update_comp_amount(self, record_id, comp_amount):
        """更新补偿金额"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE refund_records SET comp_amount=? WHERE id=?', (comp_amount, record_id))
        self.conn.commit()
        return cursor.rowcount > 0

    def get_store_id_by_name(self, store_name):
        """根据店铺名称获取店铺ID"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT id FROM stores WHERE store_name = ? AND IFNULL(sync_deleted, 0)=0', (store_name,))
        result = cursor.fetchone()
        if result and isinstance(result, (tuple, list)) and len(result) > 0:
            return result[0] if isinstance(result[0], int) else int(result[0])
        return None

    def update_record_partial(self, record_id, **kwargs):
        """智能增量更新记录：只更新提供的字段，保护未提供的字段"""
        if not kwargs:
            return False
        
        # 构建动态SQL更新语句
        set_clauses = []
        params = []
        
        # 支持的字段映射
        field_mapping = {
            'store_id': 'store_id',
            'order_no': 'order_no', 
            'spec_name': 'spec_name',
            'spec_code': 'spec_code',
            'reason': 'reason',
            'refund_amount': 'refund_amount',
            'cancel': 'cancel',
            'compensate': 'compensate',
            'comp_amount': 'comp_amount',
            'reject': 'reject',
            'reject_result': 'reject_result',
            'notes': 'notes',
            'order_status': 'order_status',
            'after_sale_status': 'after_sale_status',
            'refund_apply_time': 'refund_apply_time',
            'refund_agree_time': 'refund_agree_time',
            'record_date': 'record_date',
            'real_refund_reason': 'real_refund_reason',
            'real_refund_reason_detail': 'real_refund_reason_detail',
            'real_refund_reason_updated_at': 'real_refund_reason_updated_at',
            'real_refund_reason_note_hash': 'real_refund_reason_note_hash',
            'quality_refund_reason': 'quality_refund_reason',
            'quality_not_cancelled_reason': 'quality_not_cancelled_reason',
            'quality_refund_reason_detail': 'quality_refund_reason_detail',
            'quality_refund_reason_updated_at': 'quality_refund_reason_updated_at',
            'quality_refund_reason_note_hash': 'quality_refund_reason_note_hash'
        }
        
        # 处理每个提供的字段
        for field, value in kwargs.items():
            if field in field_mapping:
                # 处理布尔值转换为整数
                if field in ['cancel', 'compensate', 'reject']:
                    value = 1 if value else 0
                set_clauses.append(f"{field_mapping[field]}=?")
                params.append(value)
        
        if not set_clauses:
            return False
            
        # 执行更新
        cursor = self.conn.cursor()
        if 'notes' in kwargs:
            old_note_row = cursor.execute('SELECT notes FROM refund_records WHERE id=?', (record_id,)).fetchone()
            old_note = str(old_note_row[0] or '') if old_note_row else ''
            new_note = str(kwargs.get('notes') or '')
            if old_note != new_note:
                set_clauses.extend([
                    'real_refund_reason=?',
                    'real_refund_reason_detail=?',
                    'real_refund_reason_updated_at=?',
                    'real_refund_reason_note_hash=?',
                    'quality_refund_reason=?',
                    'quality_not_cancelled_reason=?',
                    'quality_refund_reason_detail=?',
                    'quality_refund_reason_updated_at=?',
                    'quality_refund_reason_note_hash=?'
                ])
                params.extend(['', '', '', '', '', '', '', '', ''])

        # 添加记录ID作为WHERE条件。必须在备注变更追加清空归因字段之后添加，
        # 否则 WHERE id 会绑定到错误参数，导致备注保存不生效。
        params.append(record_id)
        sql = f"UPDATE refund_records SET {', '.join(set_clauses)} WHERE id=?"
        cursor.execute(sql, params)
        self.conn.commit()
        
        return cursor.rowcount > 0

    def update_quality_refund_reason(self, record_id, quality_reason, not_cancelled_reason, detail="", note_hash="", updated_at=""):
        """更新单条记录的品质退款原因分析结果。"""
        cursor = self.conn.cursor()
        cursor.execute(
            '''
            UPDATE refund_records SET
                quality_refund_reason = ?,
                quality_not_cancelled_reason = ?,
                quality_refund_reason_detail = ?,
                quality_refund_reason_note_hash = ?,
                quality_refund_reason_updated_at = ?
            WHERE id = ?
            ''',
            (
                str(quality_reason or '').strip(),
                str(not_cancelled_reason or '').strip(),
                str(detail or '').strip(),
                str(note_hash or '').strip(),
                str(updated_at or '').strip(),
                record_id,
            )
        )
        self.conn.commit()
        return cursor.rowcount > 0

    def get_quality_not_cancelled_reason_categories(self, active_only=True):
        """获取未撤销原因分类列表。"""
        cursor = self.conn.cursor()
        query = '''
            SELECT id, category_name, status, sort_order, created_at, updated_at
            FROM quality_not_cancelled_reason_categories
            WHERE IFNULL(sync_deleted, 0)=0
        '''
        if active_only:
            query += " AND status = 'ACTIVE'"
        query += " ORDER BY sort_order ASC, id ASC"
        cursor.execute(query)
        return [
            {
                "id": row[0],
                "category_name": row[1],
                "status": row[2],
                "sort_order": row[3],
                "created_at": row[4],
                "updated_at": row[5],
            }
            for row in cursor.fetchall()
        ]

    def save_quality_not_cancelled_reason_categories(self, categories):
        """保存未撤销原因分类。支持字符串列表和配置字典列表。"""
        cursor = self.conn.cursor()
        cleaned = []
        seen = set()
        for index, item in enumerate(categories or []):
            if isinstance(item, dict):
                text = str(item.get("category_name") or item.get("name") or "").strip()
                status = str(item.get("status") or "ACTIVE").strip() or "ACTIVE"
                sort_order = int(item.get("sort_order", index) or 0)
            else:
                text = str(item or "").strip()
                status = "ACTIVE"
                sort_order = index
            if not text or text in seen:
                continue
            seen.add(text)
            cleaned.append({
                "category_name": text,
                "status": status,
                "sort_order": sort_order,
            })

        for item in cleaned:
            cursor.execute(
                '''
                INSERT INTO quality_not_cancelled_reason_categories (category_name, status, sort_order)
                VALUES (?, ?, ?)
                ON CONFLICT(category_name) DO UPDATE SET
                    status=excluded.status,
                    sort_order=excluded.sort_order,
                    updated_at=CURRENT_TIMESTAMP
                ''',
                (item["category_name"], item["status"], item["sort_order"])
            )
        self.conn.commit()
        return cleaned

    def update_quality_not_cancelled_reason(self, record_id, not_cancelled_reason, detail="", note_hash="", updated_at=""):
        """更新单条记录的未撤销原因分析结果，不覆盖旧的品质退款原因字段。"""
        cursor = self.conn.cursor()
        cursor.execute(
            '''
            UPDATE refund_records SET
                quality_not_cancelled_reason = ?,
                quality_refund_reason_detail = ?,
                quality_refund_reason_note_hash = ?,
                quality_refund_reason_updated_at = ?
            WHERE id = ?
            ''',
            (
                str(not_cancelled_reason or '').strip(),
                str(detail or '').strip(),
                str(note_hash or '').strip(),
                str(updated_at or '').strip(),
                record_id,
            )
        )
        self.conn.commit()
        return cursor.rowcount > 0

    def delete_record(self, record_id):
        """删除退款记录，返回是否成功（增强错误处理）"""
        try:
            # 检查记录ID是否有效
            if record_id is None:
                return False
                
            cursor = self.conn.cursor()
            return self._soft_delete('refund_records', 'id=?', (record_id,)) > 0
        except Exception as e:
            print(f"删除记录 {record_id} 时数据库错误: {e}")
            return False

    def get_record_by_id(self, record_id):
        """根据ID获取记录"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT r.id, r.order_no, r.spec_name, r.spec_code, r.reason,
                   r.real_refund_reason, r.real_refund_reason_detail, r.real_refund_reason_updated_at, r.real_refund_reason_note_hash,
                   r.quality_refund_reason, r.quality_not_cancelled_reason, r.quality_refund_reason_detail, r.quality_refund_reason_updated_at, r.quality_refund_reason_note_hash,
                   r.refund_amount, r.cancel, r.compensate, r.comp_amount, 
                   r.reject, r.reject_result, r.notes, r.order_status, r.after_sale_status, r.refund_apply_time, r.refund_agree_time, r.record_date, s.store_name, r.store_id
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE r.id=? AND IFNULL(r.sync_deleted, 0)=0 AND IFNULL(s.sync_deleted, 0)=0
        ''', (record_id,))
        row = cursor.fetchone()
        if row:
            return {
                'id': row[0], 'order_no': row[1], 'spec_name': row[2] or '', 'spec_code': row[3] or '',
                'reason': row[4], 'real_refund_reason': row[5] or '', 'real_refund_reason_detail': row[6] or '',
                'real_refund_reason_updated_at': row[7] or '', 'real_refund_reason_note_hash': row[8] or '',
                'quality_refund_reason': row[9] or '', 'quality_not_cancelled_reason': row[10] or '',
                'quality_refund_reason_detail': row[11] or '', 'quality_refund_reason_updated_at': row[12] or '',
                'quality_refund_reason_note_hash': row[13] or '',
                'refund_amount': row[14], 'cancel': bool(row[15]), 'compensate': bool(row[16]), 'comp_amount': row[17],
                'reject': bool(row[18]), 'reject_result': row[19], 'notes': row[20],
                'order_status': row[21], 'after_sale_status': row[22],
                'refund_apply_time': row[23] or '', 'refund_agree_time': row[24] or '',
                'record_date': row[25], 'store_name': row[26], 'store_id': row[27]
            }
        return None

    def is_order_no_exists(self, order_no):
        """检查订单号是否已存在"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT id FROM refund_records WHERE order_no = ? AND IFNULL(sync_deleted, 0)=0', (order_no,))
        return cursor.fetchone() is not None

    def get_record_by_order_no(self, order_no):
        """根据订单号获取记录"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT r.id, r.order_no, r.spec_name, r.spec_code, r.reason,
                   r.real_refund_reason, r.real_refund_reason_detail, r.real_refund_reason_updated_at, r.real_refund_reason_note_hash,
                   r.quality_refund_reason, r.quality_not_cancelled_reason, r.quality_refund_reason_detail, r.quality_refund_reason_updated_at, r.quality_refund_reason_note_hash,
                   r.refund_amount, r.cancel, r.compensate, r.comp_amount, 
                   r.reject, r.reject_result, r.notes, r.order_status, r.after_sale_status, r.refund_apply_time, r.refund_agree_time, r.record_date, s.store_name, r.store_id
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE r.order_no=? AND IFNULL(r.sync_deleted, 0)=0 AND IFNULL(s.sync_deleted, 0)=0
        ''', (order_no,))
        row = cursor.fetchone()
        if row:
            return {
                'id': row[0], 'order_no': row[1], 'spec_name': row[2] or '', 'spec_code': row[3] or '',
                'reason': row[4], 'real_refund_reason': row[5] or '', 'real_refund_reason_detail': row[6] or '',
                'real_refund_reason_updated_at': row[7] or '', 'real_refund_reason_note_hash': row[8] or '',
                'quality_refund_reason': row[9] or '', 'quality_not_cancelled_reason': row[10] or '',
                'quality_refund_reason_detail': row[11] or '', 'quality_refund_reason_updated_at': row[12] or '',
                'quality_refund_reason_note_hash': row[13] or '',
                'refund_amount': row[14], 'cancel': bool(row[15]), 'compensate': bool(row[16]), 'comp_amount': row[17],
                'reject': bool(row[18]), 'reject_result': row[19], 'notes': row[20],
                'order_status': row[21], 'after_sale_status': row[22],
                'refund_apply_time': row[23] or '', 'refund_agree_time': row[24] or '',
                'record_date': row[25], 'store_name': row[26], 'store_id': row[27]
            }
        return None

    # ==================== 驳回倒计时状态管理 ====================

    def save_reject_countdown(self, order_no, store_name, current_round, end_time):
        """保存驳回倒计时状态到数据库"""
        cursor = self.conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO reject_countdown (order_no, store_name, current_round, end_time)
            VALUES (?, ?, ?, ?)
        ''', (order_no, store_name, current_round, end_time.isoformat()))
        self.conn.commit()

    def get_reject_countdown(self, order_no):
        """获取指定订单的驳回倒计时状态"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT order_no, store_name, current_round, end_time
            FROM reject_countdown
            WHERE order_no = ?
        ''', (order_no,))
        row = cursor.fetchone()
        if row:
            return {
                'order_no': row[0],
                'store_name': row[1],
                'current_round': row[2],
                'end_time': datetime.fromisoformat(row[3])
            }
        return None

    def get_all_active_reject_countdowns(self):
        """获取所有活动的驳回倒计时（未过期的）"""
        cursor = self.conn.cursor()
        now = datetime.now().isoformat()
        cursor.execute('''
            SELECT order_no, store_name, current_round, end_time
            FROM reject_countdown
            WHERE end_time > ? AND IFNULL(sync_deleted, 0)=0
        ''', (now,))
        rows = cursor.fetchall()
        result = []
        for row in rows:
            result.append({
                'order_no': row[0],
                'store_name': row[1],
                'current_round': row[2],
                'end_time': datetime.fromisoformat(row[3])
            })
        return result

    def delete_reject_countdown(self, order_no):
        """删除指定订单的驳回倒计时状态"""
        self._soft_delete('reject_countdown', 'order_no = ?', (order_no,))

    def clear_expired_reject_countdowns(self):
        """清理已过期的驳回倒计时记录"""
        now = datetime.now().isoformat()
        self._soft_delete('reject_countdown', 'end_time < ?', (now,))

    def search_records(self, order_no='', reason='全部', cancel='全部',
                       reject='全部', start_date=None, end_date=None,
                       store_name='全部', real_reason=''):
        """根据条件搜索记录，返回结果列表"""
        cursor = self.conn.cursor()
        query = '''
            SELECT r.id, r.order_no, r.spec_name, r.spec_code, r.reason,
                   r.real_refund_reason, r.real_refund_reason_detail, r.real_refund_reason_updated_at, r.real_refund_reason_note_hash,
                   r.quality_refund_reason, r.quality_not_cancelled_reason, r.quality_refund_reason_detail, r.quality_refund_reason_updated_at, r.quality_refund_reason_note_hash,
                   r.refund_amount, r.cancel, r.compensate, r.comp_amount,
                   r.reject, r.reject_result, r.notes, r.order_status, r.after_sale_status, r.refund_apply_time, r.refund_agree_time, r.record_date, s.store_name, r.store_id
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE IFNULL(r.sync_deleted, 0)=0 AND IFNULL(s.sync_deleted, 0)=0
        '''
        params = []
        if order_no:
            # 智能模糊搜索：去除输入中的空格、换行符等特殊字符，支持部分匹配
            # 清理用户输入：去除空格、换行符、制表符等
            order_no_cleaned = re.sub(r'\s+', '', order_no)  # 去除所有空白字符
            
            # 如果清理后还有内容，进行模糊搜索
            if order_no_cleaned:
                # 同时清理数据库中的订单号进行匹配
                query += ' AND REPLACE(REPLACE(r.order_no, " ", ""), "\n", "") LIKE ?'
                params.append(f'%{order_no_cleaned}%')
        if reason != '全部':
            if isinstance(reason, list) and len(reason) > 0:
                # 多选情况：使用 IN 查询
                placeholders = ','.join(['?'] * len(reason))
                query += f' AND r.reason IN ({placeholders})'
                params.extend(reason)
            else:
                # 单选情况
                query += ' AND r.reason = ?'
                params.append(reason)
        real_reason = str(real_reason or '').strip()
        if real_reason:
            query += ' AND r.real_refund_reason LIKE ?'
            params.append(f'%{real_reason}%')
        if cancel != '全部':
            query += ' AND r.cancel = ?'
            params.append(1 if cancel == '是' else 0)
        if reject != '全部':
            query += ' AND r.reject = ?'
            params.append(1 if reject == '是' else 0)
        if store_name != '全部':
            query += ' AND s.store_name = ?'
            params.append(store_name)
        if start_date:
            query += ' AND r.record_date >= ?'
            params.append(start_date)
        if end_date:
            query += ' AND r.record_date <= ?'
            params.append(end_date)
        query += ' ORDER BY r.record_date DESC, r.id DESC'
        cursor.execute(query, params)
        rows = cursor.fetchall()
        results = []
        for row in rows:
            results.append({
                'id': row[0], 'order_no': row[1], 'spec_name': row[2] or '', 'spec_code': row[3] or '',
                'reason': row[4], 'real_refund_reason': row[5] or '', 'real_refund_reason_detail': row[6] or '',
                'real_refund_reason_updated_at': row[7] or '', 'real_refund_reason_note_hash': row[8] or '',
                'quality_refund_reason': row[9] or '', 'quality_not_cancelled_reason': row[10] or '',
                'quality_refund_reason_detail': row[11] or '', 'quality_refund_reason_updated_at': row[12] or '',
                'quality_refund_reason_note_hash': row[13] or '',
                'refund_amount': row[14],
                'cancel': bool(row[15]), 'compensate': bool(row[16]), 'comp_amount': row[17],
                'reject': bool(row[18]), 'reject_result': row[19], 'notes': row[20],
                'order_status': row[21], 'after_sale_status': row[22],
                'refund_apply_time': row[23] or '', 'refund_agree_time': row[24] or '',
                'record_date': row[25], 'store_name': row[26], 'store_id': row[27]
            })
        return results

    def get_records_by_filters(self, store_id=None, start_date=None, end_date=None, reasons=None, order_no=None):
        """根据筛选条件获取记录"""
        cursor = self.conn.cursor()
        
        query = '''
            SELECT r.id, r.order_no, r.spec_name, r.spec_code, r.reason,
                   r.real_refund_reason, r.real_refund_reason_detail, r.real_refund_reason_updated_at, r.real_refund_reason_note_hash,
                   r.quality_refund_reason, r.quality_not_cancelled_reason, r.quality_refund_reason_detail, r.quality_refund_reason_updated_at, r.quality_refund_reason_note_hash,
                   r.refund_amount, r.cancel, r.compensate, r.comp_amount,
                   r.reject, r.reject_result, r.notes, r.order_status, r.after_sale_status, r.refund_apply_time, r.refund_agree_time, r.record_date, s.store_name, r.store_id
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE IFNULL(r.sync_deleted, 0)=0 AND IFNULL(s.sync_deleted, 0)=0
        '''
        params = []
        
        if store_id is not None:
            query += ' AND r.store_id = ?'
            params.append(store_id)
            
        if start_date:
            query += ' AND r.record_date >= ?'
            params.append(start_date)
            
        if end_date:
            query += ' AND r.record_date <= ?'
            params.append(end_date)
            
        if reasons:
            placeholders = ','.join(['?'] * len(reasons))
            query += f' AND r.reason IN ({placeholders})'
            params.extend(reasons)
            
        if order_no:
            query += ' AND r.order_no LIKE ?'
            params.append(f'%{order_no}%')
            
        query += ' ORDER BY r.record_date DESC, r.id DESC'
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        results = []
        for row in rows:
            results.append({
                'id': row[0], 'order_no': row[1], 'spec_name': row[2] or '', 'spec_code': row[3] or '',
                'reason': row[4], 'real_refund_reason': row[5] or '', 'real_refund_reason_detail': row[6] or '',
                'real_refund_reason_updated_at': row[7] or '', 'real_refund_reason_note_hash': row[8] or '',
                'quality_refund_reason': row[9] or '', 'quality_not_cancelled_reason': row[10] or '',
                'quality_refund_reason_detail': row[11] or '', 'quality_refund_reason_updated_at': row[12] or '',
                'quality_refund_reason_note_hash': row[13] or '',
                'refund_amount': row[14],
                'cancel': bool(row[15]), 'compensate': bool(row[16]), 'comp_amount': row[17],
                'reject': bool(row[18]), 'reject_result': row[19], 'notes': row[20],
                'order_status': row[21], 'after_sale_status': row[22],
                'refund_apply_time': row[23] or '', 'refund_agree_time': row[24] or '',
                'record_date': row[25], 'store_name': row[26], 'store_id': row[27]
            })
        return results

    def get_refund_stats_by_store(self, store_id, start_date, end_date, reasons=None):
        """获取单个店铺的退款统计"""
        cursor = self.conn.cursor()
        
        # 基础查询
        query = '''
            SELECT 
                COUNT(*) as total_count,
                SUM(refund_amount) as total_amount,
                SUM(CASE WHEN reason IN ('商品腐败、变质、包装胀气等', '商品破损/压坏', '质量问题', '大小/规格/重量等与商品描述不符', '品种/标签/图片/包装等与商品描述不符', '货物与描述不符') THEN 1 ELSE 0 END) as quality_count,
                SUM(CASE WHEN reason = '其他' THEN 1 ELSE 0 END) as other_count,
                SUM(CASE WHEN cancel = 1 AND reason IN ('商品腐败、变质、包装胀气等', '商品破损/压坏', '质量问题', '大小/规格/重量等与商品描述不符', '品种/标签/图片/包装等与商品描述不符', '货物与描述不符') THEN 1 ELSE 0 END) as canceled_quality_count,
                SUM(CASE WHEN compensate = 1 THEN comp_amount ELSE 0 END) as comp_total_amount,
                SUM(CASE WHEN reject = 1 AND reject_result = '成功' THEN 1 ELSE 0 END) as reject_success_count,
                SUM(CASE WHEN reject = 1 THEN 1 ELSE 0 END) as reject_total_count
            FROM refund_records 
            WHERE store_id = ? AND record_date BETWEEN ? AND ? AND IFNULL(sync_deleted, 0)=0
        '''
        
        params = [store_id, start_date, end_date]
        
        # 添加退款原因筛选
        if reasons:
            placeholders = ','.join(['?'] * len(reasons))
            query += f' AND reason IN ({placeholders})'
            params.extend(reasons)
        
        cursor.execute(query, params)
        row = cursor.fetchone()
        
        if not row:
            return {}
            
        total_count = row[0] or 0
        total_amount = row[1] or 0
        quality_count = row[2] or 0
        other_count = row[3] or 0
        canceled_quality_count = row[4] or 0
        comp_total_amount = row[5] or 0
        reject_success_count = row[6] or 0
        reject_total_count = row[7] or 0
        
        # 计算各种比率
        total_refund_rate = (total_count / 100) * 100 if total_count > 0 else 0
        refund_amount_ratio = (total_amount / 10000) * 100 if total_amount > 0 else 0
        apply_quality_rate = (quality_count / 100) * 100 if quality_count > 0 else 0
        actual_quality_rate = ((quality_count - canceled_quality_count) / 100) * 100 if quality_count > 0 else 0
        quality_cancel_rate = (canceled_quality_count / quality_count) * 100 if quality_count > 0 else 0
        reject_success_rate = (reject_success_count / reject_total_count) * 100 if reject_total_count > 0 else 0
        
        # 获取退款原因排名
        reason_query = '''
            SELECT reason, COUNT(*) as count
            FROM refund_records
            WHERE store_id = ? AND record_date BETWEEN ? AND ? AND IFNULL(sync_deleted, 0)=0
            GROUP BY reason
            ORDER BY count DESC
            LIMIT 1
        '''
        cursor.execute(reason_query, (store_id, start_date, end_date))
        reason_row = cursor.fetchone()
        
        top_reason = reason_row[0] if reason_row else ""
        top_reason_count = reason_row[1] if reason_row else 0
        top_reason_ratio = (top_reason_count / total_count) * 100 if total_count > 0 else 0
        
        return {
            'quality_refund_count': quality_count,
            'other_refund_count': other_count,
            'canceled_quality_count': canceled_quality_count,
            'total_refund_rate': round(total_refund_rate, 2),
            'total_refund_amount': round(total_amount, 2),
            'refund_amount_ratio': round(refund_amount_ratio, 2),
            'quality_after_sales_amount': round(quality_count * 50, 2),  # 假设平均50元
            'other_after_sales_amount': round(other_count * 30, 2),     # 假设平均30元
            'apply_quality_rate': round(apply_quality_rate, 2),
            'actual_quality_rate': round(actual_quality_rate, 2),
            'quality_cancel_rate': round(quality_cancel_rate, 2),
            'top_refund_reason': top_reason,
            'top_reason_count': top_reason_count,
            'top_reason_ratio': round(top_reason_ratio, 2),
            'comp_total_amount': round(comp_total_amount, 2),
            'reject_success_rate': round(reject_success_rate, 2)
        }

    def get_refund_stats_all_stores(self, start_date, end_date, reasons=None):
        """获取所有店铺的汇总退款统计"""
        cursor = self.conn.cursor()
        
        # 获取所有店铺的汇总数据
        query = '''
            SELECT 
                COUNT(*) as total_count,
                SUM(refund_amount) as total_amount,
                SUM(CASE WHEN reason IN ('商品腐败、变质、包装胀气等', '商品破损/压坏', '质量问题', '大小/规格/重量等与商品描述不符', '品种/标签/图片/包装等与商品描述不符', '货物与描述不符') THEN 1 ELSE 0 END) as quality_count,
                SUM(CASE WHEN reason = '其他' THEN 1 ELSE 0 END) as other_count,
                SUM(CASE WHEN cancel = 1 AND reason IN ('商品腐败、变质、包装胀气等', '商品破损/压坏', '质量问题', '大小/规格/重量等与商品描述不符', '品种/标签/图片/包装等与商品描述不符', '货物与描述不符') THEN 1 ELSE 0 END) as canceled_quality_count,
                SUM(CASE WHEN compensate = 1 THEN comp_amount ELSE 0 END) as comp_total_amount,
                SUM(CASE WHEN reject = 1 AND reject_result = '成功' THEN 1 ELSE 0 END) as reject_success_count,
                SUM(CASE WHEN reject = 1 THEN 1 ELSE 0 END) as reject_total_count
            FROM refund_records 
            WHERE record_date BETWEEN ? AND ? AND IFNULL(sync_deleted, 0)=0
        '''
        
        params = [start_date, end_date]
        
        # 添加退款原因筛选
        if reasons:
            placeholders = ','.join(['?'] * len(reasons))
            query += f' AND reason IN ({placeholders})'
            params.extend(reasons)
        
        cursor.execute(query, params)
        row = cursor.fetchone()
        
        if not row:
            return {}
            
        total_count = row[0] or 0
        total_amount = row[1] or 0
        quality_count = row[2] or 0
        other_count = row[3] or 0
        canceled_quality_count = row[4] or 0
        comp_total_amount = row[5] or 0
        reject_success_count = row[6] or 0
        reject_total_count = row[7] or 0
        
        # 计算各种比率
        total_refund_rate = (total_count / 100) * 100 if total_count > 0 else 0
        refund_amount_ratio = (total_amount / 10000) * 100 if total_amount > 0 else 0
        apply_quality_rate = (quality_count / 100) * 100 if quality_count > 0 else 0
        actual_quality_rate = ((quality_count - canceled_quality_count) / 100) * 100 if quality_count > 0 else 0
        quality_cancel_rate = (canceled_quality_count / quality_count) * 100 if quality_count > 0 else 0
        reject_success_rate = (reject_success_count / reject_total_count) * 100 if reject_total_count > 0 else 0
        
        # 获取退款原因排名
        reason_query = '''
            SELECT reason, COUNT(*) as count
            FROM refund_records
            WHERE record_date BETWEEN ? AND ? AND IFNULL(sync_deleted, 0)=0
            GROUP BY reason
            ORDER BY count DESC
            LIMIT 1
        '''
        cursor.execute(reason_query, (start_date, end_date))
        reason_row = cursor.fetchone()
        
        top_reason = reason_row[0] if reason_row else ""
        top_reason_count = reason_row[1] if reason_row else 0
        top_reason_ratio = (top_reason_count / total_count) * 100 if total_count > 0 else 0
        
        return {
            'quality_refund_count': quality_count,
            'other_refund_count': other_count,
            'canceled_quality_count': canceled_quality_count,
            'total_refund_rate': round(total_refund_rate, 2),
            'total_refund_amount': round(total_amount, 2),
            'refund_amount_ratio': round(refund_amount_ratio, 2),
            'quality_after_sales_amount': round(quality_count * 50, 2),
            'other_after_sales_amount': round(other_count * 30, 2),
            'apply_quality_rate': round(apply_quality_rate, 2),
            'actual_quality_rate': round(actual_quality_rate, 2),
            'quality_cancel_rate': round(quality_cancel_rate, 2),
            'top_refund_reason': top_reason,
            'top_reason_count': top_reason_count,
            'top_reason_ratio': round(top_reason_ratio, 2),
            'comp_total_amount': round(comp_total_amount, 2),
            'reject_success_rate': round(reject_success_rate, 2)
        }

    def save_api_config(self, api_url, api_key, model):
        """保存API配置"""
        cursor = self.conn.cursor()
        
        # 检查是否已有配置
        cursor.execute('SELECT id FROM api_config LIMIT 1')
        existing_config = cursor.fetchone()
        
        if existing_config:
            # 更新现有配置
            cursor.execute('''
                UPDATE api_config 
                SET api_url=?, api_key=?, model=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (api_url, api_key, model, existing_config[0]))
        else:
            # 插入新配置
            cursor.execute('''
                INSERT INTO api_config (api_url, api_key, model)
                VALUES (?, ?, ?)
            ''', (api_url, api_key, model))
        
        self.conn.commit()
        return True

    def load_api_config(self):
        """加载API配置"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT api_url, api_key, model FROM api_config LIMIT 1')
        row = cursor.fetchone()
        
        if row:
            return {
                'api_url': row[0],
                'api_key': row[1],
                'model': row[2]
            }
        else:
            # 返回默认配置
            return {
                'api_url': 'https://api.deepseek.com/v1/chat/completions',
                'api_key': '',
                'model': 'deepseek-chat'
            }


class CloudSyncDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.db = parent.db
        self.setWindowTitle("云同步")
        self.setMinimumWidth(520)
        self._init_ui()
        self.load_config()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.secret_id_edit = QLineEdit()
        self.secret_key_edit = QLineEdit()
        self.secret_key_edit.setEchoMode(QLineEdit.Password)
        self.bucket_edit = QLineEdit()
        self.region_edit = QLineEdit()
        self.prefix_edit = QLineEdit()

        self.secret_id_edit.setPlaceholderText("SecretId")
        self.secret_key_edit.setPlaceholderText("SecretKey")
        self.bucket_edit.setPlaceholderText("bucketname-appid")
        self.region_edit.setPlaceholderText("ap-guangzhou")
        self.prefix_edit.setPlaceholderText("shouhou-sync/")

        form.addRow("SecretId：", self.secret_id_edit)
        form.addRow("SecretKey：", self.secret_key_edit)
        form.addRow("Bucket：", self.bucket_edit)
        form.addRow("Region：", self.region_edit)
        form.addRow("云端前缀：", self.prefix_edit)
        layout.addLayout(form)

        self.status_label = QLabel("配置只保存在本机，用于手动上传和下载增量数据。")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)

        button_layout = QHBoxLayout()
        self.save_btn = QPushButton("保存配置")
        self.test_btn = QPushButton("测试连接")
        self.upload_btn = QPushButton("上传增量")
        self.download_btn = QPushButton("下载增量")
        self.close_btn = QPushButton("关闭")

        for button in [self.save_btn, self.test_btn, self.upload_btn, self.download_btn, self.close_btn]:
            button.setMinimumHeight(32)
            button_layout.addWidget(button)

        layout.addLayout(button_layout)
        self.save_btn.clicked.connect(self.save_config)
        self.test_btn.clicked.connect(self.test_connection)
        self.upload_btn.clicked.connect(self.upload_incremental)
        self.download_btn.clicked.connect(self.download_incremental)
        self.close_btn.clicked.connect(self.reject)

    def load_config(self):
        config = self.db.load_cloud_sync_config()
        self.secret_id_edit.setText(config.get('secret_id', ''))
        self.secret_key_edit.setText(config.get('secret_key', ''))
        self.bucket_edit.setText(config.get('bucket', ''))
        self.region_edit.setText(config.get('region', ''))
        self.prefix_edit.setText(config.get('prefix', 'shouhou-sync/'))

    def collect_config(self):
        prefix = self.prefix_edit.text().strip() or 'shouhou-sync/'
        return {
            'secret_id': self.secret_id_edit.text().strip(),
            'secret_key': self.secret_key_edit.text().strip(),
            'bucket': self.bucket_edit.text().strip(),
            'region': self.region_edit.text().strip(),
            'prefix': prefix,
        }

    def save_config(self):
        self.db.save_cloud_sync_config(self.collect_config())
        self.status_label.setText("配置已保存。")
        QMessageBox.information(self, "云同步", "配置已保存")

    def _service(self):
        config = self.collect_config()
        self.db.save_cloud_sync_config(config)
        return CloudSyncService(self.db, config)

    def _set_busy(self, busy):
        for button in [self.save_btn, self.test_btn, self.upload_btn, self.download_btn, self.close_btn]:
            button.setEnabled(not busy)
        QApplication.processEvents()

    def _run_with_progress(self, title, message, operation):
        progress = QProgressDialog(message, None, 0, 0, self)
        progress.setWindowTitle(title)
        progress.setWindowModality(Qt.ApplicationModal)
        progress.setMinimumDuration(0)
        progress.show()
        self._set_busy(True)
        try:
            QApplication.processEvents()
            return operation()
        finally:
            progress.close()
            self._set_busy(False)

    def test_connection(self):
        try:
            self._run_with_progress("云同步", "正在测试腾讯云 COS 连接...", lambda: self._service().test_connection())
            self.status_label.setText("连接测试成功。")
            QMessageBox.information(self, "云同步", "连接测试成功")
        except Exception as exc:
            self.status_label.setText(f"连接失败：{exc}")
            QMessageBox.warning(self, "云同步", f"连接失败：{exc}")

    def upload_incremental(self):
        try:
            result = self._run_with_progress("云同步", "正在上传增量数据...", lambda: self._service().upload_incremental())
            uploaded = result.get('uploaded', 0)
            deduped = result.get('deduped', 0)
            index_count = result.get('index_count', 0)
            snapshot_count = result.get('snapshot_count', 0)
            message = (
                f"上传完成：{uploaded} 条变更，清理重复 {deduped} 条，云端索引 {index_count} 条，云端快照 {snapshot_count} 条。"
                if uploaded or deduped or index_count or snapshot_count
                else "没有需要上传的增量数据。"
            )
            self.status_label.setText(message)
            QMessageBox.information(self, "云同步", message)
        except Exception as exc:
            self.status_label.setText(f"上传失败：{exc}")
            QMessageBox.warning(self, "云同步", f"上传失败：{exc}")

    def download_incremental(self):
        try:
            result = self._run_with_progress("云同步", "正在下载并合并增量数据...", lambda: self._service().download_incremental())
            downloaded = result.get('downloaded', 0)
            packages = result.get('packages', 0)
            snapshot_restored = result.get('snapshot_restored', 0)
            cloud_missing_deleted = result.get('cloud_missing_deleted', 0)
            deduped = result.get('deduped', 0)
            self.parent.refresh_after_cloud_sync()
            message = (
                f"下载完成：合并 {downloaded} 条变更，处理 {packages} 个增量包，"
                f"快照恢复 {snapshot_restored} 条，云端缺失删除 {cloud_missing_deleted} 条，清理重复 {deduped} 条。"
            )
            if downloaded == 0 and snapshot_restored == 0 and cloud_missing_deleted == 0 and deduped == 0:
                message = "没有需要下载的增量数据。"
                if not result.get('index_found', False):
                    message += " 云端暂无订单索引，未执行云端缺失删除。"
                elif not result.get('snapshot_found', False):
                    message += " 云端暂无订单快照，请在数据完整的电脑上传一次。"
            self.status_label.setText(message)
            QMessageBox.information(self, "云同步", message)
        except Exception as exc:
            self.status_label.setText(f"下载失败：{exc}")
            QMessageBox.warning(self, "云同步", f"下载失败：{exc}")

# ---------------------------- AI分析独立窗口 ---------------------------------
class AIAnalysisWindow(QWidget):
    def __init__(self, panel_widget, parent=None):
        super().__init__(None)
        self.panel_widget = panel_widget
        self.owner = parent
        self._manual_close_in_progress = False
        self.setWindowTitle("AI分析与图表数据")
        self.setWindowFlag(Qt.Window, True)
        self.resize(960, 720)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.addWidget(self.panel_widget)

    def closeEvent(self, event):
        if event.spontaneous():
            self._manual_close_in_progress = True
            event.accept()
            return
        event.ignore()

    def hideEvent(self, event):
        super().hideEvent(event)
        if not self._manual_close_in_progress:
            QTimer.singleShot(0, self._restore_visibility)

    def showEvent(self, event):
        self._manual_close_in_progress = False
        super().showEvent(event)

    def _restore_visibility(self):
        if self.isVisible() or self._manual_close_in_progress:
            return
        self.show()
        self.raise_()
        self.activateWindow()


class DailyWorkSummaryDialog(QDialog):
    """当前筛选范围内的日报/工作总结窗口。"""

    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self.latest_result = ""
        self.setWindowTitle("工作总结")
        self.resize(980, 760)
        self._setup_ui()
        self.refresh_preview()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        self.filter_label = QLabel("")
        self.filter_label.setWordWrap(True)
        self.filter_label.setStyleSheet(
            "font-family: 'Microsoft YaHei'; font-size: 13px; font-weight: bold; "
            "color: #1F2937; background-color: #EEF5FF; border: 1px solid #C9DAF8; "
            "border-radius: 5px; padding: 6px 8px;"
        )
        layout.addWidget(self.filter_label)

        text_style = """
            QTextEdit {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 5px;
                padding: 8px;
                line-height: 1.5;
            }
        """

        manual_label = QLabel("今日补充工作内容")
        manual_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px; font-weight: bold;")
        layout.addWidget(manual_label)

        self.manual_input = QTextEdit()
        self.manual_input.setPlaceholderText("可以填写今天完成的工作、领航员标红问题、需要晚班客服跟踪的事项等。")
        self.manual_input.setFont(QFont("Microsoft YaHei", 10))
        self.manual_input.setMinimumHeight(100)
        self.manual_input.textChanged.connect(self.refresh_preview)
        layout.addWidget(self.manual_input, 1)

        result_label = QLabel("AI专业总结")
        result_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px; font-weight: bold;")
        layout.addWidget(result_label)

        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        self.result_text.setFont(QFont("Microsoft YaHei", 10))
        self.result_text.setStyleSheet(text_style)
        self.result_text.setMarkdown("点击“AI生成专业总结”后，将在这里显示日报/工作总结。")
        layout.addWidget(self.result_text, 5)

        button_layout = QHBoxLayout()
        self.generate_btn = QPushButton("AI生成专业总结")
        self.copy_btn = QPushButton("复制总结")
        self.debug_prompt_btn = QPushButton("调试提示词")
        close_btn = QPushButton("关闭")

        self.generate_btn.clicked.connect(self.generate_summary)
        self.copy_btn.clicked.connect(self.copy_result)
        self.debug_prompt_btn.clicked.connect(self.show_debug_prompt)
        close_btn.clicked.connect(self.accept)

        button_layout.addWidget(self.generate_btn)
        button_layout.addWidget(self.copy_btn)
        button_layout.addWidget(self.debug_prompt_btn)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)

    def refresh_preview(self):
        if not self.main_window:
            return
        self.filter_label.setText(f"当前筛选：{self.main_window.get_current_filter_summary_text()}")

    def generate_summary(self):
        if self.main_window:
            self.main_window.generate_daily_work_summary_for_dialog(self)

    def show_debug_prompt(self):
        if self.main_window:
            self.main_window.show_daily_work_summary_debug_prompt(self)

    def set_result(self, markdown_text):
        self.latest_result = str(markdown_text or "")
        self.result_text.setMarkdown(self.latest_result or "AI未返回内容。")

    def copy_result(self):
        text = self.latest_result or self.result_text.toPlainText()
        if not text.strip():
            QMessageBox.information(self, "提示", "当前没有可复制的总结内容")
            return
        QApplication.clipboard().setText(self._plain_text_for_wechat(text))
        QMessageBox.information(self, "复制成功", "工作总结已复制到剪贴板")

    @staticmethod
    def _plain_text_for_wechat(text):
        """把Markdown结果清洗成适合微信聊天粘贴的纯文本。"""
        cleaned_lines = []
        for raw_line in str(text or "").splitlines():
            line = raw_line.strip()
            if not line:
                cleaned_lines.append("")
                continue
            line = re.sub(r'^\s{0,3}#{1,6}\s*', '', line)
            line = re.sub(r'^\s*[-*+]\s+', '', line)
            line = re.sub(r'^\s*\d+[.)、]\s+', '', line)
            line = re.sub(r'\*\*(.*?)\*\*', r'\1', line)
            line = re.sub(r'__(.*?)__', r'\1', line)
            line = re.sub(r'\*(.*?)\*', r'\1', line)
            line = re.sub(r'_(.*?)_', r'\1', line)
            line = re.sub(r'`([^`]*)`', r'\1', line)
            if '|' in line:
                line = re.sub(r'\s*\|\s*', '  ', line).strip()
            if re.fullmatch(r'[-:\s|]+', line):
                continue
            cleaned_lines.append(line)

        text = "\n".join(cleaned_lines)
        text = re.sub(r'\n{3,}', '\n\n', text)
        return text.strip()

    def get_user_text(self):
        return self.manual_input.toPlainText().strip()


# ---------------------------- 主窗口类 ---------------------------------
class RefundManager(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.current_record_id = None  # 当前编辑的记录ID（用于更新）
        self.highlighted_orders = set()  # 刚导入需要高亮的订单号集合
        self.selected_reasons = set()  # 多选退款原因集合
        self.store_settings = {}  # 店铺基本信息设置
        self._last_import_undo_data = None  # 最近一次导入的撤销信息
        
        # AI分析器
        self.ai_analyzer = AIAnalyzer()
        
        # 加载API配置
        self.load_api_config()
        
        # 性能优化：初始化定时器（避免重复创建）
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        
        # 性能优化：数据缓存
        self._cached_records = None  # 缓存搜索结果
        self._last_search_params = None  # 上次搜索参数
        self.latest_summary_snapshot = None
        self.latest_summary_history_id = None
        self.latest_summary_debug_data = {}
        self.db.ensure_default_real_refund_reason_categories(self._get_default_real_reason_category_configs())
        self._merge_default_real_reason_keywords()
        
        self.init_ui()
        # 初始化店铺设置
        self.load_store_settings()
        self.load_stores()
        self._auto_fill_missing_spec_codes_from_names()
        self.load_table_data()
        self.setup_shortcuts()
        
        # ==================== 系统托盘功能 ====================
        self._init_system_tray()
        
        # ==================== 自动更新检查 ====================
        # 程序启动后延迟检查更新
        if ENABLE_AUTO_UPDATE:
            self.update_checker = UpdateChecker()
            self.update_checker.update_available.connect(self.show_update_dialog)
            QTimer.singleShot(UPDATE_CHECK_DELAY * 1000, self.check_for_updates)
        
        # ==================== 驳回流程管理 ====================
        self.reject_manager = RejectProcessManager(self.db)
        self.reject_manager.countdown_finished.connect(self.on_reject_countdown_finished)
        self.reject_manager.countdown_updated.connect(self.on_reject_countdown_updated)
        self.reject_manager.reminder_48h_triggered.connect(self.show_48h_reminder)
        
        # 从数据库恢复倒计时状态（软件重启后）
        self.reject_manager.restore_countdowns_from_db()
        
        # 更新表格显示恢复的倒计时
        self.restore_reject_display_from_db()
    
    def check_for_updates(self):
        """检查更新"""
        try:
            self.update_checker.check_for_updates()
        except Exception as e:
            print(f"启动更新检查时出错: {e}")
    
    def show_update_dialog(self, update_info):
        """显示更新对话框"""
        try:
            dialog = UpdateDialog(
                CURRENT_VERSION,
                update_info['version'],
                update_info['notes'],
                update_info['url'],
                parent=self
            )
            dialog.exec_()
        except Exception as e:
            QMessageBox.warning(self, "更新检查", f"显示更新对话框时出错: {e}")

    def init_ui(self):
        self.setWindowTitle(f"电商售后品质退款管理工具 v{CURRENT_VERSION}")
        # 【窗口默认尺寸设置】第451行 - 修改这里的数字来改变窗口默认大小
        self.resize(1700, 950)  # 窗口宽度1700像素，高度950像素
        self.setMinimumSize(0, 0)  # 设置窗口最小尺寸，允许适当缩小
        
        # 创建菜单栏
        self._create_menu_bar()
        
        # 应用护眼配色样式表
        self.apply_stylesheet()

        # 中央控件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局：三列分割器（左侧栏 / 中间主区域 / 右侧统计栏）
        main_splitter = QSplitter(Qt.Horizontal)
        main_splitter.setChildrenCollapsible(False)
        main_splitter.setHandleWidth(8)
        main_splitter.setStretchFactor(0, 0)
        main_splitter.setStretchFactor(1, 1)
        main_splitter.setStretchFactor(2, 0)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(6, 6, 6, 6)
        main_layout.addWidget(main_splitter)

        # 左侧栏：搜索筛选 + 快捷日期
        left_sidebar_splitter = QSplitter(Qt.Vertical)
        left_sidebar_splitter.setChildrenCollapsible(False)
        left_sidebar_splitter.setHandleWidth(6)
        left_sidebar_splitter.setStretchFactor(0, 1)
        left_sidebar_splitter.setStretchFactor(1, 0)

        # 中间列：主要功能区 + 订单记录表格
        center_splitter = QSplitter(Qt.Vertical)
        center_splitter.setChildrenCollapsible(False)
        center_splitter.setHandleWidth(8)
        center_splitter.setStretchFactor(0, 0)
        center_splitter.setStretchFactor(1, 1)

        # 保存分割器引用，用于固定默认尺寸
        self.main_splitter = main_splitter
        self.top_splitter = center_splitter
        self.bottom_splitter = left_sidebar_splitter
        
        # 左上角：信息录入区（使用UI文件加载）
        self.input_panel = QGroupBox()
        loadUi(get_resource_path("input_panel.ui"), self.input_panel)
        self.input_panel.setTitle("主要功能区")
        
        # 设置对象名称，用于样式表选择器
        self.input_panel.setObjectName("InputPanel")
        
        # 应用多巴胺配色方案
        self._apply_dopamine_styles()
        
        # 连接信号和槽
        self._connect_input_signals()
        
        # 连接导入导出按钮
        self._connect_import_export_buttons()
        self._optimize_input_panel_layout()

        # AI分析与图表数据窗口（按钮入口保留在主要功能区，窗口不占主布局）
        self.ai_chart_group = QGroupBox("AI分析与图表数据")
        ai_chart_layout = QVBoxLayout()
        self.ai_chart_group.setLayout(ai_chart_layout)
        self.ai_tab_widget = QTabWidget()
        ai_chart_layout.addWidget(self.ai_tab_widget)

        report_tab = QWidget()
        report_layout = QVBoxLayout(report_tab)
        report_layout.setContentsMargins(6, 6, 6, 6)
        
        # AI分析功能区域
        ai_analysis_layout = QHBoxLayout()
        
        # AI分析按钮
        self.ai_analyze_btn = QPushButton("AI分析")
        self.ai_analyze_btn.setStyleSheet("""
            QPushButton {
                font-size: 14px; 
                padding: 6px 12px;
                background-color: #9C27B0;
                color: white;
                border: 1px solid #7B1FA2;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
            QPushButton:pressed {
                background-color: #6A1B9A;
            }
        """)
        self.ai_analyze_btn.clicked.connect(self.ai_analyze_data)
        ai_analysis_layout.addWidget(self.ai_analyze_btn)

        self.orange_button_style = """
            QPushButton {
                font-size: 14px;
                padding: 6px 12px;
                background-color: #FF9800;
                color: white;
                border: 1px solid #F57C00;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
            QPushButton:pressed {
                background-color: #EF6C00;
            }
        """

        # API设置按钮
        self.api_settings_btn = QPushButton("API设置")
        self.api_settings_btn.setStyleSheet("""
            QPushButton {
                font-size: 14px; 
                padding: 6px 12px;
                background-color: #2196F3;
                color: white;
                border: 1px solid #1976D2;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
        """)
        self.api_settings_btn.clicked.connect(self.show_api_settings_dialog)
        ai_analysis_layout.addWidget(self.api_settings_btn)
        
        ai_analysis_layout.addStretch()
        report_layout.addLayout(ai_analysis_layout)
        
        # 图表区域
        self.chart_widget = ChartWidget(self, self.db)
        report_layout.addWidget(self.chart_widget, 1)
        self.ai_tab_widget.addTab(report_tab, "报告")

        summary_tab = QWidget()
        summary_layout = QVBoxLayout(summary_tab)
        summary_layout.setContentsMargins(6, 6, 6, 6)

        summary_action_layout = QHBoxLayout()
        self.summary_analyze_btn = QPushButton("生成总结")
        self.summary_analyze_btn.setStyleSheet(self.ai_analyze_btn.styleSheet())
        self.summary_analyze_btn.clicked.connect(self.generate_summary_analysis)
        summary_action_layout.addWidget(self.summary_analyze_btn)

        self.real_reason_category_btn = QPushButton("本地分类管理")
        self.real_reason_category_btn.setStyleSheet(self.api_settings_btn.styleSheet())
        self.real_reason_category_btn.clicked.connect(self.open_local_reason_category_manager)
        summary_action_layout.addWidget(self.real_reason_category_btn)

        self.real_reason_assign_btn = QPushButton("本地归因")
        self.real_reason_assign_btn.setStyleSheet(self.ai_analyze_btn.styleSheet())
        self.real_reason_assign_btn.clicked.connect(self.assign_real_refund_reasons)
        summary_action_layout.addWidget(self.real_reason_assign_btn)

        self.real_reason_manual_btn = QPushButton("手动归因")
        self.real_reason_manual_btn.setStyleSheet(self.api_settings_btn.styleSheet())
        self.real_reason_manual_btn.clicked.connect(self.open_manual_real_reason_assignment)
        summary_action_layout.addWidget(self.real_reason_manual_btn)

        self.real_reason_view_btn = QPushButton("查看品质退款")
        self.real_reason_view_btn.setStyleSheet(self.orange_button_style)
        self.real_reason_view_btn.clicked.connect(self.show_quality_refund_reason_view)
        summary_action_layout.addWidget(self.real_reason_view_btn)

        self.current_range_reason_btn = QPushButton("当前范围归因")
        self.current_range_reason_btn.setStyleSheet(self.orange_button_style)
        self.current_range_reason_btn.clicked.connect(self.open_current_range_reason_assignment)
        summary_action_layout.addWidget(self.current_range_reason_btn)

        self.summary_export_btn = QPushButton("导出总结")
        self.summary_export_btn.setStyleSheet(self.api_settings_btn.styleSheet())
        self.summary_export_btn.clicked.connect(self.export_summary_excel)
        summary_action_layout.addWidget(self.summary_export_btn)

        self.summary_history_btn = QPushButton("历史记录")
        self.summary_history_btn.setStyleSheet(self.orange_button_style)
        self.summary_history_btn.clicked.connect(self.open_summary_history)
        summary_action_layout.addWidget(self.summary_history_btn)

        summary_action_layout.addStretch()
        summary_layout.addLayout(summary_action_layout)

        self.summary_result_text = QTextEdit()
        self.summary_result_text.setReadOnly(True)
        self.summary_result_text.setFont(QFont("Microsoft YaHei", 10))
        self.summary_result_text.setStyleSheet("""
            QTextEdit {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 5px;
                padding: 10px;
                line-height: 1.6;
            }
        """)
        self.summary_result_text.setMarkdown(
            "## 总结分析\n\n先点“本地分类管理”维护分类和关键词，再点“本地归因”对当前筛选范围内备注做本地关键词匹配。之后“生成总结”和导出会优先复用本地已归因结果。"
        )
        summary_layout.addWidget(self.summary_result_text, 1)
        self.latest_summary_snapshot = None
        self.last_real_reason_category_debug = {}
        self.last_real_reason_assignment_debug = {}
        self.ai_tab_widget.addTab(summary_tab, "总结")
        
        # 右侧：店铺信息区
        store_info_group = QGroupBox("店铺信息与统计")
        store_info_layout = QVBoxLayout()
        store_info_layout.setContentsMargins(6, 6, 6, 6)
        store_info_layout.setSpacing(4)
        store_info_group.setLayout(store_info_layout)

        stats_column = QVBoxLayout()
        stats_column.setContentsMargins(0, 0, 0, 0)
        stats_column.setSpacing(4)

        base_cell_style = (
            "font-family: 'Microsoft YaHei'; font-size: 11px; font-weight: bold; "
            "color: #2D3748; background-color: #f8f9fa; border: 1px solid #d8dee6; "
            "border-radius: 5px; padding: 2px 6px;"
        )
        strong_cell_style = (
            "font-family: 'Microsoft YaHei'; font-size: 11px; font-weight: bold; "
            "color: #1F2937; background-color: #EEF5FF; border: 1px solid #C9DAF8; "
            "border-radius: 5px; padding: 2px 6px;"
        )
        budget_cell_style = (
            "font-family: 'Microsoft YaHei'; font-size: 11px; font-weight: bold; "
            "color: #D64545; background-color: #FFF5F5; border: 1px solid #F2CACA; "
            "border-radius: 5px; padding: 2px 6px;"
        )

        def add_info_cell(initial_text, attr_name, style=base_cell_style):
            value_label = QLabel(initial_text)
            value_label.setStyleSheet(style)
            value_label.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            value_label.setWordWrap(False)
            value_label.setMinimumHeight(26)
            value_label.setMaximumHeight(26)
            setattr(self, attr_name, value_label)
            stats_column.addWidget(value_label)
            return value_label

        add_info_cell("当前店铺：未选择", "current_store_label", strong_cell_style)
        add_info_cell("当前范围单量：0", "orders_label")
        add_info_cell("当前范围销售额：¥0.0", "sales_label")
        add_info_cell("退款预算：¥0.0", "daily_budget_remaining_label", budget_cell_style)
        add_info_cell("品质退款：0单", "quality_refund_count_label")
        add_info_cell("其他退款：0单", "other_refund_count_label")
        add_info_cell("撤销品质：0单", "canceled_quality_count_label")
        add_info_cell("总退款率：0.00%", "total_refund_rate_label")
        add_info_cell("售后总额：¥0.0", "total_after_sales_label")
        add_info_cell("金额占比：0.00%", "refund_ratio_label")
        add_info_cell("品质售后：¥0.0", "quality_after_sales_amount_label")
        add_info_cell("其他售后：¥0.0", "other_after_sales_amount_label")
        add_info_cell("申请品质率：0.00%", "quality_apply_rate_label")
        add_info_cell("实际品质率：0.00%", "quality_actual_rate_label")
        add_info_cell("撤销率：0.00%", "quality_cancel_rate_label")
        add_info_cell("最多原因：无数据", "top_refund_reason_label")
        add_info_cell("出现次数：0", "top_reason_count_label")
        add_info_cell("占比：0.0%", "top_reason_ratio_label")
        store_info_layout.addLayout(stats_column)

        self.store_settings_btn = QPushButton("店铺设置")
        self.store_settings_btn.setStyleSheet("""
            QPushButton {
                font-family: 'Microsoft YaHei';
                font-size: 11px;
                border: 1px solid #1976D2;
                border-radius: 6px;
                background-color: #2196F3;
                color: white;
                font-weight: bold;
                min-height: 30px;
                padding: 4px 8px;
            }
            QPushButton:hover {
                background-color: #1976D2;
                border-color: #1565C0;
            }
            QPushButton:pressed {
                background-color: #1565C0;
                border-color: #0D47A1;
            }
        """)
        self.store_settings_btn.setToolTip("店铺基本信息设置")
        self.store_settings_btn.clicked.connect(self.open_store_settings)
        store_info_layout.addWidget(self.store_settings_btn)

        self.daily_work_summary_btn = QPushButton("工作总结")
        self.daily_work_summary_btn.setStyleSheet(self.orange_button_style)
        self.daily_work_summary_btn.setToolTip("根据当前筛选范围生成日报/工作总结")
        self.daily_work_summary_btn.clicked.connect(self.show_daily_work_summary_dialog)
        store_info_layout.addWidget(self.daily_work_summary_btn)

        store_info_layout.addStretch()
        
        # 左下角：搜索筛选区 - 使用UI文件
        search_group = loadUi(get_resource_path("search_panel.ui"))
        
        # 获取UI文件中的控件引用
        self.search_store_combo = search_group.findChild(QComboBox, "search_store_combo")
        self.search_order_edit = search_group.findChild(QLineEdit, "search_order_edit")
        self.search_reason_btn = search_group.findChild(QPushButton, "search_reason_btn")
        self.search_real_reason_edit = search_group.findChild(QLineEdit, "search_real_reason_edit")
        self.start_date_edit = search_group.findChild(QDateEdit, "start_date_edit")
        self.end_date_edit = search_group.findChild(QDateEdit, "end_date_edit")
        self.search_cancel_combo = search_group.findChild(QComboBox, "search_cancel_combo")
        self.search_reject_combo = search_group.findChild(QComboBox, "search_reject_combo")
        
        # 获取按钮引用
        reset_btn = search_group.findChild(QPushButton, "reset_btn")
        show_all_btn = search_group.findChild(QPushButton, "show_all_btn")
        
        # 加载独立的快捷日期UI文件
        quick_date_group = loadUi(get_resource_path("quick_date_panel.ui"))
        today_btn = quick_date_group.findChild(QPushButton, "today_btn")
        yesterday_btn = quick_date_group.findChild(QPushButton, "yesterday_btn")
        prev_day_btn = quick_date_group.findChild(QPushButton, "prev_day_btn")
        week_btn = quick_date_group.findChild(QPushButton, "week_btn")
        month_btn = quick_date_group.findChild(QPushButton, "month_btn")
        full_week_btn = quick_date_group.findChild(QPushButton, "full_week_btn")
        full_month_btn = quick_date_group.findChild(QPushButton, "full_month_btn")
        all_time_btn = quick_date_group.findChild(QPushButton, "all_time_btn")
        
        # 设置控件初始值
        self.search_store_combo.addItem("全部")
        
        def on_store_changed(store_name):
            # 加载对应店铺的设置
            self.load_store_settings()
            # 触发搜索更新
            self.on_search_changed()
        
        self.search_store_combo.currentTextChanged.connect(on_store_changed)
        
        # 设置订单号输入框
        self.search_order_edit.textChanged.connect(self.on_search_changed)
        self.search_order_edit.mousePressEvent = self.search_order_mouse_press

        if self.search_real_reason_edit is not None:
            self.search_real_reason_edit.textChanged.connect(self.on_search_changed)
        
        # 设置退款原因多选控件
        reasons = ["商品腐败、变质、包装胀气等", "商品破损/压坏", "质量问题", "大小/规格/重量等与商品描述不符", "品种/标签/图片/包装等与商品描述不符", "货物与描述不符", "生产日期/保质期与商品描述不符", "其他"]
        self.search_reason_dropdown = MultiSelectComboBox()
        self.search_reason_dropdown.addItems(reasons)
        self.search_reason_dropdown.itemsChanged.connect(self.on_search_changed)
        self.search_reason_dropdown.setMinimumHeight(32)

        # 找到退款原因按钮所在的位置，将其替换为多选控件
        search_group.layout().removeWidget(self.search_reason_btn)
        self.search_reason_btn.setParent(None)

        # 添加多选控件到新的单列布局位置
        search_group.layout().addWidget(self.search_reason_dropdown, 5, 0)
                
        print(f"[DEBUG] 搜索筛选区退款原因多选控件已设置，选项数量: {len(reasons)}")
        
        # 设置日期选择器
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate())
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.start_date_edit.dateChanged.connect(self.on_search_changed)
        
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.end_date_edit.dateChanged.connect(self.on_search_changed)
        
        # 设置其他筛选条件
        self.search_cancel_combo.addItems(["全部", "是", "否"])
        self.search_cancel_combo.currentTextChanged.connect(self.on_search_changed)

        self.search_reject_combo.addItems(["全部", "是", "否"])
        self.search_reject_combo.currentTextChanged.connect(self.on_search_changed)
        
        # 连接按钮信号
        reset_btn.clicked.connect(self.reset_search)
        show_all_btn.clicked.connect(self.show_all_records)
        today_btn.clicked.connect(lambda: self.set_quick_date(0))
        yesterday_btn.clicked.connect(lambda: self.set_quick_date(1))
        prev_day_btn.clicked.connect(self.previous_day)
        week_btn.clicked.connect(lambda: self.set_quick_date(7))
        month_btn.clicked.connect(lambda: self.set_quick_date(30))
        full_week_btn.clicked.connect(self.set_last_full_week)
        full_month_btn.clicked.connect(self.set_last_full_month)
        all_time_btn.clicked.connect(self.show_all_time)
        # 右下角：订单记录表格 - 使用UI文件
        table_group = loadUi(get_resource_path("table_panel.ui"))
        
        # 获取UI文件中的表格引用
        self.table = table_group.findChild(QTableWidget, "order_table")
        
        # 获取UI文件中的调试标签引用
        self.debug_label = table_group.findChild(QLabel, "debug_label")
        
        # 设置表格基本属性
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels(["店铺名称", "订单号", "规格编码", "退款原因", "退款金额", "撤销", "打款补偿", "补偿金额", "驳回", "驳回结果", "登记日期", "备注"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)  # 设置扩展选择模式，支持多选和Ctrl+A
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 禁用编辑，使用双击切换功能
        
        # 设置列宽自适应模式
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)  # 设置为交互模式
        
        # 为订单号、退款原因列设置特殊拉伸模式，确保字符显示完整
        self.table.setColumnWidth(1, 200)  # 订单号列设置较宽宽度
        self.table.setColumnWidth(2, 75)   # 规格编码
        self.table.setColumnWidth(3, 245)  # 退款原因列设置较宽宽度
        
        # 其他列使用默认宽度
        self.table.setColumnWidth(0, 120)  # 店铺名称
        self.table.setColumnWidth(4, 95)   # 退款金额
        self.table.setColumnWidth(5, 57)   # 撤销
        self.table.setColumnWidth(6, 77)   # 打款补偿
        self.table.setColumnWidth(7, 92)   # 补偿金额
        self.table.setColumnWidth(8, 58)   # 驳回
        self.table.setColumnWidth(9, 100)  # 驳回结果
        self.table.setColumnWidth(10, 96)  # 登记日期
        
        # 设置列宽调整策略
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 订单号：根据内容调整
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 规格编码：根据内容调整
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # 退款原因：根据内容调整
        header.setSectionResizeMode(11, QHeaderView.Stretch)  # 备注：完全自动拉伸
        
        # 设置自定义的编辑检查函数
        self.table.setItemDelegate(CustomItemDelegate(self))
        self.table.itemDoubleClicked.connect(self.on_item_double_clicked)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        # 连接单元格编辑完成信号
        self.table.cellChanged.connect(self.on_cell_changed)
        
        # AI分析窗口独立显示，不占用主窗口主布局
        self.ai_window = AIAnalysisWindow(self.ai_chart_group, self)

        # 将区域添加到新的三列布局中
        left_sidebar_splitter.addWidget(search_group)
        left_sidebar_splitter.addWidget(quick_date_group)

        center_splitter.addWidget(self.input_panel)
        center_splitter.addWidget(table_group)

        main_splitter.addWidget(left_sidebar_splitter)
        main_splitter.addWidget(center_splitter)
        main_splitter.addWidget(store_info_group)

        # 设置各板块的最小尺寸，确保布局稳定
        main_splitter.setMinimumSize(1000, 700)
        left_sidebar_splitter.setMinimumWidth(138)
        left_sidebar_splitter.setMaximumWidth(138)
        left_sidebar_splitter.setSizes([420, 320])
        center_splitter.setMinimumWidth(760)
        store_info_group.setMinimumWidth(170)
        store_info_group.setMaximumWidth(170)
        self.input_panel.setMinimumWidth(720)
        self.input_panel.setFixedHeight(240)
        table_group.setMinimumSize(720, 320)
        search_group.setMinimumWidth(138)
        search_group.setMaximumWidth(138)
        search_group.setMinimumHeight(390)
        quick_date_group.setMinimumWidth(138)
        quick_date_group.setMaximumWidth(138)
        quick_date_group.setMinimumHeight(320)

        # 底部状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # 加载保存的界面设置
        self.load_window_settings()
        center_splitter.setSizes([240, 2000])

        # 初始化店铺信息下拉框
        self.load_store_info_combo()

    def _create_menu_bar(self):
        """创建菜单栏"""
        # 创建帮助菜单
        help_menu = QMenu("帮助", self)
        
        # 连接检测菜单项
        connection_check_action = QAction("连接检测", self)
        connection_check_action.triggered.connect(self.show_help_dialog)
        help_menu.addAction(connection_check_action)
        
        # 检查更新菜单项
        check_update_action = QAction("检查更新", self)
        check_update_action.triggered.connect(self.manual_check_update)
        help_menu.addAction(check_update_action)
        
        # 分隔线
        help_menu.addSeparator()
        
        # 关于菜单项
        about_action = QAction("关于", self)
        about_action.triggered.connect(self.show_about_dialog)
        help_menu.addAction(about_action)
        
        # 将帮助菜单添加到菜单栏
        self.menuBar().addMenu(help_menu)
    
    def manual_check_update(self):
        """手动检查更新"""
        try:
            # 创建进度对话框
            progress = QProgressDialog("正在检查更新...", None, 0, 0, self)
            progress.setWindowTitle("检查更新")
            progress.setWindowModality(Qt.WindowModal)
            progress.setCancelButton(None)
            progress.show()
            
            # 创建更新检查器
            self.manual_update_checker = UpdateChecker()
            self.manual_update_checker.update_available.connect(
                lambda info: self._on_manual_update_found(info, progress)
            )
            self.manual_update_checker.check_finished.connect(
                lambda: self._on_manual_check_finished(progress)
            )
            
            # 开始检查
            self.manual_update_checker.check_for_updates()
            
        except Exception as e:
            QMessageBox.warning(self, "检查更新", f"检查更新时出错: {e}")
    
    def _on_manual_update_found(self, update_info, progress):
        """手动检查时发现更新"""
        progress.close()
        self.show_update_dialog(update_info)
    
    def _on_manual_check_finished(self, progress):
        """手动检查完成（无更新）"""
        progress.close()
        QMessageBox.information(
            self,
            "检查更新",
            f"当前已是最新版本 (v{CURRENT_VERSION})"
        )
    
    def show_help_dialog(self):
        """显示帮助与连接检测对话框"""
        try:
            help_dialog = HelpDialog(self, GITHUB_API_URL)
            help_dialog.exec_()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"无法打开帮助对话框: {e}")
    
    def show_about_dialog(self):
        """显示关于对话框"""
        QMessageBox.about(
            self,
            "关于",
            f"""<h2>电商售后品质退款管理工具</h2>
            <p><b>版本:</b> v{CURRENT_VERSION}</p>
            <p><b>功能:</b> 管理售后退款记录、AI智能分析、数据可视化</p>
            <p><b>更新:</b> 支持自动在线更新</p>
            <hr>
            <p style='color: #666;'>如有问题请联系开发者</p>
            """
        )

    def _init_system_tray(self):
        """初始化系统托盘图标"""
        # 检查系统是否支持托盘
        if not QSystemTrayIcon.isSystemTrayAvailable():
            QMessageBox.critical(self, "系统托盘", "您的系统不支持系统托盘功能")
            return
        
        # 创建托盘图标
        self.tray_icon = QSystemTrayIcon(self)
        
        # 使用苹果emoji作为托盘图标
        # 创建一个包含emoji的图像
        from PyQt5.QtGui import QPixmap, QPainter, QFont
        
        # 创建64x64像素的图像
        pixmap = QPixmap(64, 64)
        pixmap.fill(Qt.transparent)  # 透明背景
        
        # 在图像上绘制苹果emoji
        painter = QPainter(pixmap)
        painter.setFont(QFont("Segoe UI Emoji", 48))  # Windows系统使用Segoe UI Emoji字体
        painter.drawText(pixmap.rect(), Qt.AlignCenter, "🍎")
        painter.end()
        
        # 设置图标
        self.tray_icon.setIcon(QIcon(pixmap))
        
        # 设置托盘提示文本
        self.tray_icon.setToolTip(f"电商售后品质退款管理工具 v{CURRENT_VERSION}")
        
        # 创建托盘菜单
        tray_menu = QMenu()
        
        # 显示主窗口
        show_action = QAction("显示主窗口", self)
        show_action.triggered.connect(self.showNormal)
        tray_menu.addAction(show_action)
        
        tray_menu.addSeparator()
        
        # 退出程序
        quit_action = QAction("退出", self)
        quit_action.triggered.connect(self._quit_application)
        tray_menu.addAction(quit_action)
        
        # 设置托盘菜单
        self.tray_icon.setContextMenu(tray_menu)
        
        # 连接托盘图标激活信号（单击/双击）
        self.tray_icon.activated.connect(self._on_tray_activated)
        
        # 显示托盘图标
        self.tray_icon.show()
        
        print("[DEBUG] 系统托盘图标已初始化")
    
    def _on_tray_activated(self, reason):
        """托盘图标被激活时的处理"""
        # reason 1 = 单击, 2 = 双击, 3 = 右键点击
        if reason == QSystemTrayIcon.DoubleClick or reason == QSystemTrayIcon.Trigger:
            # 双击或单击显示主窗口
            self.showNormal()
            self.activateWindow()
    
    def _quit_application(self):
        """完全退出应用程序"""
        # 隐藏托盘图标
        if hasattr(self, 'tray_icon') and self.tray_icon:
            self.tray_icon.hide()
        
        # 退出应用程序
        QApplication.instance().quit()
    
    def _apply_dopamine_styles(self):
        """应用多巴胺配色方案到信息录入区"""
        try:
            # 读取多巴胺配色样式表
            with open(get_resource_path("dopamine_styles.qss"), "r", encoding="utf-8") as f:
                dopamine_styles = f.read()
            
            # 应用样式表到信息录入区
            self.input_panel.setStyleSheet(dopamine_styles)
            print("[DEBUG] 多巴胺配色方案已应用到信息录入区")
            
        except Exception as e:
            print(f"[DEBUG] 应用多巴胺配色方案失败: {e}")
            # 如果样式表文件不存在，使用默认样式
            default_styles = """
                QGroupBox#InputPanel {
                    background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                      stop: 0 #FF6B9D, stop: 0.5 #FFD166, stop: 1 #06D6A0);
                    border: 2px solid #118AB2;
                    border-radius: 15px;
                    font-family: "Microsoft YaHei";
                    font-weight: bold;
                    font-size: 10px;
                }
                QPushButton {
                    background-color: #FFD166;
                    border: 2px solid #EF476F;
                    border-radius: 8px;
                    color: #073B4C;
                    font-family: "Microsoft YaHei";
                    font-weight: bold;
                    font-size: 10px;
                    padding: 5px 10px;
                }
            """
            self.input_panel.setStyleSheet(default_styles)

    def _connect_input_signals(self):
        """连接信息录入区UI文件的信号和槽"""
        # 店铺相关控件
        self.store_combo = self.input_panel.findChild(QComboBox, "store_combo")
        self.add_store_btn = self.input_panel.findChild(QPushButton, "add_store_btn")
        self.edit_store_btn = self.input_panel.findChild(QPushButton, "edit_store_btn")
        self.delete_store_btn = self.input_panel.findChild(QPushButton, "delete_store_btn")
        
        # 订单和金额相关控件
        self.order_no_edit = self.input_panel.findChild(QLineEdit, "order_no_edit")
        self.refund_amount_edit = self.input_panel.findChild(QLineEdit, "refund_amount_edit")
        
        # 复选框控件
        self.cancel_check = self.input_panel.findChild(QCheckBox, "cancel_check")
        self.compensate_check = self.input_panel.findChild(QCheckBox, "compensate_check")
        self.reject_check = self.input_panel.findChild(QCheckBox, "reject_check")
        
        # 补偿金额和驳回结果
        self.comp_amount_edit = self.input_panel.findChild(QLineEdit, "comp_amount_edit")
        self.reject_result_combo = self.input_panel.findChild(QComboBox, "reject_result_combo")
        
        # 退款原因和日期
        self.reason_combo = self.input_panel.findChild(QComboBox, "reason_combo")
        self.record_date_edit = self.input_panel.findChild(QDateEdit, "record_date_edit")
        
        # 调试信息：检查控件是否找到
        print(f"[DEBUG] 信息录入区退款原因下拉框找到: {self.reason_combo is not None}")
        if self.reason_combo:
            print(f"[DEBUG] 退款原因下拉框对象类型: {type(self.reason_combo)}")
        
        # 备注
        self.notes_edit = self.input_panel.findChild(QLineEdit, "notes_edit")
        
        # 操作按钮
        self.add_btn = self.input_panel.findChild(QPushButton, "add_btn")
        self.update_btn = self.input_panel.findChild(QPushButton, "update_btn")
        self.clear_btn = self.input_panel.findChild(QPushButton, "clear_btn")
        
        # 连接信号
        if self.store_combo:
            self.store_combo.currentTextChanged.connect(self.on_store_combo_changed)
        if self.add_store_btn:
            self.add_store_btn.clicked.connect(self.add_store_dialog)
        if self.edit_store_btn:
            self.edit_store_btn.clicked.connect(self.edit_store_dialog)
        if self.delete_store_btn:
            self.delete_store_btn.clicked.connect(self.delete_store_dialog)
        
        if self.compensate_check:
            self.compensate_check.stateChanged.connect(self.toggle_comp_amount)
        if self.reject_check:
            self.reject_check.stateChanged.connect(self.toggle_reject_result)
            
        if self.add_btn:
            self.add_btn.clicked.connect(self.add_record)
        if self.update_btn:
            self.update_btn.clicked.connect(self.update_record)
        if self.clear_btn:
            self.clear_btn.clicked.connect(self.clear_input)
        
        # 设置鼠标点击事件
        if self.order_no_edit:
            self.order_no_edit.mousePressEvent = self.order_no_mouse_press
        if self.refund_amount_edit:
            self.refund_amount_edit.mousePressEvent = self.refund_amount_mouse_press
        if self.comp_amount_edit:
            self.comp_amount_edit.mousePressEvent = self.comp_amount_mouse_press
            
        # 初始化控件状态
        if self.comp_amount_edit:
            self.comp_amount_edit.setEnabled(False)
        if self.reject_result_combo is not None:
            self.reject_result_combo.setEnabled(False)
            
        # 设置退款原因选项
        print(f"[DEBUG] 开始设置退款原因选项，self.reason_combo: {self.reason_combo}")
        print(f"[DEBUG] self.reason_combo is None: {self.reason_combo is None}")
        print(f"[DEBUG] bool(self.reason_combo): {bool(self.reason_combo)}")
        
        # 使用更明确的条件判断
        if self.reason_combo is not None:
            print(f"[DEBUG] 退款原因下拉框存在，开始设置选项")
            reasons = ["商品腐败、变质、包装胀气等", "商品破损/压坏", "质量问题", 
                      "大小/规格/重量等与商品描述不符", "品种/标签/图片/包装等与商品描述不符", 
                      "货物与描述不符", "生产日期/保质期与商品描述不符", "其他"]
            print(f"[DEBUG] 退款原因列表: {reasons}")
            self.reason_combo.clear()  # 先清空现有选项
            self.reason_combo.addItems(reasons)
            print(f"[DEBUG] 信息录入区退款原因已设置，选项数量: {self.reason_combo.count()}")
        else:
            print(f"[DEBUG] 退款原因下拉框未找到，无法设置选项")
            
        # 设置驳回结果选项
        if self.reject_result_combo is not None:
            self.reject_result_combo.clear()
            self.reject_result_combo.addItems(["-", "驳回成功", "驳回失败"])
            self.reject_result_combo.setCurrentIndex(0)
            
        # 设置日期为今天
        if self.record_date_edit:
            self.record_date_edit.setDate(QDate.currentDate())

        self._configure_search_store_combo()

    def _connect_import_export_buttons(self):
        """连接导入导出按钮的信号和槽"""
        # 查找导入导出按钮
        self.import_btn = self.input_panel.findChild(QPushButton, "import_btn")
        self.export_btn = self.input_panel.findChild(QPushButton, "export_btn")
        self.cloud_sync_btn = self.input_panel.findChild(QPushButton, "cloud_sync_btn")
        
        # 调试信息
        print(f"[DEBUG] 导入按钮找到: {self.import_btn is not None}")
        print(f"[DEBUG] 导出按钮找到: {self.export_btn is not None}")
        print(f"[DEBUG] 云同步按钮找到: {self.cloud_sync_btn is not None}")
        
        # 连接信号
        if self.import_btn:
            self.import_btn.clicked.connect(self.import_excel)
            print("[DEBUG] 导入按钮信号已连接")
        if self.export_btn:
            self.export_btn.clicked.connect(self.export_excel)
            print("[DEBUG] 导出按钮信号已连接")
        if self.cloud_sync_btn:
            self.cloud_sync_btn.clicked.connect(self.open_cloud_sync_dialog)
            print("[DEBUG] 云同步按钮信号已连接")

        self._setup_ai_window_button()

    def _setup_ai_window_button(self):
        """将AI窗口入口集成到主要功能区底部操作行"""
        if hasattr(self, 'open_ai_window_btn'):
            return

        placeholder = self.input_panel.findChild(QWidget, "ai_button_placeholder")
        if not placeholder or not self.add_btn or not self.update_btn or not self.clear_btn:
            return

        self.open_ai_window_btn = QPushButton("分析")
        self.open_ai_window_btn.setMinimumHeight(30)
        self.open_ai_window_btn.clicked.connect(self.open_ai_window)
        self.reidentify_spec_code_btn = QPushButton("识别规格")
        self.reidentify_spec_code_btn.setMinimumHeight(30)
        self.reidentify_spec_code_btn.setToolTip("重新识别所有已录入规格名称订单的规格编码")
        self.reidentify_spec_code_btn.clicked.connect(self.reidentify_spec_codes_from_names)
        placeholder_layout = placeholder.layout()
        if placeholder_layout is None:
            placeholder_layout = QHBoxLayout(placeholder)
            placeholder_layout.setContentsMargins(0, 0, 0, 0)
            placeholder_layout.setSpacing(4)
        else:
            placeholder_layout.setSpacing(4)
        placeholder_layout.addWidget(self.open_ai_window_btn)
        placeholder_layout.addWidget(self.reidentify_spec_code_btn)

    def _optimize_input_panel_layout(self):
        """压缩主要功能区控件尺寸，使其适配更窄的中间列。"""
        panel_layout = self.input_panel.layout()
        if panel_layout:
            if hasattr(panel_layout, "setHorizontalSpacing"):
                panel_layout.setHorizontalSpacing(4)
            if hasattr(panel_layout, "setVerticalSpacing"):
                panel_layout.setVerticalSpacing(4)
            panel_layout.setContentsMargins(8, 8, 8, 8)

        label_names = [
            "store_label", "order_label", "refund_label", "comp_label",
            "reject_result_label", "reason_label", "date_label", "notes_label"
        ]
        for name in label_names:
            label = self.input_panel.findChild(QLabel, name)
            if not label:
                continue
            font = QFont(label.font())
            font.setPointSize(10)
            label.setFont(font)
            label.setMinimumHeight(30)
            label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        input_widgets = [
            self.store_combo,
            self.order_no_edit,
            self.refund_amount_edit,
            self.comp_amount_edit,
            self.reject_result_combo,
            self.reason_combo,
            self.record_date_edit,
        ]
        for widget in input_widgets:
            if not widget:
                continue
            font = QFont(widget.font())
            font.setPointSize(10)
            widget.setFont(font)
            widget.setMinimumHeight(32)
            widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        if self.comp_amount_edit:
            self.comp_amount_edit.setFixedWidth(65)
            self.comp_amount_edit.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        if self.refund_amount_edit:
            self.refund_amount_edit.setFixedWidth(65)
            self.refund_amount_edit.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        if self.notes_edit:
            self.notes_edit.setMinimumWidth(260)

        if self.notes_edit:
            font = QFont(self.notes_edit.font())
            font.setPointSize(10)
            self.notes_edit.setFont(font)
            self.notes_edit.setMinimumHeight(30)
            self.notes_edit.setMaximumHeight(32)
            self.notes_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        for checkbox in [self.cancel_check, self.compensate_check, self.reject_check]:
            if not checkbox:
                continue
            font = QFont(checkbox.font())
            font.setPointSize(10)
            checkbox.setFont(font)
            checkbox.setMinimumHeight(30)
            checkbox.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        compact_buttons = [
            self.add_store_btn, self.edit_store_btn, self.delete_store_btn,
            self.import_btn, self.export_btn, getattr(self, 'cloud_sync_btn', None), self.add_btn,
            self.update_btn, self.clear_btn
        ]
        for button in compact_buttons:
            if not button:
                continue
            font = QFont(button.font())
            font.setPointSize(10)
            button.setFont(font)
            button.setMinimumHeight(30)
            button.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
            button.setStyleSheet((button.styleSheet() or "") + "padding: 2px 8px;")

        for button in [
            getattr(self, 'open_ai_window_btn', None),
            getattr(self, 'reidentify_spec_code_btn', None)
        ]:
            if not button:
                continue
            font = QFont(button.font())
            font.setPointSize(10)
            button.setFont(font)
            button.setMinimumHeight(30)
            button.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
            button.setStyleSheet((button.styleSheet() or "") + "padding: 2px 8px;")

        placeholder = self.input_panel.findChild(QWidget, "ai_button_placeholder")
        if placeholder:
            placeholder.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)

    def _configure_search_store_combo(self):
        """优化搜索区店铺下拉框的可读性和弹出列表宽度。"""
        if not hasattr(self, 'search_store_combo') or self.search_store_combo is None:
            return

        combo = self.search_store_combo
        font = QFont(combo.font())
        font.setPointSize(13)
        combo.setFont(font)
        combo.setMinimumHeight(36)

        view = combo.view()
        if view is not None:
            view.setFont(font)
            view.setStyleSheet(
                "QListView { font-size: 14px; padding: 4px; }"
                "QListView::item { min-height: 30px; padding: 4px 10px; }"
            )

        max_text_width = 0
        font_metrics = combo.fontMetrics()
        for index in range(combo.count()):
            max_text_width = max(max_text_width, font_metrics.horizontalAdvance(combo.itemText(index)))

        popup_width = max(combo.width() + 30, max_text_width + 56)
        combo.setMinimumWidth(max(combo.minimumWidth(), min(popup_width - 30, 220)))
        if view is not None:
            view.setMinimumWidth(popup_width)

    def open_ai_window(self):
        """打开独立的AI分析窗口"""
        if not hasattr(self, 'ai_window') or self.ai_window is None:
            return

        self.ai_window.show()
        self.ai_window.raise_()
        self.ai_window.activateWindow()

    def open_cloud_sync_dialog(self):
        """打开腾讯云 COS 增量同步窗口。"""
        dialog = CloudSyncDialog(self)
        dialog.exec_()

    def upload_cloud_sync_shortcut(self):
        """Ctrl+S 快捷上传云端增量。"""
        config = self.db.load_cloud_sync_config()
        missing = [
            name for name, label in [
                ('secret_id', 'SecretId'),
                ('secret_key', 'SecretKey'),
                ('bucket', 'Bucket'),
                ('region', 'Region'),
            ]
            if not str(config.get(name) or '').strip()
        ]
        if missing:
            QMessageBox.warning(self, "云同步", "请先点击“云同步”填写并保存腾讯云配置。")
            return

        progress = QProgressDialog("正在上传增量数据...", None, 0, 0, self)
        progress.setWindowTitle("云同步")
        progress.setWindowModality(Qt.ApplicationModal)
        progress.setMinimumDuration(0)
        progress.show()
        QApplication.processEvents()
        try:
            result = CloudSyncService(self.db, config).upload_incremental()
            uploaded = result.get('uploaded', 0)
            deduped = result.get('deduped', 0)
            index_count = result.get('index_count', 0)
            snapshot_count = result.get('snapshot_count', 0)
            message = (
                f"上传完成：{uploaded} 条变更，清理重复 {deduped} 条，云端索引 {index_count} 条，云端快照 {snapshot_count} 条。"
                if uploaded or deduped or index_count or snapshot_count
                else "没有需要上传的增量数据。"
            )
            self.status_bar.showMessage(message, 3000)
            self.show_tooltip(message, "rgba(76, 175, 80, 0.95)" if uploaded else "rgba(255, 193, 7, 0.95)", 1500)
        except Exception as exc:
            QMessageBox.warning(self, "云同步", f"上传失败：{exc}")
        finally:
            progress.close()

    def refresh_after_cloud_sync(self):
        """云端增量合并后刷新主界面数据。"""
        self._cached_records = None
        self._last_search_params = None
        self.load_stores()
        if hasattr(self, 'search_store_combo'):
            current = self.search_store_combo.currentText()
            self.search_store_combo.blockSignals(True)
            self.search_store_combo.clear()
            self.search_store_combo.addItem("全部")
            for _, store_name in self.db.get_stores():
                self.search_store_combo.addItem(store_name)
            if current:
                index = self.search_store_combo.findText(current)
                self.search_store_combo.setCurrentIndex(index if index >= 0 else 0)
            self.search_store_combo.blockSignals(False)
        self.load_table_data(force_reload=True)
        self.update_store_stats_display()
        if hasattr(self, 'status_bar'):
            self.status_bar.showMessage("云同步数据已刷新", 3000)

    def ensure_ai_window_visible(self):
        """确保AI分析窗口保持可见，除非用户手动关闭。"""
        if not hasattr(self, 'ai_window') or self.ai_window is None:
            return
        if not self.ai_window.isVisible():
            self.ai_window.show()
        self.ai_window.raise_()
        self.ai_window.activateWindow()

    def undo_last_import(self):
        """撤销最近一次导入（Ctrl+Z）。"""
        if not self._last_import_undo_data:
            self.show_tooltip("没有可撤销的导入记录", "rgba(255, 193, 7, 0.95)", 1500)
            return

        undo_info = self._last_import_undo_data
        created_ids = undo_info.get('created_ids', [])
        updated_records = undo_info.get('updated_records', [])

        if not created_ids and not updated_records:
            self.show_tooltip("没有可撤销的导入记录", "rgba(255, 193, 7, 0.95)", 1500)
            self._last_import_undo_data = None
            return

        confirm = QMessageBox.question(
            self,
            "撤销导入",
            f"将撤销最近一次导入：删除 {len(created_ids)} 条新增记录，恢复 {len(updated_records)} 条覆盖记录。\n是否继续？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if confirm != QMessageBox.Yes:
            return

        restored_count = 0
        deleted_count = 0
        failed_count = 0

        for record in reversed(updated_records):
            try:
                self.db.update_record(
                    record['id'],
                    record['store_id'],
                    record['order_no'],
                    record['reason'],
                    record['refund_amount'],
                    record['cancel'],
                    record['compensate'],
                    record['comp_amount'],
                    record['reject'],
                    record['reject_result'],
                    record['notes'],
                    record['record_date'],
                    record.get('order_status', ''),
                    record.get('after_sale_status', ''),
                    record.get('spec_name', ''),
                    record.get('spec_code', ''),
                    record.get('refund_apply_time', ''),
                    record.get('refund_agree_time', '')
                )
                restored_count += 1
            except Exception as e:
                failed_count += 1
                print(f"[UNDO IMPORT] 恢复记录失败 {record.get('id')}: {e}")

        for record_id in reversed(created_ids):
            try:
                if self.db.delete_record(record_id):
                    deleted_count += 1
                else:
                    failed_count += 1
            except Exception as e:
                failed_count += 1
                print(f"[UNDO IMPORT] 删除导入记录失败 {record_id}: {e}")

        self._cached_records = None
        self._last_search_params = None
        self.load_table_data(force_reload=True)
        self.table.viewport().update()
        QApplication.processEvents()

        if failed_count == 0:
            self.show_tooltip(
                f"已撤销导入 删除{deleted_count}条 恢复{restored_count}条",
                "rgba(76, 175, 80, 0.95)",
                1800
            )
            self._last_import_undo_data = None
        else:
            QMessageBox.warning(
                self,
                "撤销导入",
                f"撤销已完成，但有 {failed_count} 条记录处理失败。\n"
                f"已删除：{deleted_count} 条\n已恢复：{restored_count} 条"
            )

    def on_store_combo_changed(self, store_name):
        """信息录入区店铺选择变化"""
        # 不再同步到搜索筛选区，保持两个区域独立
        
        # 启用/禁用编辑和删除按钮
        if store_name and store_name != "请先添加店铺":
            # 启用编辑和删除按钮（只有在选择真实店铺时）
            if hasattr(self, 'edit_store_btn') and hasattr(self, 'delete_store_btn'):
                self.edit_store_btn.setEnabled(True)
                self.delete_store_btn.setEnabled(True)
        else:
            # 禁用编辑和删除按钮
            if hasattr(self, 'edit_store_btn') and hasattr(self, 'delete_store_btn'):
                self.edit_store_btn.setEnabled(False)
                self.delete_store_btn.setEnabled(False)

    def sync_store_selection(self, store_name):
        """同步所有店铺选择框"""
        # 同步搜索区的店铺选择
        if hasattr(self, 'search_store_combo'):
            index = self.search_store_combo.findText(store_name)
            if index >= 0:
                self.search_store_combo.setCurrentIndex(index)
        
        # 同步店铺信息区的店铺显示（现在直接显示搜索筛选区选择的店铺）
        # 加载对应店铺的设置
        self.load_store_settings()
        self.update_store_stats_display()

    def save_window_settings(self):
        """空方法，已删除记忆功能"""
        # 不再保存窗口设置，使用固定的默认值
        pass

    def load_window_settings(self):
        """设置固定的默认窗口设置（删除记忆功能）"""
        # 直接设置固定的默认值，不使用记忆功能
        self.main_splitter.setSizes([138, 1142, 170])
        self.top_splitter.setSizes([305, 525])
        self.bottom_splitter.setSizes([420, 320])

    def closeEvent(self, event):
        """窗口关闭事件，保存设置并实现最小化到托盘"""
        # 先保存窗口设置
        self.save_window_settings()
        
        # 检查是否支持系统托盘，如果支持则最小化到托盘而不是关闭
        if hasattr(self, 'tray_icon') and self.tray_icon:
            # 隐藏主窗口
            self.hide()
            
            # 显示气泡提示（仅第一次关闭时显示）
            if not hasattr(self, '_tray_notification_shown'):
                self.tray_icon.showMessage(
                    "售后管理工具",
                    "程序已最小化到系统托盘，双击图标可恢复窗口",
                    QSystemTrayIcon.Information,
                    3000  # 显示3秒
                )
                self._tray_notification_shown = True
            
            # 忽略关闭事件，不退出程序
            event.ignore()
            print("[DEBUG] 窗口已最小化到系统托盘")
        else:
            # 如果没有托盘图标，正常关闭
            print("[DEBUG] 没有托盘图标，正常关闭程序")
            # 关闭数据库连接
            self.db.close()
            event.accept()

    def show_bubble_message(self, message):
        """显示淡入淡出气泡消息"""
        bubble = BubbleMessage(message, self)
        bubble.show_bubble()

    def load_store_settings(self):
        """从数据库加载店铺设置"""
        # 获取当前选择的店铺
        current_store = None
        if hasattr(self, 'search_store_combo') and self.search_store_combo.currentText():
            current_store = self.search_store_combo.currentText()
        
        # 只在程序启动时设置默认店铺（第一次调用时）
        if not hasattr(self, '_store_settings_initialized'):
            self._store_settings_initialized = True
            if not current_store:
                # 默认选择"全部"店铺
                current_store = "全部"
                # 如果搜索筛选区已初始化，更新选择
                if hasattr(self, 'search_store_combo'):
                    index = self.search_store_combo.findText("全部")
                    if index >= 0:
                        self.search_store_combo.setCurrentIndex(index)
        
        if current_store and current_store != "全部":
            # 获取店铺ID
            stores = self.db.get_stores()
            store_id = None
            for sid, sname in stores:
                if sname == current_store:
                    store_id = sid
                    break
            
            if store_id:
                # 从数据库加载设置
                db_settings = self.db.get_store_settings(store_id)
                if db_settings:
                    self.store_settings = db_settings
                else:
                    # 如果没有设置，使用默认值
                    self.store_settings = {'daily_orders': 0, 'daily_sales': 0.0, 'refund_budget': 0.0}
        else:
            # 如果选择了"全部"店铺，从全局设置加载
            self.store_settings = self.db.get_global_settings()

    def load_store_info_combo(self):
        """加载店铺信息下拉框（现在使用信息录入区的店铺选择）"""
        # 不再需要这个功能，因为店铺信息区直接显示信息录入区选择的店铺
        pass

    def on_store_info_changed(self, store_name):
        """店铺信息选择变化"""
        if store_name:
            # 获取预估订单量
            estimated_orders = self.db.get_estimated_orders(store_name)
            self.estimated_orders_edit.setText(str(estimated_orders) if estimated_orders > 0 else "")
            
            # 更新店铺统计
            self.update_store_stats(store_name)

    def update_store_stats(self, store_name):
        """更新店铺统计信息"""
        if not store_name:
            return
            
        # 获取店铺退款统计（排除撤销订单）
        stats = self.db.get_store_refund_stats(store_name)
        refund_count = stats['refund_count']
        total_refund = stats['total_refund']
        total_comp = stats['total_comp']
        
        # 获取预估订单量（实时更新到数据库）
        estimated_text = self.estimated_orders_edit.text().strip()
        estimated_orders = 0
        if estimated_text:
            try:
                estimated_orders = int(estimated_text)
                if estimated_orders > 0:
                    # 实时保存到数据库
                    self.db.set_estimated_orders(store_name, estimated_orders)
            except ValueError:
                estimated_orders = self.db.get_estimated_orders(store_name)
        else:
            estimated_orders = self.db.get_estimated_orders(store_name)
        
        # 计算退款率
        refund_rate = 0.0
        if estimated_orders > 0:
            refund_rate = (refund_count / estimated_orders) * 100
        
        # 更新显示
        if hasattr(self, 'store_stats_label'):
            self.store_stats_label.setText(
                f"有效退款：{refund_count}单 | 退款率：{refund_rate:.2f}% | 总金额：¥{total_refund + total_comp:.2f}"
            )

    def update_refund_rate(self):
        """更新退款率显示（现在使用信息录入区的店铺选择）"""
        store_name = self.store_combo.currentText()
        if store_name:
            self.update_store_stats(store_name)

    def open_add_store_dialog(self):
        """打开添加店铺对话框"""
        dialog = AddStoreDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            store_name = dialog.get_store_name()
            if store_name:
                # 这里可以添加保存店铺到数据库的逻辑
                QMessageBox.information(self, "添加成功", f"店铺 '{store_name}' 已添加")
                # 刷新店铺列表
                self.load_stores()

    def open_store_settings(self):
        """打开店铺基本信息设置对话框"""
        # 确保当前店铺设置已从数据库加载
        self.load_store_settings()
        dialog = StoreSettingsDialog(self)
        dialog.load_settings(self.store_settings)
        if dialog.exec_() == QDialog.Accepted:
            # 设置已保存，更新显示
            self.update_store_stats_display()

    def update_store_stats_display(self):
        """更新店铺统计信息显示"""
        # 更新当前店铺名称显示（使用搜索筛选区的店铺选择）
        current_store = self.search_store_combo.currentText() if self.search_store_combo.currentText() else "未选择"
        self.current_store_label.setText(f"当前店铺：{current_store}")

        if self._should_skip_store_stats_calculation():
            self._set_store_stats_skipped_state()
            return
        
        orders_sales = self.calculate_orders_and_sales()
        self.orders_label.setText(f"当前范围单量：{self._format_metric_int(orders_sales['orders'])}")
        self.sales_label.setText(f"当前范围销售额：¥{orders_sales['sales']:.1f}")
        
        # 更新日退款预算剩余
        daily_budget_remaining = self.calculate_daily_budget_remaining()
        self.daily_budget_remaining_label.setText(f"退款预算：¥{daily_budget_remaining:.1f}")
        
        # 更新增强的退款统计信息
        enhanced_stats = self.calculate_enhanced_refund_stats()
        
        # 更新品质退款相关统计
        quality_stats = self.calculate_quality_refund_stats()
        
        # 更新售后金额相关统计
        refund_stats = self.calculate_refund_amount_stats()
        
        self.quality_refund_count_label.setText(f"品质退款：{enhanced_stats['quality_refund_count']}单")
        self.other_refund_count_label.setText(f"其他退款：{enhanced_stats['other_refund_count']}单")
        self.canceled_quality_count_label.setText(f"撤销品质：{enhanced_stats['canceled_quality_count']}单")
        self.total_refund_rate_label.setText(f"总退款率：{enhanced_stats['total_refund_rate']:.2f}%")

        self.total_after_sales_label.setText(f"售后总额：¥{refund_stats['total_refund']:.1f}")
        self.refund_ratio_label.setText(f"金额占比：{refund_stats['refund_ratio']:.2f}%")
        self.quality_after_sales_amount_label.setText(f"品质售后：¥{enhanced_stats['quality_after_sales_amount']:.1f}")
        self.other_after_sales_amount_label.setText(f"其他售后：¥{enhanced_stats['other_after_sales_amount']:.1f}")

        self.quality_apply_rate_label.setText(f"申请品质率：{quality_stats['apply_rate']:.2f}%")
        self.quality_actual_rate_label.setText(f"实际品质率：{quality_stats['actual_rate']:.2f}%")
        self.quality_cancel_rate_label.setText(f"撤销率：{quality_stats['cancel_rate']:.2f}%")

        self.top_refund_reason_label.setText(f"最多原因：{enhanced_stats['top_refund_reason']}")
        self.top_reason_count_label.setText(f"出现次数：{enhanced_stats['top_reason_count']}")
        self.top_reason_ratio_label.setText(f"占比：{enhanced_stats['top_reason_ratio']:.1f}%")

    @staticmethod
    def _normalize_reason(reason):
        return str(reason or "").strip()

    @classmethod
    def _is_quality_reason(cls, reason):
        return cls._normalize_reason(reason) not in ("", "其他")

    @staticmethod
    def _normalize_reject_result_value(result):
        text = str(result or "").strip()
        if text in ("驳回成功", "成功"):
            return "驳回成功"
        if text in ("驳回失败", "失败"):
            return "驳回失败"
        return text

    @classmethod
    def _display_reject_result_value(cls, record):
        text = cls._normalize_reject_result_value(record.get('reject_result'))
        if not bool(record.get('reject')):
            return text
        if text in ("", "-", "无", "None"):
            return "驳回中"
        return text

    @classmethod
    def _is_reject_success_record(cls, record):
        return bool(record.get('reject')) and cls._normalize_reject_result_value(record.get('reject_result')) == "驳回成功"

    @classmethod
    def _is_reject_failure_record(cls, record):
        return bool(record.get('reject')) and cls._normalize_reject_result_value(record.get('reject_result')) == "驳回失败"

    @classmethod
    def _has_compensation_record(cls, record):
        return bool(record.get('compensate')) and cls._safe_float(record.get('comp_amount', 0)) > 0

    @classmethod
    def _is_effective_refund_record(cls, record):
        return (not bool(record.get('cancel'))) and not cls._is_reject_success_record(record)

    @staticmethod
    def _weekly_to_daily_avg(value):
        """将用户录入的7天总值换算为日均值。"""
        try:
            return float(value) / 7
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _format_metric_value(value):
        """格式化数量显示，避免无意义的小数尾数。"""
        return f"{float(value):.2f}".rstrip('0').rstrip('.')

    @staticmethod
    def _format_metric_int(value):
        """将单量显示为整数。"""
        try:
            return str(int(round(float(value))))
        except (ValueError, TypeError):
            return "0"

    def get_entered_week_metrics(self):
        """获取用户填写的近7天单量和销售额。"""
        if not self.store_settings:
            return {"orders": 0.0, "sales": 0.0}

        return {
            "orders": self._safe_float(self.store_settings.get('daily_orders', 0)),
            "sales": self._safe_float(self.store_settings.get('daily_sales', 0.0)),
        }

    @staticmethod
    def _safe_float(value):
        """安全转换为浮点数。"""
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _is_all_time_range(self):
        """是否处于“全部时间”范围。"""
        return (
            self.start_date_edit.date() == QDate(2000, 1, 1) and
            self.end_date_edit.date() == QDate(2100, 12, 31)
        )

    def _should_skip_store_stats_calculation(self):
        """全部店铺 + 全部时间时，跳过店铺信息与统计板块计算。"""
        current_store = self.search_store_combo.currentText()
        return current_store == "全部" and self._is_all_time_range()

    def _set_store_stats_skipped_state(self):
        """设置店铺信息与统计板块为跳过计算的展示状态。"""
        self.current_store_label.setText("当前店铺：全部")
        self.orders_label.setText("当前范围单量：--")
        self.sales_label.setText("当前范围销售额：--")
        self.daily_budget_remaining_label.setText("退款预算：--")
        self.quality_refund_count_label.setText("品质退款：--")
        self.other_refund_count_label.setText("其他退款：--")
        self.canceled_quality_count_label.setText("撤销品质：--")
        self.total_refund_rate_label.setText("总退款率：--")
        self.total_after_sales_label.setText("售后总额：--")
        self.refund_ratio_label.setText("金额占比：--")
        self.quality_after_sales_amount_label.setText("品质售后：--")
        self.other_after_sales_amount_label.setText("其他售后：--")
        self.quality_apply_rate_label.setText("申请品质率：--")
        self.quality_actual_rate_label.setText("实际品质率：--")
        self.quality_cancel_rate_label.setText("撤销率：--")
        self.top_refund_reason_label.setText("最多原因：--")
        self.top_reason_count_label.setText("出现次数：--")
        self.top_reason_ratio_label.setText("占比：--")

    def _create_search_signal_blockers(self):
        """批量修改筛选条件时，阻止重复触发搜索。"""
        widgets = [
            getattr(self, 'search_store_combo', None),
            getattr(self, 'search_order_edit', None),
            getattr(self, 'search_real_reason_edit', None),
            getattr(self, 'search_cancel_combo', None),
            getattr(self, 'search_reject_combo', None),
            getattr(self, 'start_date_edit', None),
            getattr(self, 'end_date_edit', None),
            getattr(self, 'search_reason_dropdown', None),
        ]
        return [QSignalBlocker(widget) for widget in widgets if widget is not None]

    def calculate_daily_budget_remaining(self):
        """计算退款预算剩余（周销售额录入，预算按日均口径参与统计）。"""
        if not self.store_settings:
            return 0.0
        
        # 获取筛选的天数
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        days_count = (end_date - start_date).days + 1  # 计算筛选的天数
        
        # 获取用户设置的日退款预算
        daily_refund_budget = self.store_settings.get('refund_budget', 0.0)
        
        # 计算多天的总预算
        total_refund_budget = daily_refund_budget * days_count
        
        # 计算筛选期间的总退款金额
        filtered_records = self.get_filtered_records()
        total_refund = 0.0
        for record in filtered_records:
            if self._has_compensation_record(record):
                total_refund += self._safe_float(record.get('comp_amount', 0))

            if self._is_effective_refund_record(record):
                total_refund += self._safe_float(record.get('refund_amount', 0))
        
        return total_refund_budget - total_refund

    def calculate_orders_and_sales(self):
        """计算订单量和销售金额（用户录入7天总值，统计时自动换算为日均）。"""
        try:
            if not self.store_settings:
                return {"orders": 0, "sales": 0.0}
            
            # 获取筛选的天数
            start_date = self.start_date_edit.date().toPyDate()
            end_date = self.end_date_edit.date().toPyDate()
            days_count = (end_date - start_date).days + 1  # 计算筛选的天数
            
            # 安全地获取用户设置的周订单量和周销售额，并换算为日均值
            daily_orders_str = self.store_settings.get('daily_orders', '0')
            daily_sales_str = self.store_settings.get('daily_sales', '0.0')
            
            # 转换为数值类型，处理可能的异常
            try:
                daily_orders = self._weekly_to_daily_avg(float(daily_orders_str)) if daily_orders_str else 0.0
            except (ValueError, TypeError):
                daily_orders = 0.0
                
            try:
                daily_sales = self._weekly_to_daily_avg(float(daily_sales_str)) if daily_sales_str else 0.0
            except (ValueError, TypeError):
                daily_sales = 0.0
            
            # 计算多天的总订单量和总销售金额
            total_orders = daily_orders * days_count
            total_sales = daily_sales * days_count
            
            return {"orders": total_orders, "sales": total_sales}
            
        except Exception as e:
            # 如果出现任何异常，返回默认值
            print(f"计算订单量和销售金额时出错: {e}")
            return {"orders": 0, "sales": 0.0}

    def calculate_today_refund_amount(self, date):
        """计算指定日期的退款金额（基于当前筛选条件）"""
        # 使用与表格相同的筛选条件获取记录
        filtered_records = self.get_filtered_records()
        
        # 筛选指定日期的记录
        today_str = date.strftime('%Y-%m-%d')
        today_records = [r for r in filtered_records if r['record_date'] == today_str]
        
        if not today_records:
            return 0.0
        
        # 计算今天的退款金额（使用与退款金额统计相同的逻辑）
        today_refund = 0.0
        for record in today_records:
            if self._has_compensation_record(record):
                today_refund += self._safe_float(record.get('comp_amount', 0))

            if self._is_effective_refund_record(record):
                today_refund += self._safe_float(record.get('refund_amount', 0))
        
        return today_refund

    def calculate_quality_refund_stats(self):
        """计算品质退款相关统计（基于当前筛选条件，支持多天筛选）"""
        # 使用与表格相同的筛选条件获取记录
        filtered_records = self.get_filtered_records()
        
        if not filtered_records:
            return {'apply_rate': 0.0, 'actual_rate': 0.0, 'cancel_rate': 0.0}
        
        # 获取筛选的天数
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        days_count = (end_date - start_date).days + 1  # 计算筛选的天数
        
        # 计算品质退款相关统计
        # 顾客申请品质退款率：品质退款订单数 ÷ (用户设置的周单量÷7 × 筛选天数)
        daily_orders = self._weekly_to_daily_avg(self.store_settings.get('daily_orders', 0))
        total_orders = daily_orders * days_count  # 多天的总订单量
        
        quality_refund_orders = [r for r in filtered_records if self._is_quality_reason(r.get('reason'))]
        apply_quality_count = len(quality_refund_orders)
        apply_rate = (apply_quality_count / total_orders * 100) if total_orders > 0 else 0.0
        
        # 实际计入品质退款率：减去已撤销和驳回成功的订单
        actual_quality_count = sum(1 for record in quality_refund_orders if self._is_effective_refund_record(record))
        actual_rate = (actual_quality_count / total_orders * 100) if total_orders > 0 else 0.0
        
        # 品质退款撤销率：已撤销的品质退款订单数 ÷ 总品质退款订单数
        canceled_quality_count = sum(1 for r in quality_refund_orders if r.get('cancel'))
        cancel_rate = (canceled_quality_count / apply_quality_count * 100) if apply_quality_count > 0 else 0.0
        
        return {
            'apply_rate': apply_rate,    # 顾客申请品质退款率
            'actual_rate': actual_rate,  # 实际计入品质退款率
            'cancel_rate': cancel_rate   # 品质退款撤销率
        }

    def calculate_refund_amount_stats(self):
        """计算退款金额相关统计（基于当前筛选条件，支持多天筛选）"""
        # 使用与表格相同的筛选条件获取记录
        filtered_records = self.get_filtered_records()
        
        if not filtered_records:
            return {'total_refund': 0.0, 'refund_ratio': 0.0}
        
        # 获取筛选的天数
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        days_count = (end_date - start_date).days + 1  # 计算筛选的天数
        
        # 计算退款金额
        total_refund = 0.0
        for record in filtered_records:
            if self._has_compensation_record(record):
                total_refund += self._safe_float(record.get('comp_amount', 0))

            if self._is_effective_refund_record(record):
                total_refund += self._safe_float(record.get('refund_amount', 0))
        
        # 计算退款金额占比：退款金额 ÷ (用户设置的周销售额÷7 × 筛选天数)
        daily_sales = self._weekly_to_daily_avg(self.store_settings.get('daily_sales', 0.0))
        total_sales = daily_sales * days_count  # 多天的总销售额
        refund_ratio = (total_refund / total_sales * 100) if total_sales > 0 else 0.0
        
        return {
            'total_refund': total_refund,  # 退款金额
            'refund_ratio': refund_ratio  # 退款金额占比
        }

    def calculate_enhanced_refund_stats(self):
        """计算增强的退款统计信息（基于当前筛选条件）"""
        # 使用与表格相同的筛选条件获取记录
        filtered_records = self.get_filtered_records()
        
        if not filtered_records:
            return {
                'quality_refund_count': 0,
                'other_refund_count': 0,
                'canceled_quality_count': 0,
                'total_refund_rate': 0.0,
                'quality_after_sales_amount': 0.0,
                'other_after_sales_amount': 0.0,
                'top_refund_reason': '无数据',
                'top_reason_count': 0,
                'top_reason_ratio': 0.0
            }
        
        # 获取筛选的天数
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        days_count = (end_date - start_date).days + 1  # 计算筛选的天数
        
        # 计算总订单数
        daily_orders = self._weekly_to_daily_avg(self.store_settings.get('daily_orders', 0))
        total_orders = daily_orders * days_count  # 多天的总订单量
        
        # 统计退款原因出现次数
        reason_counts = {}
        
        # 初始化统计变量
        quality_refund_count = 0
        other_refund_count = 0
        canceled_quality_count = 0
        quality_after_sales_amount = 0.0
        other_after_sales_amount = 0.0
        
        # 遍历所有记录进行统计
        for record in filtered_records:
            reason = record['reason']
            
            # 统计退款原因出现次数
            reason_counts[reason] = reason_counts.get(reason, 0) + 1
            
            # 判断是否为品质退款
            is_quality_refund = self._is_quality_reason(reason)
            
            # 统计数量
            if is_quality_refund:
                quality_refund_count += 1
                # 统计撤销的品质退款数量
                if record['cancel']:
                    canceled_quality_count += 1
            else:
                other_refund_count += 1
            
            order_after_sales_amount = 0.0
            if self._is_effective_refund_record(record):
                order_after_sales_amount += self._safe_float(record.get('refund_amount', 0))
            if self._has_compensation_record(record):
                order_after_sales_amount += self._safe_float(record.get('comp_amount', 0))

            if is_quality_refund:
                quality_after_sales_amount += order_after_sales_amount
            else:
                other_after_sales_amount += order_after_sales_amount
        
        # 计算总退款率
        total_refund_count = quality_refund_count + other_refund_count
        total_refund_rate = (total_refund_count / total_orders * 100) if total_orders > 0 else 0.0
        
        # 找出退款最多的原因
        top_refund_reason = '无数据'
        top_reason_count = 0
        top_reason_ratio = 0.0
        
        if reason_counts:
            top_refund_reason = max(reason_counts, key=reason_counts.get)
            top_reason_count = reason_counts[top_refund_reason]
            top_reason_ratio = (top_reason_count / total_refund_count * 100) if total_refund_count > 0 else 0.0
        
        return {
            'quality_refund_count': quality_refund_count,        # 品质退款数量
            'other_refund_count': other_refund_count,            # 其他退款数量
            'canceled_quality_count': canceled_quality_count,    # 撤销的品质退款数量
            'total_refund_rate': total_refund_rate,              # 总退款率
            'quality_after_sales_amount': quality_after_sales_amount,  # 品质售后金额
            'other_after_sales_amount': other_after_sales_amount,      # 其他售后金额
            'top_refund_reason': top_refund_reason,              # 退款最多的原因
            'top_reason_count': top_reason_count,                # 最多原因出现次数
            'top_reason_ratio': top_reason_ratio                 # 最多原因占比
        }

    def update_total_amount_display(self):
        """更新右上角全局统计显示"""
        # 获取所有记录
        records = self.db.get_all_records()
        
        # 计算总金额（排除撤销订单）
        total_refund = sum(r['refund_amount'] for r in records if not r['cancel'])
        total_comp = sum(r['comp_amount'] for r in records if not r['cancel'])
        total_amount = total_refund + total_comp
        
        # 计算总退款率
        total_refund_count = sum(1 for r in records if not r['cancel'])
        total_estimated_orders = self.get_total_estimated_orders()
        
        total_refund_rate = 0.0
        if total_estimated_orders > 0:
            total_refund_rate = (total_refund_count / total_estimated_orders) * 100
        
        # 更新右上角全局统计显示
        if hasattr(self, 'global_stats_label'):
            self.global_stats_label.setText(f"总金额：¥{total_amount:,.2f} | 总退款率：{total_refund_rate:.2f}%")

    def get_total_estimated_orders(self):
        """获取所有店铺的预估订单量总和"""
        stores = self.db.get_stores()
        total_estimated = 0
        for store_id, store_name in stores:
            estimated = self.db.get_estimated_orders(store_name)
            total_estimated += estimated
        return total_estimated

    def apply_stylesheet(self):
        """应用极简风格样式表"""
        stylesheet = """
        /* 主窗口背景色 */
        QMainWindow {
            background-color: #F8F9FA;  /* 极浅灰色背景 */
        }
        
        /* 中央控件背景 */
        QWidget {
            background-color: #F8F9FA;
            color: #212529;  /* 深灰文字 */
        }
        
        /* 分组框样式 */
        QGroupBox {
            font-weight: bold;
            font-size: 12px;
            border: 1px solid #DEE2E6;  /* 浅灰边框 */
            border-radius: 6px;
            margin-top: 8px;
            padding-top: 8px;
            background-color: #FFFFFF;  /* 白色背景 */
        }
        
        QGroupBox::title {
            subcontrol-origin: margin;
            subcontrol-position: top center;
            padding: 0 6px;
            background-color: #6C757D;  /* 中灰标题背景 */
            color: white;
            border-radius: 3px;
        }
        
        /* 按钮样式 */
        QPushButton {
            background-color: #6C757D;  /* 中灰背景 */
            color: white;
            border: none;
            padding: 6px 12px;
            border-radius: 3px;
            font-weight: normal;
            min-width: 80px;
        }
        
        QPushButton:hover {
            background-color: #5A6268;  /* 深灰悬停 */
        }
        
        QPushButton:pressed {
            background-color: #495057;  /* 更深灰按下 */
        }
        
        /* 重要操作按钮特殊样式 */
        QPushButton[important="true"] {
            background-color: #DC3545;  /* 红色强调 */
        }
        
        QPushButton[important="true"]:hover {
            background-color: #C82333;
        }
        
        /* 输入框样式 */
        QLineEdit, QComboBox, QDateEdit {
            padding: 6px;
            border: 1px solid #CED4DA;
            border-radius: 4px;
            background-color: white;
            selection-background-color: #6C757D;
        }
        
        QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
            border: 2px solid #6C757D;
        }
        
        /* 表格样式 - 已由table_panel.ui文件控制，此处删除相关设置 */
        
        /* 复选框样式 */
        QCheckBox {
            spacing: 8px;
        }
        
        QCheckBox::indicator {
            width: 16px;
            height: 16px;
        }
        
        QCheckBox::indicator:unchecked {
            border: 1px solid #CCCCCC;
            background-color: white;
            border-radius: 2px;
        }
        
        QCheckBox::indicator:checked {
            border: 1px solid #2E8B57;
            background-color: #2E8B57;
            border-radius: 2px;
        }
        
        /* 状态栏样式 */
        QStatusBar {
            background-color: #2E8B57;
            color: white;
            padding: 4px;
        }
        
        /* 标签样式 */
        QLabel {
            color: #333333;
            font-weight: normal;
        }
        """
        self.setStyleSheet(stylesheet)

    def setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+S"), self, self.upload_cloud_sync_shortcut)
        QShortcut(QKeySequence("Ctrl+E"), self, self.export_excel)
        QShortcut(QKeySequence("Ctrl+F"), self, lambda: self.search_order_edit.setFocus())
        QShortcut(QKeySequence("Ctrl+D"), self, self.delete_record)
        QShortcut(QKeySequence("Ctrl+Z"), self, self.undo_last_import)

    def toggle_comp_amount(self, state):
        self.comp_amount_edit.setEnabled(state == Qt.Checked)
        if state != Qt.Checked:
            self.comp_amount_edit.clear()

    def toggle_reject_result(self, state):
        """控制驳回结果下拉框的可用性"""
        self.reject_result_combo.setEnabled(state == Qt.Checked)
        if state == Qt.Checked:
            if self.reject_result_combo.currentIndex() <= 0:
                self.reject_result_combo.setCurrentIndex(1)
        else:
            self.reject_result_combo.setCurrentIndex(0)  # 重置为占位值

    def load_stores(self):
        """加载店铺列表到所有下拉框"""
        # 检查控件是否存在（避免在UI初始化完成前调用）
        if not hasattr(self, 'store_combo') or self.store_combo is None:
            print("[DEBUG] store_combo 尚未初始化，跳过加载店铺列表")
            return
        if not hasattr(self, 'search_store_combo') or self.search_store_combo is None:
            print("[DEBUG] search_store_combo 尚未初始化，跳过加载店铺列表")
            return
        
        # 清空所有店铺下拉框
        self.store_combo.clear()
        self.search_store_combo.clear()
            
        stores = self.db.get_stores()
        self.search_store_combo.addItem("全部")
        for store_id, store_name in stores:
            self.store_combo.addItem(store_name, store_id)
            self.search_store_combo.addItem(store_name, store_id)
        
        # 如果有店铺，信息录入区选择第一个，搜索筛选区选择"全部"
        if self.store_combo.count() > 0:
            # 信息录入区选择第一个店铺
            current_store = self.store_combo.currentText()
            # 搜索筛选区选择"全部"
            self.search_store_combo.setCurrentIndex(0)  # 0是"全部"选项
            # 不再同步店铺信息显示，保持两个区域独立

        self._configure_search_store_combo()
        
        if self.store_combo.count() == 0:
            self.store_combo.addItem("请先添加店铺", None)

    def add_store_dialog(self):
        """添加店铺对话框"""
        name, ok = QInputDialog.getText(self, "添加店铺", "店铺名称：")
        if ok and name.strip():
            name = name.strip()
            if self.db.add_store(name):
                self.load_stores()
                self.show_tooltip(f"店铺 {name} 已添加", "rgba(76, 175, 80, 0.95)", 1500)  # 绿色气泡显示1.5秒
            else:
                QMessageBox.warning(self, "错误", f"店铺 '{name}' 已存在！")

    def edit_store_dialog(self):
        """修改店铺名称对话框"""
        current_store = self.store_combo.currentText()
        if not current_store or current_store == "请先添加店铺":
            QMessageBox.warning(self, "错误", "请先选择一个店铺！")
            return
        
        # 获取当前店铺ID
        stores = self.db.get_stores()
        store_id = None
        for sid, sname in stores:
            if sname == current_store:
                store_id = sid
                break
        
        if not store_id:
            QMessageBox.warning(self, "错误", "未找到选中的店铺！")
            return
        
        new_name, ok = QInputDialog.getText(self, "修改店铺名称", "新店铺名称：", text=current_store)
        if ok and new_name.strip():
            new_name = new_name.strip()
            if new_name == current_store:
                QMessageBox.information(self, "提示", "店铺名称未改变！")
                return
            
            if self.db.update_store_name(store_id, new_name):
                self.load_stores()
                # 更新当前选择
                index = self.store_combo.findText(new_name)
                if index >= 0:
                    self.store_combo.setCurrentIndex(index)
                self.show_tooltip(f"店铺名称已修改为 {new_name}", "rgba(33, 150, 243, 0.95)", 1500)  # 蓝色气泡
            else:
                QMessageBox.warning(self, "错误", f"店铺名称 '{new_name}' 已存在或修改失败！")

    def delete_store_dialog(self):
        """删除店铺对话框"""
        current_store = self.store_combo.currentText()
        if not current_store or current_store == "请先添加店铺":
            QMessageBox.warning(self, "错误", "请先选择一个店铺！")
            return
        
        # 获取当前店铺ID
        stores = self.db.get_stores()
        store_id = None
        for sid, sname in stores:
            if sname == current_store:
                store_id = sid
                break
        
        if not store_id:
            QMessageBox.warning(self, "错误", "未找到选中的店铺！")
            return
        
        # 确认删除对话框
        reply = QMessageBox.question(self, "确认删除", 
                                    f"确定要删除店铺 '{current_store}' 吗？\n\n⚠️ 警告：删除后该店铺的所有退款记录也将被删除！\n此操作不可撤销！",
                                    QMessageBox.Yes | QMessageBox.No, 
                                    QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            if self.db.delete_store(store_id):
                # 刷新店铺列表
                self.load_stores()
                
                # 强制刷新订单记录表格（清除所有缓存和筛选条件）
                if hasattr(self, 'load_table_data'):
                    self.load_table_data(force_reload=True)
                
                # 刷新搜索筛选区的店铺选择
                if hasattr(self, 'search_store_combo'):
                    # 重新加载店铺列表，搜索筛选区的下拉框会自动更新
                    self.load_stores()
                    self.search_store_combo.setCurrentIndex(0)  # 重置为"全部"
                
                # 重置信息录入区的选择
                if self.store_combo.count() > 0:
                    self.store_combo.setCurrentIndex(0)
                else:
                    self.store_combo.addItem("请先添加店铺", None)
                    self.store_combo.setCurrentIndex(0)
                
                # 禁用编辑和删除按钮
                if hasattr(self, 'edit_store_btn') and hasattr(self, 'delete_store_btn'):
                    self.edit_store_btn.setEnabled(False)
                    self.delete_store_btn.setEnabled(False)
                
                # 刷新所有统计信息
                if hasattr(self, 'update_store_stats_display'):
                    self.update_store_stats_display()
                
                if hasattr(self, 'update_status_bar'):
                    self.update_status_bar()
                
                if hasattr(self, 'update_total_amount_display'):
                    self.update_total_amount_display()
                
                self.show_tooltip(f"店铺 {current_store} 及其所有数据已删除", "rgba(244, 67, 54, 0.95)", 2000)  # 红色气泡
            else:
                QMessageBox.warning(self, "错误", "删除店铺失败！")

    def get_current_date(self):
        return datetime.now().strftime("%Y-%m-%d")

    def update_debug_label(self, record_count, order_no, reason, store_name):
        """更新调试标签显示当前筛选结果"""
        debug_text = f"表格区域 - 当前显示 {record_count} 条订单记录"
        
        # 如果有筛选条件，显示筛选信息
        conditions = []
        if order_no:
            conditions.append(f"订单号: {order_no}")
        if reason and reason != "全部":
            conditions.append(f"退款原因: {reason}")
        if store_name and store_name != "全部":
            conditions.append(f"店铺: {store_name}")
        
        if conditions:
            debug_text += f" | 筛选条件: {' | '.join(conditions)}"
        
        self.debug_label.setText(debug_text)

    def parse_date_string(self, date_str):
        """解析多种日期格式，支持带时间格式，返回标准格式 YYYY-MM-DD"""
        date_str = str(date_str).strip()
        
        # 如果已经是标准格式，直接返回
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except:
            pass
        
        # 0. 处理带时间的格式：2026-03-16 09:47:44、2026/03/16 09:47:44、2026.03.16 09:47:44
        time_formats = [
            '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y.%m.%d %H:%M:%S',
            '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M', '%Y.%m.%d %H:%M'
        ]
        
        for fmt in time_formats:
            try:
                parsed_datetime = datetime.strptime(date_str, fmt)
                return parsed_datetime.strftime('%Y-%m-%d')
            except:
                continue
        
        # 1. 处理斜杠分隔格式：2026/3/2、2026/03/02、3/13、3/14
        if '/' in date_str:
            parts = date_str.split('/')
            if len(parts) == 3:
                # 格式：2026/3/2 或 2026/03/02
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return f"{year:04d}-{month:02d}-{day:02d}"
            elif len(parts) == 2:
                # 格式：3/13、3/14（自动识别今年年份）
                current_year = datetime.now().year
                month = int(parts[0])
                day = int(parts[1])
                return f"{current_year:04d}-{month:02d}-{day:02d}"
        
        # 2. 处理点分隔格式：3.13、3.14、3.15（自动识别今年年份）
        elif '.' in date_str:
            parts = date_str.split('.')
            if len(parts) == 2:
                current_year = datetime.now().year
                month = int(parts[0])
                day = int(parts[1])
                return f"{current_year:04d}-{month:02d}-{day:02d}"
            elif len(parts) == 3:
                # 格式：2026.3.2 或 2026.03.02
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return f"{year:04d}-{month:02d}-{day:02d}"
        
        # 3. 处理横杠分隔格式：3-13、3-14、2026-3-2
        elif '-' in date_str:
            parts = date_str.split('-')
            if len(parts) == 2:
                # 格式：3-13、3-14（自动识别今年年份）
                current_year = datetime.now().year
                month = int(parts[0])
                day = int(parts[1])
                return f"{current_year:04d}-{month:02d}-{day:02d}"
            elif len(parts) == 3:
                # 格式：2026-3-2 或 2026-03-02
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return f"{year:04d}-{month:02d}-{day:02d}"
        
        # 4. 处理中文格式：2026年3月2日、3月13日、3月14日
        if '年' in date_str and '月' in date_str and '日' in date_str:
            import re
            match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str)
            if match:
                year = int(match.group(1))
                month = int(match.group(2))
                day = int(match.group(3))
                return f"{year:04d}-{month:02d}-{day:02d}"
            
            match = re.search(r'(\d{1,2})月(\d{1,2})日', date_str)
            if match:
                current_year = datetime.now().year
                month = int(match.group(1))
                day = int(match.group(2))
                return f"{current_year:04d}-{month:02d}-{day:02d}"
        
        # 5. 处理无分隔符格式：20260302、0302（自动识别今年年份）
        if date_str.isdigit():
            if len(date_str) == 8:
                # 格式：20260302
                year = int(date_str[:4])
                month = int(date_str[4:6])
                day = int(date_str[6:8])
                return f"{year:04d}-{month:02d}-{day:02d}"
            elif len(date_str) == 4:
                # 格式：0302（3月2日）、0313（3月13日）
                current_year = datetime.now().year
                month = int(date_str[:2])
                day = int(date_str[2:4])
                return f"{current_year:04d}-{month:02d}-{day:02d}"
        
        # 6. 尝试常见日期格式解析
        common_formats = [
            '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y',
            '%Y-%m-%d', '%m-%d-%Y', '%d-%m-%Y',
            '%Y.%m.%d', '%m.%d.%Y', '%d.%m.%Y',
            '%Y年%m月%d日', '%m月%d日',
            '%Y%m%d'
        ]
        
        for fmt in common_formats:
            try:
                parsed_date = datetime.strptime(date_str, fmt)
                return parsed_date.strftime('%Y-%m-%d')
            except:
                continue
        
        # 如果所有格式都解析失败，返回当前日期
        return self.get_current_date()

    def _extract_mapped_value(self, row, column_mapping, field_name, default=''):
        """根据列映射读取 Excel 行值。"""
        actual_column = column_mapping.get(field_name)
        if not actual_column:
            return default
        return row.get(actual_column, default)

    def _parse_import_date_value(self, raw_value):
        """解析导入日期，失败时返回今天，避免写入空日期导致记录不可见。"""
        if raw_value in (None, ''):
            return self.get_current_date()

        if isinstance(raw_value, datetime):
            return raw_value.strftime('%Y-%m-%d')

        try:
            return self.parse_date_string(raw_value)
        except Exception:
            return self.get_current_date()

    def _parse_import_datetime_value(self, raw_value):
        """解析导入时间，成功时保留到秒，失败返回空字符串。"""
        if raw_value in (None, ''):
            return ''

        if isinstance(raw_value, datetime):
            return raw_value.strftime('%Y-%m-%d %H:%M:%S')

        text = str(raw_value).strip()
        if not text:
            return ''

        formats = [
            '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y.%m.%d %H:%M:%S',
            '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M', '%Y.%m.%d %H:%M',
            '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
            '%m-%d %H:%M:%S', '%m/%d %H:%M:%S', '%m.%d %H:%M:%S',
            '%m-%d %H:%M', '%m/%d %H:%M', '%m.%d %H:%M',
            '%H:%M:%S', '%H:%M'
        ]
        for fmt in formats:
            try:
                parsed = datetime.strptime(text, fmt)
                if fmt.startswith('%m'):
                    parsed = parsed.replace(year=datetime.now().year)
                elif fmt.startswith('%H'):
                    now = datetime.now()
                    parsed = parsed.replace(year=now.year, month=now.month, day=now.day)
                return parsed.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                continue
        return ''

    @staticmethod
    def _format_export_process_time(value):
        if not value:
            return ''
        try:
            parsed = datetime.strptime(str(value), '%Y-%m-%d %H:%M:%S')
            return parsed.strftime('%m-%d %H:%M:%S')
        except Exception:
            return ''

    @staticmethod
    def _format_refund_process_duration(apply_time, agree_time):
        if not apply_time or not agree_time:
            return ''
        try:
            start = datetime.strptime(str(apply_time), '%Y-%m-%d %H:%M:%S')
            end = datetime.strptime(str(agree_time), '%Y-%m-%d %H:%M:%S')
        except Exception:
            return ''
        seconds = int((end - start).total_seconds())
        if seconds < 0:
            return ''
        if seconds < 60:
            return f"{seconds}秒"

        hours, remainder = divmod(seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        if hours:
            parts = [f"{hours}小时"]
            if minutes:
                parts.append(f"{minutes}分")
            if seconds:
                parts.append(f"{seconds}秒")
            return ''.join(parts)
        if seconds:
            return f"{minutes}分{seconds}秒"
        return f"{minutes}分钟"

    def _resolve_import_record_date(self, row, column_mapping):
        """严格按用户确认后的映射列读取登记日期。"""
        raw_value = self._extract_mapped_value(row, column_mapping, '登记日期', '')
        return self._parse_import_date_value(raw_value)

    def _coerce_import_bool(self, value):
        if isinstance(value, str):
            return value.strip() in ['是', 'True', 'true', '1', 'TRUE']
        return bool(value)

    def _coerce_import_float(self, value, default=0.0):
        try:
            return float(value) if value not in (None, '') else default
        except Exception:
            return default

    def _render_table_row(self, row, rec):
        """渲染单行表格数据。"""
        store_color = self.db.get_store_color(rec['store_name'])

        store_item = QTableWidgetItem(rec['store_name'])
        if store_color:
            store_item.setBackground(QColor(store_color))
        self.table.setItem(row, 0, store_item)

        order_item = QTableWidgetItem(rec['order_no'])
        if store_color:
            order_item.setBackground(QColor(store_color))
        self.table.setItem(row, 1, order_item)

        spec_code_text = str(rec.get('spec_code') or '').strip() or '-'
        spec_item = QTableWidgetItem(spec_code_text)
        spec_item.setTextAlignment(Qt.AlignCenter)
        spec_item.setToolTip(str(rec.get('spec_name') or '').strip() or "未识别规格名称")
        if store_color:
            spec_item.setBackground(QColor(store_color))
        self.table.setItem(row, 2, spec_item)

        reason_item = QTableWidgetItem(rec['reason'])
        reason_item.setToolTip(self._build_reason_tooltip_text(rec))
        if store_color:
            reason_item.setBackground(QColor(store_color))
        self.table.setItem(row, 3, reason_item)

        amount_item = QTableWidgetItem(f"¥{rec['refund_amount']:.2f}")
        amount_item.setTextAlignment(Qt.AlignCenter)
        if store_color:
            amount_item.setBackground(QColor(store_color))
        self.table.setItem(row, 4, amount_item)

        cancel_text = "是" if rec['cancel'] else "否"
        cancel_item = QTableWidgetItem(cancel_text)
        cancel_item.setBackground(QColor("#4CAF50" if rec['cancel'] else "#F44336"))
        cancel_item.setForeground(QColor("white"))
        cancel_item.setTextAlignment(Qt.AlignCenter)
        if not rec['cancel']:
            reason_text = str(rec.get('quality_not_cancelled_reason') or '').strip() or "未明确说明未撤销原因"
            tooltip_lines = [f"未撤销原因：{reason_text}"]
            detail_text = str(rec.get('quality_refund_reason_detail') or '').strip()
            updated_at = str(rec.get('quality_refund_reason_updated_at') or '').strip()
            if detail_text:
                tooltip_lines.append(f"识别说明：{detail_text}")
            if updated_at:
                tooltip_lines.append(f"识别时间：{updated_at}")
            cancel_item.setToolTip("\n".join(tooltip_lines))
        self.table.setItem(row, 5, cancel_item)

        comp_text = "是" if rec['compensate'] else "否"
        comp_item = QTableWidgetItem(comp_text)
        comp_item.setBackground(QColor("#4CAF50" if rec['compensate'] else "#F44336"))
        comp_item.setForeground(QColor("white"))
        comp_item.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(row, 6, comp_item)

        comp_amount_item = QTableWidgetItem(f"¥{rec['comp_amount']:.2f}")
        comp_amount_item.setTextAlignment(Qt.AlignCenter)
        if store_color:
            comp_amount_item.setBackground(QColor(store_color))
        self.table.setItem(row, 7, comp_amount_item)

        reject_text = "是" if rec['reject'] else "否"
        reject_item = QTableWidgetItem(reject_text)
        reject_item.setBackground(QColor("#4CAF50" if rec['reject'] else "#F44336"))
        reject_item.setForeground(QColor("white"))
        reject_item.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(row, 8, reject_item)

        reject_result_item = QTableWidgetItem(self._display_reject_result_value(rec))
        reject_result_item.setTextAlignment(Qt.AlignCenter)
        if store_color:
            reject_result_item.setBackground(QColor(store_color))
        self.table.setItem(row, 9, reject_result_item)

        date_item = QTableWidgetItem(rec['record_date'])
        if store_color:
            date_item.setBackground(QColor(store_color))
        self.table.setItem(row, 10, date_item)

        notes_item = QTableWidgetItem(rec['notes'])
        if store_color:
            notes_item.setBackground(QColor(store_color))
        self.table.setItem(row, 11, notes_item)

        self._bind_record_id_to_row(row, rec.get('id'))

        if rec['order_no'] in self.highlighted_orders:
            for col in range(12):
                if col in [5, 6, 8, 9]:
                    continue
                if self.table.item(row, col):
                    self.table.item(row, col).setBackground(QColor("#FFD700"))

    def _populate_table(self, records, update_chart=True):
        """统一填充表格数据，避免 load_table_data/show_all_time 分叉。"""
        try:
            self.table.cellChanged.disconnect(self.on_cell_changed)
        except TypeError:
            pass

        current_row_count = self.table.rowCount()
        new_row_count = len(records)
        if new_row_count != current_row_count:
            self.table.setRowCount(new_row_count)

        for row, rec in enumerate(records):
            if self._should_update_row(row, rec):
                self._render_table_row(row, rec)
            else:
                self._bind_record_id_to_row(row, rec.get('id'))

        self.table.cellChanged.connect(self.on_cell_changed)
        self._restore_reject_display_after_load()
        self._update_all_statistics(records)
        if update_chart:
            self.update_current_chart(records)

    def _sync_cached_record(self, updated_record):
        """同步缓存中的单条记录，避免局部更新后又被旧缓存覆盖。"""
        if not updated_record or self._cached_records is None:
            return

        for idx, record in enumerate(self._cached_records):
            if record.get('id') == updated_record.get('id'):
                self._cached_records[idx] = updated_record
                return

    def _refresh_statistics_incremental(self):
        """局部编辑后只重算统计与图表，不重绘整张表。"""
        records = self.get_filtered_records()
        self._cached_records = records
        self._last_search_params = self._get_current_search_params()
        self.update_statusbar(records)
        self.update_total_amount_display()
        self.update_store_stats_display()
        self.update_current_chart(records)

    def _refresh_row_by_record_id(self, record_id, refresh_statistics=True):
        """按记录ID重渲染单行，找不到行时退回整表刷新。"""
        row = self.get_row_from_record_id(record_id)
        record = self.db.get_record_by_id(record_id)
        if row is None or not record:
            self._cached_records = None
            self._last_search_params = None
            self.load_table_data(force_reload=True)
            return False

        try:
            self.table.cellChanged.disconnect(self.on_cell_changed)
        except TypeError:
            pass

        self._render_table_row(row, record)
        self.table.cellChanged.connect(self.on_cell_changed)
        self._restore_reject_display_after_load()
        self._sync_cached_record(record)

        if refresh_statistics:
            self._refresh_statistics_incremental()
        return True

    def add_record(self):
        """添加记录"""
        try:
            store_id = self.store_combo.currentData()
            if store_id is None:
                QMessageBox.warning(self, "警告", "请选择店铺！")
                return
            order_no = self.order_no_edit.text().strip()
            if not order_no:
                QMessageBox.warning(self, "警告", "订单号不能为空！")
                return
            reason = self.reason_combo.currentText()
            if not reason:
                QMessageBox.warning(self, "警告", "请选择退款原因！")
                return
            try:
                refund_amount = float(self.refund_amount_edit.text().strip())
            except ValueError:
                QMessageBox.warning(self, "警告", "退款金额必须为有效数字！")
                return
            cancel = self.cancel_check.isChecked()
            compensate = self.compensate_check.isChecked()
            comp_amount = 0.0
            if compensate:
                try:
                    comp_amount_text = self.comp_amount_edit.text().strip()
                    comp_amount = float(comp_amount_text) if comp_amount_text else 0.0
                except ValueError:
                    QMessageBox.warning(self, "警告", "补偿金额必须为有效数字！")
                    return
            
            reject = self.reject_check.isChecked()
            reject_result = ""
            if reject:
                reject_result = self.reject_result_combo.currentText()
                if reject_result == "-":
                    reject_result = "驳回成功"
            
            notes = self.notes_edit.text().strip()
            
            record_date = self.get_current_date()
            existing = self.db.get_record_by_order_no(order_no)
            if existing:
                QMessageBox.warning(self, "警告", f"订单号 {order_no} 已存在，无法重复添加！")
                return

            self.db.add_record(store_id, order_no, reason, refund_amount, cancel, compensate, comp_amount, reject, reject_result, notes, record_date)
            self.show_tooltip("已添加", "rgba(76, 175, 80, 0.95)", 1000)
            self.clear_input()
            self.load_table_data(force_reload=True)

        except Exception as e:
            import traceback
            print(f"[ERROR] add_record: 捕获到异常: {type(e).__name__}: {e}")
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"添加记录时发生错误：{type(e).__name__}: {e}")
            return

    def update_record(self):
        """更新记录"""
        if self.current_record_id is None:
            QMessageBox.warning(self, "警告", "请先在表格中选中要更新的记录！")
            return
        store_id = self.store_combo.currentData()
        if store_id is None:
            QMessageBox.warning(self, "警告", "请选择店铺！")
            return
        order_no = self.order_no_edit.text().strip()
        if not order_no:
            QMessageBox.warning(self, "警告", "订单号不能为空！")
            return
        reason = self.reason_combo.currentText()
        if not reason:
            QMessageBox.warning(self, "警告", "请选择退款原因！")
            return
        try:
            refund_amount = float(self.refund_amount_edit.text())
        except ValueError:
            QMessageBox.warning(self, "警告", "退款金额必须为有效数字！")
            return
        cancel = self.cancel_check.isChecked()
        compensate = self.compensate_check.isChecked()
        comp_amount = 0.0
        if compensate:
            try:
                comp_amount = float(self.comp_amount_edit.text()) if self.comp_amount_edit.text() else 0.0
            except ValueError:
                QMessageBox.warning(self, "警告", "补偿金额必须为有效数字！")
                return
        # 驳回相关字段
        reject = self.reject_check.isChecked()
        reject_result = ""
        if reject:
            reject_result = self.reject_result_combo.currentText()
            if reject_result == "-":
                reject_result = "驳回成功"
        
        notes = self.notes_edit.text().strip()
        
        existing_record = self.db.get_record_by_id(self.current_record_id)
        record_date = self.get_current_date()

        self.db.update_record(
            self.current_record_id, store_id, order_no, reason, refund_amount, cancel,
            compensate, comp_amount, reject, reject_result, notes, record_date,
            existing_record.get('order_status', '') if existing_record else '',
            existing_record.get('after_sale_status', '') if existing_record else '',
            existing_record.get('spec_name', '') if existing_record else '',
            existing_record.get('spec_code', '') if existing_record else ''
        )
        self.show_tooltip("已更新", "rgba(76, 175, 80, 0.95)", 1000)  # 绿色气泡显示1秒
        
        # 不清空输入区域，保持当前记录显示
        # 强制刷新表格数据，让用户看到更新效果
        self.load_table_data(force_reload=True)
        
        # 重新选中当前记录，让用户看到更新后的状态
        self._select_current_record_after_update()

    def delete_record(self):
        """删除选中的记录（支持多选删除）"""
        # 获取所有选中的行
        selected_rows = self.table.selectionModel().selectedRows()
        
        if not selected_rows:
            QMessageBox.warning(self, "警告", "请先在表格中选中要删除的记录！")
            return
        
        # 获取选中行的记录ID
        record_ids = []
        for index in selected_rows:
            row = index.row()
            record_id = self.get_record_id_from_row(row)
            if record_id:
                record_ids.append(record_id)

        # 去重并保持顺序，避免重复删除同一条记录
        record_ids = list(dict.fromkeys(record_ids))
        
        if not record_ids:
            QMessageBox.warning(self, "警告", "无法获取选中记录的ID！")
            return
        
        # 确认删除对话框
        if len(record_ids) == 1:
            message = "确定要删除这条记录吗？"
        else:
            message = f"确定要删除选中的 {len(record_ids)} 条记录吗？"
            
        reply = QMessageBox.question(self, "确认删除", message,
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            total_before_delete = self.db.get_total_record_count()
            selected_visible_count = len(record_ids)

            # 批量删除记录
            success_count = 0
            failed_ids = []
            
            for record_id in record_ids:
                try:
                    if self.db.delete_record(record_id):
                        success_count += 1
                    else:
                        # 记录删除失败的ID和原因
                        failed_ids.append((record_id, "数据库删除操作返回失败"))
                except Exception as e:
                    error_msg = f"删除记录 {record_id} 时出错: {str(e)}"
                    print(error_msg)
                    failed_ids.append((record_id, error_msg))
            
            if success_count > 0:
                total_after_delete = self.db.get_total_record_count()
                remaining_db_count = max(total_after_delete, 0)

                if success_count == 1:
                    QMessageBox.information(self, "成功", "记录已删除！")
                else:
                    result_message = f"已成功删除 {success_count} 条记录！"

                    expected_after_delete = max(total_before_delete - success_count, 0)
                    if total_after_delete != expected_after_delete:
                        result_message += (
                            f"\n\n注意：删除前数据库共 {total_before_delete} 条，"
                            f"删除后还有 {total_after_delete} 条。"
                            "\n这说明这次并不是把数据库所有记录都删空了。"
                        )
                    elif remaining_db_count > 0 and selected_visible_count == self.table.rowCount():
                        result_message += (
                            f"\n\n当前表格已删空，但数据库里仍有 {remaining_db_count} 条其他筛选范围的记录。"
                            "\n如果要彻底删空，请切到“全部店铺 + 全部时间”后再删除。"
                        )

                    QMessageBox.information(self, "成功", result_message)
                
                # 清除输入并刷新表格
                self.clear_input()
                self._cached_records = None
                self._last_search_params = None
                # 强制刷新表格数据（确保删除后立即消失）
                self.load_table_data(force_reload=True)
                # 强制刷新表格显示
                self.table.viewport().update()
                
                # 如果有失败的删除，显示详细警告
                if failed_ids:
                    # 构建详细的失败信息
                    failed_info = f"成功删除 {success_count} 条记录，但 {len(failed_ids)} 条记录删除失败！\n\n"
                    failed_info += "失败记录详情：\n"
                    
                    for i, (record_id, error_msg) in enumerate(failed_ids[:5]):  # 最多显示5条
                        failed_info += f"{i+1}. 记录ID: {record_id} - 原因: {error_msg}\n"
                    
                    if len(failed_ids) > 5:
                        failed_info += f"...等{len(failed_ids) - 5}条记录失败\n"
                    
                    failed_info += "\n建议：请检查数据库连接或重启程序后重试。"
                    
                    QMessageBox.warning(self, "部分删除失败", failed_info)
            else:
                QMessageBox.warning(self, "错误", "所有记录删除失败！")

    def refund_amount_mouse_press(self, event):
        """退款金额输入框鼠标点击事件 - 只在有内容时自动全选"""
        # 只有当输入框有内容时才自动全选
        if self.refund_amount_edit.text():
            self.refund_amount_edit.selectAll()
        # 调用原始的鼠标点击事件
        QLineEdit.mousePressEvent(self.refund_amount_edit, event)

    def order_no_mouse_press(self, event):
        """订单号输入框鼠标点击事件 - 自动全选文本"""
        self.order_no_edit.selectAll()
        # 调用原始的鼠标点击事件
        QLineEdit.mousePressEvent(self.order_no_edit, event)

    def comp_amount_mouse_press(self, event):
        """补偿金额输入框鼠标点击事件 - 自动全选文本"""
        self.comp_amount_edit.selectAll()
        # 调用原始的鼠标点击事件
        QLineEdit.mousePressEvent(self.comp_amount_edit, event)

    def search_order_mouse_press(self, event):
        """搜索订单号输入框鼠标点击事件 - 自动全选文本"""
        self.search_order_edit.selectAll()
        # 调用原始的鼠标点击事件
        QLineEdit.mousePressEvent(self.search_order_edit, event)

    def clear_input(self):
        """清空输入区域"""
        self.store_combo.blockSignals(True)
        self.store_combo.setCurrentIndex(0)
        self.store_combo.blockSignals(False)
        self.order_no_edit.clear()
        self.reason_combo.setCurrentIndex(0)
        self.refund_amount_edit.clear()
        self.cancel_check.setChecked(False)
        self.compensate_check.setChecked(False)
        self.comp_amount_edit.clear()
        self.current_record_id = None
        self.table.clearSelection()

    def get_filtered_records(self):
        """获取当前筛选条件下的记录（与表格显示的数据相同）"""
        try:
            order_no = self.search_order_edit.text()
            
            reasons = []
            if hasattr(self, 'search_reason_dropdown'):
                reasons = list(self.search_reason_dropdown.selected_items)
            
            cancel = self.search_cancel_combo.currentText()
            reject = self.search_reject_combo.currentText()
            store_name = self.search_store_combo.currentText()
            start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
            end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
            real_reason = ""
            if hasattr(self, 'search_real_reason_edit') and self.search_real_reason_edit is not None:
                real_reason = self.search_real_reason_edit.text().strip()

            reason_param = "全部" if not reasons else reasons
            
            return self.db.search_records(
                order_no, reason_param, cancel, reject, start_date, end_date, store_name, real_reason
            )
        except Exception as e:
            print(f"[ERROR] get_filtered_records: 异常: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            return []

    def load_table_data(self, force_reload=False):
        """加载表格数据（根据筛选条件）"""
        print("[DEBUG] load_table_data: 开始执行")
        try:
            # 检查必要的控件是否已初始化
            if not hasattr(self, 'table') or self.table is None:
                print("[DEBUG] table 尚未初始化，跳过加载表格数据")
                return
            if not hasattr(self, 'search_order_edit') or self.search_order_edit is None:
                print("[DEBUG] search_order_edit 尚未初始化，跳过加载表格数据")
                return
            if not hasattr(self, 'search_store_combo') or self.search_store_combo is None:
                print("[DEBUG] search_store_combo 尚未初始化，跳过加载表格数据")
                return
            
            if force_reload:
                self._cached_records = None
                self._last_search_params = None
            
            current_params = self._get_current_search_params()
            if self._last_search_params == current_params and self._cached_records is not None:
                records = self._cached_records
                print("[DEBUG] load_table_data: 使用缓存数据")
            else:
                print("[DEBUG] load_table_data: 重新查询数据库")
                records = self.get_filtered_records()
                self._cached_records = records
                self._last_search_params = current_params
            
            try:
                self.table.cellChanged.disconnect(self.on_cell_changed)
            except TypeError:
                pass
            
            order_no = self.search_order_edit.text()
            
            if hasattr(self, 'search_reason_dropdown'):
                self.selected_reasons = self.search_reason_dropdown.selected_items
            else:
                self.selected_reasons = set()
            
            store_name = self.search_store_combo.currentText()
            
            self.update_debug_label(len(records), order_no, str(len(self.selected_reasons)) + "个原因", store_name)
            self._populate_table(records, update_chart=True)
            
            print("[DEBUG] load_table_data: 执行完成")

        except Exception as e:
            import traceback
            print(f"[ERROR] load_table_data: 捕获到异常: {type(e).__name__}: {e}")
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"加载表格数据时发生错误：{type(e).__name__}: {e}")

    def update_statusbar(self, records):
        """更新状态栏统计"""
        total = len(records)
        total_refund = sum(
            self._safe_float(r.get('refund_amount', 0))
            for r in records if self._is_effective_refund_record(r)
        )
        total_comp = sum(
            self._safe_float(r.get('comp_amount', 0))
            for r in records if self._has_compensation_record(r)
        )
        # 总金额：退款金额 + 补偿金额
        total_amount = total_refund + total_comp
        cancel_count = sum(1 for r in records if r['cancel'])
        # 计算驳回相关统计
        reject_count = sum(1 for r in records if r['reject'])
        reject_success_count = sum(1 for r in records if self._is_reject_success_record(r))
        reject_fail_count = sum(1 for r in records if self._is_reject_failure_record(r))
        
        self.status_bar.showMessage(
            f"记录总数: {total} | 退款总额: ¥{total_refund:,.2f} | 补偿总额: ¥{total_comp:,.2f} | "
            f"总金额: ¥{total_amount:,.2f} | 撤销订单: {cancel_count}单 | 驳回: {reject_count}单(成功:{reject_success_count}/失败:{reject_fail_count})"
        )

    def on_search_changed(self):
        """搜索条件变化时自动搜索（实时搜索）"""
        # 性能优化：复用已有的定时器，避免重复创建
        self._search_timer.stop()
        
        def update_data():
            # 性能优化：合并数据库查询，一次搜索获取所有数据
            self.load_table_data()
            # 不再单独调用 update_store_stats_display()，因为 load_table_data() 中已经包含统计更新
        
        # 性能优化：确保定时器连接正确
        try:
            self._search_timer.timeout.disconnect()
        except:
            pass
        self._search_timer.timeout.connect(update_data)
        self._search_timer.start(800)

    def _restore_reject_display_after_load(self):
        """表格加载完成后恢复进行中的驳回流程显示和48小时提醒工具提示"""
        if not hasattr(self, 'reject_manager') or not self.reject_manager:
            return
        
        # 临时断开 cellChanged 信号，避免递归
        try:
            self.table.cellChanged.disconnect(self.on_cell_changed)
        except TypeError:
            pass
        
        try:
            # 遍历所有进行中的驳回流程
            for order_no, process_info in self.reject_manager.active_processes.items():
                # 在表格中查找该订单
                for row in range(self.table.rowCount()):
                    order_item = self.table.item(row, 1)  # 订单号列
                    if order_item and order_item.text() == order_no:
                        current_round = process_info['round']
                        # 直接设置单元格文本，不调用 update_reject_result_display 避免递归
                        reject_result_item = self.table.item(row, 9)  # 驳回结果列
                        if reject_result_item:
                            if current_round == 1:
                                reject_result_item.setText("第一轮驳回中...")
                                reject_result_item.setBackground(QColor("#FFF3E0"))
                                reject_result_item.setForeground(QColor("#E65100"))
                            elif current_round == 2:
                                reject_result_item.setText("第二轮驳回中...")
                                reject_result_item.setBackground(QColor("#E3F2FD"))
                                reject_result_item.setForeground(QColor("#1565C0"))
                        break
            
            # 遍历所有表格行，为"驳回成功"且设置了48小时提醒的订单添加工具提示
            for row in range(self.table.rowCount()):
                order_item = self.table.item(row, 1)  # 订单号列
                if not order_item:
                    continue
                order_no = order_item.text()
                
                reject_result_item = self.table.item(row, 9)  # 驳回结果列
                if not reject_result_item:
                    continue
                
                result_text = reject_result_item.text()
                
                # 如果是驳回成功，检查是否有48小时提醒
                if "驳回成功" in result_text:
                    reminder_info = self.reject_manager.get_48h_reminder_info(order_no)
                    if reminder_info:
                        end_time = reminder_info['end_time']
                        remaining_hours = reminder_info['remaining_hours']
                        remind_time_str = end_time.strftime("%m-%d %H:%M")
                        tooltip = (
                            f"✅ 驳回成功\n"
                            f"⏰ 48小时提醒剩余: 约{remaining_hours}小时\n"
                            f"📅 提醒时间: {remind_time_str}\n"
                            f"⚠️ 请及时检查是否被平台介入退款\n"
                            f"💡 双击可标记平台介入退款"
                        )
                        reject_result_item.setToolTip(tooltip)
                        reject_result_item.setBackground(QColor("#E8F5E9"))
                        reject_result_item.setForeground(QColor("#2E7D32"))
                    else:
                        reject_result_item.setToolTip("✅ 驳回成功\n💡 双击可标记平台介入退款")
                        reject_result_item.setBackground(QColor("#E8F5E9"))
                        reject_result_item.setForeground(QColor("#2E7D32"))
                elif "驳回失败" in result_text:
                    reject_result_item.setToolTip("❌ 平台已介入退款")
                    reject_result_item.setBackground(QColor("#FFEBEE"))
                    reject_result_item.setForeground(QColor("#C62828"))
                    
        finally:
            # 恢复 cellChanged 信号连接
            self.table.cellChanged.connect(self.on_cell_changed)

    def _get_current_search_params(self):
        """获取当前搜索参数（用于缓存检查）"""
        # 获取多选的退款原因
        reasons = []
        if hasattr(self, 'search_reason_dropdown'):
            reasons = list(self.search_reason_dropdown.selected_items)
        real_reason = ""
        if hasattr(self, 'search_real_reason_edit') and self.search_real_reason_edit is not None:
            real_reason = self.search_real_reason_edit.text().strip()
        
        return (
            self.search_order_edit.text(),
            real_reason,
            tuple(reasons) if reasons else (),
            self.search_cancel_combo.currentText(),
            self.search_reject_combo.currentText(),
            self.search_store_combo.currentText(),
            self.start_date_edit.date().toString("yyyy-MM-dd"),
            self.end_date_edit.date().toString("yyyy-MM-dd")
        )

    def _should_update_row(self, row, record):
        """检查是否需要更新指定行（增量更新优化）"""
        if row >= self.table.rowCount():
            return True
        
        try:
            current_store = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
            current_order = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
            
            if current_store != record['store_name'] or current_order != record['order_no']:
                return True
            
            # 检查退款原因
            current_reason = self.table.item(row, 3).text() if self.table.item(row, 3) else ""
            if current_reason != record['reason']:
                return True
            
            # 检查退款金额
            current_amount_text = self.table.item(row, 4).text() if self.table.item(row, 4) else "¥0.00"
            current_amount = float(current_amount_text.replace('¥', '').replace(',', '')) if current_amount_text else 0.0
            if abs(current_amount - record['refund_amount']) > 0.01:
                return True
            
            current_cancel = self.table.item(row, 5).text() if self.table.item(row, 5) else ""
            current_compensate = self.table.item(row, 6).text() if self.table.item(row, 6) else ""
            current_reject = self.table.item(row, 8).text() if self.table.item(row, 8) else ""
            
            expected_cancel = "是" if record['cancel'] else "否"
            expected_compensate = "是" if record['compensate'] else "否"
            expected_reject = "是" if record['reject'] else "否"
            
            if (current_cancel != expected_cancel or 
                current_compensate != expected_compensate or 
                current_reject != expected_reject):
                return True

            current_cancel_tooltip = self.table.item(row, 5).toolTip() if self.table.item(row, 5) else ""
            expected_cancel_tooltip = ""
            if not record['cancel']:
                reason_text = str(record.get('quality_not_cancelled_reason') or '').strip() or "未明确说明未撤销原因"
                tooltip_lines = [f"未撤销原因：{reason_text}"]
                detail_text = str(record.get('quality_refund_reason_detail') or '').strip()
                updated_at = str(record.get('quality_refund_reason_updated_at') or '').strip()
                if detail_text:
                    tooltip_lines.append(f"识别说明：{detail_text}")
                if updated_at:
                    tooltip_lines.append(f"识别时间：{updated_at}")
                expected_cancel_tooltip = "\n".join(tooltip_lines)
            if current_cancel_tooltip != expected_cancel_tooltip:
                return True
            
            # 检查补偿金额
            current_comp_amount_text = self.table.item(row, 7).text() if self.table.item(row, 7) else "¥0.00"
            current_comp_amount = float(current_comp_amount_text.replace('¥', '').replace(',', '')) if current_comp_amount_text else 0.0
            if abs(current_comp_amount - record['comp_amount']) > 0.01:
                return True
            
            # 检查驳回结果
            current_reject_result = self.table.item(row, 9).text() if self.table.item(row, 9) else ""
            if current_reject_result != self._display_reject_result_value(record):
                return True
            
            # 检查日期
            current_date = self.table.item(row, 10).text() if self.table.item(row, 10) else ""
            if current_date != record['record_date']:
                return True
            
            current_spec_code = self.table.item(row, 2).text().strip() if self.table.item(row, 2) else "-"
            expected_spec_code = str(record.get('spec_code') or '').strip() or "-"
            if current_spec_code != expected_spec_code:
                return True

            current_reason_tooltip = self.table.item(row, 3).toolTip() if self.table.item(row, 3) else ""
            expected_reason_tooltip = self._build_reason_tooltip_text(record)
            if current_reason_tooltip != expected_reason_tooltip:
                return True

            # 检查备注
            current_notes = self.table.item(row, 11).text() if self.table.item(row, 11) else ""
            if current_notes != record['notes']:
                return True
            
            return False
        except Exception as e:
            print(f"[ERROR] _should_update_row: 异常: {type(e).__name__}: {e}")
            return True

    def _update_all_statistics(self, records):
        """合并更新所有统计信息（避免重复计算）"""
        # 更新状态栏统计
        self.update_statusbar(records)
        # 更新左下角总金额显示
        self.update_total_amount_display()
        # 更新店铺统计信息显示
        self.update_store_stats_display()
    
    def get_current_records_for_chart(self):
        """获取当前筛选条件下的记录用于图表显示"""
        # 获取当前筛选条件下的记录
        records = self.get_current_filtered_records()
        
        # 获取日期范围
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
        
        return records, start_date, end_date
    
    def update_current_chart(self, records=None):
        """更新当前图表显示"""
        if hasattr(self, 'chart_widget'):
            if records is None:
                records, start_date, end_date = self.get_current_records_for_chart()
            else:
                start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
                end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
            self.chart_widget.update_chart(records, start_date, end_date)

    def reset_search(self):
        """重置搜索条件"""
        self._search_timer.stop()
        blockers = self._create_search_signal_blockers()
        self.search_order_edit.clear()
        if hasattr(self, 'search_real_reason_edit') and self.search_real_reason_edit is not None:
            self.search_real_reason_edit.clear()
        self.search_store_combo.setCurrentIndex(0)  # 全部
        if hasattr(self, 'search_reason_dropdown'):
            self.search_reason_dropdown.clear_selection()  # 清空多选状态
        self.search_cancel_combo.setCurrentIndex(0)  # 全部
        self.search_reject_combo.setCurrentIndex(0)  # 全部
        today = QDate.currentDate()
        self.start_date_edit.setDate(today)
        self.end_date_edit.setDate(today)
        del blockers
        self.load_table_data()

    def show_all_records(self):
        """显示全部记录（清除所有筛选条件，强制重新加载）"""
        self._search_timer.stop()
        blockers = self._create_search_signal_blockers()

        # 清除所有筛选条件
        self.search_order_edit.clear()
        if hasattr(self, 'search_real_reason_edit') and self.search_real_reason_edit is not None:
            self.search_real_reason_edit.clear()
        self.search_store_combo.setCurrentIndex(0)  # 全部
        if hasattr(self, 'search_reason_dropdown'):
            self.search_reason_dropdown.clear_selection()  # 清空多选状态
        self.search_cancel_combo.setCurrentIndex(0)  # 全部
        self.search_reject_combo.setCurrentIndex(0)  # 全部
        
        # 设置日期为所有日期
        self.start_date_edit.setDate(QDate(2000, 1, 1))  # 很早的日期
        self.end_date_edit.setDate(QDate(2100, 12, 31))  # 很晚的日期
        del blockers
        
        # 强制重新加载所有数据
        self.load_table_data(force_reload=True)
        
        # 显示淡入淡出气泡提示信息
        total_count = self.table.rowCount()
        self.show_bubble_message(f"✅ 已显示全部记录！\n当前显示 {total_count} 条记录。")

    def set_quick_date(self, days):
        """快捷日期设置（近7天和近30天不包括今天）"""
        self._search_timer.stop()
        today = QDate.currentDate()
        if days == 0:  # 今天
            start = today
            end = today
        elif days == 1:  # 昨天
            start = today.addDays(-1)
            end = today.addDays(-1)
        else:
            # 近7天和近30天不包括今天，只计算完整一天的数据
            start = today.addDays(-days)  # 从昨天往前推days-1天
            end = today.addDays(-1)       # 到昨天为止
        blockers = self._create_search_signal_blockers()
        self.start_date_edit.setDate(start)
        self.end_date_edit.setDate(end)
        del blockers
        self.load_table_data()

    def set_last_full_week(self):
        """设置为上一个完整自然周（周一到周日）。"""
        self._search_timer.stop()
        today = QDate.currentDate()
        current_week_monday = today.addDays(1 - today.dayOfWeek())
        start = current_week_monday.addDays(-7)
        end = current_week_monday.addDays(-1)

        blockers = self._create_search_signal_blockers()
        self.start_date_edit.setDate(start)
        self.end_date_edit.setDate(end)
        del blockers
        self.load_table_data()

    def set_last_full_month(self):
        """设置为上一个完整自然月。"""
        self._search_timer.stop()
        today = QDate.currentDate()
        current_month_start = QDate(today.year(), today.month(), 1)
        end = current_month_start.addDays(-1)
        start = QDate(end.year(), end.month(), 1)

        blockers = self._create_search_signal_blockers()
        self.start_date_edit.setDate(start)
        self.end_date_edit.setDate(end)
        del blockers
        self.load_table_data()
    
    def show_all_time(self):
        """显示全部时间范围的记录（不触发时间曲线图自动刷新）"""
        self._search_timer.stop()
        blockers = self._create_search_signal_blockers()
        # 设置一个很大的日期范围来显示所有记录
        self.start_date_edit.setDate(QDate(2000, 1, 1))  # 很早的日期
        self.end_date_edit.setDate(QDate(2100, 12, 31))  # 很晚的日期
        del blockers
        
        # 手动加载数据，避免触发图表自动刷新
        records = self.get_filtered_records()
        self._populate_table(records, update_chart=False)
        
        # 显示提示信息
        total_count = len(records)
        self.show_bubble_message(f"📅 已显示全部时间范围的记录！\n当前显示 {total_count} 条记录。\n（时间曲线图未自动刷新）")

    def previous_day(self):
        """前一天：将当前日期范围往前移动一天"""
        current_start = self.start_date_edit.date()
        current_end = self.end_date_edit.date()
        
        # 如果开始日期和结束日期相同（单天选择）
        if current_start == current_end:
            new_date = current_start.addDays(-1)
            self.start_date_edit.setDate(new_date)
            self.end_date_edit.setDate(new_date)
        else:
            # 如果是多天选择，整体往前移动一天
            new_start = current_start.addDays(-1)
            new_end = current_end.addDays(-1)
            self.start_date_edit.setDate(new_start)
            self.end_date_edit.setDate(new_end)
        
        self.load_table_data()

    def next_day(self):
        """后一天：将当前日期范围往后移动一天"""
        current_start = self.start_date_edit.date()
        current_end = self.end_date_edit.date()
        
        # 如果开始日期和结束日期相同（单天选择）
        if current_start == current_end:
            new_date = current_start.addDays(1)
            # 检查是否超过今天
            today = QDate.currentDate()
            if new_date > today:
                new_date = today
            self.start_date_edit.setDate(new_date)
            self.end_date_edit.setDate(new_date)
        else:
            # 如果是多天选择，整体往后移动一天
            new_start = current_start.addDays(1)
            new_end = current_end.addDays(1)
            # 检查是否超过今天
            today = QDate.currentDate()
            if new_end > today:
                new_end = today
                new_start = new_end.addDays(-(current_end.daysTo(current_start)))
            self.start_date_edit.setDate(new_start)
            self.end_date_edit.setDate(new_end)
        
        self.load_table_data()

    def on_item_double_clicked(self, item):
        """双击表格项：根据列类型执行不同操作"""
        try:
            row = item.row()
            column = item.column()
            
            # 根据列类型执行不同操作
            if column == 0:  # 店铺名称列：录入信息
                self.load_record_to_input(row)
            elif column == 1:  # 订单号列：复制订单号
                self.copy_order_no(row)
            elif column == 2:  # 规格编码列：直接编辑
                self.table.editItem(item)
            elif column == 3:  # 退款原因列：无操作
                pass
            elif column == 4:  # 退款金额列：直接编辑
                self.table.editItem(item)
            elif column == 5:  # 撤销列：双击切换
                self.toggle_status_field(row, column)
            elif column == 6:  # 打款补偿列：双击切换
                self.toggle_status_field(row, column)
            elif column == 7:  # 补偿金额列：条件编辑
                if self.table.item(row, 6).text() == "是":  # 只有打款补偿为"是"时才能编辑
                    self.table.editItem(item)
            elif column == 8:  # 驳回列：双击切换
                self.toggle_status_field(row, column)
            elif column == 9:  # 驳回结果列：双击打开驳回流程管理
                self.on_reject_result_double_click(row, column)
            elif column == 10:  # 登记日期列：无操作
                pass
            elif column == 11:  # 备注列：直接编辑
                self.table.editItem(item)
        except Exception as e:
            # 捕获所有异常，防止程序崩溃
            QMessageBox.warning(self, "操作错误", f"双击操作失败：{str(e)}")
        
    def load_record_to_input(self, row):
        """将选中行的数据录入到输入框（只有双击店铺名称列时调用）"""
        # 安全检查：确保行号有效
        if row < 0 or row >= self.table.rowCount():
            return
            
        # 安全检查：只检查必要的列（前6列必须有数据，后4列可以为空）
        required_columns = [0, 1, 3, 4, 5, 6]  # 店铺名称、订单号、退款原因、退款金额、撤销、打款补偿
        for col in required_columns:
            if not self.table.item(row, col):
                QMessageBox.warning(self, "错误", f"第{col+1}列数据缺失，无法加载")
                return
        
        # 获取选中行的数据
        store_name = self.table.item(row, 0).text()
        order_no = self.table.item(row, 1).text()
        reason = self.table.item(row, 3).text()
        refund_amount_text = self.table.item(row, 4).text()
        cancel_text = self.table.item(row, 5).text()
        compensate_text = self.table.item(row, 6).text()
        comp_amount_text = self.table.item(row, 7).text()
        reject_text = self.table.item(row, 8).text()
        reject_result_text = self.table.item(row, 9).text()
        notes_text = self.table.item(row, 11).text()
        
        # 解析退款金额（去掉¥符号）
        try:
            refund_amount = float(refund_amount_text.replace('¥', '').strip())
        except:
            refund_amount = 0.0
            
        # 解析补偿金额
        try:
            comp_amount = float(comp_amount_text.replace('¥', '').strip()) if comp_amount_text else 0.0
        except:
            comp_amount = 0.0
            
        # 设置店铺
        store_index = self.store_combo.findText(store_name)
        if store_index >= 0:
            self.store_combo.setCurrentIndex(store_index)
            
        # 设置订单号
        self.order_no_edit.setText(order_no)
        
        # 设置退款原因
        reason_index = self.reason_combo.findText(reason)
        if reason_index >= 0:
            self.reason_combo.setCurrentIndex(reason_index)
        else:
            self.reason_combo.setCurrentIndex(0)
            
        # 设置退款金额
        self.refund_amount_edit.setText(f"{refund_amount:.2f}")
        
        # 设置撤销状态
        self.cancel_check.setChecked(cancel_text == "是")
        
        # 设置补偿状态和金额
        self.compensate_check.setChecked(compensate_text == "是")
        self.comp_amount_edit.setText(f"{comp_amount:.2f}" if comp_amount > 0 else "")
        
        # 设置驳回状态和结果
        self.reject_check.setChecked(reject_text == "是")
        reject_result_index = self.reject_result_combo.findText(reject_result_text)
        if reject_text != "是" or reject_result_text in ("", "无", None):
            self.reject_result_combo.setCurrentIndex(0)
        elif reject_result_index >= 0:
            self.reject_result_combo.setCurrentIndex(reject_result_index)
        else:
            self.reject_result_combo.setCurrentIndex(0)
            
        # 设置备注
        self.notes_edit.setText(notes_text)
        
        # 设置当前记录ID
        rec = self.db.get_record_by_order_no(order_no)
        if rec:
            self.current_record_id = rec['id']
        else:
            self.current_record_id = None

    def on_cell_changed(self, row, column):
        """表格单元格编辑完成时触发"""
        try:
            self.table.cellChanged.disconnect(self.on_cell_changed)
        except TypeError:
            pass
        
        try:
            item = self.table.item(row, column)
            if not item:
                return
                
            # 获取记录ID
            record_id = self.get_record_id_from_row(row)
            if not record_id:
                return
            
            # 根据列索引处理不同的字段
            if column == 2:  # 规格编码列
                self.update_spec_code(record_id, item.text())
            elif column == 11:  # 备注列
                self.update_notes(record_id, item.text())
            elif column == 4:  # 退款金额列
                self.update_refund_amount(record_id, item.text())
            elif column == 7:  # 补偿金额列
                self.update_comp_amount(record_id, item.text())
            elif column in [5, 6, 8]:  # 撤销、打款补偿、驳回状态列
                # 处理状态字段编辑：自动标准化输入
                text = item.text().strip()
                
                # 自动标准化输入
                if text.lower() in ['是', 'true', '1', 'yes', 'y', 't']:
                    item.setText("是")
                    self.update_status_field(record_id, column, "是")
                    
                    # 如果是驳回列从"否"变为"是"，触发驳回流程
                    if column == 8:
                        rec = self.db.get_record_by_id(record_id)
                        if rec and not rec['reject']:  # 之前是"否"，现在变为"是"
                            self.start_reject_process(record_id, rec['order_no'], rec['store_name'])
                            
                elif text.lower() in ['否', 'false', '0', 'no', 'n', 'f']:
                    item.setText("否")
                    self.update_status_field(record_id, column, "否")
                    
                    # 如果是驳回列从"是"变为"否"，停止驳回流程
                    if column == 8:
                        rec = self.db.get_record_by_id(record_id)
                        if rec and rec['reject']:  # 之前是"是"，现在变为"否"
                            self.reject_manager.stop_process(rec['order_no'])
                            
                else:
                    # 无效输入，恢复原值
                    rec = self.db.get_record_by_id(record_id)
                    if rec:
                        if column == 5:  # 撤销
                            original_value = "是" if rec['cancel'] else "否"
                        elif column == 6:  # 打款补偿
                            original_value = "是" if rec['compensate'] else "否"
                        elif column == 8:  # 驳回
                            original_value = "是" if rec['reject'] else "否"
                        item.setText(original_value)
                        QMessageBox.warning(self, "输入错误", "请输入'是'或'否'")
                
        finally:
            self.table.cellChanged.connect(self.on_cell_changed)

    def get_record_id_from_row(self, row):
        """根据行号获取记录ID（增强错误处理）"""
        try:
            # 检查行号是否有效
            if row < 0 or row >= self.table.rowCount():
                return None

            # 优先从表格行绑定的数据库主键中读取，避免通过订单号回查拿错记录
            for column in (0, 1):
                item = self.table.item(row, column)
                if item is None:
                    continue
                record_id = item.data(Qt.UserRole)
                if record_id not in (None, ""):
                    try:
                        return int(record_id)
                    except (TypeError, ValueError):
                        pass

            # 兼容旧数据：如果该行还没绑定主键，再回退到订单号查询
            order_no_item = self.table.item(row, 1)  # 订单号列
            if not order_no_item:
                return None

            order_no = order_no_item.text().strip()
            if not order_no:
                return None

            record = self.db.get_record_by_order_no(order_no)
            if record and 'id' in record:
                return record['id']
            return None
        except Exception as e:
            print(f"获取行 {row} 的记录ID时出错: {e}")
            return None

    def _bind_record_id_to_row(self, row, record_id):
        """将数据库主键绑定到表格行，避免后续通过订单号二次查询。"""
        if record_id in (None, ""):
            return

        for column in (0, 1):
            item = self.table.item(row, column)
            if item is not None:
                item.setData(Qt.UserRole, int(record_id))

    def toggle_status_field(self, row, column):
        """双击切换状态字段（撤销、打款补偿、驳回）"""
        try:
            # 获取记录ID
            record_id = self.get_record_id_from_row(row)
            if not record_id:
                return
                
            # 获取当前记录信息
            rec = self.db.get_record_by_id(record_id)
            if not rec:
                return
                
            # 根据列索引确定要切换的字段
            if column == 5:  # 撤销列
                new_cancel = not rec['cancel']  # 切换状态
                self.db.update_record(
                    record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                    rec['refund_amount'], new_cancel, rec['compensate'], rec['comp_amount'],
                    rec['reject'], rec['reject_result'], rec['notes'], rec['record_date'],
                    rec.get('order_status', ''), rec.get('after_sale_status', ''),
                    rec.get('spec_name', ''), rec.get('spec_code', '')
                )
            elif column == 6:  # 打款补偿列
                new_compensate = not rec['compensate']  # 切换状态
                self.db.update_record(
                    record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                    rec['refund_amount'], rec['cancel'], new_compensate, rec['comp_amount'],
                    rec['reject'], rec['reject_result'], rec['notes'], rec['record_date'],
                    rec.get('order_status', ''), rec.get('after_sale_status', ''),
                    rec.get('spec_name', ''), rec.get('spec_code', '')
                )
            elif column == 8:  # 驳回列
                new_reject = not rec['reject']  # 切换状态
                self.db.update_record(
                    record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                    rec['refund_amount'], rec['cancel'], rec['compensate'], rec['comp_amount'],
                    new_reject, rec['reject_result'], rec['notes'], rec['record_date'],
                    rec.get('order_status', ''), rec.get('after_sale_status', ''),
                    rec.get('spec_name', ''), rec.get('spec_code', '')
                )
            
            self._refresh_row_by_record_id(record_id, refresh_statistics=True)
            
        except Exception as e:
            self.load_table_data(force_reload=True)


    
    def _select_current_record_after_update(self):
        """更新记录后重新选中当前记录"""
        if self.current_record_id is None:
            return
            
        # 根据记录ID找到对应的行号
        for row in range(self.table.rowCount()):
            record_id = self.get_record_id_from_row(row)
            if record_id == self.current_record_id:
                # 选中该行
                self.table.selectRow(row)
                # 滚动到该行
                self.table.scrollToItem(self.table.item(row, 0))
                break
    
    def _update_statistics_only(self):
        """只更新统计信息，不刷新整个表格"""
        # 获取当前筛选条件下的记录
        records = self.get_filtered_records()
        # 更新状态栏统计
        self.update_statusbar(records)

    def update_status_field(self, record_id, column, value):
        """更新状态字段（撤销、打款补偿、驳回）"""
        # 获取当前记录信息
        rec = self.db.get_record_by_id(record_id)
        if not rec:
            return
            
        # 根据列索引确定要更新的字段
        if column == 5:  # 撤销列
            cancel = value.lower() in ['是', 'true', '1', 'yes']
            self.db.update_record(
                record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                rec['refund_amount'], cancel, rec['compensate'], rec['comp_amount'],
                rec['reject'], rec['reject_result'], rec['notes'], rec['record_date'],
                rec.get('order_status', ''), rec.get('after_sale_status', ''),
                rec.get('spec_name', ''), rec.get('spec_code', '')
            )
        elif column == 6:  # 打款补偿列
            compensate = value.lower() in ['是', 'true', '1', 'yes']
            self.db.update_record(
                record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                rec['refund_amount'], rec['cancel'], compensate, rec['comp_amount'],
                rec['reject'], rec['reject_result'], rec['notes'], rec['record_date'],
                rec.get('order_status', ''), rec.get('after_sale_status', ''),
                rec.get('spec_name', ''), rec.get('spec_code', '')
            )
        elif column == 8:  # 驳回列
            reject = value.lower() in ['是', 'true', '1', 'yes']
            self.db.update_record(
                record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                rec['refund_amount'], rec['cancel'], rec['compensate'], rec['comp_amount'],
                reject, rec['reject_result'], rec['notes'], rec['record_date'],
                rec.get('order_status', ''), rec.get('after_sale_status', ''),
                rec.get('spec_name', ''), rec.get('spec_code', '')
            )

        self._refresh_row_by_record_id(record_id, refresh_statistics=True)

    def update_spec_code(self, record_id, spec_code_text):
        """更新规格编码。"""
        normalized = str(spec_code_text or "").strip()
        if normalized == "-":
            normalized = ""
        if self.db.update_record_partial(record_id, spec_code=normalized):
            self._refresh_row_by_record_id(record_id, refresh_statistics=True)
            self.show_tooltip("规格编码已更新", "rgba(76, 175, 80, 0.95)", 1000)
        else:
            self.load_table_data(force_reload=True)

    def update_notes(self, record_id, notes_text):
        """更新备注，并自动使真实退款原因失效。"""
        normalized = str(notes_text or "").strip()
        if self.db.update_record_partial(record_id, notes=normalized):
            self._refresh_row_by_record_id(record_id, refresh_statistics=True)
            self.show_tooltip("备注已更新", "rgba(76, 175, 80, 0.95)", 1000)
        else:
            self.load_table_data(force_reload=True)

    def update_refund_amount(self, record_id, amount_text):
        """更新退款金额"""
        try:
            # 提取数字部分
            amount = float(amount_text.replace('¥', '').strip())
            if self.db.update_refund_amount(record_id, amount):
                self._refresh_row_by_record_id(record_id, refresh_statistics=True)
                self.show_tooltip("退款金额已更新", "rgba(76, 175, 80, 0.95)", 1000)  # 绿色气泡显示1秒
            else:
                self.load_table_data(force_reload=True)
        except ValueError:
            QMessageBox.warning(self, "错误", "请输入有效的金额数字")
            self._refresh_row_by_record_id(record_id, refresh_statistics=False)

    def update_comp_amount(self, record_id, amount_text):
        """更新补偿金额"""
        try:
            # 提取数字部分
            amount = float(amount_text.replace('¥', '').strip())
            if self.db.update_comp_amount(record_id, amount):
                self._refresh_row_by_record_id(record_id, refresh_statistics=True)
                self.show_tooltip("补偿金额已更新", "rgba(76, 175, 80, 0.95)", 1000)  # 绿色气泡显示1秒
            else:
                self.load_table_data(force_reload=True)
        except ValueError:
            QMessageBox.warning(self, "错误", "请输入有效的金额数字")
            self._refresh_row_by_record_id(record_id, refresh_statistics=False)

    def get_row_from_record_id(self, record_id):
        """根据记录ID获取行号"""
        for row in range(self.table.rowCount()):
            current_id = self.get_record_id_from_row(row)
            if current_id == record_id:
                return row
        return None

    # ==================== 驳回流程管理方法 ====================
    
    def start_reject_process(self, record_id, order_no, store_name):
        """开始驳回流程"""
        # 显示驳回选择对话框
        dialog = RejectSelectionDialog(current_round=0, parent=self)
        
        if dialog.exec_() == QDialog.Accepted:
            option = dialog.get_selected_option()
            
            if option == "first":
                # 开始第一轮驳回
                self.reject_manager.start_first_round(order_no, store_name)
                # 更新驳回结果列显示
                self.update_reject_result_display(order_no, "第一轮驳回中...")
                QMessageBox.information(self, "第一轮驳回", f"订单 {order_no} 第一轮驳回已开始\n30分钟后将提醒您继续操作")
                
            elif option == "second":
                # 开始第二轮驳回
                self.reject_manager.start_second_round(order_no, store_name)
                self.update_reject_result_display(order_no, "第二轮驳回中...")
                QMessageBox.information(self, "第二轮驳回", f"订单 {order_no} 第二轮驳回已开始\n30分钟后将提醒您继续操作")
                
            elif option == "success":
                # 驳回成功
                self.show_reject_success_dialog(record_id, order_no, store_name)
    
    def update_reject_result_display(self, order_no, text):
        """更新驳回结果列的显示文本"""
        # 找到对应的行
        for row in range(self.table.rowCount()):
            order_item = self.table.item(row, 1)  # 订单号列
            if order_item and order_item.text() == order_no:
                # 更新驳回结果列（第8列）
                result_item = QTableWidgetItem(text)
                result_item.setTextAlignment(Qt.AlignCenter)
                
                # 根据状态设置颜色
                if "第一轮" in text:
                    result_item.setBackground(QColor("#FFF3E0"))  # 浅橙色
                    result_item.setForeground(QColor("#E65100"))  # 深橙色
                elif "第二轮" in text:
                    result_item.setBackground(QColor("#E3F2FD"))  # 浅蓝色
                    result_item.setForeground(QColor("#1565C0"))  # 深蓝色
                elif "驳回成功" in text:
                    result_item.setBackground(QColor("#E8F5E9"))  # 浅绿色
                    result_item.setForeground(QColor("#2E7D32"))  # 深绿色
                elif "驳回失败" in text:
                    result_item.setBackground(QColor("#FFEBEE"))  # 浅红色
                    result_item.setForeground(QColor("#C62828"))  # 深红色
                
                # 设置工具提示（鼠标悬停显示）
                process_info = self.reject_manager.get_process_info(order_no)
                if process_info and "驳回中" in text:
                    remaining = process_info['remaining']
                    minutes = remaining // 60
                    seconds = remaining % 60
                    tooltip = f"⏳ 剩余时间: {minutes}分{seconds}秒\n💡 双击可跳过等待"
                    result_item.setToolTip(tooltip)
                elif "等待操作" in text:
                    result_item.setToolTip("⏰ 时间已到，请继续操作\n💡 双击打开选择窗口")
                elif "驳回成功" in text:
                    # 检查是否有48小时提醒
                    reminder_info = self.reject_manager.get_48h_reminder_info(order_no)
                    if reminder_info:
                        end_time = reminder_info['end_time']
                        remaining_hours = reminder_info['remaining_hours']
                        # 格式化提醒时间：月-日 时:分
                        remind_time_str = end_time.strftime("%m-%d %H:%M")
                        result_item.setToolTip(
                            f"✅ 驳回成功\n"
                            f"⏰ 48小时提醒剩余: 约{remaining_hours}小时\n"
                            f"📅 提醒时间: {remind_time_str}\n"
                            f"⚠️ 请及时检查是否被平台介入退款\n"
                            f"💡 双击可标记平台介入退款"
                        )
                    else:
                        result_item.setToolTip("✅ 驳回成功\n💡 双击可标记平台介入退款")
                elif "驳回失败" in text:
                    result_item.setToolTip("❌ 平台已介入退款")
                else:
                    result_item.setToolTip("💡 双击开始驳回流程")
                    
                self.table.setItem(row, 9, result_item)
                break
    
    def on_reject_countdown_finished(self, order_no, round_text):
        """倒计时结束时的处理"""
        # 获取订单信息
        rec = self.db.get_record_by_order_no(order_no)
        store_name = rec['store_name'] if rec else "未知店铺"
        record_id = rec['id'] if rec else None
        
        # 使用自定义对话框，让订单号可复制
        dialog = RejectCountdownFinishedDialog(order_no, store_name, round_text, self)
        result = dialog.exec_()
        
        if result == QDialog.Accepted:
            # 用户选择继续下一轮驳回
            if round_text == "第一轮":
                # 开始第二轮驳回
                self.start_reject_process(record_id, order_no, store_name)
            elif round_text == "第二轮":
                # 显示驳回成功对话框
                self.show_reject_success_dialog(record_id, order_no, store_name)
        
        # 更新显示为"等待操作"
        self.update_reject_result_display(order_no, f"{round_text}等待操作")
        
        # 强制重查数据库，避免旧缓存把界面上的驳回状态覆盖回去
        self.load_table_data(force_reload=True)
    
    def on_reject_countdown_updated(self, order_no, remaining_seconds, round_text):
        """倒计时更新时的处理 - 实时更新工具提示显示剩余时间"""
        # 找到对应的行并更新工具提示
        for row in range(self.table.rowCount()):
            order_item = self.table.item(row, 1)  # 订单号列
            if order_item and order_item.text() == order_no:
                # 获取驳回结果列的item
                result_item = self.table.item(row, 9)
                if result_item:
                    # 计算分钟和秒
                    minutes = remaining_seconds // 60
                    seconds = remaining_seconds % 60
                    # 更新工具提示
                    tooltip = f"⏳ 剩余时间: {minutes}分{seconds}秒\n💡 双击可跳过等待"
                    result_item.setToolTip(tooltip)
                break
    
    def show_reject_success_dialog(self, record_id, order_no, store_name):
        """显示驳回成功对话框"""
        dialog = RejectSuccessDialog(order_no, store_name, parent=self)
        
        if dialog.exec_() == QDialog.Accepted:
            # 更新数据库中的驳回结果
            rec = self.db.get_record_by_id(record_id)
            if rec:
                self.db.update_record(
                    record_id, rec['store_id'], order_no, rec['reason'],
                    rec['refund_amount'], rec['cancel'], rec['compensate'], rec['comp_amount'],
                    True, "驳回成功", rec['notes'], rec['record_date'],
                    rec.get('order_status', ''), rec.get('after_sale_status', ''),
                    rec.get('spec_name', ''), rec.get('spec_code', '')
                )
            
            # 更新显示
            self.update_reject_result_display(order_no, "驳回成功")
            
            # 检查是否需要设置48小时提醒
            if dialog.should_remind_48h():
                self.reject_manager.set_48h_reminder(order_no, store_name)
                QMessageBox.information(
                    self,
                    "✅ 驳回成功",
                    f"订单 {order_no} 驳回成功！\n\n已设置48小时后提醒。"
                )
            else:
                QMessageBox.information(
                    self,
                    "✅ 驳回成功",
                    f"订单 {order_no} 驳回成功！"
                )
            
            # 停止驳回流程
            self.reject_manager.stop_process(order_no)
            
            # 强制重查数据库，避免旧缓存把刚写入的驳回成功状态覆盖掉
            self.load_table_data(force_reload=True)
    
    def show_reject_success_actions_dialog(self, record_id, order_no, store_name):
        """显示驳回成功后的操作对话框（平台介入退款等）"""
        dialog = RejectSuccessActionsDialog(order_no, store_name, parent=self)
        result = dialog.exec_()
        
        if result == QDialog.Accepted:
            # 用户选择平台介入退款（标记为驳回失败）
            rec = self.db.get_record_by_id(record_id)
            if rec:
                self.db.update_record(
                    record_id, rec['store_id'], order_no, rec['reason'],
                    rec['refund_amount'], rec['cancel'], rec['compensate'], rec['comp_amount'],
                    True, "驳回失败", rec['notes'], rec['record_date'],
                    rec.get('order_status', ''), rec.get('after_sale_status', ''),
                    rec.get('spec_name', ''), rec.get('spec_code', '')
                )
            
            # 取消48小时提醒（如果有）
            if self.reject_manager.has_48h_reminder(order_no):
                self.reject_manager.stop_48h_reminder(order_no)
            
            # 更新显示
            self.update_reject_result_display(order_no, "驳回失败")
            
            QMessageBox.information(
                self,
                "已标记平台介入退款",
                f"订单 {order_no} 已标记为平台介入退款（驳回失败）"
            )
            
            # 强制重查数据库，避免旧缓存把刚写入的驳回失败状态覆盖掉
            self.load_table_data(force_reload=True)
    
    def on_reject_result_double_click(self, row, column):
        """双击驳回结果列的处理"""
        if column != 9:  # 不是驳回结果列
            return
        
        # 先检查驳回列的状态，如果为"否"则不执行任何操作
        reject_item = self.table.item(row, 8)  # 驳回列
        if not reject_item or reject_item.text() != "是":
            # 驳回状态为"否"，不执行任何操作
            return
        
        # 获取订单号
        order_item = self.table.item(row, 1)
        if not order_item:
            return
        order_no = order_item.text()
        
        # 获取驳回结果列的文本
        result_item = self.table.item(row, 9)
        result_text = result_item.text() if result_item else ""
        
        # 如果驳回成功，显示平台介入退款对话框
        if "驳回成功" in result_text:
            record_id = self.get_record_id_from_row(row)
            rec = self.db.get_record_by_id(record_id)
            if rec:
                self.show_reject_success_actions_dialog(record_id, order_no, rec['store_name'])
            return
        
        # 检查是否有进行中的驳回流程
        process_info = self.reject_manager.get_process_info(order_no)
        if not process_info:
            # 没有进行中的流程，显示选择对话框
            record_id = self.get_record_id_from_row(row)
            rec = self.db.get_record_by_id(record_id)
            if rec:
                self.start_reject_process(record_id, order_no, rec['store_name'])
            return
        
        # 有进行中的流程，显示跳过等待对话框
        dialog = RejectSkipDialog(order_no, process_info['round'], parent=self)
        
        if dialog.exec_() == QDialog.Accepted:
            # 跳过等待
            current_round = self.reject_manager.skip_wait(order_no)
            
            if current_round == 1:
                # 第一轮结束，显示选择对话框继续第二轮
                QMessageBox.information(self, "第一轮结束", "第一轮驳回等待已跳过")
                record_id = self.get_record_id_from_row(row)
                rec = self.db.get_record_by_id(record_id)
                if rec:
                    # 显示选择对话框，禁用第一轮按钮
                    dialog2 = RejectSelectionDialog(current_round=1, parent=self)
                    if dialog2.exec_() == QDialog.Accepted:
                        option = dialog2.get_selected_option()
                        if option == "second":
                            self.reject_manager.start_second_round(order_no, rec['store_name'])
                            self.update_reject_result_display(order_no, "第二轮驳回中...")
                        elif option == "success":
                            self.show_reject_success_dialog(record_id, order_no, rec['store_name'])
                            
            elif current_round == 2:
                # 第二轮结束，显示成功对话框
                QMessageBox.information(self, "第二轮结束", "第二轮驳回等待已跳过")
                record_id = self.get_record_id_from_row(row)
                rec = self.db.get_record_by_id(record_id)
                if rec:
                    self.show_reject_success_dialog(record_id, order_no, rec['store_name'])
    
    def show_48h_reminder(self, order_no, store_name):
        """显示48小时提醒（订单号可复制）"""
        dialog = Reminder48hDialog(order_no, store_name, self)
        dialog.exec_()

    def restore_reject_display_from_db(self):
        """从数据库恢复驳回显示状态（软件启动时调用）"""
        # 获取所有活动的倒计时
        active_countdowns = self.db.get_all_active_reject_countdowns()
        
        for countdown in active_countdowns:
            order_no = countdown['order_no']
            current_round = countdown['current_round']
            
            # 更新表格显示
            if current_round == 1:
                self.update_reject_result_display(order_no, "第一轮驳回中...")
            elif current_round == 2:
                self.update_reject_result_display(order_no, "第二轮驳回中...")

    def on_item_clicked(self, item):
        """单击表格项：自动录入订单信息到输入框"""
        row = item.row()
        
        # 获取选中行的数据
        store_name = self.table.item(row, 0).text()
        order_no = self.table.item(row, 1).text()
        reason = self.table.item(row, 3).text()
        refund_amount_text = self.table.item(row, 4).text()
        cancel_text = self.table.item(row, 5).text()
        compensate_text = self.table.item(row, 6).text()
        comp_amount_text = self.table.item(row, 7).text()
        
        # 解析退款金额（去掉¥符号）
        try:
            refund_amount = float(refund_amount_text.replace('¥', '').strip())
        except:
            refund_amount = 0.0
            
        # 解析补偿金额
        try:
            comp_amount = float(comp_amount_text.replace('¥', '').strip()) if comp_amount_text else 0.0
        except:
            comp_amount = 0.0
            
        # 设置店铺
        store_index = self.store_combo.findText(store_name)
        if store_index >= 0:
            self.store_combo.setCurrentIndex(store_index)
            
        # 设置订单号
        self.order_no_edit.setText(order_no)
        
        # 设置退款原因
        reason_index = self.reason_combo.findText(reason)
        if reason_index >= 0:
            self.reason_combo.setCurrentIndex(reason_index)
        else:
            self.reason_combo.setCurrentIndex(0)
            
        # 设置退款金额
        self.refund_amount_edit.setText(f"{refund_amount:.2f}")
        
        # 设置撤销状态
        self.cancel_check.setChecked(cancel_text == "是")
        
        # 设置补偿状态和金额
        self.compensate_check.setChecked(compensate_text == "是")
        self.comp_amount_edit.setText(f"{comp_amount:.2f}" if comp_amount > 0 else "")
        
        # 记录当前记录ID
        rec = self.db.get_record_by_order_no(order_no)
        if rec:
            self.current_record_id = rec['id']
        else:
            self.current_record_id = None
            
        # 点击表格后清除高亮
        if self.highlighted_orders:
            self.highlighted_orders.clear()
            # 立即刷新表格显示，清除高亮
            self.load_table_data()

    # 双击功能已改为单击，此方法不再使用
    # def on_row_double_clicked(self, item):
    #     """双击行：填充到输入区"""
    #     row = item.row()
    #     store_name = self.table.item(row, 0).text()
    #     order_no = self.table.item(row, 1).text()
    #     reason = self.table.item(row, 2).text()
    #     refund_amount_text = self.table.item(row, 3).text().replace('¥', '').replace(',', '')
    #     try:
    #         refund_amount = float(refund_amount_text)
    #     except:
    #         refund_amount = 0.0
    #     cancel_text = self.table.item(row, 4).text()
    #     compensate_text = self.table.item(row, 5).text()
    #     comp_amount_text = self.table.item(row, 6).text().replace('¥', '').replace(',', '')
    #     try:
    #         comp_amount = float(comp_amount_text)
    #     except:
    #         comp_amount = 0.0

    #     # 设置店铺
    #     stores = self.db.get_stores()
    #     for idx, (sid, sname) in enumerate(stores):
    #         if sname == store_name:
    #             self.store_combo.setCurrentIndex(idx)
    #             break
    #     self.order_no_edit.setText(order_no)
    #     # 退款原因
    #     idx = self.reason_combo.findText(reason)
    #     if idx >= 0:
    #         self.reason_combo.setCurrentIndex(idx)
    #     else:
    #         # 如果原因不在列表中，添加并选中
    #         self.reason_combo.addItem(reason)
    #         self.reason_combo.setCurrentText(reason)
    #     self.refund_amount_edit.setText(str(refund_amount))
    #     self.cancel_check.setChecked(cancel_text == "是")
    #     self.compensate_check.setChecked(compensate_text == "是")
    #     if compensate_text == "是":
    #         self.comp_amount_edit.setEnabled(True)
    #         self.comp_amount_edit.setText(str(comp_amount) if comp_amount != 0 else "")
    #     else:
    #         self.comp_amount_edit.setEnabled(False)
    #         self.comp_amount_edit.clear()
    #     # 记录当前编辑的ID
    #     order_no = self.table.item(row, 1).text()
    #     rec = self.db.get_record_by_order_no(order_no)
    #     if rec:
    #         self.current_record_id = rec['id']
    #     else:
    #         self.current_record_id = None

    def show_context_menu(self, pos):
        """显示右键菜单"""
        item = self.table.itemAt(pos)
        
        # 创建自定义右键菜单
        menu = QMenu(self)
        
        if item is not None:
            # 如果点击了具体行，显示行操作菜单
            row = item.row()
            order_no = self.table.item(row, 1).text()
            store_name = self.table.item(row, 0).text()
            
            copy_order_action = QAction("复制订单号", self)
            copy_order_action.triggered.connect(lambda: self.copy_to_clipboard(order_no))
            copy_store_action = QAction("复制店铺名称", self)
            copy_store_action.triggered.connect(lambda: self.copy_to_clipboard(store_name))
            edit_action = QAction("编辑记录", self)
            edit_action.triggered.connect(lambda: self.on_item_clicked(self.table.item(row, 0)))
            delete_action = QAction("删除记录", self)
            delete_action.triggered.connect(self.delete_record)
            
            menu.addAction(copy_order_action)
            menu.addAction(copy_store_action)
            menu.addAction(edit_action)
            menu.addAction(delete_action)
            menu.addSeparator()
        
        # 添加全选当前筛选订单功能（无论是否点击具体行都显示）
        select_all_action = QAction("全选当前筛选订单", self)
        select_all_action.triggered.connect(self.select_all_filtered_orders)
        menu.addAction(select_all_action)
        
        menu.exec_(self.table.mapToGlobal(pos))

    def select_all_filtered_orders(self):
        """全选当前筛选出来的所有订单"""
        self.table.selectAll()
        selected_count = len(self.table.selectedItems()) // self.table.columnCount()
        self.show_tooltip(f"已选择 {selected_count} 条", "rgba(0, 120, 212, 0.95)", 1000)  # 蓝色气泡显示1秒





    def copy_to_clipboard(self, text):
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        self.show_tooltip("已复制", "rgba(76, 175, 80, 0.95)", 1000)  # 绿色气泡显示1秒

    def clear_highlight(self):
        """清除刚导入订单的金色高亮（不是清除用户鼠标选中的高亮）"""
        # 清除高亮订单集合（只清除刚导入订单的金色高亮标记）
        if hasattr(self, 'highlighted_orders'):
            self.highlighted_orders.clear()
            print("[DEBUG] 已清除高亮订单集合")
        
        # 重新加载表格数据，清除金色高亮显示
        self.load_table_data()
        
        # 显示绿色提示
        self.show_tooltip("已清除高亮", "rgba(76, 175, 80, 0.95)", 1000)

    def refresh_table_format(self):
        """刷新表格格式，清除导入后的高亮显示"""
        # 清除高亮订单集合
        if hasattr(self, 'highlighted_orders'):
            self.highlighted_orders.clear()
            print("[DEBUG] 已清除高亮订单集合")
        
        # 重新加载表格数据，清除高亮显示
        self.load_table_data()
        
        # 显示丝滑的气泡提示"已刷新"
        self.show_refresh_tooltip()

    # ---------------------------- 导入导出功能 ---------------------------------
    def _get_export_headers(self, include_process_analysis=False):
        headers = ["店铺名称", "订单号", "规格编码", "退款原因", "退款金额", "撤销", "打款补偿", "补偿金额", "驳回", "驳回结果", "登记日期", "备注"]
        if include_process_analysis:
            headers.extend(["申请时间", "同意退款时间", "退款处理时长"])
        return headers

    def _collect_current_export_rows(self):
        """收集当前表格可见行，导出严格跟随当前筛选结果。"""
        export_rows = []
        for row_idx in range(self.table.rowCount()):
            row_data = []
            record = None
            for col in range(12):
                item = self.table.item(row_idx, col)
                text = item.text() if item else ""

                if col in [4, 7]:  # 退款金额、补偿金额
                    text = text.replace('¥', '').replace(',', '')

                if col == 10:  # 登记日期
                    record_id = self.get_record_id_from_row(row_idx)
                    if record_id:
                        record = self.db.get_record_by_id(record_id)
                        if record and record.get('record_date'):
                            text = record['record_date']

                row_data.append(text)

            export_rows.append({
                "store_name": row_data[0] or "未知店铺",
                "row_data": row_data,
                "refund_apply_time": (record or {}).get("refund_apply_time", ""),
                "refund_agree_time": (record or {}).get("refund_agree_time", ""),
            })
        return export_rows

    def _group_export_rows_by_store(self, export_rows):
        grouped_rows = {}
        for row in export_rows:
            store_name = row.get("store_name") or "未知店铺"
            grouped_rows.setdefault(store_name, []).append(row)
        return grouped_rows

    def _choose_export_options(self, has_multiple_stores):
        """选择订单导出选项。"""
        dialog = QDialog(self)
        dialog.setWindowTitle("导出选项")
        layout = QVBoxLayout(dialog)

        mode_combo = None
        if has_multiple_stores:
            layout.addWidget(QLabel("当前导出内容包含多个店铺，请选择导出方式："))
            mode_combo = QComboBox()
            mode_combo.addItem("单表格模式", "single_table")
            mode_combo.addItem("独立Sheet模式", "independent_sheet")
            layout.addWidget(mode_combo)

        process_check = QCheckBox("追加申请时间、同意退款时间和退款处理时长")
        layout.addWidget(process_check)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec_() != QDialog.Accepted:
            return None

        return {
            "mode": mode_combo.currentData() if mode_combo else "single_table",
            "include_process_analysis": process_check.isChecked(),
        }

    def _check_export_file_available(self, file_path):
        if not os.path.exists(file_path):
            return True
        try:
            with open(file_path, 'a', encoding='utf-8'):
                pass
            return True
        except PermissionError:
            QMessageBox.warning(
                self,
                "文件被占用",
                f"文件 '{os.path.basename(file_path)}' 正在被其他程序使用！\n\n请先关闭该文件，然后重试。"
            )
            return False

    def _write_refund_export_sheet(self, ws, export_rows, include_process_analysis=False):
        headers = self._get_export_headers(include_process_analysis)
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        for row in export_rows:
            row_data = list(row["row_data"])
            if include_process_analysis:
                apply_time = row.get("refund_apply_time", "")
                agree_time = row.get("refund_agree_time", "")
                row_data.extend([
                    self._format_export_process_time(apply_time),
                    self._format_export_process_time(agree_time),
                    self._format_refund_process_duration(apply_time, agree_time),
                ])
            ws.append(row_data)

        for row_idx in range(2, len(export_rows) + 2):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = center_alignment
                cell.border = thin_border

                if col_idx in [5, 8]:  # 退款金额、补偿金额
                    try:
                        if cell.value not in (None, ""):
                            cell.value = float(cell.value)
                    except Exception:
                        pass

                if col_idx == 11:  # 登记日期
                    try:
                        if cell.value:
                            date_obj = datetime.strptime(str(cell.value), '%Y-%m-%d')
                            cell.value = date_obj
                            cell.number_format = 'YYYY-MM-DD'
                    except Exception:
                        pass

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    def export_excel(self):
        """导出当前表格数据到Excel"""
        export_rows = self._collect_current_export_rows()
        if not export_rows:
            QMessageBox.information(self, "提示", "没有数据可导出")
            return

        grouped_rows = self._group_export_rows_by_store(export_rows)
        export_options = self._choose_export_options(len(grouped_rows) > 1)
        if not export_options:
            return
        export_mode = export_options["mode"]
        include_process_analysis = export_options["include_process_analysis"]

        # 选择保存路径
        default_name = f"退款记录_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "导出Excel", default_name, "Excel文件 (*.xlsx)")
        if not file_path:
            return

        try:
            if not self._check_export_file_available(file_path):
                return

            wb = openpyxl.Workbook()
            if export_mode == "independent_sheet" and len(grouped_rows) > 1:
                wb.remove(wb.active)
                used_names = set()
                for store_name, rows_for_store in grouped_rows.items():
                    ws = wb.create_sheet(self._safe_excel_sheet_name(store_name, used_names))
                    self._write_refund_export_sheet(ws, rows_for_store, include_process_analysis)
            else:
                ws = wb.active
                ws.title = "退款记录"
                if len(grouped_rows) > 1:
                    ordered_rows = []
                    for rows_for_store in grouped_rows.values():
                        ordered_rows.extend(rows_for_store)
                else:
                    ordered_rows = export_rows
                self._write_refund_export_sheet(ws, ordered_rows, include_process_analysis)

            wb.save(file_path)
            self.show_tooltip("导出成功", "rgba(76, 175, 80, 0.95)", 1500)  # 绿色气泡显示1.5秒
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败：{str(e)}")

    def _write_summary_metric_rows(self, ws, start_row, title, metrics):
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=start_row, column=1).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        row = start_row + 1
        for label, value in metrics:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        return row

    def _write_summary_category_block(self, ws, start_row, title, categories, empty_text="无"):
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=start_row, column=1).fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        row = start_row + 1
        if not categories:
            ws.cell(row=row, column=1, value=empty_text)
            return row + 1

        ws.cell(row=row, column=1, value="分类")
        ws.cell(row=row, column=2, value="数量")
        ws.cell(row=row, column=3, value="占比")
        row += 1
        for category in categories:
            ws.cell(row=row, column=1, value=category.get("name", "未分类"))
            ws.cell(row=row, column=2, value=category.get("count", 0))
            ws.cell(row=row, column=3, value=f"{self._safe_float(category.get('ratio', 0)):.2f}%")
            row += 1
        return row

    @staticmethod
    def _safe_excel_sheet_name(name, used_names=None):
        used_names = used_names if used_names is not None else set()
        text = re.sub(r'[\[\]\:\*\?\/\\]', '_', str(name or "Sheet").strip()) or "Sheet"
        text = text[:31]
        base = text or "Sheet"
        candidate = base
        suffix = 1
        while candidate in used_names:
            suffix_text = f"_{suffix}"
            candidate = f"{base[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        used_names.add(candidate)
        return candidate

    @staticmethod
    def _get_summary_export_date_range(summary_snapshot, fallback_metrics=None):
        metrics = fallback_metrics or {}
        date_range = metrics.get("date_range", {}) if isinstance(metrics, dict) else {}
        if not date_range:
            stores = summary_snapshot.get("stores", []) if isinstance(summary_snapshot, dict) else []
            if stores:
                date_range = (stores[0].get("metrics", {}) or {}).get("date_range", {})
        start_date = date_range.get("start_date", "")
        end_date = date_range.get("end_date", "")
        if start_date or end_date:
            return f"{start_date} 至 {end_date}"
        return ""

    @staticmethod
    def _format_summary_basic_stats_date_range(summary_snapshot, fallback_metrics=None):
        metrics = fallback_metrics or {}
        date_range = metrics.get("date_range", {}) if isinstance(metrics, dict) else {}
        if not date_range:
            stores = summary_snapshot.get("stores", []) if isinstance(summary_snapshot, dict) else []
            if stores:
                date_range = (stores[0].get("metrics", {}) or {}).get("date_range", {})

        def format_date(value):
            text = str(value or "").strip()
            try:
                parsed = datetime.strptime(text, "%Y-%m-%d")
                return f"{parsed.month}月{parsed.day}日"
            except ValueError:
                try:
                    parsed = datetime.strptime(text, "%Y/%m/%d")
                    return f"{parsed.month}月{parsed.day}日"
                except ValueError:
                    return text

        start_date = format_date(date_range.get("start_date", ""))
        end_date = format_date(date_range.get("end_date", ""))
        if start_date or end_date:
            return f"{start_date}至{end_date}"
        return ""

    def _build_summary_export_metric_rows(self, metrics):
        return [
            ("当前范围单量", self._format_metric_int(metrics.get("orders", 0))),
            ("当前范围销售金额", round(metrics.get("sales", 0), 2)),
            ("退款预算金额", round(metrics.get("refund_budget_remaining", 0), 2)),
            ("退款单量", metrics.get("record_count", metrics.get("quality_refund_count", 0) + metrics.get("other_refund_count", 0))),
            ("总退款率", f"{metrics.get('total_refund_rate', 0):.2f}%"),
            ("品质退款单量", metrics.get("quality_refund_count", 0)),
            ("顾客申请申请品质退款订单的比例", f"{metrics.get('quality_apply_rate', 0):.2f}%"),
            ("撤销品质单量", metrics.get("canceled_quality_count", metrics.get("quality_cancel_count", 0))),
            ("顾客申请品质退款订单的比例（已撤销）", f"{metrics.get('quality_cancel_rate', 0):.2f}%"),
            ("品质退款实际单量", metrics.get("quality_actual_count", 0)),
            ("申请品质退款订单的比例（未撤销）", f"{metrics.get('quality_actual_rate', 0):.2f}%"),
            ("其他退款单量", metrics.get("other_refund_count", 0)),
            ("有效退款金额", round(metrics.get("effective_refund_amount", 0), 2)),
            ("补偿金额", round(metrics.get("compensation_amount", 0), 2)),
            ("售后总金额", round(metrics.get("total_after_sales", 0), 2)),
            ("金额占比", f"{metrics.get('refund_ratio', 0):.2f}%"),
            ("品质售后金额", round(metrics.get("quality_after_sales_amount", 0), 2)),
            ("其他售后金额", round(metrics.get("other_after_sales_amount", 0), 2)),
            ("品质退款申请单量", metrics.get("quality_apply_count", 0)),
            ("品质退款撤销单量", metrics.get("quality_cancel_count", 0)),
            ("有备注订单单量", metrics.get("note_count", 0)),
            ("无备注订单单量", metrics.get("no_note_count", 0)),
            ("备注率", f"{metrics.get('note_rate', 0):.2f}%"),
            ("无备注率", f"{metrics.get('no_note_rate', 0):.2f}%"),
            ("最多原因", metrics.get("top_refund_reason", "无数据")),
            ("最多原因出现次数", metrics.get("top_reason_count", 0)),
            ("最多原因占比", f"{metrics.get('top_reason_ratio', 0):.2f}%"),
        ]

    def _get_stable_spec_fill(self, spec_code):
        spec_text = str(spec_code or "-").strip() or "-"
        digest = hashlib.md5(spec_text.encode("utf-8")).hexdigest()
        # Keep colors light enough for black text while remaining stable per spec.
        red = 220 + int(digest[0:2], 16) % 30
        green = 220 + int(digest[2:4], 16) % 30
        blue = 220 + int(digest[4:6], 16) % 30
        color = f"{red:02X}{green:02X}{blue:02X}"
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _get_stable_store_fill(self, store_name):
        store_text = str(store_name or "未知店铺").strip() or "未知店铺"
        digest = hashlib.md5(store_text.encode("utf-8")).hexdigest()
        # Keep the upgraded sheet grouped visually by store while preserving readability.
        red = 220 + int(digest[0:2], 16) % 30
        green = 220 + int(digest[2:4], 16) % 30
        blue = 220 + int(digest[4:6], 16) % 30
        color = f"{red:02X}{green:02X}{blue:02X}"
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _write_real_reason_distribution_blocks(self, ws, start_row, analysis, border, center):
        section_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        data_fill = PatternFill(start_color="F3F8EF", end_color="F3F8EF", fill_type="solid")
        row = start_row

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        title_cell = ws.cell(row=row, column=1, value="总体真实退款原因分布")
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = section_fill
        title_cell.alignment = center
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        headers = ["真实退款原因", "数量", "占比"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border
        row += 1

        categories = analysis.get("overall_categories", []) if isinstance(analysis, dict) else []
        if categories:
            for category in categories:
                values = [
                    category.get("name", "未分类"),
                    category.get("count", 0),
                    f"{self._safe_float(category.get('ratio', 0)):.2f}%",
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.fill = data_fill
                    cell.alignment = center
                    cell.border = border
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=1, value="无")
            cell.fill = data_fill
            cell.alignment = center
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
            row += 1

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        title_cell = ws.cell(row=row, column=1, value="按规格编码分布")
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = section_fill
        title_cell.alignment = center
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        headers = ["规格编码", "真实退款原因", "数量", "占比"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border
        row += 1

        spec_categories = analysis.get("spec_categories", []) if isinstance(analysis, dict) else []
        spec_rows = []
        for spec_item in spec_categories:
            spec = str(spec_item.get("spec", "-") or "-").strip() or "-"
            for category in spec_item.get("categories", []):
                spec_rows.append((
                    spec,
                    category.get("name", "未分类"),
                    category.get("count", 0),
                    f"{self._safe_float(category.get('ratio', 0)):.2f}%",
                ))

        if spec_rows:
            grouped_rows = {}
            for spec, reason_name, count, ratio in spec_rows:
                grouped_rows.setdefault(spec, []).append((reason_name, count, ratio))

            for spec, rows in grouped_rows.items():
                group_start_row = row
                group_fill = self._get_stable_spec_fill(spec)
                for index, (reason_name, count, ratio) in enumerate(rows):
                    values = [
                        spec if index == 0 else "",
                        reason_name,
                        count,
                        ratio,
                    ]
                    ws.row_dimensions[row].height = 24
                    for col, value in enumerate(values, start=1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.fill = group_fill
                        cell.alignment = center
                        cell.border = border
                    row += 1

                group_end_row = row - 1
                if group_end_row > group_start_row:
                    ws.merge_cells(start_row=group_start_row, start_column=1, end_row=group_end_row, end_column=1)
                    merged_cell = ws.cell(row=group_start_row, column=1)
                    merged_cell.value = spec
                    merged_cell.fill = group_fill
                    merged_cell.alignment = center
                    merged_cell.border = border
                    for merged_row in range(group_start_row, group_end_row + 1):
                        ws.cell(row=merged_row, column=1).fill = group_fill
                        ws.cell(row=merged_row, column=1).border = border
                        ws.cell(row=merged_row, column=1).alignment = center
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=1, value="无")
            cell.fill = data_fill
            cell.alignment = center
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
            row += 1

        return row

    def _write_summary_basic_stats_sheet(self, ws, summary_snapshot):
        thin_side = Side(style="thin", color="B7B7B7")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        data_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        total_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        key_labels = {
            "总退款率",
            "顾客申请申请品质退款订单的比例",
            "顾客申请品质退款订单的比例（已撤销）",
            "申请品质退款订单的比例（未撤销）",
            "有效退款金额",
            "补偿金额",
            "售后总金额",
            "金额占比",
        }

        stores = summary_snapshot.get("stores", []) if isinstance(summary_snapshot, dict) else []
        metric_rows = self._build_summary_export_metric_rows(stores[0].get("metrics", {}) if stores else {})
        headers = ["时间", "店铺名称"] + [label for label, _value in metric_rows]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[1].height = 32

        row = 2
        for store in stores:
            metrics = store.get("metrics", {})
            date_range = self._format_summary_basic_stats_date_range(summary_snapshot, metrics)
            values = [date_range, store.get("store_name", "未知店铺")]
            values.extend(value for _label, value in self._build_summary_export_metric_rows(metrics))
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = data_fill
                cell.font = Font(size=11, bold=headers[col - 1] in key_labels)
                cell.alignment = center
                cell.border = border
            ws.row_dimensions[row].height = 24
            row += 1

        totals = summary_snapshot.get("totals") if isinstance(summary_snapshot, dict) else None
        if totals:
            date_range = self._format_summary_basic_stats_date_range(summary_snapshot, stores[0].get("metrics", {}) if stores else totals)
            values = [date_range, totals.get("store_name", "全部总和")]
            values.extend(value for _label, value in self._build_summary_export_metric_rows(totals))
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.fill = total_fill
                cell.font = Font(size=11, bold=True)
                cell.alignment = center
                cell.border = border
            ws.row_dimensions[row].height = 24

        ws.freeze_panes = "C2"
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            if col == 1:
                ws.column_dimensions[letter].width = 24
            elif col == 2:
                ws.column_dimensions[letter].width = 22
            else:
                ws.column_dimensions[letter].width = min(max(len(str(headers[col - 1])) + 4, 14), 34)

    def _write_summary_overall_real_reason_sheet(self, ws, summary_snapshot):
        thin_side = Side(style="thin", color="B7B7B7")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        data_fill = PatternFill(start_color="F3F8EF", end_color="F3F8EF", fill_type="solid")
        headers = ["时间", "店铺", "退款原因", "退款数量", "占比"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[1].height = 32

        stores = summary_snapshot.get("stores", []) if isinstance(summary_snapshot, dict) else []
        row = 2
        for store in stores:
            metrics = store.get("metrics", {})
            date_range = self._format_summary_basic_stats_date_range(summary_snapshot, metrics)
            store_name = store.get("store_name", "未知店铺")
            analysis = store.get("real_reason_analysis", {})
            categories = analysis.get("overall_categories", []) if isinstance(analysis, dict) else []
            if not categories:
                categories = [{"name": "无", "count": 0, "ratio": 0}]

            group_start_row = row
            for category in categories:
                values = [
                    date_range,
                    store_name,
                    category.get("name", "未分类"),
                    category.get("count", 0),
                    f"{self._safe_float(category.get('ratio', 0)):.2f}%",
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.fill = data_fill
                    cell.alignment = center
                    cell.border = border
                ws.row_dimensions[row].height = 24
                row += 1

            group_end_row = row - 1
            if group_end_row > group_start_row:
                for col in (1, 2):
                    ws.merge_cells(
                        start_row=group_start_row,
                        start_column=col,
                        end_row=group_end_row,
                        end_column=col
                    )
                    cell = ws.cell(row=group_start_row, column=col)
                    cell.alignment = center
                    cell.border = border
                    cell.fill = data_fill
                    for merged_row in range(group_start_row, group_end_row + 1):
                        ws.cell(row=merged_row, column=col).alignment = center
                        ws.cell(row=merged_row, column=col).border = border
                        ws.cell(row=merged_row, column=col).fill = data_fill

        ws.freeze_panes = "A2"
        widths = [24, 22, 30, 14, 14]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _write_summary_spec_real_reason_sheet(self, ws, summary_snapshot):
        thin_side = Side(style="thin", color="B7B7B7")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        headers = ["时间", "店铺", "规格编码", "退款原因", "数量", "占比"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[1].height = 32

        stores = summary_snapshot.get("stores", []) if isinstance(summary_snapshot, dict) else []
        row = 2
        for store in stores:
            metrics = store.get("metrics", {})
            date_range = self._format_summary_basic_stats_date_range(summary_snapshot, metrics)
            store_name = store.get("store_name", "未知店铺")
            analysis = store.get("real_reason_analysis", {})
            spec_categories = analysis.get("spec_categories", []) if isinstance(analysis, dict) else []
            if not spec_categories:
                spec_categories = [{"spec": "-", "categories": [{"name": "无", "count": 0, "ratio": 0}]}]

            for spec_item in spec_categories:
                spec = str(spec_item.get("spec", "-") or "-").strip() or "-"
                categories = spec_item.get("categories", []) or [{"name": "无", "count": 0, "ratio": 0}]
                group_start_row = row
                group_fill = self._get_stable_spec_fill(spec)
                for category in categories:
                    values = [
                        date_range,
                        store_name,
                        spec,
                        category.get("name", "未分类"),
                        category.get("count", 0),
                        f"{self._safe_float(category.get('ratio', 0)):.2f}%",
                    ]
                    for col, value in enumerate(values, start=1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.fill = group_fill
                        cell.alignment = center
                        cell.border = border
                    ws.row_dimensions[row].height = 24
                    row += 1

                group_end_row = row - 1
                if group_end_row > group_start_row:
                    for col in (1, 2, 3):
                        ws.merge_cells(
                            start_row=group_start_row,
                            start_column=col,
                            end_row=group_end_row,
                            end_column=col
                        )
                        cell = ws.cell(row=group_start_row, column=col)
                        cell.alignment = center
                        cell.border = border
                        cell.fill = group_fill
                        for merged_row in range(group_start_row, group_end_row + 1):
                            ws.cell(row=merged_row, column=col).alignment = center
                            ws.cell(row=merged_row, column=col).border = border
                            ws.cell(row=merged_row, column=col).fill = group_fill

        ws.freeze_panes = "A2"
        widths = [24, 22, 24, 30, 14, 14]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    @staticmethod
    def _parse_summary_date_text(date_text):
        text = str(date_text or "").strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def _get_summary_single_week_start(self, summary_snapshot):
        """筛选范围必须完整落在同一自然周内，才返回该周周一。"""
        filters = summary_snapshot.get("filters", {}) if isinstance(summary_snapshot, dict) else {}
        start_text = filters.get("start_date", "")
        end_text = filters.get("end_date", "")
        if not start_text or not end_text:
            metrics = {}
            stores = summary_snapshot.get("stores", []) if isinstance(summary_snapshot, dict) else []
            if stores:
                metrics = stores[0].get("metrics", {}) or {}
            elif isinstance(summary_snapshot.get("totals"), dict):
                metrics = summary_snapshot.get("totals", {}) or {}
            date_range = metrics.get("date_range", {}) if isinstance(metrics, dict) else {}
            start_text = date_range.get("start_date", "")
            end_text = date_range.get("end_date", "")

        start_date = self._parse_summary_date_text(start_text)
        end_date = self._parse_summary_date_text(end_text)
        if not start_date or not end_date:
            return ""

        start_monday = start_date - timedelta(days=start_date.weekday())
        end_monday = end_date - timedelta(days=end_date.weekday())
        if start_monday != end_monday:
            return ""
        return start_monday.strftime("%Y-%m-%d")

    def _build_summary_spec_order_map(self, summary_snapshot):
        week_start = self._get_summary_single_week_start(summary_snapshot)
        if not week_start:
            return {}

        result = {}
        stores = summary_snapshot.get("stores", []) if isinstance(summary_snapshot, dict) else []
        for store in stores:
            store_name = str(store.get("store_name") or "").strip()
            if not store_name:
                continue
            store_id = self.db.get_store_id_by_name(store_name)
            if not store_id:
                continue
            items = self.db.get_store_weekly_spec_orders_by_week(store_id, week_start)
            result[store_name] = {
                str(item.get("spec_code") or "").strip(): int(item.get("order_count") or 0)
                for item in items
                if str(item.get("spec_code") or "").strip()
            }
        return result

    def _write_summary_spec_real_reason_upgrade_sheet(self, ws, summary_snapshot):
        thin_side = Side(style="thin", color="B7B7B7")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        headers = ["时间", "店铺", "规格编码", "规格订单量", "退款原因", "数量", "占比"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[1].height = 32

        spec_order_map = self._build_summary_spec_order_map(summary_snapshot)
        stores = summary_snapshot.get("stores", []) if isinstance(summary_snapshot, dict) else []
        row = 2
        for store in stores:
            metrics = store.get("metrics", {})
            date_range = self._format_summary_basic_stats_date_range(summary_snapshot, metrics)
            store_name = store.get("store_name", "未知店铺")
            analysis = store.get("real_reason_analysis", {})
            spec_categories = analysis.get("spec_categories", []) if isinstance(analysis, dict) else []
            if not spec_categories:
                spec_categories = [{"spec": "-", "categories": [{"name": "无", "count": 0}]}]

            store_spec_orders = spec_order_map.get(store_name, {})
            store_fill = self._get_stable_store_fill(store_name)
            for spec_item in spec_categories:
                spec = str(spec_item.get("spec", "-") or "-").strip() or "-"
                categories = spec_item.get("categories", []) or [{"name": "无", "count": 0}]
                group_start_row = row
                raw_order_count = store_spec_orders.get(spec) if spec != "-" else None
                order_count = raw_order_count if raw_order_count and raw_order_count > 0 else None

                for category in categories:
                    refund_count = int(category.get("count") or 0)
                    ratio_text = f"{(refund_count / order_count * 100):.2f}%" if order_count else ""
                    values = [
                        date_range,
                        store_name,
                        spec,
                        order_count if order_count else "",
                        category.get("name", "未分类"),
                        refund_count,
                        ratio_text,
                    ]
                    for col, value in enumerate(values, start=1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.fill = store_fill
                        cell.alignment = center
                        cell.border = border
                    ws.row_dimensions[row].height = 24
                    row += 1

                group_end_row = row - 1
                if group_end_row > group_start_row:
                    for col in (1, 2, 3, 4):
                        ws.merge_cells(
                            start_row=group_start_row,
                            start_column=col,
                            end_row=group_end_row,
                            end_column=col
                        )
                        cell = ws.cell(row=group_start_row, column=col)
                        cell.alignment = center
                        cell.border = border
                        cell.fill = store_fill
                        for merged_row in range(group_start_row, group_end_row + 1):
                            ws.cell(row=merged_row, column=col).alignment = center
                            ws.cell(row=merged_row, column=col).border = border
                            ws.cell(row=merged_row, column=col).fill = store_fill

        ws.freeze_panes = "A2"
        widths = [24, 22, 24, 16, 30, 14, 14]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _write_summary_export_sheet(self, ws, real_reason_analysis=None):
        thin_side = Side(style="thin", color="B7B7B7")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row = 1
        if real_reason_analysis:
            row = self._write_real_reason_distribution_blocks(ws, row, real_reason_analysis, border, center)
        else:
            ws.cell(row=row, column=1, value="无")
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=1).border = border

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 42
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value is not None:
                    cell.alignment = center

    def _write_quality_refund_order_details_sheet(self, ws, summary_snapshot):
        thin_side = Side(style="thin", color="B7B7B7")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        summary_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        summary_data_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        empty_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

        detail_rows = []
        summary_rows = []
        if isinstance(summary_snapshot, dict):
            detail_rows = summary_snapshot.get("quality_refund_order_details") or []
            summary_rows = summary_snapshot.get("quality_not_cancelled_reason_summary") or []
        if not summary_rows and detail_rows:
            summary_rows = self._build_quality_not_cancelled_reason_summary(detail_rows)

        detail_header_row = 2
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        detail_title_cell = ws.cell(row=1, column=1, value="品质退款订单明细")
        detail_title_cell.fill = summary_header_fill
        detail_title_cell.font = Font(size=12, bold=True, color="FFFFFF")
        detail_title_cell.alignment = center
        for col in range(1, 8):
            ws.cell(row=1, column=col).fill = summary_header_fill
            ws.cell(row=1, column=col).border = border

        headers = ["订单号", "店铺", "规格编码", "退款原因", "备注", "未撤销原因", "分析说明"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=detail_header_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[detail_header_row].height = 30

        detail_end_row = detail_header_row
        if detail_rows:
            indexed_rows = [
                (index, item)
                for index, item in enumerate(detail_rows)
            ]
            indexed_rows.sort(key=lambda pair: (
                str(pair[1].get("store_name") or "未知店铺"),
                str(pair[1].get("spec_code") or "-"),
                pair[0],
                str(pair[1].get("order_no") or ""),
            ))
            group_ranges = {}
            for row_index, (_original_index, item) in enumerate(indexed_rows, start=detail_header_row + 1):
                store_name = str(item.get("store_name") or "未知店铺")
                spec_code = str(item.get("spec_code") or "-").strip() or "-"
                store_fill = self._get_stable_store_fill(store_name)
                values = [
                    item.get("order_no", ""),
                    store_name,
                    spec_code,
                    item.get("refund_reason", ""),
                    item.get("notes", ""),
                    item.get("not_cancelled_reason", ""),
                    item.get("analysis_detail", ""),
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_index, column=col, value=value)
                    cell.fill = store_fill
                    cell.alignment = left_wrap if col in (5, 7) else center
                    cell.border = border
                ws.row_dimensions[row_index].height = 48
                store_key = ("store", store_name)
                spec_key = ("spec", store_name, spec_code)
                group_ranges.setdefault(store_key, [row_index, row_index])[1] = row_index
                group_ranges.setdefault(spec_key, [row_index, row_index])[1] = row_index
                detail_end_row = row_index

            for (group_type, *_group_values), (start_row, end_row) in group_ranges.items():
                if end_row <= start_row:
                    continue
                col = 2 if group_type == "store" else 3
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                cell = ws.cell(row=start_row, column=col)
                cell.alignment = center
                cell.border = border
                for merged_row in range(start_row, end_row + 1):
                    ws.cell(row=merged_row, column=col).alignment = center
                    ws.cell(row=merged_row, column=col).border = border
        else:
            empty_row = detail_header_row + 1
            ws.merge_cells(start_row=empty_row, start_column=1, end_row=empty_row, end_column=len(headers))
            cell = ws.cell(row=empty_row, column=1, value="当前筛选范围无未撤销品质退款订单")
            cell.fill = empty_fill
            cell.alignment = center
            cell.border = border
            for col in range(1, len(headers) + 1):
                ws.cell(row=empty_row, column=col).fill = empty_fill
                ws.cell(row=empty_row, column=col).border = border
            ws.row_dimensions[empty_row].height = 28
            detail_end_row = empty_row

        summary_title_row = detail_end_row + 3
        ws.merge_cells(start_row=summary_title_row, start_column=1, end_row=summary_title_row, end_column=3)
        title_cell = ws.cell(row=summary_title_row, column=1, value="品质退款未撤销原因占比")
        title_cell.fill = summary_header_fill
        title_cell.font = Font(size=12, bold=True, color="FFFFFF")
        title_cell.alignment = center
        for col in range(1, 4):
            ws.cell(row=summary_title_row, column=col).fill = summary_header_fill
            ws.cell(row=summary_title_row, column=col).border = border

        summary_header_row = summary_title_row + 1
        summary_headers = ["未撤销原因", "数量", "占比"]
        for col, header in enumerate(summary_headers, start=1):
            cell = ws.cell(row=summary_header_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.alignment = center
            cell.border = border

        row_index = summary_header_row + 1
        if summary_rows:
            for item in summary_rows:
                values = [
                    item.get("name", "未明确说明未撤销原因"),
                    item.get("count", 0),
                    f"{self._safe_float(item.get('ratio', 0)):.2f}%",
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_index, column=col, value=value)
                    cell.fill = summary_data_fill
                    cell.alignment = left_wrap if col == 1 else center
                    cell.border = border
                ws.row_dimensions[row_index].height = 28
                row_index += 1
        else:
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            cell = ws.cell(row=row_index, column=1, value="当前筛选范围无未撤销品质退款订单")
            cell.fill = empty_fill
            cell.alignment = center
            cell.border = border
            for col in range(1, 4):
                ws.cell(row=row_index, column=col).fill = empty_fill
                ws.cell(row=row_index, column=col).border = border
            ws.row_dimensions[row_index].height = 28

        ws.freeze_panes = None
        widths = [22, 24, 16, 28, 62, 34, 62]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def export_summary_excel(self, checked=False, snapshot=None):
        """导出总结快照到Excel。"""
        summary_snapshot = snapshot or self.latest_summary_snapshot
        dialog_parent = self.ai_window if hasattr(self, 'ai_window') and self.ai_window else self
        if not summary_snapshot:
            QMessageBox.information(dialog_parent, "提示", "请先生成总结或从历史中打开一份总结")
            self.ensure_ai_window_visible()
            return

        default_name = f"本地总结_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(dialog_parent, "导出总结Excel", default_name, "Excel文件 (*.xlsx)")
        if not file_path:
            self.ensure_ai_window_visible()
            return

        try:
            wb = openpyxl.Workbook()
            default_ws = wb.active
            wb.remove(default_ws)
            used_names = set()

            stores = summary_snapshot.get("stores", [])
            basic_ws = wb.create_sheet(self._safe_excel_sheet_name("基础统计", used_names))
            self._write_summary_basic_stats_sheet(basic_ws, summary_snapshot)
            overall_reason_ws = wb.create_sheet(self._safe_excel_sheet_name("总体真实退款原因分布", used_names))
            self._write_summary_overall_real_reason_sheet(overall_reason_ws, summary_snapshot)
            spec_reason_ws = wb.create_sheet(self._safe_excel_sheet_name("按规格编码分布", used_names))
            self._write_summary_spec_real_reason_sheet(spec_reason_ws, summary_snapshot)
            spec_reason_upgrade_ws = wb.create_sheet(self._safe_excel_sheet_name("按规格编码分布 升级版", used_names))
            self._write_summary_spec_real_reason_upgrade_sheet(spec_reason_upgrade_ws, summary_snapshot)
            quality_detail_ws = wb.create_sheet(self._safe_excel_sheet_name("品质退款订单明细", used_names))
            self._write_quality_refund_order_details_sheet(quality_detail_ws, summary_snapshot)

            for store in stores:
                sheet_name = self._safe_excel_sheet_name(store.get("store_name", "未知店铺"), used_names)
                ws = wb.create_sheet(sheet_name)
                self._write_summary_export_sheet(
                    ws,
                    store.get("real_reason_analysis", {})
                )

            totals = summary_snapshot.get("totals")
            if totals:
                ws = wb.create_sheet(self._safe_excel_sheet_name("全部总和", used_names))
                self._write_summary_export_sheet(
                    ws,
                    summary_snapshot.get("overall_real_reason_analysis", {})
                )

            if not wb.sheetnames:
                ws = wb.create_sheet("本地总结")
                self._write_summary_export_sheet(ws, {})

            wb.save(file_path)
            self.show_tooltip("总结导出成功", "rgba(76, 175, 80, 0.95)", 1500)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出总结失败：{e}")
        finally:
            self.ensure_ai_window_visible()

    def open_summary_history(self):
        """打开总结历史记录窗口。"""
        dialog_parent = self.ai_window if hasattr(self, 'ai_window') and self.ai_window else self
        dialog = SummaryHistoryDialog(self.db, self, dialog_parent)
        if dialog.exec_() == QDialog.Accepted and dialog.selected_snapshot:
            self.latest_summary_history_id = dialog.selected_history_id
            self.display_summary_snapshot(dialog.selected_snapshot)
        self.ensure_ai_window_visible()

    @staticmethod
    def normalize_header_text(text):
        """标准化表头文本，便于做别名和关键词匹配。"""
        return re.sub(r'[\s\-_（）()【】\[\]:：/\\]+', '', str(text or '').strip()).lower()

    def suggest_column_mapping(self, headers, column_configs):
        """自动推荐 Excel 列映射。"""
        normalized_headers = {header: self.normalize_header_text(header) for header in headers if header}
        mapping = {}

        for config in column_configs:
            target_name = config['target']
            aliases = [self.normalize_header_text(alias) for alias in config.get('aliases', []) if alias]
            keywords = [self.normalize_header_text(keyword) for keyword in config.get('keywords', []) if keyword]

            matched_header = None

            # 先走精确别名匹配
            for header, normalized in normalized_headers.items():
                if normalized in aliases:
                    matched_header = header
                    break

            # 再走关键词命中数匹配
            if not matched_header and keywords:
                scored_headers = []
                for header, normalized in normalized_headers.items():
                    score = sum(1 for keyword in keywords if keyword and keyword in normalized)
                    if score > 0:
                        scored_headers.append((score, len(normalized), header))

                if scored_headers:
                    scored_headers.sort(key=lambda item: (-item[0], item[1]))
                    matched_header = scored_headers[0][2]

            if matched_header:
                mapping[target_name] = matched_header

        return mapping

    def _get_hidden_import_time_mapping(self, headers):
        """自动识别隐藏时间字段，不放入列映射确认窗口。"""
        return self.suggest_column_mapping(headers, [
            {
                'target': '申请时间',
                'aliases': ['申请时间', '申请退款时间', '退款申请时间', '售后申请时间'],
                'keywords': ['申请时间', '申请退款时间', '退款申请时间', '售后申请时间'],
            },
            {
                'target': '同意退款时间',
                'aliases': ['同意退款时间', '同意时间', '退款同意时间', '同意售后时间'],
                'keywords': ['同意退款时间', '同意时间', '退款同意时间', '同意售后时间'],
            },
        ])

    def check_required_columns(self, headers, required_config):
        """检查必要列：支持模糊匹配"""
        missing_columns = []
        column_mapping = {}
        
        for col_config in required_config:
            if isinstance(col_config, str):
                # 简单字符串匹配
                if col_config not in headers:
                    missing_columns.append(col_config)
                else:
                    column_mapping[col_config] = col_config
            elif isinstance(col_config, dict):
                # 模糊匹配配置
                target_name = col_config['target']
                suggested_mapping = self.suggest_column_mapping(headers, [col_config])
                
                if target_name in suggested_mapping:
                    column_mapping[target_name] = suggested_mapping[target_name]
                else:
                    missing_columns.append(target_name)
        
        return missing_columns, column_mapping

    def import_excel(self):
        """导入Excel文件（自动识别列映射并允许用户确认/手动调整）"""
        file_path, _ = QFileDialog.getOpenFileName(self, "导入订单", "", "Excel文件 (*.xlsx)")
        if not file_path:
            return

        # 解析Excel
        data_rows = []
        column_mapping = {}
        hidden_time_mapping = {}
        try:
            if file_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
                hidden_time_mapping = self._get_hidden_import_time_mapping(headers)

                # 自动识别字段，并允许用户确认/手动调整
                column_configs = [
                    {
                        'target': '店铺名称',
                        'aliases': ['店铺名称', '店铺名', '店名', '门店名称', '门店'],
                        'keywords': ['店铺', '店名', '门店'],
                        'required': False
                    },
                    {
                        'target': '订单号',
                        'aliases': ['订单号', '订单编号', '订单编码', '单号', '订单id', 'orderid', 'orderno'],
                        'keywords': ['订单号', '订单编号', '订单', '单号', '编号', '编码', 'order'],
                        'required': True
                    },
                    {
                        'target': '规格名称',
                        'aliases': ['规格名称', '商品规格', '规格', '商品名称', 'sku名称', '规格描述'],
                        'keywords': ['规格名称', '商品规格', '规格', '商品名称', 'sku'],
                        'required': False
                    },
                    {
                        'target': '退款原因',
                        'aliases': ['退款原因', '原因', '退款理由', '原因说明'],
                        'keywords': ['退款原因', '原因', '理由'],
                        'required': True
                    },
                    {
                        'target': '退款金额',
                        'aliases': ['退款金额', '退款金额(元)', '金额', '退款钱数', '退款费用'],
                        'keywords': ['退款金额', '退款', '金额', 'amount'],
                        'required': True
                    },
                    {
                        'target': '撤销',
                        'aliases': ['撤销', '是否撤销', '取消', '撤单'],
                        'keywords': ['撤销', '取消', '撤单'],
                        'required': False
                    },
                    {
                        'target': '打款补偿',
                        'aliases': ['打款补偿', '是否打款补偿', '打款', '补偿', '赔付'],
                        'keywords': ['打款', '补偿', '赔付', '赔偿'],
                        'required': False
                    },
                    {
                        'target': '补偿金额',
                        'aliases': ['补偿金额', '赔付金额', '赔偿金额', '补偿费用'],
                        'keywords': ['补偿金额', '赔付金额', '赔偿金额', '补偿'],
                        'required': False
                    },
                    {
                        'target': '驳回',
                        'aliases': ['驳回', '是否驳回', '拒绝', '不通过'],
                        'keywords': ['驳回', '拒绝', '不通过'],
                        'required': False
                    },
                    {
                        'target': '驳回结果',
                        'aliases': ['驳回结果', '处理结果', '结果', '审核结果'],
                        'keywords': ['驳回结果', '处理结果', '审核结果', '结果'],
                        'required': False
                    },
                    {
                        'target': '备注',
                        'aliases': ['备注', '说明', '注释', '备注说明', 'note'],
                        'keywords': ['备注', '说明', '注释', 'note'],
                        'required': False
                    },
                    {
                        'target': '订单状态',
                        'aliases': ['订单状态', '发货状态', '物流状态'],
                        'keywords': ['订单状态', '发货状态', '物流状态', '发货'],
                        'required': False
                    },
                    {
                        'target': '售后状态',
                        'aliases': ['售后状态', '退款状态', '售后单状态'],
                        'keywords': ['售后状态', '退款状态', '售后单状态', '售后'],
                        'required': False
                    },
                    {
                        'target': '登记日期',
                        'aliases': ['登记日期', '登记时间', '日期', '创建日期', '创建时间', '下单日期'],
                        'keywords': ['登记日期', '登记时间', '创建日期', '创建时间', '下单日期', 'date'],
                        'required': False
                    }
                ]

                # 检查必要列：根据搜索筛选区选择动态调整
                current_search_store = self.search_store_combo.currentText()
                if current_search_store and current_search_store != "全部":
                    # 选择了具体店铺，店铺名称列可选，登记日期可手动确认
                    required_fields = ['订单号', '退款原因', '退款金额', '登记日期']
                else:
                    # 选择了"全部"，店铺名称列和登记日期都是必填
                    required_fields = ['店铺名称', '订单号', '退款原因', '退款金额', '登记日期']

                suggested_mapping = self.suggest_column_mapping(headers, column_configs)
                mapping_dialog = ColumnMappingDialog(
                    headers=headers,
                    column_configs=column_configs,
                    initial_mapping=suggested_mapping,
                    required_fields=required_fields,
                    parent=self
                )
                if mapping_dialog.exec_() != QDialog.Accepted:
                    return

                column_mapping = mapping_dialog.get_mapping()
                missing_required = [
                    field for field in required_fields
                    if field not in column_mapping and not (field == '登记日期' and '同意退款时间' in hidden_time_mapping)
                ]
                if missing_required:
                    QMessageBox.critical(self, "错误", f"缺少必要字段：{', '.join(missing_required)}")
                    return
                
                # 读取数据行，读取所有列（不仅仅是必要列）
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):  # 空行跳过
                        continue
                    row_dict = {}
                    for idx, val in enumerate(row):
                        if idx < len(headers):
                            header_name = headers[idx]
                            # 读取所有列，而不仅仅是必要列
                            row_dict[header_name] = val
                    data_rows.append(row_dict)
            else:
                QMessageBox.critical(self, "错误", "不支持的文件格式")
                return
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取文件失败：{str(e)}")
            return

        if not data_rows:
            QMessageBox.information(self, "提示", "Excel中没有数据")
            return

        # 显示导入文件基本信息
        total_rows = len(data_rows)
        QMessageBox.information(self, "导入文件信息", 
                               f"Excel文件包含 {total_rows} 条数据\n\n"
                               f"开始导入处理...")

        # 处理导入
        success_count = 0
        overwrite_count = 0
        skip_count = 0
        fail_count = 0
        duplicate_count = 0
        status_filtered_skip_count = 0
        imported_record_dates = []
        self.highlighted_orders.clear()
        import_created_ids = []
        import_updated_records = {}
        
        # 收集所有重复订单信息
        duplicate_orders = []  # 存储重复订单信息
        valid_rows = []  # 存储有效的行数据
        
        # 第一步：合并Excel文件中的重复订单（同一个订单号出现多次）
        merged_data_rows = []
        order_no_groups = {}
        
        # 按订单号分组，识别Excel中的重复订单
        for row_idx, row in enumerate(data_rows):
            try:
                order_no = str(self._extract_mapped_value(row, column_mapping, '订单号', '')).strip()
                if '订单状态' in column_mapping:
                    order_status = str(self._extract_mapped_value(row, column_mapping, '订单状态', '')).strip()
                    if order_status != '已发货':
                        status_filtered_skip_count += 1
                        continue
                
                if order_no:
                    if order_no not in order_no_groups:
                        order_no_groups[order_no] = []
                    order_no_groups[order_no].append((row_idx, row))
            except:
                pass
        
        # 处理重复订单合并
        merge_info = []
        for order_no, rows in order_no_groups.items():
            if len(rows) > 1:
                # 发现重复订单，进行金额合并
                total_refund_amount = 0.0
                total_comp_amount = 0.0
                first_row_data = None
                first_active_row_data = None
                canceled_row_count = 0
                
                for row_idx, row in rows:
                    refund_amount = self._coerce_import_float(self._extract_mapped_value(row, column_mapping, '退款金额', 0), 0.0)
                    comp_amount = self._coerce_import_float(self._extract_mapped_value(row, column_mapping, '补偿金额', 0), 0.0)
                    row_after_sale_status = str(self._extract_mapped_value(row, column_mapping, '售后状态', '')).strip()
                    row_is_canceled = '售后状态' in column_mapping and row_after_sale_status == '已撤销'
                    if row_is_canceled:
                        canceled_row_count += 1
                        refund_amount = 0.0
                    elif first_active_row_data is None:
                        first_active_row_data = row
                    
                    total_refund_amount += refund_amount
                    total_comp_amount += comp_amount
                    
                    # 保存第一个订单的数据作为基础
                    if first_row_data is None:
                        first_row_data = row
                
                if first_row_data:
                    base_row = first_active_row_data if first_active_row_data is not None else first_row_data
                    merged_row = base_row.copy()
                    # 更新合并后的金额
                    if '退款金额' in column_mapping:
                        actual_amount_col = column_mapping['退款金额']
                        merged_row[actual_amount_col] = total_refund_amount
                    if '补偿金额' in column_mapping:
                        actual_comp_amount_col = column_mapping['补偿金额']
                        merged_row[actual_comp_amount_col] = total_comp_amount
                    if '售后状态' in column_mapping:
                        actual_after_sale_col = column_mapping['售后状态']
                        merged_row[actual_after_sale_col] = '已撤销' if total_refund_amount <= 0 and canceled_row_count == len(rows) else ''
                    
                    # 添加合并备注
                    if '备注' in column_mapping:
                        actual_notes_col = column_mapping['备注']
                        original_notes = merged_row.get(actual_notes_col, '')
                        merge_note = f"合并了{len(rows)}条重复记录，退款金额合计：{total_refund_amount:.2f}元"
                        if total_comp_amount > 0:
                            merge_note += f"，补偿金额合计：{total_comp_amount:.2f}元"
                        if canceled_row_count > 0:
                            merge_note += f"，其中{canceled_row_count}条已撤销未计入退款金额"
                        
                        if original_notes:
                            merged_row[actual_notes_col] = f"{original_notes} | {merge_note}"
                        else:
                            merged_row[actual_notes_col] = merge_note
                    
                    merged_data_rows.append(merged_row)
                    merge_info.append(f"订单号 {order_no}: 合并{len(rows)}条记录，退款金额={total_refund_amount:.2f}元")
            else:
                # 没有重复，直接添加
                merged_data_rows.append(rows[0][1])
        
        # 显示合并信息（如果有重复订单）
        if merge_info:
            merge_summary = f"发现并合并了 {len(merge_info)} 个重复订单：\n\n"
            merge_summary += "\n".join(merge_info)
            QMessageBox.information(self, "重复订单合并", merge_summary)
        
        # 第二步：预处理所有数据，收集重复订单信息（针对软件数据库中的重复）
        for row in merged_data_rows:
            try:
                store_name = str(self._extract_mapped_value(row, column_mapping, '店铺名称', '')).strip()
                
                # 如果店铺名称为空，检查搜索筛选区是否选择了具体店铺
                if not store_name:
                    # 获取当前搜索筛选区选择的店铺
                    current_search_store = self.search_store_combo.currentText()
                    if current_search_store and current_search_store != "全部":
                        # 使用搜索筛选区选择的店铺名称
                        store_name = current_search_store
                    else:
                        # 没有选择具体店铺，跳过该行
                        fail_count += 1
                        continue
                
                order_no = str(self._extract_mapped_value(row, column_mapping, '订单号', '')).strip()
                if not order_no:
                    fail_count += 1
                    continue
                
                reason = str(self._extract_mapped_value(row, column_mapping, '退款原因', '')).strip()
                if not reason:
                    fail_count += 1
                    continue
                
                # 定义品质退款原因列表（除了"其他"之外的所有原因）
                quality_reasons = ["商品腐败、变质、包装胀气等", "商品破损/压坏", "质量问题", 
                                  "大小/规格/重量等与商品描述不符", "品种/标签/图片/包装等与商品描述不符", "货物与描述不符"]
                
                # 如果导入的原因不在品质退款列表中，则归类为"其他"
                if reason not in quality_reasons and reason != "其他":
                    reason = "其他"
                
                refund_amount = self._coerce_import_float(self._extract_mapped_value(row, column_mapping, '退款金额', None), None)
                if refund_amount is None:
                    fail_count += 1
                    continue
                
                compensate = self._coerce_import_bool(self._extract_mapped_value(row, column_mapping, '打款补偿', '否'))
                comp_amount = self._coerce_import_float(self._extract_mapped_value(row, column_mapping, '补偿金额', 0), 0.0)
                cancel = self._coerce_import_bool(self._extract_mapped_value(row, column_mapping, '撤销', '否'))
                order_status = str(self._extract_mapped_value(row, column_mapping, '订单状态', '')).strip()
                after_sale_status = str(self._extract_mapped_value(row, column_mapping, '售后状态', '')).strip()
                if '售后状态' in column_mapping:
                    cancel = after_sale_status == '已撤销'
                reject = self._coerce_import_bool(self._extract_mapped_value(row, column_mapping, '驳回', '否'))
                reject_result = self._extract_mapped_value(row, column_mapping, '驳回结果', '')
                if isinstance(reject_result, str):
                    reject_result = reject_result.strip()
                else:
                    reject_result = str(reject_result) if reject_result else ''
                
                # 如果驳回为否，则驳回结果强制设置为"无"
                if not reject:
                    reject_result = "无"
                
                notes = self._extract_mapped_value(row, column_mapping, '备注', '')
                if isinstance(notes, str):
                    notes = notes.strip()
                else:
                    notes = str(notes) if notes else ''

                spec_name = self._extract_mapped_value(row, column_mapping, '规格名称', '')
                spec_name, auto_spec_code = self._extract_spec_info(spec_name)
                
                refund_apply_time = self._parse_import_datetime_value(
                    self._extract_mapped_value(row, hidden_time_mapping, '申请时间', '')
                )
                refund_agree_time = self._parse_import_datetime_value(
                    self._extract_mapped_value(row, hidden_time_mapping, '同意退款时间', '')
                )
                record_date = self._resolve_import_record_date(row, column_mapping)
                if '登记日期' not in column_mapping and refund_agree_time:
                    record_date = refund_agree_time[:10]

                # 智能店铺名称识别策略
                # 1. 首先检查订单号是否在软件数据库中存在
                existing = self.db.get_record_by_order_no(order_no)
                
                if existing:
                    # 订单号存在：使用软件中已有的店铺名称（增量存储策略）
                    store_name = existing['store_name']  # 使用软件中已有的店铺名称
                    store_id = existing['store_id']
                else:
                    # 订单号不存在：检查Excel表格是否有店铺名称列
                    if '店铺名称' in column_mapping and store_name:
                        # Excel有店铺名称列：使用Excel中的店铺名称
                        pass
                    else:
                        # Excel没有店铺名称列：使用当前搜索筛选选择的店铺名称
                        current_search_store = self.search_store_combo.currentText()
                        if current_search_store and current_search_store != "全部":
                            store_name = current_search_store
                        else:
                            # 没有选择具体店铺，跳过该行
                            fail_count += 1
                            continue
                    
                    # 获取或创建店铺
                    store_id = None
                    stores = self.db.get_stores()
                    for sid, sname in stores:
                        if sname == store_name:
                            store_id = sid
                            break
                    if store_id is None:
                        # 自动添加店铺
                        store_id = self.db.add_store(store_name)
                        if store_id is None:
                            fail_count += 1
                            continue
                        self.load_stores()  # 刷新下拉框

                # 检查订单号是否存在（再次检查，因为上面可能已经获取了existing）
                if existing:
                    # 记录识别到的字段信息（用于增量覆盖）
                    detected_fields = {}
                    
                    # 退款金额必须更新（变量字段）
                    detected_fields['refund_amount'] = refund_amount
                    
                    # 只更新识别到的字段
                    if '退款原因' in column_mapping:
                        detected_fields['reason'] = reason
                    if '撤销' in column_mapping:
                        detected_fields['cancel'] = cancel
                    if '打款补偿' in column_mapping:
                        detected_fields['compensate'] = compensate
                    if '补偿金额' in column_mapping:
                        detected_fields['comp_amount'] = comp_amount
                    if '驳回' in column_mapping:
                        detected_fields['reject'] = reject
                    if '驳回结果' in column_mapping and reject_result:
                        detected_fields['reject_result'] = reject_result
                    if '备注' in column_mapping:
                        detected_fields['notes'] = notes
                    if '规格名称' in column_mapping and spec_name:
                        detected_fields['spec_name'] = spec_name
                        if not existing.get('spec_code'):
                            detected_fields['spec_code'] = auto_spec_code
                    if '订单状态' in column_mapping:
                        detected_fields['order_status'] = order_status
                    if '售后状态' in column_mapping:
                        detected_fields['after_sale_status'] = after_sale_status
                    if '登记日期' in column_mapping and record_date:
                        detected_fields['record_date'] = record_date
                    if '申请时间' in hidden_time_mapping and refund_apply_time:
                        detected_fields['refund_apply_time'] = refund_apply_time
                    if '同意退款时间' in hidden_time_mapping and refund_agree_time:
                        detected_fields['refund_agree_time'] = refund_agree_time
                    
                    # 比较数据是否一致（只比较识别到的字段）
                    same = True
                    for field, new_value in detected_fields.items():
                        if field == 'refund_amount':
                            if abs(existing['refund_amount'] - new_value) >= 0.01:
                                same = False
                                break
                        elif field == 'cancel' or field == 'compensate' or field == 'reject':
                            if existing[field] != new_value:
                                same = False
                                break
                        elif field == 'comp_amount':
                            if abs(existing['comp_amount'] - new_value) >= 0.01:
                                same = False
                                break
                        else:
                            if existing[field] != new_value:
                                same = False
                                break
                    if same:
                        skip_count += 1
                        continue
                    else:
                        # 记录重复订单信息（包含识别到的字段信息）
                        duplicate_orders.append({
                            'order_no': order_no,
                            'existing_data': existing,
                            'new_data': {
                                'store_id': store_id,
                                'order_no': order_no,
                                'reason': reason,
                                'refund_amount': refund_amount,
                                'cancel': cancel,
                                'compensate': compensate,
                                'comp_amount': comp_amount,
                                'reject': reject,
                                'reject_result': reject_result,
                                'notes': notes,
                                'spec_name': spec_name,
                                'spec_code': auto_spec_code,
                                'order_status': order_status,
                                'after_sale_status': after_sale_status,
                                'refund_apply_time': refund_apply_time,
                                'refund_agree_time': refund_agree_time,
                                'record_date': record_date
                            },
                            'detected_fields': detected_fields  # 记录识别到的字段信息
                        })
                else:
                    # 新增订单，直接添加到有效行
                    valid_rows.append({
                        'store_id': store_id,
                        'order_no': order_no,
                        'reason': reason,
                        'refund_amount': refund_amount,
                        'cancel': cancel,
                        'compensate': compensate,
                        'comp_amount': comp_amount,
                        'reject': reject,
                        'reject_result': reject_result,
                        'notes': notes,
                        'spec_name': spec_name,
                        'spec_code': auto_spec_code,
                        'order_status': order_status,
                        'after_sale_status': after_sale_status,
                        'refund_apply_time': refund_apply_time,
                        'refund_agree_time': refund_agree_time,
                        'record_date': record_date
                    })
            except Exception as e:
                fail_count += 1
                print(f"导入错误：{e}")
        
        # 第二步：如果有重复订单，提供详细处理选项
        if duplicate_orders:
            duplicate_count = len(duplicate_orders)
            
            # 创建详细的选择对话框
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("发现重复订单")
            msg_box.setIcon(QMessageBox.Question)
            
            # 显示详细的重复订单信息（包含店铺名称）
            duplicate_info = f"发现 {duplicate_count} 条重复订单（订单号已存在但数据不一致）\n\n"
            duplicate_info += f"重复订单示例：\n"
            for i, dup in enumerate(duplicate_orders[:5]):
                existing_store = dup['existing_data']['store_name']
                new_store = dup['new_data']['store_name'] if 'store_name' in dup['new_data'] else "导入文件中的店铺"
                duplicate_info += f"{i+1}. 订单号：{dup['order_no']} | 现有店铺：{existing_store} | 导入店铺：{new_store}\n"
            
            if duplicate_count > 5:
                duplicate_info += f"...等{duplicate_count}个订单\n"
            
            duplicate_info += "\n请选择处理方式："
            msg_box.setText(duplicate_info)
            
            # 添加自定义按钮（支持换行）
            overwrite_all_btn = msg_box.addButton("覆盖所有\n重复订单", QMessageBox.YesRole)
            skip_all_btn = msg_box.addButton("跳过所有\n重复订单", QMessageBox.NoRole)
            review_each_btn = msg_box.addButton("逐条查看\n并选择", QMessageBox.ActionRole)
            cancel_btn = msg_box.addButton("取消导入", QMessageBox.RejectRole)
            
            # 设置按钮样式（支持换行和变大）
            for btn in [overwrite_all_btn, skip_all_btn, review_each_btn, cancel_btn]:
                btn.setStyleSheet("""
                    QPushButton {
                        font-size: 12px;
                        padding: 8px 12px;
                        min-height: 40px;
                        min-width: 100px;
                    }
                """)
            
            msg_box.setDefaultButton(overwrite_all_btn)
            msg_box.exec_()
            
            clicked_button = msg_box.clickedButton()
            
            if clicked_button == overwrite_all_btn:
                # 智能增量覆盖所有重复订单
                current_search_store = self.search_store_combo.currentText()
                for dup in duplicate_orders:
                    existing_store = dup['existing_data']['store_name']
                    new_store = dup['new_data']['store_name'] if 'store_name' in dup['new_data'] else ""
                    
                    # 如果店铺名称不一致且当前搜索筛选选择了具体店铺，使用当前店铺
                    if existing_store != new_store and current_search_store and current_search_store != "全部":
                        # 获取当前店铺的ID
                        stores = self.db.get_stores()
                        current_store_id = None
                        for sid, sname in stores:
                            if sname == current_search_store:
                                current_store_id = sid
                                break
                        
                        if current_store_id:
                            # 使用当前搜索筛选的店铺
                            dup['new_data']['store_id'] = current_store_id
                            dup['new_data']['store_name'] = current_search_store
                    
                    # 智能增量更新：只更新识别到的字段
                    update_fields = {}
                    
                    # 退款金额必须更新（变量字段）
                    update_fields['refund_amount'] = dup['new_data']['refund_amount']
                    
                    # 只更新识别到的字段
                    if 'detected_fields' in dup:
                        for field, value in dup['detected_fields'].items():
                            if field != 'refund_amount':  # 退款金额已经单独处理
                                update_fields[field] = value
                    
                    # 使用智能增量更新函数
                    if update_fields:
                        import_updated_records.setdefault(dup['existing_data']['id'], dup['existing_data'].copy())
                        self.db.update_record_partial(dup['existing_data']['id'], **update_fields)
                    
                    overwrite_count += 1
                    imported_record_dates.append(dup['new_data']['record_date'])
                    self.highlighted_orders.add(dup['order_no'])
            elif clicked_button == skip_all_btn:
                # 跳过所有重复订单
                skip_count += duplicate_count
            elif clicked_button == review_each_btn:
                # 逐条查看重复订单
                for dup in duplicate_orders:
                    existing = dup['existing_data']
                    new_data = dup['new_data']
                    
                    # 显示详细的对比信息（包含店铺名称不一致处理）
                    comparison_info = f"订单号：{dup['order_no']}\n\n"
                    comparison_info += "【现有数据】\n"
                    comparison_info += f"店铺：{existing['store_name']}\n"
                    comparison_info += f"退款原因：{existing['reason']}\n"
                    comparison_info += f"退款金额：¥{existing['refund_amount']}\n"
                    comparison_info += f"登记日期：{existing['record_date']}\n\n"
                    
                    comparison_info += "【导入数据】\n"
                    comparison_info += f"店铺：{new_data['store_name']}\n"
                    comparison_info += f"退款原因：{new_data['reason']}\n"
                    comparison_info += f"退款金额：¥{new_data['refund_amount']}\n"
                    comparison_info += f"登记日期：{new_data['record_date']}\n\n"
                    
                    # 添加店铺名称不一致提示
                    if existing['store_name'] != new_data['store_name']:
                        current_search_store = self.search_store_combo.currentText()
                        if current_search_store and current_search_store != "全部":
                            comparison_info += f"⚠️ 店铺名称不一致，将使用当前筛选的店铺：{current_search_store}\n\n"
                        else:
                            comparison_info += f"⚠️ 店铺名称不一致：现有({existing['store_name']}) vs 导入({new_data['store_name']})\n\n"
                    
                    comparison_info += "请选择处理方式："
                    
                    review_msg_box = QMessageBox(self)
                    review_msg_box.setWindowTitle("重复订单处理")
                    review_msg_box.setIcon(QMessageBox.Question)
                    review_msg_box.setText(comparison_info)
                    
                    overwrite_btn = review_msg_box.addButton("覆盖现有\n数据", QMessageBox.YesRole)
                    skip_btn = review_msg_box.addButton("跳过此\n订单", QMessageBox.NoRole)
                    review_msg_box.addButton("取消剩余\n导入", QMessageBox.RejectRole)
                    
                    # 设置按钮样式
                    for btn in [overwrite_btn, skip_btn]:
                        btn.setStyleSheet("""
                            QPushButton {
                                font-size: 12px;
                                padding: 8px 12px;
                                min-height: 40px;
                                min-width: 80px;
                            }
                        """)
                    
                    review_msg_box.setDefaultButton(overwrite_btn)
                    review_msg_box.exec_()
                    
                    clicked_review_button = review_msg_box.clickedButton()
                    
                    if clicked_review_button == overwrite_btn:
                        # 智能增量覆盖此订单（处理店铺名称不一致）
                        current_search_store = self.search_store_combo.currentText()
                        if existing['store_name'] != new_data['store_name'] and current_search_store and current_search_store != "全部":
                            # 获取当前店铺的ID
                            stores = self.db.get_stores()
                            current_store_id = None
                            for sid, sname in stores:
                                if sname == current_search_store:
                                    current_store_id = sid
                                    break
                            
                            if current_store_id:
                                # 使用当前搜索筛选的店铺
                                new_data['store_id'] = current_store_id
                                new_data['store_name'] = current_search_store
                        
                        # 智能增量更新：只更新识别到的字段
                        update_fields = {}
                        
                        # 退款金额必须更新（变量字段）
                        update_fields['refund_amount'] = new_data['refund_amount']
                        
                        # 只更新识别到的字段
                        if 'detected_fields' in dup:
                            for field, value in dup['detected_fields'].items():
                                if field != 'refund_amount':  # 退款金额已经单独处理
                                    update_fields[field] = value
                        
                        # 使用智能增量更新函数
                        if update_fields:
                            import_updated_records.setdefault(existing['id'], existing.copy())
                            self.db.update_record_partial(existing['id'], **update_fields)
                        
                        overwrite_count += 1
                        imported_record_dates.append(new_data['record_date'])
                        self.highlighted_orders.add(dup['order_no'])
                    elif clicked_review_button == skip_btn:
                        # 跳过此订单
                        skip_count += 1
                    else:
                        # 取消剩余导入
                        skip_count += len(duplicate_orders) - duplicate_orders.index(dup) - 1
                        break
            else:
                # 取消导入
                QMessageBox.information(self, "导入取消", "导入操作已取消")
                return
        
        # 第三步：处理新增订单
        for row_data in valid_rows:
            try:
                record_id = self.db.add_record(row_data['store_id'],
                                              row_data['order_no'],
                                              row_data['reason'],
                                              row_data['refund_amount'],
                                              row_data['cancel'],
                                              row_data['compensate'],
                                              row_data['comp_amount'],
                                              row_data['reject'],
                                              row_data['reject_result'],
                                              row_data['notes'],
                                              row_data['record_date'],
                                              row_data.get('order_status', ''),
                                              row_data.get('after_sale_status', ''),
                                              row_data.get('spec_name', ''),
                                              row_data.get('spec_code', ''),
                                              row_data.get('refund_apply_time', ''),
                                              row_data.get('refund_agree_time', ''))
                import_created_ids.append(record_id)
                success_count += 1
                imported_record_dates.append(row_data['record_date'])
                self.highlighted_orders.add(row_data['order_no'])
            except Exception as e:
                fail_count += 1
                print(f"新增订单错误：{e}")

        # 显示详细的导入结果
        total_processed = success_count + overwrite_count + skip_count + fail_count
        self._last_import_undo_data = None
        if import_created_ids or import_updated_records:
            self._last_import_undo_data = {
                'created_ids': import_created_ids,
                'updated_records': list(import_updated_records.values())
            }
        
        # 创建详细的导入结果对话框
        result_msg = f"导入完成！\n\n"
        result_msg += f"📊 导入统计：\n"
        result_msg += f"• 文件总数据：{total_rows} 条\n"
        result_msg += f"• 成功导入：{success_count} 条\n"
        result_msg += f"• 覆盖重复：{overwrite_count} 条\n"
        result_msg += f"• 跳过重复：{skip_count} 条\n"
        result_msg += f"• 状态过滤跳过：{status_filtered_skip_count} 条\n"
        result_msg += f"• 导入失败：{fail_count} 条\n\n"
        
        if duplicate_count > 0:
            result_msg += f"⚠️ 发现重复订单：{duplicate_count} 条\n"
        
        if fail_count > 0:
            result_msg += f"❌ 失败原因：数据格式错误或必填字段缺失\n"
        
        if success_count + overwrite_count > 0:
            result_msg += f"✅ 成功处理：{success_count + overwrite_count} 条数据已保存\n"
            result_msg += "↩️ 可按 Ctrl+Z 撤销最近一次导入"
        
        # 显示详细结果对话框
        QMessageBox.information(self, "导入结果", result_msg)
        
        # 同时显示气泡提示
        if success_count == 0 and overwrite_count == 0 and skip_count == 0 and fail_count == 0:
            self.show_tooltip("没有导入数据", "rgba(255, 193, 7, 0.95)", 1500)  # 黄色气泡显示1.5秒
        else:
            self.show_tooltip(f"导入完成 {success_count + overwrite_count}条", "rgba(76, 175, 80, 0.95)", 1500)  # 绿色气泡显示1.5秒
        
        # 强制清除所有缓存，确保数据完全刷新
        self._cached_records = None
        self._last_search_params = None
        self.load_table_data(force_reload=True)
        
        # 强制刷新表格显示并处理所有挂起的事件
        self.table.viewport().update()
        QApplication.processEvents()  # 处理所有挂起的事件，确保界面完全更新
        
        # 检查导入的记录是否显示
        displayed_count = self.table.rowCount()
        imported_count = success_count + overwrite_count
        
        # 如果导入的记录没有显示，自动切换到本次导入的日期范围
        if imported_count > 0 and displayed_count == 0:
            valid_import_dates = sorted(date for date in imported_record_dates if date)
            if valid_import_dates:
                blockers = self._create_search_signal_blockers()
                self.start_date_edit.setDate(QDate.fromString(valid_import_dates[0], "yyyy-MM-dd"))
                self.end_date_edit.setDate(QDate.fromString(valid_import_dates[-1], "yyyy-MM-dd"))
                del blockers
                self.load_table_data(force_reload=True)
                displayed_count = self.table.rowCount()
            QMessageBox.information(self, "导入提示", 
                                  f"✅ 成功导入 {imported_count} 条记录！\n"
                                  f"已自动切换到本次导入的日期范围。\n"
                                  f"当前显示 {displayed_count} 条记录。")
        elif imported_count > 0:
            QMessageBox.information(self, "导入成功", 
                                  f"✅ 成功导入 {imported_count} 条记录！\n"
                                  f"当前显示 {displayed_count} 条记录。")
        
        # 设置一个定时器，在用户点击表格后清除高亮（在on_item_clicked中处理）

    def _check_store_exists(self, store_name):
        """检查店铺名称是否存在"""
        try:
            if not self.db or not self.db.conn:
                return False
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM stores WHERE store_name = ?', (store_name,))
            result = cursor.fetchone()
            if result and isinstance(result, (tuple, list)) and len(result) > 0:
                return result[0] > 0
            return False
        except:
            return False

    def _check_reason_exists(self, reason):
        """检查退款原因是否存在"""
        try:
            if not self.db or not self.db.conn:
                return False
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM refund_records WHERE reason = ?', (reason,))
            result = cursor.fetchone()
            if result and isinstance(result, (tuple, list)) and len(result) > 0:
                return result[0] > 0
            return False
        except:
            return False

    def check_data_consistency(self):
        """检查数据库和本地表格数据一致性（比较总记录数）"""
        try:
            # 确保数据库连接正常
            if not self.db or not self.db.conn:
                QMessageBox.critical(self, "数据库错误", "数据库连接异常，请重启程序")
                return
            
            # 获取数据库总记录数（所有记录，不考虑筛选条件）
            total_db_count = self.db.get_total_record_count()
            
            # 获取本地表格显示的总行数（当前显示的所有记录）
            local_count = self.table.rowCount() if hasattr(self, 'table') else 0
            
            # 获取本地表格所有记录数（不管筛选不筛选，所有存在的记录）
            all_local_records = self.db.get_all_records() if hasattr(self, 'db') else []
            all_local_count = len(all_local_records)
            
            # 显示核对结果（简化显示，只显示总条数）
            result_msg = f"📊 数据核对结果\n\n"
            result_msg += f"• 数据库总记录数：{total_db_count} 条\n"
            result_msg += f"• 当前显示的条数：{local_count} 条\n"
            result_msg += f"• 本地表格所有记录数：{all_local_count} 条\n\n"
            
            if total_db_count == local_count:
                result_msg += "✅ 数据一致！数据库和本地表格记录数匹配。"
                QMessageBox.information(self, "数据核对", result_msg)
            else:
                result_msg += f"⚠️ 数据不一致！相差 {abs(total_db_count - local_count)} 条记录。\n\n"
                
                if total_db_count > local_count:
                    result_msg += f"数据库中有 {total_db_count - local_count} 条记录未在本地显示。\n"
                    result_msg += "可能原因：数据缓存问题或筛选条件导致记录被隐藏。"
                else:
                    result_msg += f"本地表格显示 {local_count - total_db_count} 条记录在数据库中不存在。\n"
                    result_msg += "可能原因：数据未保存或数据库连接问题。"
                
                # 提供同步选项
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("数据不一致")
                msg_box.setIcon(QMessageBox.Warning)
                msg_box.setText(result_msg)
                
                # 添加自定义按钮（支持换行的大按钮）
                sync_btn = msg_box.addButton("同步到本地表格\n（清除所有筛选）", QMessageBox.YesRole)
                sync_btn.setMinimumWidth(180)  # 设置按钮最小宽度
                
                force_sync_btn = msg_box.addButton("强制全局同步\n（清理所有不一致）", QMessageBox.ActionRole)
                force_sync_btn.setMinimumWidth(180)
                
                cleanup_btn = msg_box.addButton("清理数据库\n孤儿记录", QMessageBox.ActionRole)
                cleanup_btn.setMinimumWidth(180)
                
                refresh_btn = msg_box.addButton("刷新表格", QMessageBox.NoRole)
                refresh_btn.setMinimumWidth(120)
                
                cancel_btn = msg_box.addButton("取消", QMessageBox.RejectRole)
                cancel_btn.setMinimumWidth(120)
                
                msg_box.setDefaultButton(sync_btn)
                msg_box.exec_()
                
                clicked_button = msg_box.clickedButton()
                
                if clicked_button == sync_btn:
                    # 强制重新加载表格数据，清除所有筛选条件
                    if hasattr(self, '_cached_records'):
                        self._cached_records = None  # 清除缓存
                    if hasattr(self, '_last_search_params'):
                        self._last_search_params = None  # 清除搜索参数缓存
                    
                    # 清除所有筛选条件
                    self.search_order_edit.clear()
                    if hasattr(self, 'search_reason_dropdown'):
                        self.search_reason_dropdown.clear_selection()
                    self.search_cancel_combo.setCurrentText('全部')
                    self.search_reject_combo.setCurrentText('全部')
                    self.search_store_combo.setCurrentText('全部')
                    
                    # 强制重新加载所有数据（从数据库下载到本地）
                    if hasattr(self, 'load_table_data'):
                        self.load_table_data(force_reload=True)
                    
                    # 重新检查一致性
                    new_local_count = self.table.rowCount() if hasattr(self, 'table') else 0
                    new_all_local_records = self.db.get_all_records() if hasattr(self, 'db') else []
                    new_all_local_count = len(new_all_local_records)
                    if new_local_count == total_db_count:
                        QMessageBox.information(self, "同步成功", 
                                               f"✅ 数据同步完成！\n\n"
                                               f"数据库数据已下载到本地表格。\n"
                                               f"当前显示的条数：{new_local_count} 条\n"
                                               f"本地表格所有记录数：{new_all_local_count} 条\n"
                                               f"与数据库一致。")
                    else:
                        # 如果仍然不一致，显示调试信息
                        debug_records = self.db.debug_database_records()
                        debug_info = f"数据库中有 {len(debug_records)} 条记录：\n"
                        for record in debug_records:
                            debug_info += f"ID:{record['id']} 订单:{record['order_no']} 店铺:{record['store_name']}\n"
                        
                        QMessageBox.warning(self, "同步失败", 
                                           f"同步后仍然不一致。\n"
                                           f"数据库：{total_db_count}条，当前显示：{new_local_count}条\n\n"
                                           f"调试信息：\n{debug_info}")
                elif clicked_button == force_sync_btn:
                    # 强制全局同步：彻底清理所有不一致数据
                    sync_result = self.db.force_global_sync()
                    
                    # 清除所有筛选条件并刷新表格
                    if hasattr(self, '_cached_records'):
                        self._cached_records = None
                    if hasattr(self, '_last_search_params'):
                        self._last_search_params = None
                    
                    self.search_order_edit.clear()
                    if hasattr(self, 'search_reason_dropdown'):
                        self.search_reason_dropdown.clear_selection()
                    self.search_cancel_combo.setCurrentText('全部')
                    self.search_reject_combo.setCurrentText('全部')
                    self.search_store_combo.setCurrentText('全部')
                    
                    if hasattr(self, 'load_table_data'):
                        self.load_table_data()
                    
                    # 重新检查一致性
                    new_total_db_count = self.db.get_total_record_count()
                    new_local_count = self.table.rowCount() if hasattr(self, 'table') else 0
                    new_all_local_records = self.db.get_all_records() if hasattr(self, 'db') else []
                    new_all_local_count = len(new_all_local_records)
                    
                    if sync_result['total_cleaned'] > 0:
                        QMessageBox.information(self, "强制同步完成", 
                                               f"✅ 强制全局同步完成！\n\n"
                                               f"清理统计：\n"
                                               f"• 孤儿记录：{sync_result['orphan_count']} 条\n"
                                               f"• 重复记录：{sync_result['duplicate_count']} 条\n"
                                               f"• 无效数据：{sync_result['invalid_count']} 条\n"
                                               f"• 总计清理：{sync_result['total_cleaned']} 条\n\n"
                                               f"同步后：\n"
                                               f"• 数据库总记录数：{new_total_db_count} 条\n"
                                               f"• 当前显示的条数：{new_local_count} 条\n"
                                               f"• 本地表格所有记录数：{new_all_local_count} 条")
                    else:
                        QMessageBox.information(self, "无需清理", "数据库中没有发现不一致数据。")
                elif clicked_button == cleanup_btn:
                    # 清理数据库孤儿记录
                    deleted_count = self.db.cleanup_orphan_records()
                    
                    # 重新检查一致性
                    new_total_db_count = self.db.get_total_record_count()
                    new_local_count = self.table.rowCount() if hasattr(self, 'table') else 0
                    new_all_local_records = self.db.get_all_records() if hasattr(self, 'db') else []
                    new_all_local_count = len(new_all_local_records)
                    
                    if deleted_count > 0:
                        QMessageBox.information(self, "清理完成", 
                                               f"成功清理 {deleted_count} 条孤儿记录！\n\n"
                                               f"清理后：\n"
                                               f"• 数据库总记录数：{new_total_db_count} 条\n"
                                               f"• 当前显示的条数：{new_local_count} 条\n"
                                               f"• 本地表格所有记录数：{new_all_local_count} 条")
                    else:
                        QMessageBox.information(self, "无需清理", "数据库中没有发现孤儿记录。")
                elif clicked_button == refresh_btn:
                    # 简单刷新表格
                    if hasattr(self, 'load_table_data'):
                        self.load_table_data()
                    QMessageBox.information(self, "刷新完成", "表格已刷新")
        
        except Exception as e:
            # 更详细的错误信息
            import traceback
            error_details = traceback.format_exc()
            QMessageBox.critical(self, "核对错误", 
                               f"数据核对过程中发生错误：{str(e)}\n\n错误详情：\n{error_details}")

    def show_theme_settings(self):
        """显示主题设置对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("主题设置")
        dialog.setFixedSize(400, 400)
        
        layout = QVBoxLayout(dialog)
        
        # 字体颜色设置
        layout.addWidget(QLabel("字体颜色设置："))
        
        # 字体颜色预览
        font_color_layout = QHBoxLayout()
        font_color_layout.addWidget(QLabel("当前字体颜色："))
        self.font_color_preview = QLabel("示例文本")
        self.font_color_preview.setFixedSize(80, 30)
        self.font_color_preview.setStyleSheet("color: black; border: 1px solid black; padding: 5px;")
        font_color_layout.addWidget(self.font_color_preview)
        
        # 字体颜色选择按钮
        self.font_color_picker_btn = QPushButton("选择字体颜色")
        self.font_color_picker_btn.clicked.connect(self.pick_font_color)
        font_color_layout.addWidget(self.font_color_picker_btn)
        
        layout.addLayout(font_color_layout)
        
        # 选中行颜色设置
        layout.addWidget(QLabel("\n选中行颜色设置："))
        
        # 当前颜色预览
        color_layout = QHBoxLayout()
        color_layout.addWidget(QLabel("当前选中行颜色："))
        self.color_preview = QLabel()
        self.color_preview.setFixedSize(50, 30)
        self.color_preview.setStyleSheet("background-color: #87CEEB; border: 1px solid black;")
        color_layout.addWidget(self.color_preview)
        
        # 颜色选择按钮
        self.color_picker_btn = QPushButton("选择颜色")
        self.color_picker_btn.clicked.connect(self.pick_selection_color)
        color_layout.addWidget(self.color_picker_btn)
        
        layout.addLayout(color_layout)
        
        # 店铺颜色管理
        layout.addWidget(QLabel("\n店铺颜色管理："))
        
        # 店铺颜色列表
        self.store_color_list = QListWidget()
        layout.addWidget(self.store_color_list)
        
        # 店铺颜色操作按钮
        store_color_layout = QHBoxLayout()
        self.set_store_color_btn = QPushButton("设置店铺颜色")
        self.set_store_color_btn.clicked.connect(self.set_store_color)
        store_color_layout.addWidget(self.set_store_color_btn)
        
        self.clear_store_color_btn = QPushButton("清除店铺颜色")
        self.clear_store_color_btn.clicked.connect(self.clear_store_color)
        store_color_layout.addWidget(self.clear_store_color_btn)
        
        layout.addLayout(store_color_layout)
        
        # 确定按钮
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(dialog.accept)
        layout.addWidget(ok_btn)
        
        # 加载店铺颜色列表
        self.load_store_colors()
        
        dialog.exec_()

    def pick_font_color(self):
        """选择字体颜色"""
        color = QColorDialog.getColor(QColor("black"), self, "选择字体颜色")
        if color.isValid():
            # 更新字体颜色预览
            self.font_color_preview.setStyleSheet(f"color: {color.name()}; border: 1px solid black; padding: 5px;")
            # 表格样式已由table_panel.ui文件控制，不再动态设置

    def pick_selection_color(self):
        """选择选中行颜色"""
        color = QColorDialog.getColor(QColor("#87CEEB"), self, "选择选中行颜色")
        if color.isValid():
            # 更新预览
            self.color_preview.setStyleSheet(f"background-color: {color.name()}; border: 1px solid black;")
            # 表格样式已由table_panel.ui文件控制，不再动态设置

    def load_store_colors(self):
        """加载店铺颜色列表（显示店铺颜色）"""
        self.store_color_list.clear()
        stores = self.db.get_stores()
        for store_id, store_name in stores:
            # 获取店铺颜色
            store_color = self.db.get_store_color(store_name)
            
            item = QListWidgetItem(f"{store_name}")
            
            # 如果店铺有设置颜色，显示对应的背景色
            if store_color:
                item.setBackground(QColor(store_color))
                # 根据背景色深浅调整文字颜色，确保可读性
                color = QColor(store_color)
                if color.lightness() > 128:  # 浅色背景用黑色文字
                    item.setForeground(QColor("black"))
                else:  # 深色背景用白色文字
                    item.setForeground(QColor("white"))
            
            self.store_color_list.addItem(item)

    def set_store_color(self):
        """设置店铺颜色"""
        current_item = self.store_color_list.currentItem()
        if current_item:
            store_name = current_item.text()
            color = QColorDialog.getColor(QColor("#FFFFFF"), self, f"选择 {store_name} 的颜色")
            if color.isValid():
                # 保存店铺颜色到数据库
                if self.db.set_store_color(store_name, color.name()):
                    QMessageBox.information(self, "提示", f"已为店铺 {store_name} 设置颜色：{color.name()}")
                    # 刷新表格以显示新颜色
                    self.load_table_data()
                    # 刷新店铺列表以显示新颜色
                    self.load_store_colors()

    def clear_store_color(self):
        """清除店铺颜色"""
        current_item = self.store_color_list.currentItem()
        if current_item:
            store_name = current_item.text()
            # 清除店铺颜色
            if self.db.clear_store_color(store_name):
                QMessageBox.information(self, "提示", f"已清除店铺 {store_name} 的颜色设置")
                # 刷新表格以清除颜色
                self.load_table_data()
                # 刷新店铺列表以清除颜色显示
                self.load_store_colors()






                self.db.update_record(record['id'], record['store_id'], record['order_no'], 
                                     record['reason'], record['refund_amount'], 
                                     record['cancel'], record['compensate'], record['comp_amount'],
                                     new_value == "是", reject_result, record['notes'], 
                                     record['record_date'],
                                     record.get('order_status', ''), record.get('after_sale_status', ''),
                                     record.get('spec_name', ''), record.get('spec_code', ''))
        
        # 使用activated信号而不是currentTextChanged，避免频繁触发
        combo.activated.connect(lambda index: on_selection_changed(combo.itemText(index)))

    def show_reject_result_dropdown(self, row, column):
        """显示驳回结果列下拉框选择"""
        # 创建下拉框
        combo = QComboBox()
        combo.addItems(["驳回成功", "驳回失败"])
        
        # 设置当前值
        current_text = self.table.item(row, column).text()
        current_index = combo.findText(current_text)
        if current_index >= 0:
            combo.setCurrentIndex(current_index)
        
        # 显示下拉框
        self.table.setCellWidget(row, column, combo)
        combo.showPopup()
        
        # 为下拉框安装事件过滤器，处理点击空白处关闭
        combo.installEventFilter(self)
        
        # 当下拉框选择改变时更新数据
        def on_selection_changed(new_value):
            self.table.removeCellWidget(row, column)
            self.table.item(row, column).setText(new_value)
            # 保持当前行的选中状态，不清除焦点和选中
            # 强制刷新表格，确保样式更新
            self.table.viewport().update()
            
            # 更新数据库
            order_no = self.table.item(row, 1).text()
            record = self.db.get_record_by_order_no(order_no)
            if record:
                self.db.update_record(record['id'], record['store_id'], record['order_no'], 
                                     record['reason'], record['refund_amount'], 
                                     record['cancel'], record['compensate'], record['comp_amount'],
                                     record['reject'], new_value, record['notes'], 
                                     record['record_date'],
                                     record.get('order_status', ''), record.get('after_sale_status', ''),
                                     record.get('spec_name', ''), record.get('spec_code', ''))
        
        # 使用activated信号而不是currentTextChanged，避免频繁触发
        combo.activated.connect(lambda index: on_selection_changed(combo.itemText(index)))

    def eventFilter(self, obj, event):
        """事件过滤器：处理下拉框点击空白处关闭"""
        if isinstance(obj, QComboBox):
            if event.type() == event.MouseButtonPress:
                # 获取全局鼠标位置
                global_pos = event.globalPos()
                # 获取下拉框的全局位置
                combo_global_rect = QRect(obj.mapToGlobal(QPoint(0, 0)), obj.size())
                
                # 检查点击是否在下拉框外部
                if not combo_global_rect.contains(global_pos):
                    # 直接移除当前下拉框
                    for row in range(self.table.rowCount()):
                        for col in range(self.table.columnCount()):
                            if self.table.cellWidget(row, col) == obj:
                                self.table.removeCellWidget(row, col)
                                # 保持当前行的选中状态，不清除焦点和选中
                                # 强制刷新表格，确保样式更新
                                self.table.viewport().update()
                                return True
        return super().eventFilter(obj, event)

    def copy_order_no(self, row):
        """复制订单号到剪贴板并显示提示气泡"""
        try:
            # 获取订单号
            order_item = self.table.item(row, 1)  # 第1列是订单号
            if order_item:
                order_no = order_item.text()
                
                # 复制到剪贴板
                clipboard = QApplication.clipboard()
                clipboard.setText(order_no)
                
                # 显示提示气泡
                self.show_copy_tooltip(order_no)
                
        except Exception as e:
            QMessageBox.warning(self, "复制失败", f"复制订单号失败：{str(e)}")
    
    def show_refresh_tooltip(self):
        """显示刷新成功的丝滑气泡提示"""
        # 创建提示标签
        tooltip = QLabel("已刷新", self)
        tooltip.setStyleSheet("""
            QLabel {
                background-color: rgba(0, 120, 212, 0.95);
                color: white;
                padding: 10px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid rgba(255, 255, 255, 0.3);
            }
        """)
        tooltip.setAlignment(Qt.AlignCenter)
        tooltip.adjustSize()
        
        # 设置位置（在窗口底部中间显示）
        window_width = self.width()
        window_height = self.height()
        tooltip_x = (window_width - tooltip.width()) // 2
        tooltip_y = window_height - tooltip.height() - 50  # 距离底部50像素
        tooltip.move(tooltip_x, tooltip_y)
        
        # 显示提示
        tooltip.show()
        tooltip.raise_()  # 确保在最上层
        
        # 设置淡入淡出动画
        tooltip.setWindowOpacity(0.0)
        
        # 淡入动画（更快更丝滑）
        fade_in = QTimer(self)
        fade_in.setSingleShot(True)
        fade_in.timeout.connect(lambda: self.safe_fade_tooltip(tooltip, 0.0, 1.0, 200))
        fade_in.start(10)
        
        # 0.8秒后淡出并销毁（更短的显示时间）
        fade_out = QTimer(self)
        fade_out.setSingleShot(True)
        fade_out.timeout.connect(lambda: self.safe_fade_tooltip(tooltip, 1.0, 0.0, 200, True))
        fade_out.start(810)

    def show_tooltip(self, message, color="rgba(0, 120, 212, 0.95)", duration=800):
        """显示通用的淡入淡出气泡提示"""
        # 创建提示标签
        tooltip = QLabel(message, self)
        tooltip.setStyleSheet(f"""
            QLabel {{
                background-color: {color};
                color: white;
                padding: 10px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid rgba(255, 255, 255, 0.3);
            }}
        """)
        tooltip.setAlignment(Qt.AlignCenter)
        tooltip.adjustSize()
        
        # 设置位置（在窗口底部中间显示）
        window_width = self.width()
        window_height = self.height()
        tooltip_x = (window_width - tooltip.width()) // 2
        tooltip_y = window_height - tooltip.height() - 50  # 距离底部50像素
        tooltip.move(tooltip_x, tooltip_y)
        
        # 显示提示
        tooltip.show()
        tooltip.raise_()  # 确保在最上层
        
        # 设置淡入淡出动画
        tooltip.setWindowOpacity(0.0)
        
        # 淡入动画（更快更丝滑）
        fade_in = QTimer(self)
        fade_in.setSingleShot(True)
        fade_in.timeout.connect(lambda: self.safe_fade_tooltip(tooltip, 0.0, 1.0, 200))
        fade_in.start(10)
        
        # 指定时间后淡出并销毁
        fade_out = QTimer(self)
        fade_out.setSingleShot(True)
        fade_out.timeout.connect(lambda: self.safe_fade_tooltip(tooltip, 1.0, 0.0, 200, True))
        fade_out.start(duration + 10)

    def show_copy_tooltip(self, order_no):
        """显示复制成功的提示气泡"""
        # 创建提示标签
        tooltip = QLabel("已复制", self)
        tooltip.setStyleSheet("""
            QLabel {
                background-color: rgba(76, 175, 80, 0.95);
                color: white;
                padding: 12px 20px;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
            }
        """)
        tooltip.setAlignment(Qt.AlignCenter)
        tooltip.adjustSize()
        
        # 设置位置（在软件窗口中下部分居中显示）
        window_width = self.width()
        window_height = self.height()
        tooltip_x = (window_width - tooltip.width()) // 2
        tooltip_y = window_height - tooltip.height() - 50  # 距离底部50像素
        tooltip.move(tooltip_x, tooltip_y)
        
        # 显示提示
        tooltip.show()
        
        # 设置淡入淡出动画
        tooltip.setWindowOpacity(0.0)
        
        # 淡入动画
        fade_in = QTimer(self)
        fade_in.setSingleShot(True)
        fade_in.timeout.connect(lambda: self.fade_tooltip(tooltip, 0.0, 1.0, 300))
        fade_in.start(10)
        
        # 1秒后淡出并销毁
        fade_out = QTimer(self)
        fade_out.setSingleShot(True)
        fade_out.timeout.connect(lambda: self.fade_tooltip(tooltip, 1.0, 0.0, 300, True))
        fade_out.start(1100)
    
    def safe_fade_tooltip(self, tooltip, start_opacity, end_opacity, duration, destroy=False):
        """安全的淡入淡出动画效果，检查对象是否仍然存在"""
        # 检查tooltip对象是否仍然存在
        try:
            # 尝试访问tooltip的属性，如果对象已被删除会抛出异常
            if not tooltip or not hasattr(tooltip, 'setWindowOpacity'):
                return
        except RuntimeError:
            # 对象已被删除，直接返回
            return
        
        # 创建定时器实现动画效果
        timer = QTimer(self)
        timer.setInterval(16)  # 约60fps
        
        start_time = datetime.now()
        
        def update_opacity():
            # 每次更新前都检查对象是否仍然存在
            try:
                if not tooltip or not hasattr(tooltip, 'setWindowOpacity'):
                    timer.stop()
                    return
            except RuntimeError:
                timer.stop()
                return
            
            current_time = datetime.now()
            elapsed = (current_time - start_time).total_seconds() * 1000
            
            if elapsed >= duration:
                try:
                    tooltip.setWindowOpacity(end_opacity)
                    timer.stop()
                    if destroy:
                        tooltip.deleteLater()
                except RuntimeError:
                    # 对象已被删除，直接停止定时器
                    timer.stop()
                return
            
            # 计算当前透明度
            progress = elapsed / duration
            current_opacity = start_opacity + (end_opacity - start_opacity) * progress
            
            try:
                tooltip.setWindowOpacity(current_opacity)
            except RuntimeError:
                # 对象已被删除，直接停止定时器
                timer.stop()
        
        timer.timeout.connect(update_opacity)
        timer.start()

    def fade_tooltip(self, tooltip, start_opacity, end_opacity, duration, destroy=False):
        """淡入淡出动画效果"""
        # 创建定时器实现动画效果
        timer = QTimer(self)
        timer.setInterval(16)  # 约60fps
        
        start_time = datetime.now()
        
        def update_opacity():
            current_time = datetime.now()
            elapsed = (current_time - start_time).total_seconds() * 1000
            
            if elapsed >= duration:
                tooltip.setWindowOpacity(end_opacity)
                timer.stop()
                if destroy:
                    tooltip.deleteLater()
                return
            
            # 计算当前透明度
            progress = elapsed / duration
            current_opacity = start_opacity + (end_opacity - start_opacity) * progress
            tooltip.setWindowOpacity(current_opacity)
        
        timer.timeout.connect(update_opacity)
        timer.start()

    def get_record_id_by_order_no(self, order_no):
        """根据订单号获取记录ID"""
        record = self.db.get_record_by_order_no(order_no)
        return record['id'] if record else None

    def collect_analysis_data(self):
        """收集当前筛选条件下的数据用于AI分析（使用搜索筛选板块的数据）"""
        records = self.get_summary_source_records()
        summary_snapshot = self._build_local_summary_snapshot(records) if records else {"stores": [], "totals": None}
        
        # 调试信息：显示当前筛选条件
        print(f"[DEBUG 数据收集] 当前筛选条件:")
        print(f"[DEBUG 数据收集] - 店铺: {self.search_store_combo.currentText()}")
        print(f"[DEBUG 数据收集] - 日期: {self.start_date_edit.date().toString('yyyy-MM-dd')} 到 {self.end_date_edit.date().toString('yyyy-MM-dd')}")
        print(f"[DEBUG 数据收集] - 退款原因: {list(self.selected_reasons)}")
        print(f"[DEBUG 数据收集] - 订单号筛选: {self.search_order_edit.text()}")
        print(f"[DEBUG 数据收集] - 表格显示行数: {self.table.rowCount()}")
        print(f"[DEBUG 数据收集] - 数据库返回记录数: {len(records)}")
        
        # 构建分析数据（优化格式，按店铺分类）
        analysis_data = {
            "analysis_period": {
                "start_date": self.start_date_edit.date().toString("yyyy-MM-dd"),
                "end_date": self.end_date_edit.date().toString("yyyy-MM-dd")
            },
            "store_settings": {
                "current_store": self.search_store_combo.currentText(),
                "filter_summary": self.get_current_filter_summary_text()
            },
            "refund_stats": summary_snapshot.get("totals") or {},
            "orders_by_store": {}
        }
        
        # 按店铺分类组织订单数据（使用搜索筛选板块的店铺选择）
        current_store = self.search_store_combo.currentText()
        
        # 自动检测当前筛选条件下的店铺分布
        store_names = set()
        for record in records:
            store_name = record.get("store_name", "未知店铺")
            store_names.add(store_name)
        
        print(f"[DEBUG] 当前筛选条件下检测到 {len(store_names)} 个店铺: {list(store_names)}")
        
        # 按店铺分类组织订单数据
        print(f"[DEBUG 数据分类] 开始处理 {len(records)} 条记录")
        
        for i, record in enumerate(records):
            store_name = record.get("store_name", "未知店铺")
            
            # 调试每个记录的店铺信息
            if i < 5:  # 只显示前5条记录的详细信息
                print(f"[DEBUG 数据分类] 记录 {i+1}: 店铺='{store_name}', 订单号='{record.get('order_no', '')}', 原因='{record.get('reason', '')}'")
            
            if store_name not in analysis_data["orders_by_store"]:
                analysis_data["orders_by_store"][store_name] = []
                print(f"[DEBUG 数据分类] 发现新店铺: {store_name}")
            
            # 发送完整的退款相关信息（除订单号外）
            order_data = {
                "reason": record.get("reason", ""),
                "refund_amount": float(record.get("refund_amount", 0)),
                "cancel": "是" if record.get("cancel", 0) else "否",
                "compensate": "是" if record.get("compensate", 0) else "否",
                "comp_amount": float(record.get("comp_amount", 0)),
                "reject": "是" if record.get("reject", 0) else "否",
                "reject_result": record.get("reject_result", "无"),
                "notes": record.get("notes", "")  # 备注最重要，包含产品型号和问题描述
            }
            
            analysis_data["orders_by_store"][store_name].append(order_data)
        
        # 检查最终分类结果
        print(f"[DEBUG 数据分类] 最终分类结果: {len(analysis_data['orders_by_store'])} 个店铺")
        for store_name, orders in analysis_data["orders_by_store"].items():
            print(f"[DEBUG 数据分类] 店铺 '{store_name}': {len(orders)} 条订单")
        
        # 统计信息
        total_orders = sum(len(orders) for orders in analysis_data["orders_by_store"].values())
        analysis_data["total_orders_count"] = total_orders
        analysis_data["store_count"] = len(analysis_data["orders_by_store"])
        
        print(f"[DEBUG] 数据收集完成: {total_orders} 个订单, {len(analysis_data['orders_by_store'])} 个店铺")
        
        return analysis_data

    def get_current_filtered_records(self):
        """获取当前筛选条件下的订单记录"""
        return self.get_summary_source_records()

    def get_current_store_stats(self):
        """获取当前店铺的统计信息"""
        current_store = self.search_store_combo.currentText()
        
        # 获取退款原因筛选
        selected_reasons = list(self.selected_reasons)
        
        if current_store == "全部":
            # 获取所有店铺的汇总统计
            return self.get_all_stores_stats(selected_reasons)
        else:
            # 获取当前店铺的统计
            return self.get_single_store_stats(current_store, selected_reasons)

    def get_single_store_stats(self, store_name, selected_reasons=None):
        """获取单个店铺的统计信息"""
        store_id = self.db.get_store_id_by_name(store_name)
        if not store_id:
            return {}
        
        # 获取日期范围
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
        
        # 获取店铺设置，并将周录入值换算为日均口径供统计/分析使用
        raw_store_settings = self.db.get_store_settings(store_id) or {}
        store_settings = {
            "current_store": store_name,
            "daily_orders": self._weekly_to_daily_avg(raw_store_settings.get("daily_orders", 0)),
            "daily_sales": self._weekly_to_daily_avg(raw_store_settings.get("daily_sales", 0.0)),
            "refund_budget_remaining": raw_store_settings.get("refund_budget", 0.0),
        }
        
        # 获取退款统计（支持退款原因筛选）
        refund_stats = self.db.get_refund_stats_by_store(
            store_id, start_date, end_date, selected_reasons
        )
        
        return {
            "store_settings": store_settings,
            "refund_stats": refund_stats
        }

    def get_all_stores_stats(self, selected_reasons=None):
        """获取所有店铺的汇总统计信息"""
        # 获取日期范围
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
        
        # 获取所有店铺的汇总统计（支持退款原因筛选）
        refund_stats = self.db.get_refund_stats_all_stores(start_date, end_date, selected_reasons)
        
        # 获取所有店铺的设置并汇总
        stores = self.db.get_stores()
        total_daily_orders = 0.0
        total_daily_sales = 0.0
        total_refund_budget = 0.0
        
        for store_id, store_name in stores:
            store_settings = self.db.get_store_settings(store_id) or {}
            total_daily_orders += self._weekly_to_daily_avg(store_settings.get("daily_orders", 0))
            total_daily_sales += self._weekly_to_daily_avg(store_settings.get("daily_sales", 0.0))
            total_refund_budget += store_settings.get("refund_budget", 0.0)
        
        return {
            "store_settings": {
                "current_store": "全部店铺",
                "daily_orders": total_daily_orders,
                "daily_sales": total_daily_sales,
                "refund_budget_remaining": total_refund_budget
            },
            "refund_stats": refund_stats
        }

    def get_current_filter_context(self):
        """获取当前搜索筛选条件。"""
        reasons = []
        if hasattr(self, 'search_reason_dropdown'):
            reasons = list(self.search_reason_dropdown.selected_items)
        real_reason = ""
        if hasattr(self, 'search_real_reason_edit') and self.search_real_reason_edit is not None:
            real_reason = self.search_real_reason_edit.text().strip()

        return {
            "store": self.search_store_combo.currentText(),
            "order_no": self.search_order_edit.text().strip(),
            "real_reason": real_reason,
            "cancel": self.search_cancel_combo.currentText(),
            "reject": self.search_reject_combo.currentText(),
            "start_date": self.start_date_edit.date().toString("yyyy-MM-dd"),
            "end_date": self.end_date_edit.date().toString("yyyy-MM-dd"),
            "reasons": reasons,
        }

    def get_current_filter_summary_text(self):
        """格式化当前筛选条件摘要。"""
        context = self.get_current_filter_context()
        reasons_text = "全部" if not context["reasons"] else "、".join(context["reasons"])
        real_reason_text = context["real_reason"] or "全部"
        return (
            f"{context['start_date']} 至 {context['end_date']} | "
            f"店铺：{context['store']} | "
            f"撤销：{context['cancel']} | 驳回：{context['reject']} | "
            f"退款原因：{reasons_text} | "
            f"真实退款原因：{real_reason_text}"
        )

    def get_summary_source_records(self):
        """总结分析与表格完全同源的数据入口。"""
        return self.get_filtered_records()

    def _get_summary_days_count(self):
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        return (end_date - start_date).days + 1

    def _get_store_summary_settings(self, store_name):
        """获取单店铺统计所需的设置。"""
        store_id = self.db.get_store_id_by_name(store_name)
        if not store_id:
            return {"daily_orders": 0.0, "daily_sales": 0.0, "refund_budget": 0.0}

        raw_settings = self.db.get_store_settings(store_id) or {}
        return {
            "daily_orders": self._weekly_to_daily_avg(raw_settings.get("daily_orders", 0)),
            "daily_sales": self._weekly_to_daily_avg(raw_settings.get("daily_sales", 0.0)),
            "refund_budget": self._safe_float(raw_settings.get("refund_budget", 0.0)),
        }

    def _build_top_reason_stats(self, records):
        reason_counts = {}
        for record in records:
            reason = self._normalize_reason(record.get("reason")) or "未填写"
            reason_counts[reason] = reason_counts.get(reason, 0) + 1

        if not reason_counts:
            return {"reason": "无数据", "count": 0, "ratio": 0.0}

        top_reason = max(reason_counts, key=reason_counts.get)
        top_count = reason_counts[top_reason]
        ratio = (top_count / len(records) * 100) if records else 0.0
        return {"reason": top_reason, "count": top_count, "ratio": ratio}

    def _build_note_presence_stats(self, records):
        """统计备注填写情况。"""
        note_records = []
        no_note_records = []
        for record in records:
            notes = str(record.get("notes", "") or "").strip()
            order_no = str(record.get("order_no", "") or "").strip()
            item = {
                "order_no": order_no,
                "reason": self._normalize_reason(record.get("reason")),
                "notes": notes,
            }
            if notes:
                note_records.append(item)
            else:
                no_note_records.append(item)

        total = len(records)
        note_count = len(note_records)
        no_note_count = len(no_note_records)
        return {
            "note_count": note_count,
            "no_note_count": no_note_count,
            "note_rate": (note_count / total * 100) if total > 0 else 0.0,
            "no_note_rate": (no_note_count / total * 100) if total > 0 else 0.0,
            "note_order_nos": [item["order_no"] for item in note_records if item["order_no"]],
            "no_note_order_nos": [item["order_no"] for item in no_note_records if item["order_no"]],
        }

    @staticmethod
    def _round_up_to_tens(value):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return 0
        return int(math.ceil(number / 10.0) * 10)

    @classmethod
    def _extract_spec_code_from_name(cls, spec_name):
        text = str(spec_name or "").strip()
        if not text:
            return ""

        length_tokens = [
            match.group(1)
            for match in re.finditer(r'(?<![\d.])(100|\d{1,2}(?:\.\d)?)(?![\d.])', text)
        ]
        length_values = [
            float(token)
            for token in length_tokens
            if 20 <= float(token) <= 100
        ]
        if not length_values:
            return ""
        max_length_value = max(length_values)
        length_code = cls._round_up_to_tens(max_length_value)
        if length_code == 100:
            length_code = 10

        jin_value = cls._extract_spec_jin_value(text)
        if length_code <= 0 or jin_value not in {3, 5, 6, 7, 10}:
            return ""
        return f"{length_code}{jin_value}"

    @staticmethod
    def _extract_spec_jin_value(text):
        jin_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:斤|jin(?![A-Za-z]))', text, re.IGNORECASE)
        if jin_match:
            try:
                return int(math.ceil(float(jin_match.group(1))))
            except (TypeError, ValueError):
                return 0

        chinese_jin_map = {
            "一": 1,
            "壹": 1,
            "三": 3,
            "叁": 3,
            "五": 5,
            "伍": 5,
            "六": 6,
            "陆": 6,
            "七": 7,
            "柒": 7,
            "十": 10,
            "拾": 10,
            "一十": 10,
            "壹拾": 10,
        }
        chinese_match = re.search(r'(壹拾|一十|[一壹三叁五伍六陆七柒十拾])\s*斤', text)
        if not chinese_match:
            return 0
        return chinese_jin_map.get(chinese_match.group(1), 0)

    @classmethod
    def _extract_spec_info(cls, spec_name):
        normalized_name = str(spec_name or "").strip()
        return normalized_name, cls._extract_spec_code_from_name(normalized_name)

    def _auto_fill_missing_spec_codes_from_names(self):
        """启动时自动补齐已录入规格名称但缺少规格编码的历史记录。"""
        try:
            records = self.db.get_records_missing_spec_code_with_name()
            if not records:
                return 0

            updated_count = 0
            for record in records:
                spec_name = str(record.get('spec_name') or '').strip()
                spec_code = self._extract_spec_code_from_name(spec_name)
                if not spec_code:
                    continue
                if self.db.update_record_partial(record.get('id'), spec_code=spec_code):
                    updated_count += 1

            if updated_count:
                self._cached_records = None
                self._last_search_params = None
                print(f"✅ 启动自动识别规格编码 {updated_count} 条")
            return updated_count
        except Exception as e:
            print(f"启动自动识别规格编码失败: {e}")
            return 0

    def reidentify_spec_codes_from_names(self):
        """手动重新识别所有已录入规格名称订单的规格编码。"""
        try:
            records = self.db.get_records_with_spec_name()
            if not records:
                QMessageBox.information(self, "识别规格", "没有找到已录入规格名称的订单")
                return

            updated_count = 0
            recognized_count = 0
            unchanged_count = 0
            unrecognized_count = 0
            for record in records:
                spec_name = str(record.get('spec_name') or '').strip()
                new_spec_code = self._extract_spec_code_from_name(spec_name)
                if not new_spec_code:
                    unrecognized_count += 1
                    continue

                recognized_count += 1
                old_spec_code = str(record.get('spec_code') or '').strip()
                if old_spec_code == new_spec_code:
                    unchanged_count += 1
                    continue

                if self.db.update_record_partial(record.get('id'), spec_code=new_spec_code):
                    updated_count += 1

            if updated_count:
                self._cached_records = None
                self._last_search_params = None
                self.load_table_data(force_reload=True)

            QMessageBox.information(
                self,
                "识别规格",
                "规格编码重新识别完成\n\n"
                f"扫描规格名称：{len(records)} 条\n"
                f"成功识别：{recognized_count} 条\n"
                f"更新编码：{updated_count} 条\n"
                f"无需更新：{unchanged_count} 条\n"
                f"未识别：{unrecognized_count} 条"
            )
        except Exception as e:
            QMessageBox.critical(self, "识别规格失败", f"重新识别规格编码时发生错误：{e}")

    @staticmethod
    def _extract_spec_candidate(notes):
        text = str(notes or "")
        match = re.search(r'(?<!\d)(\d{3,4})(?!\d)', text)
        return match.group(1) if match else ""

    @staticmethod
    def _normalize_note_text(notes):
        text = re.sub(r'\s+', ' ', str(notes or "").strip())
        return text[:200]

    @staticmethod
    def _contains_halfway_return_keyword(notes):
        return "已拦截" in str(notes or "")

    def _aggregate_note_entries(self, entries, include_reason=False):
        """聚合备注，减少AI输入规模。"""
        aggregated = {}
        for entry in entries:
            note_text = self._normalize_note_text(entry.get("notes", ""))
            if not note_text:
                continue
            spec_candidate = str(entry.get("spec_code") or "").strip() or self._extract_spec_candidate(note_text)
            key_parts = [note_text, spec_candidate]
            if include_reason:
                key_parts.append(self._normalize_reason(entry.get("quality_reason", "")))
            key = "||".join(key_parts)
            bucket = aggregated.setdefault(key, {
                "notes": note_text,
                "spec_candidate": spec_candidate,
                "count": 0,
                "quality_reason": self._normalize_reason(entry.get("quality_reason", "")) if include_reason else "",
                "contains_halfway_return": self._contains_halfway_return_keyword(note_text),
            })
            bucket["count"] += 1

        result = list(aggregated.values())
        result.sort(key=lambda item: (-item["count"], item["notes"]))
        return result

    def _normalize_real_reason_note_text(self, notes):
        return re.sub(r'\s+', ' ', str(notes or "").strip())

    def _build_real_reason_note_hash(self, notes):
        normalized = self._normalize_real_reason_note_text(notes)
        return hashlib.md5(normalized.encode('utf-8')).hexdigest() if normalized else ""

    @staticmethod
    def _normalize_real_reason_value(value):
        text = str(value or "").strip()
        if text.lower() in ("none", "null", "nan", "未生成真实退款原因"):
            return ""
        return text

    def _is_real_reason_stale(self, record):
        notes = str(record.get("notes", "") or "").strip()
        if not notes:
            return False
        current_hash = self._build_real_reason_note_hash(notes)
        stored_hash = str(record.get("real_refund_reason_note_hash") or "").strip()
        if not stored_hash:
            return True
        return current_hash != stored_hash

    def _build_real_reason_assignment_diagnostics(self, records):
        total_records = len(records)
        noted_records = 0
        empty_reason_count = 0
        unresolved_count = 0
        stale_count = 0
        valid_assigned_count = 0
        candidate_count = 0

        for record in records:
            notes = str(record.get("notes", "") or "").strip()
            if not notes:
                continue
            noted_records += 1
            real_reason = self._normalize_real_reason_value(record.get("real_refund_reason"))
            is_stale = self._is_real_reason_stale(record)

            if not real_reason:
                empty_reason_count += 1
                candidate_count += 1
                continue

            if real_reason == "未归因":
                unresolved_count += 1
                candidate_count += 1
                continue

            if is_stale:
                stale_count += 1
                candidate_count += 1
                continue

            valid_assigned_count += 1

        return {
            "total_records": total_records,
            "noted_records": noted_records,
            "empty_reason_count": empty_reason_count,
            "unresolved_count": unresolved_count,
            "stale_count": stale_count,
            "valid_assigned_count": valid_assigned_count,
            "candidate_count": candidate_count,
        }

    def _build_real_reason_category_generation_payload(self, records):
        existing_categories = [item.get("category_name", "") for item in self.db.get_real_refund_reason_categories()]
        aggregated = self._aggregate_note_entries(
            [{"notes": record.get("notes", "")} for record in records if str(record.get("notes", "")).strip()],
            include_reason=False
        )
        notes = [{"notes": item.get("notes", ""), "count": item.get("count", 0)} for item in aggregated]
        return {
            "existing_categories": existing_categories,
            "notes": notes,
        }

    def _build_real_reason_assignment_candidates(self, records):
        candidates = []
        for record in records:
            notes = str(record.get("notes", "") or "").strip()
            if not notes:
                continue
            real_reason = self._normalize_real_reason_value(record.get("real_refund_reason"))
            if real_reason and real_reason != "未归因" and not self._is_real_reason_stale(record):
                continue
            candidates.append(record)
        return candidates

    def _build_real_reason_assignment_payload(self, records, category_names):
        grouped = {}
        for record in records:
            note_text = self._normalize_real_reason_note_text(record.get("notes", ""))
            if not note_text:
                continue
            bucket = grouped.setdefault(note_text, {"notes": note_text, "count": 0, "record_ids": []})
            bucket["count"] += 1
            bucket["record_ids"].append(record.get("id"))

        note_items = list(grouped.values())
        note_items.sort(key=lambda item: (-item["count"], item["notes"]))
        payload = {
            "existing_categories": list(category_names or []),
            "notes": [{"notes": item["notes"], "count": item["count"]} for item in note_items],
        }
        return payload, note_items

    def _build_real_reason_analysis(self, records):
        total_count = len(records)
        reason_counts = {}
        reason_examples = {}
        spec_counts = {}
        classified_count = 0

        for record in records:
            notes = self._normalize_real_reason_note_text(record.get("notes", ""))
            category = str(record.get("real_refund_reason") or "").strip()
            if not category:
                continue
            classified_count += 1
            reason_counts[category] = reason_counts.get(category, 0) + 1
            if notes:
                samples = reason_examples.setdefault(category, [])
                if notes not in samples and len(samples) < 2:
                    samples.append(notes)

            spec_key = str(record.get("spec_code") or "").strip() or "-"
            spec_bucket = spec_counts.setdefault(spec_key, {})
            spec_bucket[category] = spec_bucket.get(category, 0) + 1

        overall_categories = []
        for name, count in sorted(reason_counts.items(), key=lambda item: (-item[1], item[0])):
            overall_categories.append({
                "name": name,
                "count": count,
                "ratio": (count / classified_count * 100) if classified_count else 0.0,
                "examples": reason_examples.get(name, []),
            })

        spec_categories = []
        for spec, categories in sorted(spec_counts.items(), key=lambda item: item[0]):
            spec_total = sum(categories.values())
            category_items = []
            for name, count in sorted(categories.items(), key=lambda item: (-item[1], item[0])):
                category_items.append({
                    "name": name,
                    "count": count,
                    "ratio": (count / spec_total * 100) if spec_total else 0.0,
                })
            spec_categories.append({
                "spec": spec,
                "total_count": spec_total,
                "categories": category_items,
            })

        return {
            "classified_count": classified_count,
            "unclassified_count": max(total_count - classified_count, 0),
            "classified_ratio": (classified_count / total_count * 100) if total_count else 0.0,
            "unclassified_ratio": ((total_count - classified_count) / total_count * 100) if total_count else 0.0,
            "overall_categories": overall_categories,
            "spec_categories": spec_categories,
        }

    def _build_quality_saved_reason_analysis(self, records):
        target_records = [record for record in records if self._is_quality_refund_analysis_record(record)]
        classified_count = 0
        quality_counts = {}
        not_cancelled_counts = {}
        quality_examples = {}
        not_cancelled_examples = {}

        for record in target_records:
            quality_reason = str(record.get("quality_refund_reason") or "").strip()
            not_cancelled_reason = str(record.get("quality_not_cancelled_reason") or "").strip()
            notes = self._normalize_real_reason_note_text(record.get("notes", ""))
            if quality_reason or not_cancelled_reason:
                classified_count += 1
            if quality_reason:
                quality_counts[quality_reason] = quality_counts.get(quality_reason, 0) + 1
                if notes:
                    examples = quality_examples.setdefault(quality_reason, [])
                    if notes not in examples and len(examples) < 2:
                        examples.append(notes)
            if not_cancelled_reason:
                not_cancelled_counts[not_cancelled_reason] = not_cancelled_counts.get(not_cancelled_reason, 0) + 1
                if notes:
                    examples = not_cancelled_examples.setdefault(not_cancelled_reason, [])
                    if notes not in examples and len(examples) < 2:
                        examples.append(notes)

        def build_categories(counts, examples_by_name):
            total = sum(counts.values())
            return [
                {
                    "name": name,
                    "count": count,
                    "ratio": (count / total * 100) if total else 0.0,
                    "examples": examples_by_name.get(name, []),
                }
                for name, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
            ]

        total_count = len(target_records)
        return {
            "total_count": total_count,
            "classified_count": classified_count,
            "unclassified_count": max(total_count - classified_count, 0),
            "classified_ratio": (classified_count / total_count * 100) if total_count else 0.0,
            "quality_refund_categories": build_categories(quality_counts, quality_examples),
            "not_cancelled_categories": build_categories(not_cancelled_counts, not_cancelled_examples),
        }

    def _build_store_summary_stats(self, store_name, records):
        """构建单店铺总结统计。"""
        days_count = self._get_summary_days_count()
        settings = self._get_store_summary_settings(store_name)
        orders = settings["daily_orders"] * days_count
        sales = settings["daily_sales"] * days_count
        refund_budget_remaining = settings["refund_budget"] * days_count
        note_presence = self._build_note_presence_stats(records)

        quality_records = [record for record in records if self._is_quality_reason(record.get("reason"))]
        other_records = [record for record in records if not self._is_quality_reason(record.get("reason"))]
        effective_quality_records = [record for record in quality_records if self._is_effective_refund_record(record)]

        quality_after_sales_amount = 0.0
        other_after_sales_amount = 0.0
        effective_refund_amount = 0.0
        compensation_amount = 0.0

        for record in records:
            amount = 0.0
            if self._is_effective_refund_record(record):
                refund_value = self._safe_float(record.get("refund_amount", 0))
                effective_refund_amount += refund_value
                amount += refund_value

            if self._has_compensation_record(record):
                comp_value = self._safe_float(record.get("comp_amount", 0))
                compensation_amount += comp_value
                amount += comp_value

            if self._is_quality_reason(record.get("reason")):
                quality_after_sales_amount += amount
            else:
                other_after_sales_amount += amount

        after_sales_total = effective_refund_amount + compensation_amount
        refund_budget_remaining -= after_sales_total

        top_reason = self._build_top_reason_stats(records)
        quality_apply_count = len(quality_records)
        quality_cancel_count = sum(1 for record in quality_records if record.get("cancel"))
        quality_actual_count = len(effective_quality_records)
        total_refund_count = len(records)

        total_refund_rate = (total_refund_count / orders * 100) if orders > 0 else 0.0
        refund_ratio = (after_sales_total / sales * 100) if sales > 0 else 0.0
        apply_rate = (quality_apply_count / orders * 100) if orders > 0 else 0.0
        actual_rate = (quality_actual_count / orders * 100) if orders > 0 else 0.0
        cancel_rate = (quality_cancel_count / quality_apply_count * 100) if quality_apply_count > 0 else 0.0

        return {
            "store_name": store_name,
            "date_range": {
                "start_date": self.start_date_edit.date().toString("yyyy-MM-dd"),
                "end_date": self.end_date_edit.date().toString("yyyy-MM-dd"),
            },
            "record_count": len(records),
            "orders": orders,
            "sales": sales,
            "refund_budget_remaining": refund_budget_remaining,
            "quality_refund_count": quality_apply_count,
            "other_refund_count": len(other_records),
            "canceled_quality_count": quality_cancel_count,
            "total_refund_rate": total_refund_rate,
            "total_after_sales": after_sales_total,
            "refund_ratio": refund_ratio,
            "quality_after_sales_amount": quality_after_sales_amount,
            "other_after_sales_amount": other_after_sales_amount,
            "quality_apply_rate": apply_rate,
            "quality_actual_rate": actual_rate,
            "quality_cancel_rate": cancel_rate,
            "top_refund_reason": top_reason["reason"],
            "top_reason_count": top_reason["count"],
            "top_reason_ratio": top_reason["ratio"],
            "effective_refund_amount": effective_refund_amount,
            "compensation_amount": compensation_amount,
            "quality_apply_count": quality_apply_count,
            "quality_cancel_count": quality_cancel_count,
            "quality_actual_count": quality_actual_count,
            "note_count": note_presence["note_count"],
            "no_note_count": note_presence["no_note_count"],
            "note_rate": note_presence["note_rate"],
            "no_note_rate": note_presence["no_note_rate"],
        }

    def _build_total_summary_stats(self, store_summaries, all_records):
        """构建多店汇总统计。"""
        if len(store_summaries) <= 1:
            return None

        top_reason = self._build_top_reason_stats(all_records)
        orders = sum(store["orders"] for store in store_summaries)
        sales = sum(store["sales"] for store in store_summaries)
        total_after_sales = sum(store["total_after_sales"] for store in store_summaries)
        quality_apply_count = sum(store["quality_apply_count"] for store in store_summaries)
        quality_cancel_count = sum(store["quality_cancel_count"] for store in store_summaries)
        quality_actual_count = sum(store["quality_actual_count"] for store in store_summaries)
        total_refund_count = len(all_records)
        note_presence = self._build_note_presence_stats(all_records)

        return {
            "store_name": "全部总和",
            "record_count": total_refund_count,
            "orders": orders,
            "sales": sales,
            "refund_budget_remaining": sum(store["refund_budget_remaining"] for store in store_summaries),
            "quality_refund_count": quality_apply_count,
            "other_refund_count": sum(store["other_refund_count"] for store in store_summaries),
            "canceled_quality_count": quality_cancel_count,
            "total_refund_rate": (total_refund_count / orders * 100) if orders > 0 else 0.0,
            "total_after_sales": total_after_sales,
            "refund_ratio": (total_after_sales / sales * 100) if sales > 0 else 0.0,
            "quality_after_sales_amount": sum(store["quality_after_sales_amount"] for store in store_summaries),
            "other_after_sales_amount": sum(store["other_after_sales_amount"] for store in store_summaries),
            "quality_apply_rate": (quality_apply_count / orders * 100) if orders > 0 else 0.0,
            "quality_actual_rate": (quality_actual_count / orders * 100) if orders > 0 else 0.0,
            "quality_cancel_rate": (quality_cancel_count / quality_apply_count * 100) if quality_apply_count > 0 else 0.0,
            "top_refund_reason": top_reason["reason"],
            "top_reason_count": top_reason["count"],
            "top_reason_ratio": top_reason["ratio"],
            "effective_refund_amount": sum(store["effective_refund_amount"] for store in store_summaries),
            "compensation_amount": sum(store["compensation_amount"] for store in store_summaries),
            "quality_apply_count": quality_apply_count,
            "quality_cancel_count": quality_cancel_count,
            "quality_actual_count": quality_actual_count,
            "note_count": note_presence["note_count"],
            "no_note_count": note_presence["no_note_count"],
            "note_rate": note_presence["note_rate"],
            "no_note_rate": note_presence["no_note_rate"],
        }

    def _build_quality_unreversed_notes_payload(self, store_name, records):
        note_records = []
        no_note_count = 0
        for record in records:
            if not self._is_quality_reason(record.get("reason")):
                continue
            if not self._is_effective_refund_record(record):
                continue
            notes = str(record.get("notes", "") or "").strip()
            if not notes:
                no_note_count += 1
                continue
            note_records.append({
                "notes": notes,
                "spec_code": str(record.get("spec_code") or "").strip(),
            })

        aggregated_records = self._aggregate_note_entries(note_records, include_reason=False)

        return {
            "store_name": store_name,
            "record_count": len(note_records),
            "aggregated_record_count": len(aggregated_records),
            "skipped_no_note_count": no_note_count,
            "records": aggregated_records,
        }

    def _build_other_reason_notes_payload(self, store_name, records):
        note_records = []
        no_note_count = 0
        halfway_return_count = 0
        for record in records:
            if self._normalize_reason(record.get("reason")) != "其他":
                continue
            notes = str(record.get("notes", "") or "").strip()
            if not notes:
                no_note_count += 1
                continue
            if self._contains_halfway_return_keyword(notes):
                halfway_return_count += 1
                continue
            note_records.append({
                "notes": notes,
                "spec_code": str(record.get("spec_code") or "").strip(),
            })

        aggregated_records = self._aggregate_note_entries(note_records, include_reason=False)
        total_with_note = len(note_records) + halfway_return_count

        return {
            "store_name": store_name,
            "record_count": total_with_note,
            "ai_record_count": len(note_records),
            "aggregated_record_count": len(aggregated_records),
            "skipped_no_note_count": no_note_count,
            "local_halfway_return_count": halfway_return_count,
            "local_halfway_return_ratio": (halfway_return_count / total_with_note * 100) if total_with_note > 0 else 0.0,
            "records": aggregated_records,
        }

    def _get_default_quality_note_analysis(self, payload, message="未执行备注归类"):
        return {
            "message": message,
            "record_count": payload.get("record_count", 0),
            "quality_problem_categories": [],
            "not_cancelled_reason_categories": [],
        }

    def _get_default_other_reason_analysis(self, payload, message="未执行其他原因备注复盘"):
        total = payload.get("record_count", 0)
        halfway_count = payload.get("local_halfway_return_count", 0)
        return {
            "message": message,
            "record_count": total,
            "overall_categories": [],
            "spec_categories": [],
            "halfway_return_count": halfway_count,
            "halfway_return_ratio": payload.get("local_halfway_return_ratio", 0.0),
            "unclear_count": 0,
            "unclear_ratio": 0.0,
        }

    @staticmethod
    def _normalize_display_category_name(name):
        text = str(name or "").strip()
        if text == "客服处理不及时":
            return "接线客服处理不及时"
        return text

    def _normalize_category_list(self, categories):
        normalized = []
        for category in categories or []:
            normalized.append({
                "name": self._normalize_display_category_name(category.get("name", "")),
                "count": category.get("count", 0),
                "ratio": self._safe_float(category.get("ratio", 0)),
                "examples": category.get("examples", []),
            })
        return normalized

    @staticmethod
    def _get_default_real_reason_category_configs():
        return [
            {"category_name": "腐烂变质", "keywords_text": "长毛 发霉 霉 腐烂 腐坏 烂 烂了 烂掉 烂心 烂头 烂尾 臭了 变质 坏了 坏掉 坏果 变味 有味 臭味 发酸 腐败 软烂 坏 烂果 霉变", "status": "ACTIVE", "sort_order": 0},
            {"category_name": "发芽了", "keywords_text": "发芽 发芽了 长芽 长芽了 冒芽 生芽 芽眼 芽头", "status": "ACTIVE", "sort_order": 1},
            {"category_name": "断裂破损", "keywords_text": "断 裂 压坏 破损 断了 裂开 裂口 折断 破裂 破皮 损坏", "status": "ACTIVE", "sort_order": 2},
            {"category_name": "半路退回", "keywords_text": "已拦截 拦截退回 半路退回 物流退回 中途退回 途中退回 拦截成功 快递退回", "status": "ACTIVE", "sort_order": 3},
            {"category_name": "客户主观不想要", "keywords_text": "不想要 不要了 拍错 买错 不需要 不喜欢 买多 下错单 拍多了", "status": "ACTIVE", "sort_order": 4},
            {"category_name": "外观不满意", "keywords_text": "外观 外形 不好看 太差 畸形 外观差 品相差 卖相差 难看 形状差", "status": "ACTIVE", "sort_order": 5},
            {"category_name": "少称", "keywords_text": "重量不够 少称 缺斤少两 重量少 斤两不够 称重不足 不够重", "status": "ACTIVE", "sort_order": 6},
            {"category_name": "长度不够", "keywords_text": "长度不够 长度 太短 短 小短 不够长 长度短 尺寸短", "status": "ACTIVE", "sort_order": 7},
            {"category_name": "泥巴多", "keywords_text": "土多 泥巴多 泥多 带泥 泥土多 土太多 不干净", "status": "ACTIVE", "sort_order": 8},
            {"category_name": "太细", "keywords_text": "细 太细 削皮完没有了 太小 细小 小根 个头小 规格小", "status": "ACTIVE", "sort_order": 9},
            {"category_name": "斑点", "keywords_text": "黑点 黑芯 黑斑 黑心 黑洞 斑块 坏点", "status": "ACTIVE", "sort_order": 10},
            {"category_name": "疤痕", "keywords_text": "疤痕 伤疤 疤多 划痕 刮伤 表皮伤", "status": "ACTIVE", "sort_order": 11},
            {"category_name": "挤压", "keywords_text": "尾部挤压 头部挤压 挤压 挤压严重 挤压多根 压伤 压烂 压坏 挤坏 挤烂 变形 运输挤压", "status": "ACTIVE", "sort_order": 12},
            {"category_name": "口感问题", "keywords_text": "不好吃 口感差 味道差 太硬 太软 不甜 发苦 发涩", "status": "ACTIVE", "sort_order": 13},
            {"category_name": "太湿", "keywords_text": "湿 水分大 潮湿 湿哒哒 出水 水汽", "status": "ACTIVE", "sort_order": 14},
            {"category_name": "氧化", "keywords_text": "氧化 发黑 变黑 切开黑 表面黑 氧化黑", "status": "ACTIVE", "sort_order": 15},
            {"category_name": "虫害", "keywords_text": "虫 虫眼 虫洞 虫蛀 有虫 被虫咬", "status": "ACTIVE", "sort_order": 16},
            {"category_name": "外观瑕疵", "keywords_text": "瑕疵 残次 破皮 表皮差 坑洼 凹陷", "status": "ACTIVE", "sort_order": 17},
            {"category_name": "客服问题", "keywords_text": "客服 服务 回复慢 处理慢 态度差 售后慢", "status": "ACTIVE", "sort_order": 18},
            {"category_name": "好评", "keywords_text": "好评 好评返现 返现 评价返现 晒图返现", "status": "ACTIVE", "sort_order": 19},
            {"category_name": "地标", "keywords_text": "没有地标 无地标 地标缺失 没贴标 标签缺失", "status": "ACTIVE", "sort_order": 20},
            {"category_name": "品质不符预期", "keywords_text": "品质差 质量差 不新鲜 不满意 不符合 货不对版", "status": "ACTIVE", "sort_order": 21},
            {"category_name": "拦截单", "keywords_text": "拦截 已拦截 拦截单 拦截件 半路退回 拒收退回 物流拦截", "status": "ACTIVE", "sort_order": 22},
            {"category_name": "黑洞", "keywords_text": "黑洞 空心 洞 内部空 里面黑", "status": "ACTIVE", "sort_order": 23},
            {"category_name": "太干", "keywords_text": "干 太干 干瘪 发干 缺水 皱巴", "status": "ACTIVE", "sort_order": 24},
        ]

    @staticmethod
    def _parse_real_reason_keywords(keywords_text):
        normalized = re.sub(r'[、,，;；\s]+', ' ', str(keywords_text or ""))
        return [item.strip() for item in normalized.split(" ") if item.strip()]

    @classmethod
    def _append_real_reason_keywords(cls, existing_text, keywords_to_add):
        keywords = cls._parse_real_reason_keywords(existing_text)
        seen = set(keywords)
        for keyword in keywords_to_add or []:
            keyword_text = str(keyword or "").strip()
            if keyword_text and keyword_text not in seen:
                keywords.append(keyword_text)
                seen.add(keyword_text)
        return " ".join(keywords)

    def _merge_default_real_reason_keywords(self):
        """只给已有分类追加默认关键词，不覆盖用户手工维护的词。"""
        defaults = {
            item["category_name"]: self._parse_real_reason_keywords(item.get("keywords_text", ""))
            for item in self._get_default_real_reason_category_configs()
        }
        categories = self.db.get_real_refund_reason_categories(active_only=False)
        if not categories:
            self.db.ensure_default_real_refund_reason_categories(self._get_default_real_reason_category_configs())
            return

        merged = []
        changed = False
        for item in categories:
            category_name = str(item.get("category_name") or "").strip()
            keywords_text = str(item.get("keywords_text") or "")
            merged_keywords_text = self._append_real_reason_keywords(
                keywords_text,
                defaults.get(category_name, [])
            )
            if merged_keywords_text != keywords_text:
                changed = True

            merged.append({
                "category_name": category_name,
                "keywords_text": merged_keywords_text,
                "status": item.get("status", "ACTIVE"),
                "sort_order": int(item.get("sort_order", 0) or 0),
            })

        if changed:
            self.db.save_real_refund_reason_categories(merged)

    @staticmethod
    def _keyword_matches_note(keyword, note_text):
        keyword_text = str(keyword or "").strip()
        note = str(note_text or "")
        if not keyword_text or not note:
            return False
        if keyword_text in note:
            return True
        return all(char in note for char in keyword_text)

    def _get_active_local_reason_categories(self):
        self._merge_default_real_reason_keywords()
        categories = self.db.get_real_refund_reason_categories(active_only=True)
        if not categories:
            self.db.ensure_default_real_refund_reason_categories(self._get_default_real_reason_category_configs())
            categories = self.db.get_real_refund_reason_categories(active_only=True)

        normalized = []
        for item in categories:
            normalized.append({
                "id": item.get("id"),
                "category_name": str(item.get("category_name") or "").strip(),
                "keywords_text": str(item.get("keywords_text") or ""),
                "keywords": self._parse_real_reason_keywords(item.get("keywords_text", "")),
                "status": item.get("status", "ACTIVE"),
                "sort_order": int(item.get("sort_order", 0) or 0),
            })
        normalized.sort(key=lambda item: (item["sort_order"], item["category_name"]))
        return normalized

    def _match_local_real_reason(self, notes, categories=None):
        text = self._normalize_real_reason_note_text(notes)
        if not text:
            return None, ""
        best_match = None
        for item in categories or self._get_active_local_reason_categories():
            category = item.get("category_name", "")
            if category == "未明确备注":
                continue
            matched_keywords = []
            for keyword in item.get("keywords", []):
                if self._keyword_matches_note(keyword, text):
                    matched_keywords.append(keyword)
            if not matched_keywords:
                continue
            score = len(matched_keywords)
            keyword_length_score = sum(len(keyword) for keyword in matched_keywords)
            candidate = {
                "category": category,
                "keywords": matched_keywords,
                "score": score,
                "keyword_length_score": keyword_length_score,
                "sort_order": int(item.get("sort_order", 0) or 0),
            }
            if (
                best_match is None
                or candidate["score"] > best_match["score"]
                or (
                    candidate["score"] == best_match["score"]
                    and candidate["keyword_length_score"] > best_match["keyword_length_score"]
                )
                or (
                    candidate["score"] == best_match["score"]
                    and candidate["keyword_length_score"] == best_match["keyword_length_score"]
                    and candidate["sort_order"] < best_match["sort_order"]
                )
            ):
                best_match = candidate

        if not best_match:
            return None, ""
        keywords_text = "、".join(best_match["keywords"])
        return best_match["category"], f"命中关键词：{keywords_text}"

    def _build_local_real_reason_categories_from_payload(self, payload):
        active_categories = self._get_active_local_reason_categories()
        categories = []
        examples_map = {}
        for item in payload.get("notes", []) or []:
            note_text = self._normalize_real_reason_note_text(item.get("notes", ""))
            if not note_text:
                continue
            category, _ = self._match_local_real_reason(note_text, active_categories)
            if not category:
                continue
            samples = examples_map.setdefault(category, [])
            if note_text not in samples and len(samples) < 2:
                samples.append(note_text)

        for item in active_categories:
            category = item.get("category_name", "")
            if category in examples_map:
                categories.append({"name": category, "examples": examples_map.get(category, [])})

        return {
            "message": "AI返回空结果，已切换本地兜底分类",
            "categories": categories,
        }

    def _build_local_real_reason_assignments_from_note_items(self, note_items, categories):
        allowed = set(categories or [])
        assignments = []
        keyword_hits = {}
        auto_added_categories = []
        for index, note_item in enumerate(note_items):
            note_text = self._normalize_real_reason_note_text(note_item.get("notes", ""))
            category, detail = self._match_local_real_reason(note_text)
            if category and category not in allowed:
                allowed.add(category)
                auto_added_categories.append(category)
            if not category:
                category = "未归因"
                detail = "备注无法判断"
            assignments.append({
                "index": index,
                "category": category,
                "detail": detail,
            })
            if detail.startswith("命中关键词："):
                keyword = detail.replace("命中关键词：", "", 1)
                keyword_hits[keyword] = keyword_hits.get(keyword, 0) + note_item.get("count", 0)
        return {
            "message": "AI返回空结果，已切换本地兜底归因",
            "assignments": assignments,
            "auto_added_categories": auto_added_categories,
        }, keyword_hits, auto_added_categories

    @staticmethod
    def _is_empty_real_reason_category_result(result):
        if not result or not isinstance(result, dict):
            return True
        categories = result.get("categories")
        if not categories:
            message = str(result.get("message", "") or "").strip()
            return not message or "空结果" in message or "无新增分类" in message
        return False

    @staticmethod
    def _is_empty_real_reason_assignment_result(result):
        if not result or not isinstance(result, dict):
            return True
        assignments = result.get("assignments")
        if not assignments:
            message = str(result.get("message", "") or "").strip()
            return not message or "空结果" in message or "无归因结果" in message
        return False

    def _normalize_quality_analysis_result(self, result, payload):
        normalized = dict(result or {})
        if "categories" in normalized and "not_cancelled_reason_categories" not in normalized:
            normalized["not_cancelled_reason_categories"] = normalized.get("categories", [])
        normalized["quality_problem_categories"] = self._normalize_category_list(
            normalized.get("quality_problem_categories", [])
        )
        normalized["not_cancelled_reason_categories"] = self._normalize_category_list(
            normalized.get("not_cancelled_reason_categories", [])
        )
        normalized["message"] = normalized.get("message", "AI归类完成")
        normalized["record_count"] = payload.get("record_count", 0)
        return normalized

    def _normalize_other_analysis_result(self, result, payload):
        normalized = dict(result or {})
        normalized["overall_categories"] = self._normalize_category_list(normalized.get("overall_categories", []))
        spec_categories = []
        for spec_item in normalized.get("spec_categories", []) or []:
            spec_categories.append({
                "spec": spec_item.get("spec", "未识别"),
                "categories": self._normalize_category_list(spec_item.get("categories", [])),
            })
        normalized["spec_categories"] = spec_categories
        normalized["message"] = normalized.get("message", "AI归类完成")
        normalized["record_count"] = payload.get("record_count", 0)
        local_halfway = int(payload.get("local_halfway_return_count", 0) or 0)
        normalized["halfway_return_count"] = local_halfway + int(normalized.get("halfway_return_count", 0) or 0)
        normalized["halfway_return_ratio"] = (
            normalized["halfway_return_count"] / normalized["record_count"] * 100
            if normalized["record_count"] > 0 else 0.0
        )
        normalized["unclear_count"] = normalized.get("unclear_count", 0)
        normalized["unclear_ratio"] = self._safe_float(normalized.get("unclear_ratio", 0))
        return normalized

    def _render_real_reason_analysis_lines(self, analysis):
        lines = [
            f"- 已归因：{analysis.get('classified_count', 0)}单（{analysis.get('classified_ratio', 0):.2f}%）",
            f"- 未归因：{analysis.get('unclassified_count', 0)}单（{analysis.get('unclassified_ratio', 0):.2f}%）",
            "",
            "#### 总体真实退款原因分布",
            "",
        ]
        lines.extend(self._render_category_lines(analysis.get("overall_categories", [])))
        lines.extend(["", "#### 按规格编码分布", ""])
        spec_categories = analysis.get("spec_categories", [])
        if spec_categories:
            for spec_item in spec_categories:
                lines.append(f"- 规格 {spec_item.get('spec', '-')}")
                for category in spec_item.get("categories", []):
                    lines.append(
                        f"  - {category.get('name', '未分类')}：{category.get('count', 0)}"
                        f"（{self._safe_float(category.get('ratio', 0)):.2f}%）"
                    )
        else:
            lines.append("- 无")
        return lines

    def _render_quality_saved_reason_analysis_lines(self, analysis):
        lines = [
            f"- 符合条件订单：{analysis.get('total_count', 0)}单",
            f"- 已AI分析：{analysis.get('classified_count', 0)}单（{analysis.get('classified_ratio', 0):.2f}%）",
            f"- 未分析：{analysis.get('unclassified_count', 0)}单",
            "",
            "#### 品质退款原因",
            "",
        ]
        lines.extend(self._render_category_lines(analysis.get("quality_refund_categories", [])))
        lines.extend(["", "#### 未撤销原因", ""])
        lines.extend(self._render_category_lines(analysis.get("not_cancelled_categories", [])))
        return lines

    def _build_unclassified_reason_summary(self, records):
        summary = {
            "empty_notes_count": 0,
            "no_keyword_match_count": 0,
            "no_keyword_examples": [],
            "matched_but_missing_category_count": 0,
        }
        for record in records or []:
            real_reason = self._normalize_real_reason_value(record.get("real_refund_reason"))
            if real_reason and real_reason != "未归因":
                continue
            notes = self._normalize_real_reason_note_text(record.get("notes", ""))
            if not notes:
                summary["empty_notes_count"] += 1
                continue
            category, _detail = self._match_local_real_reason(notes)
            if category:
                continue
            summary["no_keyword_match_count"] += 1
            if notes not in summary["no_keyword_examples"] and len(summary["no_keyword_examples"]) < 5:
                summary["no_keyword_examples"].append(notes)
        return summary

    def _build_reason_tooltip_text(self, record):
        real_reason_raw = self._normalize_real_reason_value(record.get('real_refund_reason'))
        real_reason = real_reason_raw or "未生成真实退款原因"
        tooltip_lines = [
            f"原始退款原因：{record.get('reason', '')}",
            f"真实退款原因：{real_reason}",
        ]
        detail_text = str(record.get('real_refund_reason_detail') or '').strip()
        updated_at = str(record.get('real_refund_reason_updated_at') or '').strip()
        if detail_text:
            tooltip_lines.append(f"归因说明：{detail_text}")
        if updated_at:
            tooltip_lines.append(f"归因时间：{updated_at}")
        return "\n".join(tooltip_lines)

    def _render_real_reason_view_markdown(self, title, analysis, records=None):
        total_records = len(records or [])
        unclassified_summary = self._build_unclassified_reason_summary(records or [])
        lines = [
            f"## {title}",
            "",
            f"- 当前筛选：{self.get_current_filter_summary_text()}",
            f"- 记录数：{total_records}",
            "",
        ]
        lines.extend(self._render_real_reason_analysis_lines(analysis))
        lines.extend([
            "",
            "#### 未归因原因说明",
            "",
            f"- 备注为空：{unclassified_summary.get('empty_notes_count', 0)}单",
            f"- 未命中任何关键词：{unclassified_summary.get('no_keyword_match_count', 0)}单",
            f"- 命中但分类表缺失：{unclassified_summary.get('matched_but_missing_category_count', 0)}单",
        ])
        examples = unclassified_summary.get("no_keyword_examples", [])
        if examples:
            lines.extend(["", "- 未命中关键词示例："])
            for example in examples:
                lines.append(f"  - {example}")
        return "\n".join(lines)

    def _get_real_reason_category_names(self):
        return [item.get("category_name", "") for item in self._get_active_local_reason_categories() if item.get("category_name")]

    def _normalize_real_reason_category_name(self, name, allowed_categories):
        text = str(name or "").strip()
        if text in allowed_categories:
            return text
        return "未归因"

    def open_local_reason_category_manager(self):
        """打开本地分类管理窗口。"""
        try:
            self.ensure_ai_window_visible()
            self._merge_default_real_reason_keywords()
            dialog_parent = self.ai_window if hasattr(self, 'ai_window') and self.ai_window else self
            dialog = LocalReasonCategoryDialog(self.db, parent=dialog_parent)
            if dialog.exec_() == QDialog.Accepted:
                categories = self.db.get_real_refund_reason_categories(active_only=False)
                self.summary_result_text.setMarkdown(
                    "## 本地分类管理已保存\n\n"
                    + "\n".join(
                        [
                            f"- {item.get('category_name', '')} | 状态：{item.get('status', '')} | 关键词：{item.get('keywords_text', '').replace(chr(10), '、') or '无'}"
                            for item in categories
                        ]
                    )
                )
                self.ai_tab_widget.setCurrentIndex(1)
                self.last_real_reason_category_debug = {
                    "categories": categories,
                    "mode": "local_category_manager",
                }
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"本地分类管理保存失败：{e}")
        finally:
            self.ai_tab_widget.setCurrentIndex(1)
            self.ensure_ai_window_visible()

    def assign_real_refund_reasons(self):
        """按当前本地分类和关键词批量归因。"""
        progress_dialog = None
        try:
            self.ensure_ai_window_visible()
            categories = self._get_active_local_reason_categories()
            if not categories:
                QMessageBox.information(self, "提示", "请先在本地分类管理里配置分类和关键词")
                return

            records = self.get_summary_source_records()
            diagnostics = self._build_real_reason_assignment_diagnostics(records)
            candidates = self._build_real_reason_assignment_candidates(records)
            self.last_real_reason_assignment_debug = {
                "local_diagnostics": diagnostics,
                "categories": categories,
            }
            if not candidates:
                QMessageBox.information(
                    self,
                    "提示",
                    "当前筛选范围内没有需要重新归因的备注。\n\n"
                    f"有备注记录：{diagnostics['noted_records']} 条\n"
                    f"空值记录：{diagnostics['empty_reason_count']} 条\n"
                    f"未归因记录：{diagnostics['unresolved_count']} 条\n"
                    f"备注失效记录：{diagnostics['stale_count']} 条\n"
                    f"已归因有效记录：{diagnostics['valid_assigned_count']} 条"
                )
                return

            progress_dialog = QProgressDialog("正在进行本地关键词归因...", None, 0, 100, self.ai_window if hasattr(self, 'ai_window') else self)
            progress_dialog.setWindowTitle("本地归因中")
            progress_dialog.setWindowModality(Qt.WindowModal)
            progress_dialog.setCancelButton(None)
            progress_dialog.show()
            QApplication.processEvents()

            progress_dialog.setValue(30)
            updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            updated_count = 0
            failed_count = 0
            keyword_hits = {}
            unclassified_examples = []
            matched_counts = {}

            for record in candidates:
                notes = self._normalize_real_reason_note_text(record.get("notes", ""))
                if not notes:
                    continue
                category, detail = self._match_local_real_reason(notes, categories)
                if not category:
                    category = "未归因"
                    detail = "备注无法判断"
                    if notes not in unclassified_examples and len(unclassified_examples) < 5:
                        unclassified_examples.append(notes)
                else:
                    matched_counts[category] = matched_counts.get(category, 0) + 1
                    keyword = detail.replace("命中关键词：", "", 1) if detail.startswith("命中关键词：") else detail
                    keyword_hits[keyword] = keyword_hits.get(keyword, 0) + 1

                note_hash = self._build_real_reason_note_hash(notes)
                if category == "未归因":
                    failed_count += 1
                if self.db.update_real_refund_reason(record.get("id"), category, detail, note_hash, updated_at):
                    updated_count += 1

            self.last_real_reason_assignment_debug = {
                "local_diagnostics": diagnostics,
                "candidate_records": len(candidates),
                "candidate_notes": len(candidates),
                "keyword_hits": keyword_hits,
                "matched_category_counts": matched_counts,
                "unclassified_examples": unclassified_examples,
                "categories": categories,
                "mode": "local_assign",
            }

            progress_dialog.setValue(100)
            self.load_table_data(force_reload=True)
            self.summary_result_text.setMarkdown(
                "## 本地归因完成\n\n"
                f"- 成功写入：{updated_count}条\n"
                f"- 归为未归因：{failed_count}条\n"
                f"- 命中分类：{len(matched_counts)}类\n"
                + ("\n".join([f"- {name}：{count}条" for name, count in matched_counts.items()]) if matched_counts else "- 无命中分类")
            )
            self.ai_tab_widget.setCurrentIndex(1)
        except Exception as e:
            QMessageBox.critical(self, "归因失败", f"本地归因失败：{e}")
        finally:
            if progress_dialog is not None:
                progress_dialog.close()
            self.ensure_ai_window_visible()

    def open_manual_real_reason_assignment(self):
        """打开手动归因窗口。"""
        try:
            if getattr(self, "manual_reason_dialog", None):
                self.manual_reason_dialog.raise_()
                self.manual_reason_dialog.activateWindow()
                return

            records = [
                record for record in self.get_summary_source_records()
                if self._normalize_real_reason_value(record.get("real_refund_reason")) in ("", "未归因")
            ]
            categories = self._get_real_reason_category_names()
            if not records:
                QMessageBox.information(self, "提示", "当前筛选范围内没有未归因记录")
                return
            if not categories:
                QMessageBox.information(self, "提示", "请先在本地分类管理里配置分类")
                return

            dialog_parent = self.ai_window if hasattr(self, 'ai_window') and self.ai_window else self
            dialog = ManualReasonAssignmentDialog(
                records,
                categories,
                parent=dialog_parent,
                manual_assign_callback=self._apply_manual_real_reason_assignment,
                ai_assign_callback=self._apply_ai_manual_real_reason_assignment,
                save_note_spec_callback=self._apply_manual_reason_note_spec_save,
                single_assign_callback=self._apply_current_range_real_reason_assignment,
            )
            dialog.setWindowModality(Qt.NonModal)
            dialog.setWindowFlag(Qt.Window, True)
            dialog.setAttribute(Qt.WA_DeleteOnClose, True)
            dialog.destroyed.connect(lambda _=None: setattr(self, "manual_reason_dialog", None))
            self.manual_reason_dialog = dialog
            dialog.show()
        except Exception as e:
            QMessageBox.critical(self, "手动归因失败", f"手动归因失败：{e}")
        finally:
            self.ensure_ai_window_visible()

    def _apply_manual_reason_note_spec_save(self, changed_records):
        """保存手动归因窗口里编辑的备注和规格编码。"""
        updated_count = 0
        note_changed_count = 0
        for item in changed_records or []:
            record = item.get("record") or {}
            record_id = record.get("id")
            if not record_id:
                continue
            new_spec_code = str(item.get("spec_code") or "").strip()
            new_notes = str(item.get("notes") or "")
            old_notes = str(record.get("notes") or "")
            notes_changed = new_notes != old_notes
            if self.db.update_record_partial(record_id, spec_code=new_spec_code, notes=new_notes):
                record["spec_code"] = new_spec_code
                record["notes"] = new_notes
                if notes_changed:
                    record["real_refund_reason"] = ""
                    record["real_refund_reason_detail"] = ""
                    record["real_refund_reason_updated_at"] = ""
                    record["real_refund_reason_note_hash"] = ""
                    note_changed_count += 1
                updated_count += 1

        if updated_count:
            self.load_table_data(force_reload=True)
            self.summary_result_text.setMarkdown(
                "## 备注/规格已保存\n\n"
                f"- 保存记录：{updated_count}条\n"
                f"- 备注变更并清空归因：{note_changed_count}条\n"
            )
            self.ai_tab_widget.setCurrentIndex(1)
        return {
            "updated_count": updated_count,
            "note_changed_count": note_changed_count,
            "message": f"已保存 {updated_count} 条；备注变更清空归因 {note_changed_count} 条",
        }

    def open_current_range_reason_assignment(self):
        """查看并修正当前筛选范围内全部订单的真实退款原因。"""
        try:
            records = self.get_summary_source_records()
            if not records:
                QMessageBox.information(self, "提示", "当前筛选范围内没有订单")
                return
            categories = self._get_real_reason_category_names()
            if not categories:
                QMessageBox.information(self, "提示", "请先在本地分类管理里配置分类")
                return
            dialog_parent = self.ai_window if hasattr(self, 'ai_window') and self.ai_window else self
            dialog = CurrentRangeReasonAssignmentDialog(
                records,
                categories,
                parent=dialog_parent,
                assign_callback=self._apply_current_range_real_reason_assignment,
            )
            dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "当前范围归因失败", f"当前范围归因失败：{e}")
        finally:
            self.ensure_ai_window_visible()

    def _apply_current_range_real_reason_assignment(self, record, selected_category):
        """当前范围归因窗口内的单行即时保存。"""
        if not record or not selected_category:
            return {"success": False, "message": "缺少订单或分类"}
        updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        note_hash = self._build_real_reason_note_hash(record.get("notes", ""))
        success = self.db.update_real_refund_reason(
            record.get("id"),
            selected_category,
            "当前范围归因手动修正",
            note_hash,
            updated_at
        )
        if success:
            self.load_table_data(force_reload=True)
            self.summary_result_text.setMarkdown(
                "## 当前范围归因已更新\n\n"
                f"- 订单号：{record.get('order_no', '')}\n"
                f"- 真实退款原因：{selected_category}\n"
            )
            self.ai_tab_widget.setCurrentIndex(1)
        return {"success": success, "message": "已保存" if success else "保存失败"}

    def _apply_manual_real_reason_assignment(self, selected_records, selected_category):
        """手动归因窗口内直接写库，不关闭窗口。"""
        if not selected_records or not selected_category:
            return {"updated_count": 0, "message": "请选择记录和目标分类"}

        updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_count = 0
        assigned_ids = []
        for record in selected_records:
            note_hash = self._build_real_reason_note_hash(record.get("notes", ""))
            if self.db.update_real_refund_reason(record.get("id"), selected_category, "手动归因", note_hash, updated_at):
                updated_count += 1
                assigned_ids.append(record.get("id"))

        self.load_table_data(force_reload=True)
        self.summary_result_text.setMarkdown(
            "## 手动归因完成\n\n"
            f"- 目标分类：{selected_category}\n"
            f"- 成功写入：{updated_count}条\n"
        )
        self.ai_tab_widget.setCurrentIndex(1)
        return {"updated_count": updated_count, "assigned_ids": assigned_ids, "message": f"已归因 {updated_count} 条"}

    def _apply_ai_manual_real_reason_assignment(self, records, progress_parent=None):
        """让AI只看备注，把手动归因窗口里的未归因记录归入现有或新增分类。"""
        records = [
            record for record in (records or [])
            if self._normalize_real_reason_note_text(record.get("notes", ""))
        ]
        if not records:
            return {"updated_count": 0, "new_categories": [], "message": "当前窗口没有可归因记录"}
        if not self.ai_analyzer.api_key:
            self.show_api_settings_dialog()
            if not self.ai_analyzer.api_key:
                return {"updated_count": 0, "new_categories": [], "message": "API未配置，已取消AI归因"}

        progress_parent = progress_parent or (self.ai_window if hasattr(self, 'ai_window') else self)
        progress_dialog = QProgressDialog("正在调用AI归因...", None, 0, 0, progress_parent)
        progress_dialog.setWindowTitle("AI归因中")
        progress_dialog.setWindowModality(Qt.WindowModal)
        progress_dialog.setCancelButton(None)
        progress_dialog.show()
        QApplication.processEvents()

        started_at = time.time()
        try:
            existing_categories = self._get_real_reason_category_names()
            notes = [self._normalize_real_reason_note_text(record.get("notes", "")) for record in records]
            payload = {
                "existing_categories": existing_categories,
                "notes": notes,
            }
            result, debug_meta = self.ai_analyzer.analyze_manual_real_reason_assignments(payload)
            assignments = result.get("assignments", []) if isinstance(result, dict) else []

            new_categories = []
            new_category_items = result.get("new_categories", []) if isinstance(result, dict) else []
            for item in new_category_items:
                if isinstance(item, dict):
                    name = str(item.get("name") or item.get("category") or "").strip()
                else:
                    name = str(item or "").strip()
                if name and name not in existing_categories and name not in new_categories and name != "未归因":
                    new_categories.append(name)

            for item in assignments:
                category = str(item.get("category", "")).strip()
                if category and category not in existing_categories and category not in new_categories and category != "未归因":
                    new_categories.append(category)

            proposed_assignments = []
            skipped_count = 0
            for item in assignments:
                try:
                    index = int(item.get("index"))
                except Exception:
                    skipped_count += 1
                    continue
                if index < 0 or index >= len(records):
                    skipped_count += 1
                    continue
                category = str(item.get("category", "")).strip()
                if not category or category == "未归因":
                    skipped_count += 1
                    continue
                if category not in existing_categories and category not in new_categories:
                    new_categories.append(category)
                record = records[index]
                proposed_assignments.append({
                    "record": record,
                    "category": category,
                    "detail": str(item.get("detail") or "AI归因").strip()[:80],
                })

            elapsed_seconds = time.time() - started_at
            progress_dialog.close()

            if not proposed_assignments:
                return {
                    "updated_count": 0,
                    "assigned_ids": [],
                    "new_categories": new_categories,
                    "categories": existing_categories,
                    "message": f"AI未返回可写入的归因结果，跳过/未匹配 {skipped_count} 条",
                }

            confirm_parent = progress_parent or (self.ai_window if hasattr(self, 'ai_window') else self)
            confirm_dialog = AIReasonAssignmentConfirmDialog(
                proposed_assignments,
                new_categories,
                elapsed_seconds,
                parent=confirm_parent
            )
            if confirm_dialog.exec_() != QDialog.Accepted:
                return {
                    "updated_count": 0,
                    "assigned_ids": [],
                    "new_categories": new_categories,
                    "categories": existing_categories,
                    "message": "用户取消，未写入AI归因结果",
                }

            if new_categories:
                current_configs = self.db.get_real_refund_reason_categories(active_only=False)
                sort_base = len(current_configs)
                for index, name in enumerate(new_categories):
                    current_configs.append({
                        "category_name": name,
                        "keywords_text": "",
                        "status": "ACTIVE",
                        "sort_order": sort_base + index,
                    })
                self.db.save_real_refund_reason_categories(current_configs)
                existing_categories = self._get_real_reason_category_names()

            updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            updated_count = 0
            assigned_ids = []
            for item in proposed_assignments:
                category = item["category"]
                if category not in existing_categories:
                    skipped_count += 1
                    continue
                record = item["record"]
                detail = item["detail"] or "AI归因"
                note_hash = self._build_real_reason_note_hash(record.get("notes", ""))
                if self.db.update_real_refund_reason(record.get("id"), category, detail, note_hash, updated_at):
                    updated_count += 1
                    assigned_ids.append(record.get("id"))

            self.load_table_data(force_reload=True)
            self.summary_result_text.setMarkdown(
                "## AI归因完成\n\n"
                f"- 成功写入：{updated_count}条\n"
                f"- 新增分类：{len(new_categories)}个"
                + (("\n" + "\n".join([f"- {name}" for name in new_categories])) if new_categories else "\n- 无新增分类")
                + f"\n- 跳过/未匹配：{skipped_count}条\n"
                + f"- API耗时：{elapsed_seconds:.2f}秒\n"
            )
            self.ai_tab_widget.setCurrentIndex(1)
            return {
                "updated_count": updated_count,
                "assigned_ids": assigned_ids,
                "new_categories": new_categories,
                "categories": existing_categories,
                "message": f"AI已归因 {updated_count} 条，新增分类 {len(new_categories)} 个，耗时 {elapsed_seconds:.2f} 秒",
            }
        finally:
            if progress_dialog is not None:
                progress_dialog.close()

    def show_real_refund_reason_view(self):
        """查看当前筛选范围内已保存的真实退款原因统计。"""
        try:
            records = self.get_summary_source_records()
            if not records:
                QMessageBox.information(self, "提示", "当前筛选条件下没有数据")
                return
            analysis = self._build_real_reason_analysis(records)
            self.summary_result_text.setMarkdown(
                self._render_real_reason_view_markdown("真实退款原因统计", analysis, records)
            )
            self.ai_tab_widget.setCurrentIndex(1)
        except Exception as e:
            QMessageBox.critical(self, "查看失败", f"查看真实退款原因失败：{e}")

    def _is_quality_refund_analysis_record(self, record):
        return (
            self._is_quality_reason(record.get("reason"))
            and not bool(record.get("cancel"))
            and not self._is_reject_success_record(record)
        )

    def _get_quality_refund_reason_records(self):
        return [
            record for record in self.get_summary_source_records()
            if self._is_quality_refund_analysis_record(record)
        ]

    def _is_quality_refund_reason_stale(self, record):
        notes = str(record.get("notes", "") or "").strip()
        if not notes:
            return False
        current_hash = self._build_real_reason_note_hash(notes)
        stored_hash = str(record.get("quality_refund_reason_note_hash") or "").strip()
        if not stored_hash:
            return True
        return current_hash != stored_hash

    def _build_quality_refund_assignment_candidates(self, records):
        candidates = []
        for record in records:
            notes = str(record.get("notes", "") or "").strip()
            if not notes:
                continue
            has_not_cancelled_reason = str(record.get("quality_not_cancelled_reason") or "").strip()
            if has_not_cancelled_reason and not self._is_quality_refund_reason_stale(record):
                continue
            candidates.append(record)
        return candidates

    def _build_quality_refund_assignment_payload(self, records):
        payload_records = []
        for index, record in enumerate(records):
            payload_records.append({
                "index": index,
                "id": record.get("id"),
                "store_name": record.get("store_name", ""),
                "spec_code": str(record.get("spec_code") or "").strip(),
                "refund_reason": str(record.get("reason") or "").strip(),
                "notes": self._normalize_real_reason_note_text(record.get("notes", "")),
            })
        existing_categories = [
            item.get("category_name", "")
            for item in self.db.get_quality_not_cancelled_reason_categories()
            if str(item.get("category_name") or "").strip()
        ]
        return {"existing_categories": existing_categories, "records": payload_records}

    def _normalize_quality_refund_assignment_item(self, item, record=None):
        try:
            index = int(item.get("index"))
        except Exception:
            index = -1
        not_cancelled_reason = self._normalize_display_category_name(item.get("not_cancelled_reason", ""))
        if not_cancelled_reason == "未明确未撤销原因":
            not_cancelled_reason = "未明确说明未撤销原因"
        detail = self._normalize_quality_not_cancelled_detail(
            item.get("detail", ""),
            not_cancelled_reason,
            record.get("notes", "") if isinstance(record, dict) else "",
        )
        return {
            "index": index,
            "not_cancelled_reason": not_cancelled_reason or "未明确说明未撤销原因",
            "detail": detail,
        }

    def _normalize_quality_not_cancelled_detail(self, detail, not_cancelled_reason="", notes=""):
        detail_text = str(detail or "").strip()
        reason_text = str(not_cancelled_reason or "").strip()
        notes_text = self._normalize_real_reason_note_text(notes)
        if not detail_text or reason_text == "未明确说明未撤销原因":
            return ""

        detail_text = re.sub(
            r'^(备注|内容|说明|用户备注|客服备注)\s*(提到|提到了|写了|写到|写到了|说了|显示|说明|描述|中说|记录|为|是)[:：，,\s]*',
            '',
            detail_text
        ).strip()
        if not detail_text:
            return ""

        quality_only_keywords = [
            "烂", "腐烂", "发霉", "长毛", "发芽", "破损", "坏了", "断裂", "变质",
            "规格不符", "个头", "口感", "不好吃", "质量问题"
        ]
        not_cancelled_keywords = [
            "客服", "接线", "超时", "未处理", "来不及", "证据", "图片", "视频",
            "仅退款", "平台", "不同意", "拒绝", "撤销", "拦截", "已退款", "保留"
        ]
        has_quality_only = any(keyword in detail_text for keyword in quality_only_keywords)
        has_not_cancelled_signal = any(keyword in detail_text for keyword in not_cancelled_keywords)
        if has_quality_only and not has_not_cancelled_signal:
            return ""
        if notes_text:
            normalized_detail = re.sub(r'\s+', '', detail_text)
            normalized_notes = re.sub(r'\s+', '', notes_text)
            if normalized_detail and (
                normalized_detail == normalized_notes
                or (len(normalized_detail) >= 6 and normalized_detail in normalized_notes)
            ):
                return ""
        return detail_text[:80]

    def _apply_quality_not_cancelled_local_fallback(self, item, record):
        notes = self._normalize_real_reason_note_text(record.get("notes", "") if isinstance(record, dict) else "")
        if not notes:
            return item
        reason = str(item.get("not_cancelled_reason") or "").strip()
        if reason and reason != "未明确说明未撤销原因":
            return item

        compact_notes = re.sub(r'\s+', '', notes)
        reject_keywords = [
            "驳回退款导致退款", "驳回导致退款", "驳回失败", "驳回后平台退款",
            "驳回后退款", "驳回被平台退款", "驳回不成功", "驳回未成功"
        ]
        video_keywords = ["多多视频导致", "多多视频售后", "多多视频"]

        if any(keyword in compact_notes for keyword in reject_keywords):
            item["not_cancelled_reason"] = "驳回失败导致退款"
            item["detail"] = ""
        elif any(keyword in compact_notes for keyword in video_keywords):
            item["not_cancelled_reason"] = "多多视频售后导致退款"
            item["detail"] = ""
        return item

    def _normalize_quality_not_cancelled_categories(self, categories):
        names = []
        seen = set()
        for item in categories or []:
            if isinstance(item, dict):
                name = item.get("category_name") or item.get("name")
            else:
                name = item
            name = self._normalize_display_category_name(name)
            if name == "未明确未撤销原因":
                name = "未明确说明未撤销原因"
            if not name or name in seen:
                continue
            seen.add(name)
            names.append(name)
        if "未明确说明未撤销原因" not in seen:
            names.append("未明确说明未撤销原因")
        return names

    def analyze_quality_refund_reasons_for_dialog(self, dialog):
        """对当前品质退款原因窗口中的候选订单执行AI逐单分类。"""
        progress_dialog = None
        try:
            records = self._get_quality_refund_reason_records()
            if not records:
                QMessageBox.information(dialog, "提示", "当前筛选条件下没有未撤销且未驳回成功的品质退款订单")
                return
            dialog_records = list(getattr(dialog, "records", []) or [])
            candidate_source = dialog_records if dialog_records else records
            candidates = [
                record for record in candidate_source
                if self._is_quality_refund_analysis_record(record)
                and str(record.get("notes", "") or "").strip()
            ]
            if not candidates:
                QMessageBox.information(dialog, "提示", "当前窗口内没有可发送给AI分析的备注")
                return
            if not self.ai_analyzer.api_key:
                self.show_api_settings_dialog()
                return

            progress_dialog = QProgressDialog("正在AI分析未撤销原因...", None, 0, 100, dialog)
            progress_dialog.setWindowTitle("AI分析中")
            progress_dialog.setWindowModality(Qt.WindowModal)
            progress_dialog.setCancelButton(None)
            progress_dialog.show()
            QApplication.processEvents()

            payload = self._build_quality_refund_assignment_payload(candidates)
            progress_dialog.setValue(30)
            result, _debug_meta = self.ai_analyzer.analyze_quality_refund_order_reasons(payload)
            assignments = result.get("assignments", []) if isinstance(result, dict) else []
            if not isinstance(assignments, list):
                raise ValueError("AI返回格式异常：assignments 不是数组")

            updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ai_categories = self._normalize_quality_not_cancelled_categories(
                result.get("categories", []) if isinstance(result, dict) else []
            )
            for raw_item in assignments:
                if isinstance(raw_item, dict):
                    assignment_name = self._normalize_display_category_name(raw_item.get("not_cancelled_reason", ""))
                    if assignment_name and assignment_name not in ai_categories:
                        ai_categories.append(assignment_name)
            existing_category_names = [
                item.get("category_name", "")
                for item in self.db.get_quality_not_cancelled_reason_categories(active_only=False)
                if str(item.get("category_name") or "").strip()
            ]
            merged_category_names = []
            for name in existing_category_names + ai_categories:
                normalized_name = self._normalize_display_category_name(name)
                if normalized_name and normalized_name not in merged_category_names:
                    merged_category_names.append(normalized_name)
            self.db.save_quality_not_cancelled_reason_categories(merged_category_names)

            updated_count = 0
            updated_record_ids = set()
            candidate_by_index = {index: record for index, record in enumerate(candidates)}
            for raw_item in assignments:
                if not isinstance(raw_item, dict):
                    continue
                raw_index = self._normalize_quality_refund_assignment_item(raw_item)["index"]
                record = candidate_by_index.get(raw_index)
                if not record:
                    continue
                item = self._normalize_quality_refund_assignment_item(raw_item, record)
                item = self._apply_quality_not_cancelled_local_fallback(item, record)
                fallback_category = item.get("not_cancelled_reason", "")
                if fallback_category and fallback_category not in merged_category_names:
                    merged_category_names.append(fallback_category)
                    self.db.save_quality_not_cancelled_reason_categories(merged_category_names)
                note_hash = self._build_real_reason_note_hash(record.get("notes", ""))
                if self.db.update_quality_not_cancelled_reason(
                    record.get("id"),
                    item["not_cancelled_reason"],
                    item["detail"],
                    note_hash,
                    updated_at,
                ):
                    updated_count += 1
                    updated_record_ids.add(record.get("id"))

            for record in candidates:
                record_id = record.get("id")
                if record_id in updated_record_ids:
                    continue
                note_hash = self._build_real_reason_note_hash(record.get("notes", ""))
                if self.db.update_quality_not_cancelled_reason(
                    record_id,
                    "未明确说明未撤销原因",
                    "",
                    note_hash,
                    updated_at,
                ):
                    updated_count += 1

            progress_dialog.setValue(100)
            progress_dialog.close()
            progress_dialog = None
            self.load_table_data(force_reload=True)
            dialog.set_records(self._get_quality_refund_reason_records())
            QMessageBox.information(dialog, "分析完成", f"AI已分析并保存 {updated_count} 条未撤销原因")
        except Exception as e:
            if progress_dialog is not None:
                progress_dialog.close()
            QMessageBox.critical(dialog, "AI分析失败", f"AI分析未撤销原因失败：{e}")

    def show_quality_refund_reason_view(self):
        """查看当前筛选范围内未撤销且未驳回成功的品质退款原因。"""
        try:
            records = self._get_quality_refund_reason_records()
            parent = self.ai_window if hasattr(self, 'ai_window') and self.ai_window else self
            dialog = QualityRefundReasonDialog(self, records, parent=parent)
            dialog.exec_()
            self.ensure_ai_window_visible()
        except Exception as e:
            QMessageBox.critical(self, "查看失败", f"查看品质退款失败：{e}")

    def _build_quality_refund_order_detail_rows(self, records):
        """构建总结导出用的未撤销品质退款订单明细。"""
        detail_rows = []
        for record in records or []:
            if not self._is_quality_refund_analysis_record(record):
                continue
            detail_rows.append({
                "order_no": str(record.get("order_no") or "").strip(),
                "store_name": str(record.get("store_name") or "").strip(),
                "spec_code": str(record.get("spec_code") or "").strip(),
                "refund_reason": self._normalize_reason(record.get("reason")),
                "notes": self._normalize_real_reason_note_text(record.get("notes", "")),
                "not_cancelled_reason": str(record.get("quality_not_cancelled_reason") or "").strip(),
                "analysis_detail": self._normalize_quality_not_cancelled_detail(
                    record.get("quality_refund_reason_detail", ""),
                    record.get("quality_not_cancelled_reason", ""),
                    record.get("notes", ""),
                ),
            })
        return detail_rows

    def _build_quality_not_cancelled_reason_summary(self, detail_rows):
        """按未撤销原因汇总当前品质退款明细占比。"""
        counts = {}
        for item in detail_rows or []:
            reason = str(item.get("not_cancelled_reason") or "").strip()
            if not reason or reason == "-":
                reason = "未明确说明未撤销原因"
            counts[reason] = counts.get(reason, 0) + 1

        total_count = sum(counts.values())
        if total_count <= 0:
            return []
        return [
            {
                "name": name,
                "count": count,
                "ratio": round(count / total_count * 100, 2),
            }
            for name, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
        ]

    def _build_local_summary_snapshot(self, records):
        """构建不依赖AI的本地总结快照。"""
        context = self.get_current_filter_context()
        grouped_records = {}
        for record in records:
            store_name = record.get("store_name", "未知店铺")
            grouped_records.setdefault(store_name, []).append(record)

        store_summaries = []
        for store_name in sorted(grouped_records.keys()):
            store_records = grouped_records[store_name]
            store_summaries.append({
                "store_name": store_name,
                "metrics": self._build_store_summary_stats(store_name, store_records),
                "real_reason_analysis": self._build_real_reason_analysis(store_records),
                "quality_saved_reason_analysis": self._build_quality_saved_reason_analysis(store_records),
            })

        quality_detail_rows = self._build_quality_refund_order_detail_rows(records)
        return {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "filters": context,
            "filter_summary": self.get_current_filter_summary_text(),
            "records_count": len(records),
            "store_count": len(store_summaries),
            "stores": store_summaries,
            "totals": self._build_total_summary_stats(
                [store["metrics"] for store in store_summaries],
                records
            ),
            "overall_real_reason_analysis": self._build_real_reason_analysis(records),
            "overall_quality_saved_reason_analysis": self._build_quality_saved_reason_analysis(records),
            "quality_refund_order_details": quality_detail_rows,
            "quality_not_cancelled_reason_summary": self._build_quality_not_cancelled_reason_summary(quality_detail_rows),
        }

    def _run_summary_ai_sections(self, snapshot, progress_dialog=None, base_progress=20, step_progress=60):
        """补充总结快照中的AI备注归类结果。"""
        stores = snapshot.get("stores", [])
        snapshot["ai_debug"] = {"stores": []}
        total_jobs = 0
        for store in stores:
            if store["quality_unreversed_note_payload"]["record_count"] > 0:
                total_jobs += 1
            if store["other_reason_note_payload"].get("ai_record_count", store["other_reason_note_payload"]["record_count"]) > 0:
                total_jobs += 1

        if total_jobs == 0:
            for store in stores:
                store["quality_unreversed_note_analysis"] = self._get_default_quality_note_analysis(
                    store["quality_unreversed_note_payload"],
                    "当前店铺没有需要归类的未撤销品质退款备注"
                )
                store["other_reason_note_analysis"] = self._get_default_other_reason_analysis(
                    store["other_reason_note_payload"],
                    "当前店铺没有“其他”原因备注需要复盘"
                )
            return

        if not self.ai_analyzer.api_key:
            for store in stores:
                store["quality_unreversed_note_analysis"] = self._get_default_quality_note_analysis(
                    store["quality_unreversed_note_payload"],
                    "未配置API，未执行备注归类"
                )
                store["other_reason_note_analysis"] = self._get_default_other_reason_analysis(
                    store["other_reason_note_payload"],
                    "未配置API，未执行其他原因备注复盘"
                )
            return

        finished_jobs = 0
        for store in stores:
            store_debug = {"store_name": store.get("store_name", ""), "quality": {}, "other": {}}
            payload = store["quality_unreversed_note_payload"]
            if payload["record_count"] > 0:
                try:
                    analysis_result, debug_meta = self.ai_analyzer.analyze_quality_unreversed_notes(payload)
                    store["quality_unreversed_note_analysis"] = self._normalize_quality_analysis_result(analysis_result, payload)
                    store_debug["quality"] = debug_meta
                except Exception as exc:
                    store["quality_unreversed_note_analysis"] = self._get_default_quality_note_analysis(payload, f"AI归类失败：{exc}")
                    store_debug["quality"] = {
                        "error": str(exc),
                        "payload_size": len(json.dumps(payload, ensure_ascii=False)),
                    }
                finished_jobs += 1
                if progress_dialog:
                    progress_dialog.setValue(base_progress + int(step_progress * finished_jobs / total_jobs))
                    QApplication.processEvents()
            else:
                store["quality_unreversed_note_analysis"] = self._get_default_quality_note_analysis(
                    payload,
                    "当前店铺没有需要归类的未撤销品质退款备注"
                )

            payload = store["other_reason_note_payload"]
            if payload.get("ai_record_count", payload["record_count"]) > 0:
                try:
                    analysis_result, debug_meta = self.ai_analyzer.analyze_other_reason_notes(payload)
                    store["other_reason_note_analysis"] = self._normalize_other_analysis_result(analysis_result, payload)
                    store_debug["other"] = debug_meta
                except Exception as exc:
                    store["other_reason_note_analysis"] = self._get_default_other_reason_analysis(payload, f"AI归类失败：{exc}")
                    store_debug["other"] = {
                        "error": str(exc),
                        "payload_size": len(json.dumps(payload, ensure_ascii=False)),
                    }
                finished_jobs += 1
                if progress_dialog:
                    progress_dialog.setValue(base_progress + int(step_progress * finished_jobs / total_jobs))
                    QApplication.processEvents()
            else:
                store["other_reason_note_analysis"] = self._get_default_other_reason_analysis(
                    payload,
                    "当前店铺其他原因备注已由本地规则完成统计" if payload.get("record_count", 0) > 0
                    else "当前店铺没有“其他”原因备注需要复盘"
                )
            snapshot["ai_debug"]["stores"].append(store_debug)

    def _render_category_lines(self, categories, value_key="count"):
        lines = []
        for category in categories or []:
            lines.append(
                f"- {category.get('name', '未分类')}：{category.get(value_key, 0)}"
                f"（{self._safe_float(category.get('ratio', 0)):.2f}%）"
            )
        return lines or ["- 无"]

    def render_summary_snapshot_markdown(self, snapshot):
        """将总结快照渲染为Markdown。"""
        lines = [
            "# 本地总结分析",
            "",
            f"- 生成时间：{snapshot.get('generated_at', '')}",
            f"- 当前筛选：{snapshot.get('filter_summary', '')}",
            f"- 记录数：{snapshot.get('records_count', 0)}",
            f"- 店铺数：{snapshot.get('store_count', 0)}",
            "",
        ]

        for store in snapshot.get("stores", []):
            metrics = store.get("metrics", {})
            lines.extend([
                f"## {store.get('store_name', '未知店铺')}",
                "",
                "### 基础统计",
                "",
                f"- 日期范围：{metrics.get('date_range', {}).get('start_date', '')} 至 {metrics.get('date_range', {}).get('end_date', '')}",
                f"- 当前范围单量：{self._format_metric_int(metrics.get('orders', 0))}",
                f"- 当前范围销售额：¥{metrics.get('sales', 0):.2f}",
                f"- 退款预算：¥{metrics.get('refund_budget_remaining', 0):.2f}",
                f"- 品质退款：{metrics.get('quality_refund_count', 0)}单",
                f"- 其他退款：{metrics.get('other_refund_count', 0)}单",
                f"- 撤销品质：{metrics.get('canceled_quality_count', 0)}单",
                f"- 总退款率：{metrics.get('total_refund_rate', 0):.2f}%",
                f"- 售后总额：¥{metrics.get('total_after_sales', 0):.2f}",
                f"- 金额占比：{metrics.get('refund_ratio', 0):.2f}%",
                f"- 品质售后：¥{metrics.get('quality_after_sales_amount', 0):.2f}",
                f"- 其他售后：¥{metrics.get('other_after_sales_amount', 0):.2f}",
                f"- 申请品质率：{metrics.get('quality_apply_rate', 0):.2f}%",
                f"- 实际品质率：{metrics.get('quality_actual_rate', 0):.2f}%",
                f"- 撤销率：{metrics.get('quality_cancel_rate', 0):.2f}%",
                f"- 品质退款申请单量：{metrics.get('quality_apply_count', 0)}单",
                f"- 品质退款撤销单量：{metrics.get('quality_cancel_count', 0)}单",
                f"- 品质退款实际单量：{metrics.get('quality_actual_count', 0)}单",
                f"- 有效退款金额：¥{metrics.get('effective_refund_amount', 0):.2f}",
                f"- 补偿金额：¥{metrics.get('compensation_amount', 0):.2f}",
                f"- 有备注订单：{metrics.get('note_count', 0)}单",
                f"- 无备注订单：{metrics.get('no_note_count', 0)}单",
                f"- 备注率：{metrics.get('note_rate', 0):.2f}%",
                f"- 无备注率：{metrics.get('no_note_rate', 0):.2f}%",
                f"- 最多原因：{metrics.get('top_refund_reason', '无数据')}",
                f"- 出现次数：{metrics.get('top_reason_count', 0)}",
                f"- 占比：{metrics.get('top_reason_ratio', 0):.2f}%",
                "",
                "### 真实退款原因统计",
                "",
            ])
            lines.extend(self._render_real_reason_analysis_lines(store.get("real_reason_analysis", {})))
            lines.extend([
                "",
                "### 品质退款原因统计",
                "",
            ])
            lines.extend(self._render_quality_saved_reason_analysis_lines(store.get("quality_saved_reason_analysis", {})))
            lines.append("")

        totals = snapshot.get("totals")
        if totals:
            lines.extend([
                "## 全部总和",
                "",
                f"- 当前范围单量：{self._format_metric_int(totals.get('orders', 0))}",
                f"- 当前范围销售额：¥{totals.get('sales', 0):.2f}",
                f"- 退款预算：¥{totals.get('refund_budget_remaining', 0):.2f}",
                f"- 品质退款：{totals.get('quality_refund_count', 0)}单",
                f"- 其他退款：{totals.get('other_refund_count', 0)}单",
                f"- 撤销品质：{totals.get('canceled_quality_count', 0)}单",
                f"- 总退款率：{totals.get('total_refund_rate', 0):.2f}%",
                f"- 售后总额：¥{totals.get('total_after_sales', 0):.2f}",
                f"- 金额占比：{totals.get('refund_ratio', 0):.2f}%",
                f"- 申请品质率：{totals.get('quality_apply_rate', 0):.2f}%",
                f"- 实际品质率：{totals.get('quality_actual_rate', 0):.2f}%",
                f"- 撤销率：{totals.get('quality_cancel_rate', 0):.2f}%",
                f"- 有备注订单：{totals.get('note_count', 0)}单",
                f"- 无备注订单：{totals.get('no_note_count', 0)}单",
                f"- 备注率：{totals.get('note_rate', 0):.2f}%",
                f"- 无备注率：{totals.get('no_note_rate', 0):.2f}%",
                "",
                "### 全部总和真实退款原因统计",
                "",
            ])
            lines.extend(self._render_real_reason_analysis_lines(snapshot.get("overall_real_reason_analysis", {})))
            lines.extend([
                "",
                "### 全部总和品质退款原因统计",
                "",
            ])
            lines.extend(self._render_quality_saved_reason_analysis_lines(snapshot.get("overall_quality_saved_reason_analysis", {})))

        return "\n".join(lines)

    def display_summary_snapshot(self, snapshot):
        """显示总结快照。"""
        self.latest_summary_snapshot = snapshot
        self.summary_result_text.setMarkdown(self.render_summary_snapshot_markdown(snapshot))
        self.ai_tab_widget.setCurrentIndex(1)

    def show_daily_work_summary_dialog(self):
        """打开当前筛选范围的工作总结窗口。"""
        try:
            parent = self.ai_window if hasattr(self, 'ai_window') and self.ai_window and self.ai_window.isVisible() else self
            dialog = DailyWorkSummaryDialog(self, parent=parent)
            dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "工作总结", f"打开工作总结窗口失败：{e}")

    def _build_quality_refund_order_lines(self, records):
        """列出当前范围内品质退款订单明细。"""
        items = []
        for record in records or []:
            if not self._is_quality_reason(record.get("reason")):
                continue
            if record.get("cancel"):
                result_text = "已撤销"
            elif self._is_reject_success_record(record):
                result_text = "驳回成功"
            else:
                result_text = "未撤销/退款保留"
                after_sale_status = str(record.get("after_sale_status") or "").strip()
                if after_sale_status:
                    result_text = f"{result_text}（{after_sale_status}）"
            items.append({
                "store_name": record.get("store_name", ""),
                "order_no": record.get("order_no", ""),
                "spec_code": str(record.get("spec_code") or "").strip(),
                "refund_reason": self._normalize_reason(record.get("reason")),
                "refund_amount": self._safe_float(record.get("refund_amount", 0)),
                "cancel": "是" if record.get("cancel") else "否",
                "reject": "是" if record.get("reject") else "否",
                "reject_result": self._display_reject_result_value(record),
                "result": result_text,
                "notes": self._normalize_real_reason_note_text(record.get("notes", "")),
                "quality_not_cancelled_reason": str(record.get("quality_not_cancelled_reason") or "").strip(),
                "quality_refund_reason_detail": str(record.get("quality_refund_reason_detail") or "").strip(),
            })
        return items

    def _build_unshipped_followup_orders(self, records):
        """列出当前范围内需要晚班客服关注的未发货订单。"""
        items = []
        for record in records or []:
            order_no = str(record.get("order_no") or "").strip()
            if not order_no:
                continue
            order_status = str(record.get("order_status") or "").strip()
            if order_status == "已发货":
                continue
            items.append({
                "store_name": record.get("store_name", ""),
                "order_no": order_no,
                "order_status": order_status or "未填写",
                "after_sale_status": str(record.get("after_sale_status") or "").strip() or "未填写",
                "refund_reason": self._normalize_reason(record.get("reason")),
                "notes": self._normalize_real_reason_note_text(record.get("notes", "")),
            })
        return items

    @staticmethod
    def _should_include_unshipped_followup(user_text):
        """只有用户明确提到发货跟踪时，才展示/发送未发货列表。"""
        text = re.sub(r'\s+', '', str(user_text or ""))
        if not text:
            return False
        keywords = [
            "未发货", "没发货", "没有发货", "待发货", "还没发货", "尚未发货",
            "发货跟踪", "跟踪发货", "晚班跟踪", "晚班客服跟踪", "发货需要跟踪",
            "未及时发货", "没有及时发货"
        ]
        return any(keyword in text for keyword in keywords)

    def _render_daily_metric_lines(self, metrics):
        if not metrics:
            return ["- 无统计数据"]
        return [
            f"- 当前范围单量：{self._format_metric_int(metrics.get('orders', 0))}",
            f"- 当前范围销售额：¥{self._safe_float(metrics.get('sales', 0)):.2f}",
            f"- 退款预算：¥{self._safe_float(metrics.get('refund_budget_remaining', 0)):.2f}",
            f"- 品质退款：{metrics.get('quality_refund_count', 0)}单",
            f"- 其他退款：{metrics.get('other_refund_count', 0)}单",
            f"- 撤销品质：{metrics.get('canceled_quality_count', 0)}单",
            f"- 总退款率：{self._safe_float(metrics.get('total_refund_rate', 0)):.2f}%",
            f"- 售后总额：¥{self._safe_float(metrics.get('total_after_sales', 0)):.2f}",
            f"- 金额占比：{self._safe_float(metrics.get('refund_ratio', 0)):.2f}%",
            f"- 申请品质率：{self._safe_float(metrics.get('quality_apply_rate', 0)):.2f}%",
            f"- 实际品质率：{self._safe_float(metrics.get('quality_actual_rate', 0)):.2f}%",
            f"- 撤销率：{self._safe_float(metrics.get('quality_cancel_rate', 0)):.2f}%",
            f"- 最多原因：{metrics.get('top_refund_reason', '无数据')}（{metrics.get('top_reason_count', 0)}单）",
        ]

    def _build_plain_store_stats_summary_lines(self, snapshot):
        """把店铺统计转成适合日报参考的大白话摘要。"""
        lines = []
        for store in snapshot.get("stores", []) or []:
            store_name = store.get("store_name", "未知店铺")
            metrics = store.get("metrics", {}) or {}
            quality_count = int(metrics.get("quality_refund_count", 0) or 0)
            after_sales = self._safe_float(metrics.get("total_after_sales", 0))
            if quality_count <= 0 and after_sales <= 0:
                continue

            orders = self._safe_float(metrics.get("orders", 0))
            sales = self._safe_float(metrics.get("sales", 0))
            cancel_count = int(metrics.get("quality_cancel_count", metrics.get("canceled_quality_count", 0)) or 0)
            actual_count = int(metrics.get("quality_actual_count", max(quality_count - cancel_count, 0)) or 0)
            apply_rate = self._safe_float(metrics.get("quality_apply_rate", 0))
            actual_rate = self._safe_float(metrics.get("quality_actual_rate", 0))
            refund_ratio = self._safe_float(metrics.get("refund_ratio", 0))
            other_count = int(metrics.get("other_refund_count", 0) or 0)

            parts = [f"{store_name}根据上周单量预估"]
            if orders > 0:
                parts.append(f"当前范围约{self._format_metric_int(orders)}单")
            if quality_count > 0:
                parts.append(f"登记品质退款{quality_count}单，预估申请品质率{apply_rate:.2f}%")
                parts.append(f"已撤销{cancel_count}单，实际预计计入{actual_count}单，实际品质率约{actual_rate:.2f}%")
            if other_count > 0:
                parts.append(f"其他退款{other_count}单")
            if after_sales > 0:
                if sales > 0:
                    parts.append(f"品质售后金额约{after_sales:.2f}元，占销售额约{refund_ratio:.2f}%")
                else:
                    parts.append(f"品质售后金额约{after_sales:.2f}元")
            top_reason = str(metrics.get("top_refund_reason") or "").strip()
            top_count = int(metrics.get("top_reason_count", 0) or 0)
            if top_reason and top_reason != "无数据" and top_count > 0:
                parts.append(f"主要登记原因是{top_reason}{top_count}单")
            lines.append("，".join(parts) + "。")

        totals = snapshot.get("totals")
        if totals:
            total_quality = int(totals.get("quality_refund_count", 0) or 0)
            total_cancel = int(totals.get("quality_cancel_count", totals.get("canceled_quality_count", 0)) or 0)
            total_actual = int(totals.get("quality_actual_count", max(total_quality - total_cancel, 0)) or 0)
            total_after_sales = self._safe_float(totals.get("total_after_sales", 0))
            if total_quality > 0 or total_after_sales > 0:
                lines.insert(
                    0,
                    "整体来看，当前筛选范围登记品质退款"
                    f"{total_quality}单，已撤销{total_cancel}单，实际预计计入{total_actual}单；"
                    f"品质售后金额约{total_after_sales:.2f}元，申请品质率约{self._safe_float(totals.get('quality_apply_rate', 0)):.2f}%，"
                    f"实际品质率约{self._safe_float(totals.get('quality_actual_rate', 0)):.2f}%。"
                )

        return lines

    def _build_plain_stats_summary_text(self, snapshot):
        lines = self._build_plain_store_stats_summary_lines(snapshot)
        return "\n".join(lines) if lines else "当前筛选范围没有明显需要写进日报的品质退款或售后金额数据。"

    def _build_quality_refund_processing_summary(self, records):
        """汇总日报里的整体品质退款处理情况。"""
        quality_records = [
            record for record in records or []
            if self._is_quality_reason(record.get("reason"))
        ]
        total_count = len(quality_records)
        cancel_count = sum(1 for record in quality_records if record.get("cancel"))
        reject_success_count = sum(1 for record in quality_records if self._is_reject_success_record(record))
        not_cancelled_count = max(total_count - cancel_count - reject_success_count, 0)
        cancel_rate = (cancel_count / total_count * 100) if total_count else 0.0
        text_parts = [
            f"整体品质退款处理：共{total_count}单",
            f"撤销{cancel_count}单",
            f"撤销率{cancel_rate:.2f}%",
        ]
        if reject_success_count > 0:
            text_parts.append(f"驳回成功{reject_success_count}单")
        text_parts.append(f"未撤销/实际计入影响{not_cancelled_count}单")
        text = "，".join(text_parts) + "。"
        return {
            "total_quality_refund_count": total_count,
            "cancel_count": cancel_count,
            "cancel_rate": round(cancel_rate, 2),
            "reject_success_count": reject_success_count,
            "not_cancelled_effective_count": not_cancelled_count,
            "summary_text": text,
        }

    def _build_quality_refund_orders_by_store(self, snapshot, quality_orders):
        """按店铺组织品质退款订单，供日报按店铺紧跟订单号输出。"""
        stats_summary_by_store = {}
        for store in snapshot.get("stores", []) or []:
            store_name = store.get("store_name", "未知店铺")
            store_snapshot = {"stores": [store], "totals": None}
            summary_lines = self._build_plain_store_stats_summary_lines(store_snapshot)
            stats_summary_by_store[store_name] = summary_lines[0] if summary_lines else ""

        grouped = {}
        for order in quality_orders or []:
            store_name = order.get("store_name", "未知店铺")
            bucket = grouped.setdefault(store_name, {
                "store_name": store_name,
                "store_summary": stats_summary_by_store.get(store_name, ""),
                "quality_refund_orders": [],
            })
            bucket["quality_refund_orders"].append(order)

        return [
            grouped[store_name]
            for store_name in sorted(grouped.keys())
        ]

    def render_daily_work_summary_preview(self, records, user_text=""):
        """渲染工作总结弹窗里的本地数据预览。"""
        records = list(records or [])
        snapshot = self._build_local_summary_snapshot(records) if records else {
            "stores": [],
            "totals": None,
            "records_count": 0,
            "store_count": 0,
        }
        quality_orders = self._build_quality_refund_order_lines(records)
        plain_stats_lines = self._build_plain_store_stats_summary_lines(snapshot)
        include_unshipped_followup = self._should_include_unshipped_followup(user_text)
        unshipped_orders = self._build_unshipped_followup_orders(records) if include_unshipped_followup else []
        lines = [
            "## 工作总结数据预览",
            "",
            f"- 当前筛选：{self.get_current_filter_summary_text()}",
            f"- 退款登记记录：{len(records)}条",
            f"- 涉及店铺：{snapshot.get('store_count', 0)}个",
            "",
        ]
        if not records:
            lines.extend([
                "当前筛选范围暂无退款登记。仍可在下方填写今日工作内容，并让AI润色成工作总结。",
                "",
            ])

        if plain_stats_lines:
            lines.extend(["### 大白话统计摘要", ""])
            lines.extend([f"- {line}" for line in plain_stats_lines])
            lines.append("")

        for store in snapshot.get("stores", []):
            metrics = store.get("metrics", {})
            lines.extend([
                f"### {store.get('store_name', '未知店铺')}",
                "",
            ])
            lines.extend(self._render_daily_metric_lines(metrics))
            lines.append("")

        totals = snapshot.get("totals")
        if totals:
            lines.extend(["### 全部总和", ""])
            lines.extend(self._render_daily_metric_lines(totals))
            lines.append("")

        lines.extend([
            "### 品质退款订单",
            "",
        ])
        if quality_orders:
            for item in quality_orders[:80]:
                lines.append(
                    f"- {item['store_name']} | {item['order_no']} | {item['refund_reason']} | "
                    f"结果：{item['result']} | 备注：{item['notes'] or '无'}"
                )
            if len(quality_orders) > 80:
                lines.append(f"- 还有 {len(quality_orders) - 80} 条品质退款订单未在预览中展开。")
        else:
            lines.append("- 当前范围无品质退款订单。")

        if include_unshipped_followup:
            lines.extend(["", "### 未发货需跟踪订单", ""])
            if unshipped_orders:
                for item in unshipped_orders[:80]:
                    lines.append(
                        f"- {item['store_name']} | {item['order_no']} | 订单状态：{item['order_status']} | "
                        f"售后状态：{item['after_sale_status']} | 备注：{item['notes'] or '无'}"
                    )
                if len(unshipped_orders) > 80:
                    lines.append(f"- 还有 {len(unshipped_orders) - 80} 条未发货订单未在预览中展开。")
            else:
                lines.append("- 当前范围无未发货跟踪订单。")

        return "\n".join(lines)

    def _compact_daily_snapshot_for_ai(self, snapshot):
        stores = []
        for store in snapshot.get("stores", []) or []:
            stores.append({
                "store_name": store.get("store_name", ""),
                "metrics": store.get("metrics", {}),
                "real_reason_analysis": store.get("real_reason_analysis", {}),
                "quality_saved_reason_analysis": store.get("quality_saved_reason_analysis", {}),
            })
        return {
            "generated_at": snapshot.get("generated_at", ""),
            "filters": snapshot.get("filters", {}),
            "filter_summary": snapshot.get("filter_summary", ""),
            "records_count": snapshot.get("records_count", 0),
            "store_count": snapshot.get("store_count", 0),
            "stores": stores,
            "totals": snapshot.get("totals"),
            "overall_real_reason_analysis": snapshot.get("overall_real_reason_analysis", {}),
            "overall_quality_saved_reason_analysis": snapshot.get("overall_quality_saved_reason_analysis", {}),
        }

    def _build_daily_work_summary_payload(self, records, user_text):
        """整理发送给AI的日报/工作总结数据。"""
        records = list(records or [])
        snapshot = self._build_local_summary_snapshot(records) if records else {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "filters": self.get_current_filter_context(),
            "filter_summary": self.get_current_filter_summary_text(),
            "records_count": 0,
            "store_count": 0,
            "stores": [],
            "totals": None,
            "overall_real_reason_analysis": {},
            "overall_quality_saved_reason_analysis": {},
        }
        quality_orders = self._build_quality_refund_order_lines(records)
        quality_orders_by_store = self._build_quality_refund_orders_by_store(snapshot, quality_orders)
        plain_stats_summary = self._build_plain_stats_summary_text(snapshot)
        quality_processing_summary = self._build_quality_refund_processing_summary(records)
        include_unshipped_followup = self._should_include_unshipped_followup(user_text)
        unshipped_orders = self._build_unshipped_followup_orders(records) if include_unshipped_followup else []
        orders_by_store = {}
        for record in records:
            store_name = record.get("store_name", "未知店铺")
            bucket = orders_by_store.setdefault(store_name, [])
            if len(bucket) >= 120:
                continue
            bucket.append({
                "order_no": record.get("order_no", ""),
                "spec_code": str(record.get("spec_code") or "").strip(),
                "refund_reason": self._normalize_reason(record.get("reason")),
                "refund_amount": self._safe_float(record.get("refund_amount", 0)),
                "comp_amount": self._safe_float(record.get("comp_amount", 0)),
                "cancel": bool(record.get("cancel")),
                "reject": bool(record.get("reject")),
                "reject_result": self._display_reject_result_value(record),
                "order_status": str(record.get("order_status") or "").strip(),
                "after_sale_status": str(record.get("after_sale_status") or "").strip(),
                "notes": self._normalize_real_reason_note_text(record.get("notes", ""))[:240],
                "real_refund_reason": str(record.get("real_refund_reason") or "").strip(),
                "quality_not_cancelled_reason": str(record.get("quality_not_cancelled_reason") or "").strip(),
            })

        return {
            "task": "生成电商售后客服日报/工作总结",
            "user_manual_work_text": str(user_text or "").strip(),
            "quality_refund_rate_rule": (
                "品质退款率标红常见原因：品质退款率增加。申请品质退款的订单都会计入品质退款率；"
                "没有撤销、没有驳回成功的订单会保留影响。需要结合备注分析为什么退款、为什么没有撤销。"
            ),
            "plain_stats_summary": plain_stats_summary,
            "quality_refund_processing_summary": quality_processing_summary,
            "current_filter": self.get_current_filter_context(),
            "filter_summary": self.get_current_filter_summary_text(),
            "summary_snapshot": self._compact_daily_snapshot_for_ai(snapshot),
            "quality_refund_orders": quality_orders[:200],
            "quality_refund_orders_truncated_count": max(len(quality_orders) - 200, 0),
            "quality_refund_orders_by_store": quality_orders_by_store,
            "include_unshipped_followup": include_unshipped_followup,
            "unshipped_followup_orders": unshipped_orders[:200],
            "unshipped_followup_orders_truncated_count": max(len(unshipped_orders) - 200, 0),
            "orders_by_store": orders_by_store,
        }

    def generate_daily_work_summary_for_dialog(self, dialog):
        """调用AI生成工作总结并写回弹窗。"""
        progress_dialog = None
        try:
            if not self.ai_analyzer.api_key:
                self.show_api_settings_dialog()
                if not self.ai_analyzer.api_key:
                    return

            records = self.get_summary_source_records()
            user_text = dialog.get_user_text() if dialog else ""
            payload = self._build_daily_work_summary_payload(records, user_text)
            progress_dialog = QProgressDialog("正在调用AI生成工作总结...", None, 0, 100, dialog or self)
            progress_dialog.setWindowTitle("AI生成中")
            progress_dialog.setWindowModality(Qt.WindowModal)
            progress_dialog.setCancelButton(None)
            progress_dialog.show()
            QApplication.processEvents()

            progress_dialog.setValue(30)
            result = self.ai_analyzer.generate_daily_work_summary(payload)
            progress_dialog.setValue(100)
            progress_dialog.close()
            progress_dialog = None

            if dialog:
                dialog.set_result(result)
        except Exception as e:
            if progress_dialog is not None:
                progress_dialog.close()
            QMessageBox.critical(dialog or self, "工作总结生成失败", f"生成工作总结时发生错误：{e}")

    def show_daily_work_summary_debug_prompt(self, dialog):
        """显示工作总结将发送给AI的完整提示词与输入数据。"""
        try:
            records = self.get_summary_source_records()
            user_text = dialog.get_user_text() if dialog else ""
            payload = self._build_daily_work_summary_payload(records, user_text)
            system_prompt = self.ai_analyzer.get_daily_work_summary_prompt()
            user_content = json.dumps(payload, ensure_ascii=False, indent=2)
            debug_content = (
                "=== System Prompt ===\n"
                f"{system_prompt}\n\n"
                "=== User JSON Payload ===\n"
                f"{user_content}"
            )

            parent = dialog or self
            debug_dialog = QDialog(parent)
            debug_dialog.setWindowTitle("工作总结调试提示词")
            debug_dialog.resize(980, 760)

            layout = QVBoxLayout(debug_dialog)
            info_label = QLabel("以下是点击“AI生成专业总结”时将发送给AI的提示词和当前输入数据：")
            info_label.setWordWrap(True)
            info_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 13px; font-weight: bold;")
            layout.addWidget(info_label)

            debug_text = QTextEdit()
            debug_text.setReadOnly(True)
            debug_text.setFont(QFont("Consolas", 9))
            debug_text.setPlainText(debug_content)
            layout.addWidget(debug_text, 1)

            button_layout = QHBoxLayout()
            copy_btn = QPushButton("复制调试内容")
            close_btn = QPushButton("关闭")
            copy_btn.clicked.connect(lambda: QApplication.clipboard().setText(debug_content))
            close_btn.clicked.connect(debug_dialog.accept)
            button_layout.addStretch()
            button_layout.addWidget(copy_btn)
            button_layout.addWidget(close_btn)
            layout.addLayout(button_layout)

            debug_dialog.exec_()
        except Exception as e:
            QMessageBox.critical(dialog or self, "调试提示词", f"生成调试提示词失败：{e}")

    def _build_summary_debug_payload(self, snapshot):
        """构建总结调试信息。"""
        debug_data = {
            "generated_at": snapshot.get("generated_at", ""),
            "filter_summary": snapshot.get("filter_summary", ""),
            "records_count": snapshot.get("records_count", 0),
            "real_reason_categories": self.db.get_real_refund_reason_categories(active_only=False),
            "last_real_reason_category_debug": getattr(self, "last_real_reason_category_debug", {}),
            "last_real_reason_assignment_debug": getattr(self, "last_real_reason_assignment_debug", {}),
            "stores": [],
        }

        for store in snapshot.get("stores", []):
            debug_data["stores"].append({
                "store_name": store.get("store_name", ""),
                "metrics": store.get("metrics", {}),
                "real_reason_analysis": store.get("real_reason_analysis", {}),
                "quality_saved_reason_analysis": store.get("quality_saved_reason_analysis", {}),
            })
        return debug_data

    def show_summary_debug_info(self):
        """显示本地分类/归因与总结快照调试信息。"""
        try:
            snapshot = self.latest_summary_snapshot
            if not snapshot:
                records = self.get_summary_source_records()
                if not records:
                    QMessageBox.information(self, "提示", "当前筛选条件下没有数据可供调试")
                    return
                snapshot = self._build_local_summary_snapshot(records)

            debug_data = self._build_summary_debug_payload(snapshot)
            debug_content = json.dumps(debug_data, ensure_ascii=False, indent=2)

            dialog_parent = self.ai_window if hasattr(self, 'ai_window') and self.ai_window else self
            debug_dialog = QDialog(dialog_parent)
            debug_dialog.setWindowTitle("总结调试信息 - 本地分类与归因")
            debug_dialog.resize(980, 760)

            layout = QVBoxLayout(debug_dialog)
            info_label = QLabel("以下内容为本地分类配置、本地归因结果与当前总结快照的调试信息：")
            info_label.setStyleSheet("font-weight: bold; font-size: 14px; margin-bottom: 10px;")
            layout.addWidget(info_label)

            debug_text = QTextEdit()
            debug_text.setReadOnly(True)
            debug_text.setFont(QFont("Consolas", 9))
            debug_text.setPlainText(debug_content)
            layout.addWidget(debug_text)

            button_layout = QHBoxLayout()
            copy_btn = QPushButton("复制内容")
            close_btn = QPushButton("关闭")
            copy_btn.clicked.connect(lambda: self.copy_to_clipboard(debug_content))
            close_btn.clicked.connect(debug_dialog.accept)
            button_layout.addWidget(copy_btn)
            button_layout.addWidget(close_btn)
            layout.addLayout(button_layout)

            debug_dialog.exec_()
            self.ensure_ai_window_visible()
        except Exception as e:
            QMessageBox.critical(self, "调试错误", f"获取总结调试信息失败：{e}")

    def generate_summary_analysis(self):
        """生成当前筛选条件下的结构化总结分析。"""
        progress_dialog = None
        try:
            self.ensure_ai_window_visible()
            records = self.get_summary_source_records()
            if not records:
                QMessageBox.information(self, "提示", "当前筛选条件下没有数据可供总结")
                return

            QApplication.setOverrideCursor(Qt.WaitCursor)
            progress_dialog = QProgressDialog("正在本地计算总结...", None, 0, 100, self.ai_window if hasattr(self, 'ai_window') else self)
            progress_dialog.setWindowTitle("本地总结中")
            progress_dialog.setWindowModality(Qt.WindowModal)
            progress_dialog.setCancelButton(None)
            progress_dialog.show()
            QApplication.processEvents()

            progress_dialog.setLabelText("正在本地统计数据与备注率...")
            progress_dialog.setValue(10)
            snapshot = self._build_local_summary_snapshot(records)

            progress_dialog.setLabelText("正在保存总结历史...")
            progress_dialog.setValue(80)
            history_id = self.db.save_ai_summary_history(snapshot.get("filter_summary", ""), snapshot)
            self.latest_summary_history_id = history_id
            snapshot["history_id"] = history_id

            progress_dialog.setValue(100)
            progress_dialog.close()
            QApplication.restoreOverrideCursor()

            self.display_summary_snapshot(snapshot)
        except Exception as e:
            if progress_dialog is not None:
                progress_dialog.close()
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "总结失败", f"生成总结时发生错误：{e}")
        finally:
            self.ensure_ai_window_visible()

    def ai_analyze_data(self):
        """执行AI数据分析"""
        try:
            print("[DEBUG] 开始AI分析流程...")
            
            # 检查API配置
            if not self.ai_analyzer.api_key:
                print("[DEBUG] API Key未配置，显示设置对话框")
                self.show_api_settings_dialog()
                return
            
            print(f"[DEBUG] API配置检查通过，API URL: {self.ai_analyzer.api_url}")
            
            # 收集数据
            QApplication.setOverrideCursor(Qt.WaitCursor)
            print("[DEBUG] 开始收集分析数据...")
            analysis_data = self.collect_analysis_data()
            print(f"[DEBUG] 数据收集完成，订单数量: {len(analysis_data.get('orders', []))}")
            
            # 检查数据量
            if len(analysis_data.get("orders", [])) == 0:
                QApplication.restoreOverrideCursor()
                print("[DEBUG] 没有数据可供分析")
                QMessageBox.information(self, "提示", "当前筛选条件下没有数据可供分析")
                return
            
            # 显示进度对话框
            progress_dialog = QProgressDialog("正在分析数据...", "取消", 0, 100, self)
            progress_dialog.setWindowTitle("AI分析中")
            progress_dialog.setWindowModality(Qt.WindowModal)
            progress_dialog.setCancelButton(None)  # 移除取消按钮
            progress_dialog.show()
            
            # 在主线程中执行AI分析（避免多线程UI问题）
            QApplication.processEvents()  # 确保进度条显示
            progress_dialog.setValue(30)
            print("[DEBUG] 进度条设置到30%，开始执行AI分析...")
            
            # 执行AI分析
            result = self.ai_analyzer.analyze_data(analysis_data)
            print(f"[DEBUG] AI分析完成，结果长度: {len(result) if result else 0}")
            
            progress_dialog.setValue(100)
            progress_dialog.close()
            QApplication.restoreOverrideCursor()
            
            # 显示分析结果
            print("[DEBUG] 显示分析结果...")
            self.show_analysis_result(result)
            print("[DEBUG] AI分析流程完成")
            
        except Exception as e:
            QApplication.restoreOverrideCursor()
            print(f"[ERROR] AI分析过程中出现异常: {str(e)}")
            print(f"[ERROR] 异常类型: {type(e).__name__}")
            import traceback
            print(f"[ERROR] 详细堆栈信息:\n{traceback.format_exc()}")
            QMessageBox.critical(self, "AI分析失败", f"分析过程中出现错误：{str(e)}")

    def load_api_config(self):
        """加载API配置"""
        config = self.db.load_api_config()
        self.ai_analyzer.set_api_config(
            config["api_url"],
            config["api_key"],
            config["model"]
        )

    def show_api_settings_dialog(self):
        """显示API设置对话框"""
        dialog = APISettingsDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            settings = dialog.get_settings()
            self.ai_analyzer.set_api_config(
                settings["api_url"],
                settings["api_key"],
                settings["model"]
            )
            # 保存到数据库
            self.db.save_api_config(
                settings["api_url"],
                settings["api_key"],
                settings["model"]
            )
            QMessageBox.information(self, "成功", "API设置已保存")

    def show_analysis_result(self, result):
        """显示分析结果"""
        dialog = AnalysisResultDialog(result, self)
        dialog.exec_()
        
    def show_debug_info(self):
        """显示调试信息 - API输入内容"""
        try:
            print("[DEBUG] 开始收集调试信息...")
            
            # 收集当前分析数据
            analysis_data = self.collect_analysis_data()
            print(f"[DEBUG] 分析数据收集完成，类型: {type(analysis_data)}")
            
            # 构建完整的API请求数据
            messages = [
                {
                    "role": "system",
                    "content": self.ai_analyzer.system_prompt if hasattr(self.ai_analyzer, 'system_prompt') else """你是一名专业的电商售后客服主管，擅长数据分析、问题归因和给出改进建议。请基于以下退款数据，以专业、清晰、有条理的方式输出分析报告。"""
                },
                {
                    "role": "user",
                    "content": json.dumps(analysis_data, ensure_ascii=False, indent=2)
                }
            ]
            
            # 显示调试信息对话框
            debug_dialog = QDialog(self)
            debug_dialog.setWindowTitle("调试信息 - API输入内容")
            debug_dialog.resize(900, 700)
            
            layout = QVBoxLayout(debug_dialog)
            
            # 添加标签说明
            info_label = QLabel("本次AI分析将发送以下数据到API：")
            info_label.setStyleSheet("font-weight: bold; font-size: 14px; margin-bottom: 10px;")
            layout.addWidget(info_label)
            
            # 显示API输入内容
            debug_text = QTextEdit()
            debug_text.setReadOnly(True)
            debug_text.setFont(QFont("Consolas", 9))
            
            # 格式化显示内容
            # 计算订单总数和店铺数量
            total_orders = analysis_data.get('total_orders_count', 0)
            store_count = analysis_data.get('store_count', 0)
            
            debug_content = f"""=== 系统提示词 ===
{messages[0]['content']}

=== 用户数据 ===
{json.dumps(analysis_data, ensure_ascii=False, indent=2)}

=== 数据统计 ===
- 订单总数: {total_orders}
- 店铺数量: {store_count}
- 数据大小: {len(json.dumps(analysis_data))} 字符
- 预计Tokens消耗: 约 {int(len(json.dumps(analysis_data)) / 4)} tokens
"""
            
            debug_text.setPlainText(debug_content)
            layout.addWidget(debug_text)
            
            # 添加按钮
            button_layout = QHBoxLayout()
            copy_btn = QPushButton("复制内容")
            close_btn = QPushButton("关闭")
            
            copy_btn.clicked.connect(lambda: self.copy_to_clipboard(debug_content))
            close_btn.clicked.connect(debug_dialog.accept)
            
            button_layout.addWidget(copy_btn)
            button_layout.addWidget(close_btn)
            layout.addLayout(button_layout)
            
            debug_dialog.exec_()
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"[DEBUG] 错误详情: {error_details}")
            QMessageBox.critical(self, "调试错误", f"获取调试信息失败: {str(e)}\n\n详细错误信息已输出到终端")
    
    def copy_to_clipboard(self, text):
        """复制文本到剪贴板"""
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        QMessageBox.information(self, "成功", "调试信息已复制到剪贴板")

# ---------------------------- 高级主题设置对话框 --------------------------------
# ---------------------------- AI分析功能相关类 ----------------------------

class AIAnalyzer:
    """AI分析器：负责API调用和响应解析"""
    
    def __init__(self, api_url=None, api_key=None, model="deepseek-chat"):
        self.api_url = api_url or "https://api.deepseek.com/v1/chat/completions"
        self.api_key = api_key
        self.model = model
        
    def set_api_config(self, api_url, api_key, model):
        """设置API配置"""
        self.api_url = api_url
        self.api_key = api_key
        self.model = model

    def _request_completion(self, messages, temperature=0.7, max_tokens=4000, response_format=None):
        """统一的聊天补全请求。"""
        print(f"[DEBUG AIAnalyzer] 开始分析数据，API URL: {self.api_url}")
        if not self.api_key:
            raise ValueError("API Key未配置，请先设置API配置")

        payload = {
            "model": self.model,
            "messages": messages,
            "temperature": temperature,
            "max_tokens": max_tokens
        }
        if response_format:
            payload["response_format"] = response_format
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        }

        try:
            started_at = time.time()
            response = requests.post(self.api_url, json=payload, headers=headers, timeout=360)
            response.raise_for_status()
            result = response.json()
            if "choices" in result and result["choices"]:
                content = result["choices"][0]["message"]["content"]
                return content, {
                    "duration_ms": int((time.time() - started_at) * 1000),
                    "payload_size": len(json.dumps(payload, ensure_ascii=False)),
                    "status_code": response.status_code,
                    "raw_response_preview": str(content or "")[:1000],
                }
            raise ValueError("API返回数据格式异常")
        except requests.exceptions.RequestException as e:
            raise Exception(f"网络请求失败: {str(e)}")
        except json.JSONDecodeError as e:
            raise Exception(f"JSON解析失败: {str(e)}")
        except Exception as e:
            raise Exception(f"AI分析失败: {str(e)}")

    @staticmethod
    def _extract_json_content(content):
        """从模型回复中提取JSON文本。"""
        text = str(content or "").strip()
        fenced_match = re.search(r"```(?:json)?\s*(\{.*\}|\[.*\])\s*```", text, re.S)
        if fenced_match:
            return fenced_match.group(1)

        start = min(
            [index for index in [text.find("{"), text.find("[")] if index != -1],
            default=-1
        )
        end_obj = text.rfind("}")
        end_arr = text.rfind("]")
        end = max(end_obj, end_arr)
        if start != -1 and end != -1 and end > start:
            return text[start:end + 1]
        return text

    @staticmethod
    def _try_plain_text_structured_fallback(content, fallback_kind):
        text = str(content or "").strip()
        if not text or ("{" in text and "}" in text):
            if fallback_kind == "real_categories":
                return {"message": "AI返回空结果，已按无新增分类处理", "categories": []}
            if fallback_kind == "real_assignments":
                return {"message": "AI返回空结果，已按无归因结果处理", "assignments": []}
            return None
        if fallback_kind == "quality" and (
            ("品质问题归类" in text and "未撤销原因归类" in text)
            or text in ("无", "品质问题归类\n无\n未撤销原因归类\n无")
        ):
            return {
                "message": "AI返回了纯文本空结果，已自动按空分类处理",
                "quality_problem_categories": [],
                "not_cancelled_reason_categories": []
            }
        if fallback_kind == "other" and (
            ("半路退回" in text and "未明确备注" in text)
            or text in ("无", "半路退回：0（0.00%）\n未明确备注：0（0.00%）")
        ):
            halfway_match = re.search(r'半路退回[:：]\s*(\d+)', text)
            unclear_match = re.search(r'未明确备注[:：]\s*(\d+)', text)
            return {
                "message": "AI返回了纯文本空结果，已自动按统计文本兜底处理",
                "overall_categories": [],
                "spec_categories": [],
                "halfway_return_count": int(halfway_match.group(1)) if halfway_match else 0,
                "halfway_return_ratio": 0.0,
                "unclear_count": int(unclear_match.group(1)) if unclear_match else 0,
                "unclear_ratio": 0.0
            }
        if fallback_kind == "real_categories" and text in ("无", "暂无", "没有"):
            return {"message": "AI返回纯文本空结果，已按无新增分类处理", "categories": []}
        if fallback_kind == "real_assignments" and text in ("无", "暂无", "没有"):
            return {"message": "AI返回纯文本空结果，已按无归因结果处理", "assignments": []}
        return None

    def analyze_structured_json(self, system_prompt, payload, max_tokens=2000, fallback_kind=None):
        """请求AI并解析结构化JSON回复。"""
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False, separators=(',', ':'))}
        ]
        content, debug_meta = self._request_completion(
            messages,
            temperature=0.1,
            max_tokens=max_tokens,
            response_format={"type": "json_object"}
        )
        json_text = self._extract_json_content(content)
        try:
            parsed = json.loads(json_text)
            debug_meta["extracted_json_preview"] = str(json_text or "")[:1000]
            debug_meta["parse_error"] = ""
            return parsed, debug_meta
        except Exception as exc:
            debug_meta["extracted_json_preview"] = str(json_text or "")[:1000]
            debug_meta["parse_error"] = str(exc)
            fallback_result = self._try_plain_text_structured_fallback(content, fallback_kind)
            if fallback_result is not None:
                debug_meta["parse_error"] = f"{exc}；已使用纯文本兜底解析"
                return fallback_result, debug_meta
            raise Exception(f"AI返回的JSON无法解析：{exc}；原始响应预览：{str(content or '')[:200]}")

    def analyze_data(self, analysis_data):
        """分析数据并返回AI响应"""
        messages = [
            {
                "role": "system",
                "content": """你是一名专业的电商售后客服主管，擅长数据分析、问题归因和给出改进建议。请基于以下退款数据，以专业、清晰、有条理的方式输出分析报告。

## 重要规则：
1. **退款类型识别**：如果退款原因为"其他"，则不是品质退款，需要重点分析用户备注中的具体问题
2. **数据准确性**：请严格基于提供的统计数据进行分析，不要自行计算或推测数据
3. **山药产品分析**：我们的产品是山药，请根据备注中的产品型号（如605=50-60cm 5斤装）分析具体问题
4. **完整输出**：确保分析报告完整，不要中途截断
5. **售后金额理解**：售后金额包括退款金额和打款补偿金额的总和

## 数据核对要求：
- **退款率计算**：退款率 = (退款订单数 / 订单量) × 100%，请使用提供的订单量和退款订单数
- **数据来源**：所有统计数据必须来自"refund_stats"字段，不要自行计算
- **订单详情**：订单详情来自"orders"字段，用于分析具体问题和备注
- **售后金额**：售后金额 = 退款金额 + 补偿金额，请使用统计板块的准确数据

## 分析要求：
- 基于实际数据进行分析，不要猜测或虚构数据
- 如果备注中包含产品型号，请分析哪种规格的山药问题最多
- 对于"其他"类型的退款，重点分析备注中的具体问题
- 确保所有统计数据和结论都基于提供的数据
- 理解"其他"退款原因：这是客服的正确操作，避免品质退款扣分
- 当前是测试阶段，部分订单可能没有备注，这是正常现象
- 输出完整的分析报告，包括：总体概况、分店铺分析、退款原因分析、产品问题分析、售后处理分析、问题总结与建议

## 输出格式要求：
请以Markdown格式输出，确保内容完整不截断。如果数据量较大，请优先保证分析结论的完整性。

请输出完整、准确的分析报告。"""
            },
            {
                "role": "user",
                "content": json.dumps(analysis_data, ensure_ascii=False, indent=2)
            }
        ]
        return self._request_completion(messages, temperature=0.7, max_tokens=4000)

    @staticmethod
    def get_daily_work_summary_prompt():
        return """你具备电商售后主管的判断能力，请把当天售后登记、店铺统计和用户补充事项整理成一份适合发到微信工作群的日报。

写作风格参考：
4.23号 处理品退8条 6条撤销的 1条已线上留言核实优惠补偿明天再次跟进 1条已线上驳回成功 给顾客协商补偿
5.6号 处理品退9条 7条撤销的售后 1条打电话让撤销售后给小额打款再忙正常驳回 1条转给我已经退款 平台极速退款
5.7号 处理品退8条 5条撤销的售后 2条驳回的 打电话不会平台留言不回 还有1条申请错了 已记录未撤销原因；分析各店铺领航员标红原因，查看各个品质退款订单，完善退款工具功能

输出要求：
1. 语气参考上面的日报，像我本人在汇报工作，但比原文更顺一点、专业一点。
2. 可以带日期标题。内容用文字段落或短行，不要表格，不要机械模板，不要写成正式公文。
3. 重点围绕品质退款、撤销/驳回/补偿协商、领航员品质退款率、线上售后问题处理、售后流程和方案优化。
4. 不要说“我是售后主管”，不要写“制表人/汇报人/AI/系统/提示词/数据字段”等字样。
5. 不要虚构订单、店铺、金额或处理结果；订单号必须原样保留。
6. 用户没有提到的事项不要生成。物流、发货、丢件、发货超时这类内容，只有 user_manual_work_text 明确提到时才可以润色输出；否则不要主动写。

内容取舍：
- 优先参考 plain_stats_summary，把本地已经整理好的店铺品质退款数据转成大白话写进日报。
- 基础数据要说人话：根据上周单量预估，今天某店铺登记品质退款几单，预估申请品质率多少，撤销几单，实际预计计入几单，实际品质率约多少，品质售后金额和占比大概多少。
- 整体品质退款处理情况必须参考 quality_refund_processing_summary，写清楚：品质退款总单数、撤销单数、撤销率、驳回成功单数、未撤销/实际计入影响单数。
- 所有退款相关表述尽量带“品质”字眼，例如“品质退款”“品质售后金额”“品质售后处理”，避免和普通退款/普通售后混淆。
- 有品质退款的店铺可以提一嘴数据；没有明显品质退款或品质售后金额的店铺不要强行写。
- 店铺输出必须按 quality_refund_orders_by_store 分组，不允许先整体说完所有店铺再统一列订单号。每个店铺先写该店铺品质退款总况，下面立刻紧跟该店铺的品质退款订单号。
- 每个店铺建议格式：
  店铺名：登记品质退款X单，已撤销X单，实际X单计入。品质售后金额X元，占销售额X%。主要原因是xxx。
  单号：xxx 处理结果：已撤销/未撤销/驳回成功 备注分析：xxx
  单号：xxx 处理结果：xxx 备注分析：xxx
- 品质退款订单不要复述备注原文，要提炼成原因分析和处理结果，例如：已撤销、驳回成功、平台极速退款、客户未回复、优惠补偿协商、申请错类等。
- 未撤销订单默认就是已经退款成功或平台介入/平台退款导致退款成功，不属于后续需要继续跟进的订单；只说明未撤销原因和实际计入影响，不要写“明天继续跟进”，除非 user_manual_work_text 明确要求继续处理某个订单。
- 品质退款订单明细必须每个单号单独一行，顺序固定为：先单号，再处理结果，再备注分析。
  固定格式：单号：xxx 处理结果：已撤销/未撤销/驳回成功 备注分析：xxx
  如果未撤销，备注分析里必须说明未撤销原因；如果已撤销或驳回成功，不要硬写“未撤销原因：无”。不要表格，不要把多个单号写在同一行。
- 如果某店铺没有驳回成功订单，不要写“没有驳回”“无驳回”“驳回成功0单”。
- 如果用户提到领航员标红、品质退款率标红、品质率异常，要解释为：申请品质退款会计入品质退款率，没有撤销、没有驳回成功的会保留影响；再结合订单情况说明哪些店铺需要重点看。
- 如果 include_unshipped_followup=true，再写用户要求跟进的未发货/物流类事项；否则完全不要写这类内容。
- 结尾写“后续优化”，不要写“后续跟进”。重点写优化售后图片、售后话术、优惠补偿方案、品质退款撤销/驳回流程、领航员品质退款率监控。

整体比简单日报更完整一些，但仍要适合微信发送，尽量控制在几段到十来行。不要输出代码块，不要提技术词。"""

    def generate_daily_work_summary(self, payload):
        """生成日报/工作总结文本。"""
        messages = [
            {
                "role": "system",
                "content": self.get_daily_work_summary_prompt()
            },
            {
                "role": "user",
                "content": json.dumps(payload, ensure_ascii=False, indent=2)
            }
        ]
        content, _debug_meta = self._request_completion(messages, temperature=0.35, max_tokens=6000)
        return content

    @staticmethod
    def get_real_reason_category_generation_prompt():
        return """只返回JSON。

任务：根据输入里的 notes 数组，总结出一套可长期复用的“真实退款原因”大类。

规则：
1. 只看备注内容，不参考其他字段。
2. 先整体归纳，再输出少量稳定的大类，不要把每条备注都单独变成类别。
3. 优先复用 existing_categories；只有确实无法归入时才新增。
4. 分类名称必须简短、稳定、适合长期复用，例如“腐烂变质”“发芽了”“断裂破损”“半路退回”“客户主观不想要”“未明确备注”。
5. 如果某些备注无法判断，也要归入“未明确备注”。
6. 不要输出解释、前言、Markdown、代码块。
7. 即使没有可新增分类，也必须返回合法JSON。

固定输出：
{
  "message":"一句话总结",
  "categories":[
    {"name":"腐烂变质","examples":["长毛了","烂掉了"]},
    {"name":"发芽了","examples":["发芽","长芽了"]}
  ]
}"""

    @staticmethod
    def get_real_reason_assignment_prompt():
        return """只返回JSON。

任务：根据输入里的 notes 数组，把每条备注归因到 existing_categories 中最匹配的一个真实退款原因。

规则：
1. 只看备注内容，不参考其他字段。
2. category 必须是 existing_categories 里的一个；如果无法判断，输出“未归因”。
3. detail 用一句很短的话概括归因依据，最多20字。
4. 不要输出解释、前言、Markdown、代码块。
5. 即使所有备注都无法判断，也必须返回合法JSON。

固定输出：
{
  "message":"一句话总结",
  "assignments":[
    {"index":0,"category":"腐烂变质","detail":"长毛发霉类"},
    {"index":1,"category":"未归因","detail":"备注无法判断"}
  ]
}"""

    @staticmethod
    def get_manual_real_reason_assignment_prompt():
        return """只返回JSON。

任务：把 notes 数组里的每条备注归因到真实退款原因分类。

规则：
1. 对订单只看备注内容，不参考订单号、店铺、金额、日期等字段。
2. existing_categories 是当前已有的真实退款原因分类名称，优先从里面选择最合适的一类。
3. 不要套用本地关键词规则；你需要根据备注语义自行判断。
4. 如果某条备注确实不属于现有分类，可以创建一个简短稳定的新分类，并在 new_categories 里返回。
5. 如果备注完全无法判断，category 返回“未归因”。
6. detail 用一句很短的话说明依据，最多20字。
7. 不要输出解释、前言、Markdown、代码块。

固定输出：
{
  "message":"一句话总结",
  "new_categories":["新分类名"],
  "assignments":[
    {"index":0,"category":"腐烂变质","detail":"备注提到腐烂"},
    {"index":1,"category":"新分类名","detail":"现有分类不匹配"}
  ]
}"""

    def analyze_real_reason_categories(self, payload):
        """基于备注生成/补充真实退款原因分类。"""
        return self.analyze_structured_json(
            self.get_real_reason_category_generation_prompt(),
            payload,
            max_tokens=1200,
            fallback_kind="real_categories"
        )

    def analyze_real_reason_assignments(self, payload):
        """基于备注为订单归因真实退款原因。"""
        return self.analyze_structured_json(
            self.get_real_reason_assignment_prompt(),
            payload,
            max_tokens=1800,
            fallback_kind="real_assignments"
        )

    def analyze_manual_real_reason_assignments(self, payload):
        """手动归因窗口内的AI归因：只看备注，允许新增分类。"""
        return self.analyze_structured_json(
            self.get_manual_real_reason_assignment_prompt(),
            payload,
            max_tokens=8000,
            fallback_kind="real_assignments"
        )

    @staticmethod
    def get_quality_unreversed_notes_prompt():
        return """只返回JSON。

输入数据只有 records 数组，每条只有 spec_code、notes、count。
任务：看“规格编码 + 备注”，做未撤销品质退款的汇总归类。

规则：
1. 先整体看完，再归类，不要一条备注生成一个类别。
2. 同义词归并，例如“长毛了/发霉了/腐坏了/烂了”归为“腐烂变质”，“长芽了/发芽了/冒芽了”归为“发芽了”。
3. 备注里既可能有品质问题，也可能有“为什么没撤销”的原因，要分别统计。
4. “客服处理不及时”统一输出为“接线客服处理不及时”。
5. 每类 examples 最多 2 条。
6. 即使没有结果，也必须返回合法JSON，数组为空，message写“无可归类备注”。
7. 不要输出解释、前言、Markdown、代码块。

固定输出：
{
  "message":"一句话总结",
  "quality_problem_categories":[
    {"name":"腐烂变质","count":1,"ratio":12.34,"examples":["605长毛了"]}
  ],
  "not_cancelled_reason_categories":[
    {"name":"接线客服处理不及时","count":1,"ratio":12.34,"examples":["..."]}
  ]
}"""

    @staticmethod
    def get_other_reason_notes_prompt():
        return """只返回JSON。

输入数据只有 records 数组，每条只有 spec_code、notes、count。
任务：看“规格编码 + 备注”，分析“退款原因=其他”的真实退款原因，按规格编码和总体分别汇总。

规则：
1. 先整体看完，再统一归类，不要一条备注生成一个类别。
2. 本地已经把“已拦截”统计为“半路退回”，不要再重复统计这部分。
3. “长毛了/发霉了/腐坏了/烂了”归为“腐烂变质”。
4. “长芽了/发芽了/冒芽了”归为“发芽了”。
5. 没有明确问题的归为“未明确备注”。
6. 每类 examples 最多 2 条。
7. 即使没有结果，也必须返回合法JSON，数组为空。
8. 不要输出解释、前言、Markdown、代码块。

固定输出：
{
  "message":"一句话总结",
  "overall_categories":[
    {"name":"腐烂变质","count":2,"ratio":50.0,"examples":["605长毛了"]}
  ],
  "spec_categories":[
    {"spec":"605","categories":[{"name":"腐烂变质","count":2,"ratio":66.67}]}
  ],
  "halfway_return_count":0,
  "halfway_return_ratio":0.0,
  "unclear_count":0,
  "unclear_ratio":0.0
}"""

    def analyze_quality_unreversed_notes(self, payload):
        """分析未撤销品质退款备注。"""
        return self.analyze_structured_json(self.get_quality_unreversed_notes_prompt(), payload, max_tokens=800, fallback_kind="quality")

    @staticmethod
    def get_quality_refund_order_reason_prompt():
        return """只返回JSON。

任务：根据 records 数组分析退款成功且未撤销的品质退款订单，识别“为什么最终退款成功/为什么没有撤销”的原因。

规则：
1. 业务前提：输入记录都已经退款成功且未撤销；不要判断是否未撤销，只判断导致退款成功/没撤销的原因。
2. 必须先整体阅读所有 records，判断本批次哪些未撤销原因出现较多，再创建少量本批次适用的分类；不要一条备注创建一个类别。
3. 分类可以比以前更长、更像大白话，但要稳定可复用，例如“已正常协商但客户拒绝沟通后平台介入退款”“已给退货退款或优惠补偿方案但平台介入退款”“客户坚持全额仅退款且平台快速介入”。
4. existing_categories 是数据库里已有分类，必须优先从里面挑选对应分类；如果订单原因和某个已有分类的核心意思相似度约50%以上，就必须复用已有分类，不要新建近似分类。只有现有分类无法覆盖主要意思、差异明显超过约50%时，才允许新增本批次高频原因分类。
5. 只识别“为什么没有撤销/为什么最终退款成功”的相关备注，忽略普通品质问题本身，例如腐烂、发芽、破损等只说明产品问题但没说明未撤销原因时，不要当作未撤销原因。
6. 不要只按关键词判断，要理解备注里的处理经过：客服是否主动电话/留言沟通、是否让客户补充照片或数量、是否给了退货退款/部分补偿/优惠方案、客户是否拒绝协商或直接申请平台介入、平台是否快速退款。
7. 平台介入本身不等于“驳回失败导致退款”。只有备注明确写了“驳回失败”“驳回退款导致退款”“驳回后平台退款”“驳回不成功”等，才归为“驳回失败导致退款”。
8. 如果备注显示客服已正常电话或留言沟通，但客户拒绝说明、拒绝协商、要求全额退款、挂断电话，随后平台介入或快速退款，应归为类似“已正常协商但客户拒绝沟通后平台介入退款”，不要归为驳回失败。
9. 如果备注显示客服已提出退货退款、部分优惠、补偿等处理方案，但客户仍申请平台介入并退款，应归为类似“已给退货退款或优惠补偿方案但平台介入退款”。
10. “多多视频导致的售后”“多多视频导致退款”归为“多多视频售后导致退款”，除非同一备注里同时明确写了驳回失败/驳回后平台退款，则优先归为“驳回失败导致退款”。
11. “客户仅退款/已申请仅退款”归为“客户已申请仅退款”；“客服超时/未及时处理/接线慢”归为“接线客服处理不及时”。
12. 参考业务规则：品质退款会影响品质退款率；平台介入可能带来纠纷退款风险。分析时要关注是否已主动服务、及时沟通、避免误会、给出合适的退货退款或补偿方案；必要时在平台判责前主动同意合理退款可以降低纠纷风险。
13. 必须给每条 records 返回一条 assignments，index 必须对应输入记录的 index。
14. 新增分类要稳定、可复用、能覆盖多条类似订单；不要为单条订单造过细分类，也不要把已有分类换一种说法重新创建。
15. not_cancelled_reason 表示为什么没有撤销或为什么最终保留品质退款，例如“驳回失败导致退款”“多多视频售后导致退款”“接线客服处理不及时”“客户已申请仅退款”“已正常协商但客户拒绝沟通后平台介入退款”“已给退货退款或优惠补偿方案但平台介入退款”。
16. 仅当备注完全看不出任何导致退款成功/未撤销的线索时，才返回“未明确说明未撤销原因”。
17. detail 写客观依据和经过，最多80字，可以大白话总结；禁止照抄整段备注，禁止写“备注提到xxx”“备注写了xxx”这类废话。
18. 备注为空、备注只描述品质问题、或备注没有明确说明未撤销原因时，detail 必须返回空字符串。
19. 示例判断：
- “拍照片发霉2根，客服打电话问是全部发霉还是只有两根，客户不说明数量、要求退全款、挂断，平台介入退款不超过5分钟”应归为“已正常协商但客户拒绝沟通后平台介入退款”，detail 写“客服核实数量但客户拒绝沟通并要求全退，随后平台快速介入退款”。
- “说是假货、臭了、难吃，客服协商退货退款或者优惠20元，平台介入退款”应归为“已给退货退款或优惠补偿方案但平台介入退款”，detail 写“客服已给退货退款或优惠补偿方案，客户仍走平台介入退款”。
20. 不要输出解释、前言、Markdown、代码块。

固定输出：
{
  "message":"一句话总结",
  "categories":[
    {"name":"已正常协商但客户拒绝沟通后平台介入退款"}
  ],
  "assignments":[
    {"index":0,"not_cancelled_reason":"已正常协商但客户拒绝沟通后平台介入退款","detail":"客服核实情况但客户拒绝继续沟通并要求全退，随后平台快速介入退款"}
  ]
}"""

    def analyze_quality_refund_order_reasons(self, payload):
        """逐单分析未撤销品质退款原因与未撤销原因。"""
        return self.analyze_structured_json(
            self.get_quality_refund_order_reason_prompt(),
            payload,
            max_tokens=5000,
            fallback_kind=None
        )

    def analyze_other_reason_notes(self, payload):
        """分析其他原因备注的真实问题分布。"""
        return self.analyze_structured_json(self.get_other_reason_notes_prompt(), payload, max_tokens=800, fallback_kind="other")


class QualityRefundReasonDialog(QDialog):
    """当前筛选范围内的品质退款查看与AI未撤销原因分析窗口。"""

    HEADERS = ["店铺", "订单号", "规格编码", "退款原因", "备注", "未撤销原因", "分析说明"]

    def __init__(self, main_window, records, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self.records = list(records or [])
        self.visible_records = []
        self.column_filters = {}
        self._loading = False
        self.setup_ui()
        self.populate_table()

    def setup_ui(self):
        self.setWindowTitle("查看品质退款")
        self.resize(1080, 720)
        layout = QVBoxLayout(self)

        self.info_label = QLabel("")
        self.info_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        layout.addWidget(self.info_label)

        self.table = QTableWidget()
        self.table.setColumnCount(len(self.HEADERS))
        self.table.setHorizontalHeaderLabels(self.HEADERS)
        self.table.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.EditKeyPressed
            | QAbstractItemView.AnyKeyPressed
        )
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setWordWrap(False)
        self.table.itemChanged.connect(self.on_item_changed)
        self.table.cellClicked.connect(self.on_cell_clicked)
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        header = self.table.horizontalHeader()
        header.sectionClicked.connect(self.show_filter_menu)
        header.setStretchLastSection(False)
        header.setMinimumSectionSize(70)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.Stretch)
        layout.addWidget(self.table)

        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.summary_text.setMaximumHeight(150)
        layout.addWidget(self.summary_text)

        button_layout = QHBoxLayout()
        self.analyze_btn = QPushButton("AI分析未撤销原因")
        self.analyze_btn.clicked.connect(self.run_ai_analysis)
        clear_ai_btn = QPushButton("清空AI结果")
        clear_ai_btn.clicked.connect(self.clear_ai_results)
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(self.analyze_btn)
        button_layout.addWidget(clear_ai_btn)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)

    def set_records(self, records):
        self.records = list(records or [])
        self.populate_table()

    def _header_text(self, col):
        suffix = " *" if col in self.column_filters else " ▼"
        return f"{self.HEADERS[col]}{suffix}"

    def _update_header_labels(self):
        self.table.setHorizontalHeaderLabels([self._header_text(col) for col in range(len(self.HEADERS))])

    def _filter_value_for_record(self, record, col):
        if col == 0:
            return str(record.get("store_name") or "").strip() or "-"
        if col == 1:
            return str(record.get("order_no") or "").strip() or "-"
        if col == 2:
            return str(record.get("spec_code") or "").strip() or "-"
        if col == 3:
            return str(record.get("reason") or "").strip() or "-"
        if col == 4:
            return "有备注" if str(record.get("notes") or "").strip() else "无备注"
        if col == 5:
            return str(record.get("quality_not_cancelled_reason") or "").strip() or "未明确说明未撤销原因"
        if col == 6:
            detail = self._display_detail(record)
            return detail if detail != "-" else "无分析说明"
        return "-"

    def _record_matches_filters(self, record, ignore_col=None):
        for col, expected in self.column_filters.items():
            if col == ignore_col:
                continue
            if self._filter_value_for_record(record, col) != expected:
                return False
        return True

    def _apply_filters(self):
        return [record for record in self.records if self._record_matches_filters(record)]

    def _records_for_filter_options(self, col):
        return [record for record in self.records if self._record_matches_filters(record, ignore_col=col)]

    def show_filter_menu(self, col):
        if col < 0 or col >= len(self.HEADERS):
            return
        menu = QMenu(self)
        all_action = QAction("全部", self)
        all_action.setCheckable(True)
        all_action.setChecked(col not in self.column_filters)
        all_action.triggered.connect(lambda _checked=False, col=col: self.clear_column_filter(col))
        menu.addAction(all_action)
        menu.addSeparator()

        values = sorted({
            self._filter_value_for_record(record, col)
            for record in self._records_for_filter_options(col)
        })
        for value in values:
            action = QAction(value, self)
            action.setCheckable(True)
            action.setChecked(self.column_filters.get(col) == value)
            action.triggered.connect(lambda _checked=False, col=col, value=value: self.set_column_filter(col, value))
            menu.addAction(action)

        header = self.table.horizontalHeader()
        pos = header.mapToGlobal(QPoint(header.sectionPosition(col), header.height()))
        menu.exec_(pos)

    def set_column_filter(self, col, value):
        self.column_filters[col] = value
        self.populate_table()

    def clear_column_filter(self, col):
        self.column_filters.pop(col, None)
        self.populate_table()

    def _make_item(self, text, editable=False):
        item = QTableWidgetItem(str(text or ""))
        if editable:
            item.setFlags(item.flags() | Qt.ItemIsEditable)
        else:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        return item

    def _display_value(self, value):
        text = str(value or "").strip()
        return text if text else "-"

    def _display_detail(self, record):
        if not self.main_window:
            return self._display_value(record.get("quality_refund_reason_detail", ""))
        detail = self.main_window._normalize_quality_not_cancelled_detail(
            record.get("quality_refund_reason_detail", ""),
            record.get("quality_not_cancelled_reason", ""),
            record.get("notes", ""),
        )
        return self._display_value(detail)

    def populate_table(self):
        self._loading = True
        self.visible_records = self._apply_filters()
        self._update_header_labels()
        self.table.setRowCount(len(self.visible_records))
        for row, record in enumerate(self.visible_records):
            values = [
                record.get("store_name", ""),
                record.get("order_no", ""),
                str(record.get("spec_code") or "").strip() or "-",
                record.get("reason", ""),
                record.get("notes", ""),
                self._display_value(record.get("quality_not_cancelled_reason", "")),
                self._display_detail(record),
            ]
            for col, value in enumerate(values):
                item = self._make_item(value, editable=(col in (2, 4, 5, 6)))
                if record.get("id") not in (None, ""):
                    item.setData(Qt.UserRole, int(record.get("id")))
                if col == 2:
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setToolTip(str(record.get("spec_name") or "").strip() or "未识别规格名称")
                elif col == 4 and str(value or "").strip():
                    item.setToolTip(str(value or ""))
                elif col == 6 and str(value or "").strip() != "-":
                    item.setToolTip(str(value or ""))
                self.table.setItem(row, col, item)
        self.info_label.setText(f"当前显示 {len(self.visible_records)} 条 / 共 {len(self.records)} 条")
        self._loading = False
        self.update_summary_text()

    def on_cell_clicked(self, row, col):
        if col != 1 or row < 0 or row >= self.table.rowCount():
            return
        item = self.table.item(row, 1)
        order_no = item.text().strip() if item else ""
        if not order_no:
            return
        QApplication.clipboard().setText(order_no)
        self.info_label.setText(f"已复制订单号：{order_no}")

    def on_cell_double_clicked(self, row, col):
        if col in (2, 4, 5, 6):
            item = self.table.item(row, col)
            if item:
                self.table.editItem(item)

    def on_item_changed(self, item):
        if self._loading or not self.main_window:
            return
        row = item.row()
        col = item.column()
        if col not in (2, 4, 5, 6):
            return
        if row < 0 or row >= len(self.visible_records):
            return
        record = self.visible_records[row]
        record_id = record.get("id")
        if record_id in (None, ""):
            return
        new_value = str(item.text() or "").strip()
        if new_value == "-":
            new_value = ""

        if col == 2:
            self.save_spec_code(record_id, new_value)
        elif col == 4:
            self.save_notes(record_id, new_value)
        elif col == 5:
            self.save_not_cancelled_reason(record_id, new_value)
        elif col == 6:
            self.save_analysis_detail(record_id, new_value)

    def _refresh_after_record_update(self, record_id, message="", refresh_statistics=True):
        if hasattr(self.main_window, "_refresh_row_by_record_id"):
            self.main_window._refresh_row_by_record_id(record_id, refresh_statistics=refresh_statistics)
        updated_records = self.main_window._get_quality_refund_reason_records()
        self.records = list(updated_records or [])
        self.populate_table()
        if message:
            self.info_label.setText(message)

    def save_spec_code(self, record_id, spec_code):
        record = next((item for item in self.records if item.get("id") == record_id), None)
        old_value = str(record.get("spec_code") or "").strip() if record else ""
        if spec_code == old_value:
            return
        if not self.main_window.db.update_record_partial(record_id, spec_code=spec_code):
            QMessageBox.warning(self, "保存失败", "规格编码保存失败")
            self.populate_table()
            return
        self._refresh_after_record_update(record_id, "规格编码已保存")

    def save_notes(self, record_id, notes):
        record = next((item for item in self.records if item.get("id") == record_id), None)
        old_value = str(record.get("notes") or "").strip() if record else ""
        if notes == old_value:
            return
        if not self.main_window.db.update_record_partial(record_id, notes=notes):
            QMessageBox.warning(self, "保存失败", "备注保存失败")
            self.populate_table()
            return
        self._refresh_after_record_update(record_id, "备注已保存，未撤销原因已清空，请重新AI分析")

    def _save_manual_quality_fields(self, record_id, **fields):
        record = next((item for item in self.records if item.get("id") == record_id), None)
        notes = record.get("notes", "") if record else ""
        updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        payload = {
            "quality_refund_reason_note_hash": self.main_window._build_real_reason_note_hash(notes),
            "quality_refund_reason_updated_at": updated_at,
        }
        payload.update(fields)
        return self.main_window.db.update_record_partial(record_id, **payload)

    def save_not_cancelled_reason(self, record_id, reason):
        record = next((item for item in self.records if item.get("id") == record_id), None)
        old_value = str(record.get("quality_not_cancelled_reason") or "").strip() if record else ""
        if reason == old_value:
            return
        if not self._save_manual_quality_fields(record_id, quality_not_cancelled_reason=reason):
            QMessageBox.warning(self, "保存失败", "未撤销原因保存失败")
            self.populate_table()
            return
        if reason:
            existing = [
                item.get("category_name", "")
                for item in self.main_window.db.get_quality_not_cancelled_reason_categories(active_only=False)
                if str(item.get("category_name") or "").strip()
            ]
            if reason not in existing:
                self.main_window.db.save_quality_not_cancelled_reason_categories(existing + [reason])
        self._refresh_after_record_update(record_id, "未撤销原因已保存")

    def save_analysis_detail(self, record_id, detail):
        record = next((item for item in self.records if item.get("id") == record_id), None)
        old_value = str(record.get("quality_refund_reason_detail") or "").strip() if record else ""
        if detail == old_value:
            return
        if not self._save_manual_quality_fields(record_id, quality_refund_reason_detail=detail):
            QMessageBox.warning(self, "保存失败", "分析说明保存失败")
            self.populate_table()
            return
        self._refresh_after_record_update(record_id, "分析说明已保存")

    def clear_ai_results(self):
        if not self.main_window:
            return
        target_records = [
            record for record in self.records
            if record.get("id") not in (None, "")
        ]
        if not target_records:
            QMessageBox.information(self, "提示", "当前窗口没有可清空的订单")
            return
        reply = QMessageBox.question(
            self,
            "清空AI结果",
            f"将清空当前窗口内 {len(target_records)} 条订单的未撤销原因和分析说明，是否继续？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        cleared_count = 0
        for record in target_records:
            record_id = record.get("id")
            if self.main_window.db.update_quality_not_cancelled_reason(record_id, "", "", "", ""):
                cleared_count += 1
                if hasattr(self.main_window, "_refresh_row_by_record_id"):
                    self.main_window._refresh_row_by_record_id(record_id, refresh_statistics=False)

        self.main_window.load_table_data(force_reload=True)
        self.set_records(self.main_window._get_quality_refund_reason_records())
        self.info_label.setText(f"已清空当前窗口AI结果：{cleared_count} 条")

    def update_summary_text(self):
        grouped = {}
        for record in self.visible_records:
            store_name = str(record.get("store_name") or "未知店铺")
            store = grouped.setdefault(store_name, {"not_cancelled": {}})
            not_cancelled_reason = str(record.get("quality_not_cancelled_reason") or "").strip() or "未明确说明未撤销原因"
            store["not_cancelled"][not_cancelled_reason] = store["not_cancelled"].get(not_cancelled_reason, 0) + 1

        lines = []
        for store_name in sorted(grouped.keys()):
            lines.append(f"{store_name}")
            not_cancelled_items = sorted(grouped[store_name]["not_cancelled"].items(), key=lambda item: (-item[1], item[0]))
            lines.append("  未撤销原因：" + ("；".join(f"{name} {count}单" for name, count in not_cancelled_items) if not_cancelled_items else "暂无"))
        self.summary_text.setPlainText("\n".join(lines) if lines else "当前没有符合条件的品质退款订单。")

    def run_ai_analysis(self):
        if not self.main_window:
            return
        self.main_window.analyze_quality_refund_reasons_for_dialog(self)


class APISettingsDialog(QDialog):
    """API设置对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.setup_ui()
        
    def setup_ui(self):
        """设置界面"""
        self.setWindowTitle("AI API设置")
        self.setFixedSize(500, 300)
        
        layout = QVBoxLayout(self)
        
        # API地址
        api_url_layout = QHBoxLayout()
        api_url_label = QLabel("API地址:")
        self.api_url_edit = QLineEdit("https://api.deepseek.com/v1/chat/completions")
        api_url_layout.addWidget(api_url_label)
        api_url_layout.addWidget(self.api_url_edit)
        layout.addLayout(api_url_layout)
        
        # API Key
        api_key_layout = QHBoxLayout()
        api_key_label = QLabel("API Key:")
        self.api_key_edit = QLineEdit()
        self.api_key_edit.setEchoMode(QLineEdit.Password)
        api_key_layout.addWidget(api_key_label)
        api_key_layout.addWidget(self.api_key_edit)
        layout.addLayout(api_key_layout)
        
        # 模型名称
        model_layout = QHBoxLayout()
        model_label = QLabel("模型名称:")
        self.model_edit = QLineEdit("deepseek-chat")
        model_layout.addWidget(model_label)
        model_layout.addWidget(self.model_edit)
        layout.addLayout(model_layout)
        
        # 按钮
        button_layout = QHBoxLayout()
        test_btn = QPushButton("检测API")
        save_btn = QPushButton("保存")
        cancel_btn = QPushButton("取消")
        test_btn.clicked.connect(self.test_api_connection)
        save_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(test_btn)
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        # 加载现有配置
        self.load_settings()
        
    def load_settings(self):
        """加载现有设置"""
        if hasattr(self.parent, 'ai_analyzer') and self.parent.ai_analyzer:
            self.api_url_edit.setText(self.parent.ai_analyzer.api_url or "")
            self.api_key_edit.setText(self.parent.ai_analyzer.api_key or "")
            self.model_edit.setText(self.parent.ai_analyzer.model or "")
            
    def get_settings(self):
        """获取设置"""
        return {
            "api_url": self.api_url_edit.text().strip(),
            "api_key": self.api_key_edit.text().strip(),
            "model": self.model_edit.text().strip()
        }

    def test_api_connection(self):
        """检测当前填写的API配置是否可用。"""
        settings = self.get_settings()
        api_url = settings["api_url"]
        api_key = settings["api_key"]
        model = settings["model"]

        if not api_url:
            QMessageBox.warning(self, "缺少配置", "请输入API地址")
            return
        if not api_key:
            QMessageBox.warning(self, "缺少配置", "请输入API Key")
            return
        if not model:
            QMessageBox.warning(self, "缺少配置", "请输入模型名称")
            return

        progress_dialog = QProgressDialog("正在检测API连接...", None, 0, 0, self)
        progress_dialog.setWindowTitle("检测API")
        progress_dialog.setWindowModality(Qt.WindowModal)
        progress_dialog.setCancelButton(None)
        progress_dialog.show()
        QApplication.processEvents()

        payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": "你是API连通性检测助手。"},
                {"role": "user", "content": "请只回复：API检测成功"}
            ],
            "temperature": 0,
            "max_tokens": 20
        }
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }

        try:
            response = requests.post(api_url, json=payload, headers=headers, timeout=30)
            status_code = response.status_code
            response.raise_for_status()

            result = response.json()
            message_content = ""
            if isinstance(result, dict):
                choices = result.get("choices") or []
                if choices and isinstance(choices[0], dict):
                    message_content = ((choices[0].get("message") or {}).get("content") or "").strip()

            progress_dialog.close()
            QMessageBox.information(
                self,
                "检测成功",
                f"API调用成功。\n\n状态码：{status_code}\n模型：{model}\n返回内容：{message_content or '已收到有效响应'}"
            )
        except requests.exceptions.HTTPError:
            progress_dialog.close()
            error_text = response.text[:1000] if 'response' in locals() and hasattr(response, 'text') else "无响应内容"
            QMessageBox.critical(
                self,
                "检测失败",
                f"API返回HTTP错误。\n\n状态码：{getattr(response, 'status_code', '未知')}\n地址：{api_url}\n模型：{model}\n\n响应内容：\n{error_text}"
            )
        except requests.exceptions.RequestException as e:
            progress_dialog.close()
            QMessageBox.critical(
                self,
                "检测失败",
                f"API请求失败：{e}\n\n请检查API地址、网络、Key和模型名称。"
            )
        except Exception as e:
            progress_dialog.close()
            QMessageBox.critical(
                self,
                "检测失败",
                f"API检测过程中发生错误：{e}"
            )


class AnalysisResultDialog(QDialog):
    """AI分析结果对话框"""
    
    def __init__(self, analysis_result, parent=None):
        super().__init__(parent)
        self.analysis_result = analysis_result
        self.setup_ui()
        
    def setup_ui(self):
        """设置界面"""
        self.setWindowTitle("AI分析结果")
        self.resize(900, 700)
        
        layout = QVBoxLayout(self)
        
        # 结果显示区域 - 支持Markdown格式
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        
        # 设置Markdown格式显示
        self.result_text.setMarkdown(self.analysis_result)
        
        # 设置字体和样式
        font = QFont("Microsoft YaHei", 10)
        self.result_text.setFont(font)
        
        # 设置样式表，美化显示效果
        self.result_text.setStyleSheet("""
            QTextEdit {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 5px;
                padding: 10px;
                line-height: 1.6;
            }
            QTextEdit:focus {
                border-color: #007bff;
            }
        """)
        
        layout.addWidget(self.result_text)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        
        copy_btn = QPushButton("复制结果")
        save_md_btn = QPushButton("保存为Markdown")
        save_html_btn = QPushButton("保存为HTML")
        close_btn = QPushButton("关闭")
        
        copy_btn.clicked.connect(self.copy_result)
        save_md_btn.clicked.connect(self.save_as_markdown)
        save_html_btn.clicked.connect(self.save_as_html)
        close_btn.clicked.connect(self.accept)
        
        button_layout.addWidget(copy_btn)
        button_layout.addWidget(save_md_btn)
        button_layout.addWidget(save_html_btn)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
        
    def copy_result(self):
        """复制结果到剪贴板"""
        clipboard = QApplication.clipboard()
        clipboard.setText(self.analysis_result)
        QMessageBox.information(self, "成功", "分析结果已复制到剪贴板")
        
    def save_as_markdown(self):
        """保存为Markdown文件"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存分析结果", "", "Markdown文件 (*.md)"
        )
        if file_path:
            try:
                # 添加文件头信息
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                md_content = f"""# AI分析报告

**生成时间**: {timestamp}  
**报告类型**: 电商售后数据分析  

---

{self.analysis_result}

---

*本报告由AI分析工具自动生成*
"""
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(md_content)
                QMessageBox.information(self, "成功", f"分析结果已保存到 {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"保存失败: {str(e)}")
        
    def save_as_html(self):
        """保存为HTML文件"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存分析结果", "", "HTML文件 (*.html)"
        )
        if file_path:
            try:
                # 使用markdown库转换为HTML
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # 添加文件头信息
                md_content = f"""# AI分析报告

**生成时间**: {timestamp}  
**报告类型**: 电商售后数据分析  

---

{self.analysis_result}

---

*本报告由AI分析工具自动生成*
"""
                
                # 转换为HTML
                html_content = markdown.markdown(md_content, extensions=['extra'])
                
                # 完整的HTML文档
                full_html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>AI分析报告</title>
    <style>
        body {{ 
            font-family: 'Microsoft YaHei', Arial, sans-serif; 
            line-height: 1.6; 
            margin: 40px; 
            max-width: 1000px;
            background-color: #f8f9fa;
        }}
        h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        h3 {{ color: #7f8c8d; }}
        pre {{ 
            background-color: #2c3e50; 
            color: #ecf0f1; 
            padding: 15px; 
            border-radius: 5px; 
            overflow-x: auto;
        }}
        code {{ background-color: #f1f2f6; padding: 2px 4px; border-radius: 3px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #3498db; color: white; }}
        tr:nth-child(even) {{ background-color: #f2f2f2; }}
        blockquote {{ 
            border-left: 4px solid #3498db; 
            margin: 20px 0; 
            padding-left: 15px; 
            color: #7f8c8d;
            font-style: italic;
        }}
        hr {{ border: 0; border-top: 2px dashed #bdc3c7; margin: 30px 0; }}
    </style>
</head>
<body>
{html_content}
</body>
</html>"""
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(full_html)
                QMessageBox.information(self, "成功", f"分析结果已保存到 {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"保存失败: {str(e)}")


# ---------------------------- 图表组件类 ---------------------------------
class ChartWidget(QWidget):
    """图表展示组件"""
    
    # 退款原因列表（固定）
    REASON_LIST = [
        "商品腐败、变质、包装胀气等",
        "商品破损/压坏", 
        "质量问题",
        "大小/规格/重量等与商品描述不符",
        "品种/标签/图片/包装等与商品描述不符",
        "货物与描述不符",
        "生产日期/保质期与商品描述不符",
        "其他"
    ]
    
    def __init__(self, parent=None, db=None):
        super().__init__(parent)
        self.db = db
        self.current_chart_index = 0  # 0:柱状图, 1:饼图, 2:曲线图
        self.chart_types = ["退款原因柱状图", "退款原因饼图", "时间曲线图"]
        # 数据缓存，供放大窗口使用
        self.current_records = []
        self.current_start_date = ""
        self.current_end_date = ""
        self.init_ui()
    
    def init_ui(self):
        """初始化界面"""
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # 顶部：切换控制区域
        control_layout = QHBoxLayout()
        
        # 左箭头按钮
        self.prev_btn = QPushButton("◀")
        self.prev_btn.setFixedSize(40, 30)
        self.prev_btn.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                font-weight: bold;
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
        """)
        self.prev_btn.clicked.connect(self.prev_chart)
        control_layout.addWidget(self.prev_btn)
        
        # 图表标题
        self.title_label = QLabel(self.chart_types[self.current_chart_index])
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50;")
        control_layout.addWidget(self.title_label, 1)
        
        # 右箭头按钮
        self.next_btn = QPushButton("▶")
        self.next_btn.setFixedSize(40, 30)
        self.next_btn.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                font-weight: bold;
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
        """)
        self.next_btn.clicked.connect(self.next_chart)
        control_layout.addWidget(self.next_btn)
        
        layout.addLayout(control_layout)
        
        # 中间：图表区域
        self.figure = Figure(figsize=(6, 4), dpi=100)  # 调整图表尺寸
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setMinimumSize(300, 200)  # 调整最小尺寸
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        layout.addWidget(self.canvas, 1)
        
        # 底部：放大按钮
        self.enlarge_btn = QPushButton("点击放大")
        self.enlarge_btn.setStyleSheet("""
            QPushButton {
                font-size: 12px;
                padding: 5px 10px;
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #2471a3;
            }
        """)
        self.enlarge_btn.clicked.connect(self.show_enlarged_window)
        layout.addWidget(self.enlarge_btn)
        
        # 初始显示空图表
        self.show_empty_chart()
    
    def update_chart(self, records, start_date, end_date):
        """根据当前图表类型更新显示"""
        print(f"[DEBUG update_chart] 开始更新图表，索引: {self.current_chart_index}, 记录数: {len(records)}")
        
        if not records:
            print("[DEBUG update_chart] 无数据，显示空图表")
            self.show_empty_chart()
            return
        
        try:
            if self.current_chart_index == 0:
                print("[DEBUG update_chart] 绘制柱状图")
                self.draw_bar_chart(records)
            elif self.current_chart_index == 1:
                print("[DEBUG update_chart] 绘制饼图")
                self.draw_pie_chart(records)
            elif self.current_chart_index == 2:
                print("[DEBUG update_chart] 绘制曲线图")
                self.draw_line_chart(records, start_date, end_date)
            else:
                print(f"[DEBUG update_chart] 未知图表索引: {self.current_chart_index}")
                self.show_empty_chart()
            
            print("[DEBUG update_chart] 图表更新完成")
        except Exception as e:
            print(f"[ERROR] 图表更新失败: {e}")
            import traceback
            traceback.print_exc()
            self.show_empty_chart()
    
    def draw_bar_chart(self, records):
        """绘制柱状图"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        # 统计每个退款原因的数量
        reason_counts = {reason: 0 for reason in self.REASON_LIST}
        for record in records:
            reason = record.get('reason', '')
            if reason in reason_counts:
                reason_counts[reason] += 1
        
        # 准备数据
        reasons = list(reason_counts.keys())
        counts = list(reason_counts.values())
        
        # 创建柱状图
        bars = ax.bar(range(len(reasons)), counts, color='#2E8B57', alpha=0.8)
        
        # 设置图表样式
        ax.set_title('退款原因分布（柱状图）', fontweight='bold')
        ax.set_xlabel('退款原因')
        ax.set_ylabel('订单数量')
        
        # 设置X轴标签（支持换行）
        formatted_reasons = []
        for reason in reasons:
            # 每6个字符换行
            formatted_reason = '\n'.join([reason[i:i+6] for i in range(0, len(reason), 6)])
            formatted_reasons.append(formatted_reason)
        
        ax.set_xticks(range(len(reasons)))
        ax.set_xticklabels(formatted_reasons, rotation=0)
        
        # 在每个柱子上方显示数字
        for i, (bar, count) in enumerate(zip(bars, counts)):
            height = bar.get_height()
            if height > 0:
                ax.text(i, height + 0.1, f'{count}', ha='center', va='bottom')
        
        # 设置Y轴范围
        ax.set_ylim(0, max(counts) * 1.1 if counts else 10)
        
        # 调整布局
        self.figure.tight_layout()
        self.canvas.draw()
    
    def draw_pie_chart(self, records):
        """绘制饼图（带小方框和箭头指示）"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        # 统计每个退款原因的数量
        reason_counts = {reason: 0 for reason in self.REASON_LIST}
        for record in records:
            reason = record.get('reason', '')
            if reason in reason_counts:
                reason_counts[reason] += 1
        
        # 过滤掉数量为0的原因
        filtered_reasons = []
        filtered_counts = []
        for reason, count in reason_counts.items():
            if count > 0:
                filtered_reasons.append(reason)
                filtered_counts.append(count)
        
        if not filtered_counts:
            self.show_empty_chart()
            return
        
        # 合并占比小于3%的原因为"其他"
        total = sum(filtered_counts)
        if total > 0:
            other_count = 0
            new_reasons = []
            new_counts = []
            
            for reason, count in zip(filtered_reasons, filtered_counts):
                percentage = (count / total) * 100
                if percentage < 3:
                    other_count += count
                else:
                    new_reasons.append(reason)
                    new_counts.append(count)
            
            if other_count > 0:
                new_reasons.append("其他")
                new_counts.append(other_count)
            
            filtered_reasons = new_reasons
            filtered_counts = new_counts
        
        # 创建饼图（不显示默认标签）
        colors = get_colormap_colors(plt.cm.Set3, len(filtered_reasons))
        # 当labels=None和autopct=None时，只返回2个值
        pie_result = ax.pie(filtered_counts, labels=None, autopct=None,
                           colors=colors, startangle=90)
        wedges = pie_result[0]
        texts = pie_result[1] if len(pie_result) > 1 else []
        
        # 设置饼图样式
        ax.set_title('退款原因分布（饼图）', fontweight='bold')
        
        # 添加自定义标签（带小方框和箭头）
        bbox_props = dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8, edgecolor="black")
        kw = dict(arrowprops=dict(arrowstyle="->", color="black"), bbox=bbox_props, zorder=0, va="center")
        
        for i, (wedge, reason, count) in enumerate(zip(wedges, filtered_reasons, filtered_counts)):
            ang = (wedge.theta2 - wedge.theta1) / 2. + wedge.theta1
            radians = math.radians(ang)
            y = math.sin(radians)
            x = math.cos(radians)
            
            # 计算百分比
            percentage = (count / total) * 100
            
            # 确定标签位置（根据角度调整）
            horizontalalignment = 'left' if x > 0 else 'right'
            connectionstyle = f"angle,angleA=0,angleB={ang}"
            kw["arrowprops"].update({"connectionstyle": connectionstyle})
            
            # 创建标签文本
            label_text = f"{reason}\n{count}单 ({percentage:.1f}%)"
            
            # 添加带箭头的标签
            x_direction = 1 if x >= 0 else -1
            ax.annotate(label_text, xy=(x, y), xytext=(1.35 * x_direction, 1.4 * y),
                       horizontalalignment=horizontalalignment, fontsize=7, **kw)
        
        # 调整布局
        self.figure.tight_layout()
        self.canvas.draw()
    
    def draw_line_chart(self, records, start_date, end_date):
        """绘制时间曲线图（智能调整显示粒度）"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        # 将日期字符串转换为datetime对象
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        except:
            self.show_empty_chart()
            return
        
        # 计算时间跨度（天数）
        time_span_days = (end_dt - start_dt).days
        
        # 智能调整显示粒度
        if time_span_days > 365:  # 大于1年
            # 从最新记录的退款日期往前数12个月
            if records:
                # 找到最新记录的日期
                latest_date = max(datetime.strptime(rec['record_date'], '%Y-%m-%d') for rec in records if rec.get('record_date'))
                # 往前数12个月
                end_dt = latest_date
                start_dt = latest_date - timedelta(days=365)  # 12个月约365天
            
            # 按月聚合数据
            date_range = []
            current_dt = start_dt.replace(day=1)  # 从月初开始
            for i in range(12):
                date_range.append(current_dt)
                # 下个月
                if current_dt.month == 12:
                    current_dt = current_dt.replace(year=current_dt.year + 1, month=1)
                else:
                    current_dt = current_dt.replace(month=current_dt.month + 1)
            
            # 按月统计退款金额总和
            monthly_reason_amounts = {}
            for record in records:
                record_date_str = record.get('record_date', '')
                reason = record.get('reason', '')
                refund_amount = record.get('refund_amount', 0)
                
                if not record_date_str or reason not in self.REASON_LIST:
                    continue
                
                try:
                    record_date = datetime.strptime(record_date_str, '%Y-%m-%d')
                    if start_dt <= record_date <= end_dt:
                        # 按月聚合
                        month_key = record_date.strftime('%Y-%m')
                        if month_key not in monthly_reason_amounts:
                            monthly_reason_amounts[month_key] = {reason: 0 for reason in self.REASON_LIST}
                        monthly_reason_amounts[month_key][reason] += refund_amount
                except:
                    continue
            
            # 为每个原因创建数据序列
            colors = get_colormap_colors(plt.cm.tab10, len(self.REASON_LIST))
            
            for i, reason in enumerate(self.REASON_LIST):
                amounts = []
                for month_dt in date_range:
                    month_key = month_dt.strftime('%Y-%m')
                    amount = monthly_reason_amounts.get(month_key, {}).get(reason, 0)
                    amounts.append(amount)
                
                # 只有当该原因有数据时才绘制
                if sum(amounts) > 0:
                    ax.plot(date_range, amounts, label=reason, color=colors[i], marker='o', markersize=3)
            
            # 设置图表样式
            ax.set_title('退款原因时间趋势（按月显示，最多12个月）', fontweight='bold')
            ax.set_xlabel('月份')
            ax.set_ylabel('退款金额（元）')
            
            # 设置X轴日期格式
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
            ax.xaxis.set_major_locator(mdates.MonthLocator())
            
        elif time_span_days <= 30:  # 30天以内
            # 按天显示
            date_range = []
            current_dt = start_dt
            while current_dt <= end_dt:
                date_range.append(current_dt)
                current_dt += timedelta(days=1)
            
            # 按天统计退款金额总和
            daily_reason_amounts = {}
            for record in records:
                record_date_str = record.get('record_date', '')
                reason = record.get('reason', '')
                refund_amount = record.get('refund_amount', 0)
                
                if not record_date_str or reason not in self.REASON_LIST:
                    continue
                
                try:
                    record_date = datetime.strptime(record_date_str, '%Y-%m-%d')
                    if start_dt <= record_date <= end_dt:
                        # 按天聚合
                        day_key = record_date.strftime('%Y-%m-%d')
                        if day_key not in daily_reason_amounts:
                            daily_reason_amounts[day_key] = {reason: 0 for reason in self.REASON_LIST}
                        daily_reason_amounts[day_key][reason] += refund_amount
                except:
                    continue
            
            # 为每个原因创建数据序列
            colors = get_colormap_colors(plt.cm.tab10, len(self.REASON_LIST))
            
            for i, reason in enumerate(self.REASON_LIST):
                amounts = []
                for day_dt in date_range:
                    day_key = day_dt.strftime('%Y-%m-%d')
                    amount = daily_reason_amounts.get(day_key, {}).get(reason, 0)
                    amounts.append(amount)
                
                # 只有当该原因有数据时才绘制
                if sum(amounts) > 0:
                    ax.plot(date_range, amounts, label=reason, color=colors[i], marker='o', markersize=3)
            
            # 设置图表样式
            ax.set_title('退款原因时间趋势（按天显示）', fontweight='bold')
            ax.set_xlabel('日期')
            ax.set_ylabel('退款金额（元）')
            
            # 设置X轴日期格式
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
            
        else:  # 30天到1年之间
            # 按月显示
            date_range = []
            current_dt = start_dt.replace(day=1)  # 从月初开始
            while current_dt <= end_dt:
                date_range.append(current_dt)
                # 下个月
                if current_dt.month == 12:
                    current_dt = current_dt.replace(year=current_dt.year + 1, month=1)
                else:
                    current_dt = current_dt.replace(month=current_dt.month + 1)
            
            # 按月统计退款金额总和
            monthly_reason_amounts = {}
            for record in records:
                record_date_str = record.get('record_date', '')
                reason = record.get('reason', '')
                refund_amount = record.get('refund_amount', 0)
                
                if not record_date_str or reason not in self.REASON_LIST:
                    continue
                
                try:
                    record_date = datetime.strptime(record_date_str, '%Y-%m-%d')
                    if start_dt <= record_date <= end_dt:
                        # 按月聚合
                        month_key = record_date.strftime('%Y-%m')
                        if month_key not in monthly_reason_amounts:
                            monthly_reason_amounts[month_key] = {reason: 0 for reason in self.REASON_LIST}
                        monthly_reason_amounts[month_key][reason] += refund_amount
                except:
                    continue
            
            # 为每个原因创建数据序列
            colors = get_colormap_colors(plt.cm.tab10, len(self.REASON_LIST))
            
            for i, reason in enumerate(self.REASON_LIST):
                amounts = []
                for month_dt in date_range:
                    month_key = month_dt.strftime('%Y-%m')
                    amount = monthly_reason_amounts.get(month_key, {}).get(reason, 0)
                    amounts.append(amount)
                
                # 只有当该原因有数据时才绘制
                if sum(amounts) > 0:
                    ax.plot(date_range, amounts, label=reason, color=colors[i], marker='o', markersize=3)
            
            # 设置图表样式
            ax.set_title('退款原因时间趋势（按月显示）', fontweight='bold')
            ax.set_xlabel('月份')
            ax.set_ylabel('退款金额（元）')
            
            # 设置X轴日期格式
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
            ax.xaxis.set_major_locator(mdates.MonthLocator())
        
        # 添加图例
        if ax.get_legend_handles_labels()[0]:
            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        
        # 旋转X轴标签
        ax.tick_params(axis='x', rotation=45)
        
        # 调整布局
        self.figure.tight_layout()
        self.canvas.draw()
    
    def show_empty_chart(self):
        """显示空数据提示"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.text(0.5, 0.5, '暂无数据', fontsize=16, ha='center', va='center', 
               transform=ax.transAxes, color='gray')
        ax.set_xticks([])
        ax.set_yticks([])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        self.canvas.draw()
    
    def prev_chart(self):
        """切换到上一个图表"""
        self.current_chart_index = (self.current_chart_index - 1) % len(self.chart_types)
        self.title_label.setText(self.chart_types[self.current_chart_index])
        print(f"[DEBUG] 切换到图表: {self.chart_types[self.current_chart_index]}")
        
        # 延迟刷新图表数据（确保数据加载完成）
        QTimer.singleShot(100, self.force_refresh_chart)  # 100毫秒延迟
    
    def next_chart(self):
        """切换到下一个图表"""
        self.current_chart_index = (self.current_chart_index + 1) % len(self.chart_types)
        self.title_label.setText(self.chart_types[self.current_chart_index])
        print(f"[DEBUG] 切换到图表: {self.chart_types[self.current_chart_index]}")
        
        # 延迟刷新图表数据（确保数据加载完成）
        QTimer.singleShot(100, self.force_refresh_chart)  # 100毫秒延迟
    
    def force_refresh_chart(self):
        """强制刷新图表数据"""
        records = []
        start_date = ""
        end_date = ""
        
        print("[DEBUG 强制刷新] 开始强制刷新图表数据")
        
        # 先显示加载状态
        self.show_loading_chart()
        
        # 获取主窗口引用（通过层层向上查找）
        main_window = self._get_main_window()
        
        if main_window:
            print("[DEBUG 强制刷新] 找到主窗口，使用主窗口方法刷新")
            # 使用主窗口的刷新机制（与筛选时间相同）
            main_window.load_table_data(force_reload=True)
            
            # 给数据加载一点时间
            QTimer.singleShot(50, lambda: self._refresh_data_with_main_window(main_window))
        else:
            print("[DEBUG 强制刷新] 未找到主窗口，使用备用方法")
            # 备用方法：直接尝试方式2
            QTimer.singleShot(50, lambda: self._refresh_data_after_delay(2))
    
    def _get_main_window(self):
        """获取主窗口引用"""
        # 方法1：尝试通过父组件链向上查找
        parent = self.parent()
        while parent:
            if hasattr(parent, 'load_table_data'):
                print(f"[DEBUG 主窗口查找] 找到主窗口: {type(parent)}")
                return parent
            parent = parent.parent()
        
        # 方法2：尝试通过应用程序查找
        app = QApplication.instance()
        if app:
            for widget in app.allWidgets():
                if hasattr(widget, 'load_table_data') and widget.isWindow():
                    print(f"[DEBUG 主窗口查找] 通过应用程序找到主窗口: {type(widget)}")
                    return widget
        
        print("[DEBUG 主窗口查找] 未找到主窗口")
        return None
    
    def _refresh_data_with_main_window(self, main_window):
        """使用主窗口刷新数据"""
        records = []
        start_date = ""
        end_date = ""
        
        print("[DEBUG 主窗口刷新] 使用主窗口方法获取数据")
        
        # 方式1：使用主窗口的标准方法
        if hasattr(main_window, 'get_current_records_for_chart'):
            try:
                records, start_date, end_date = main_window.get_current_records_for_chart()
                print(f"[DEBUG 主窗口刷新] 方式1获取到 {len(records)} 条记录")
            except Exception as e:
                print(f"[DEBUG 主窗口刷新] 方式1失败: {e}")
                records = []
        
        # 方式2：如果方式1失败，使用筛选方法
        if not records and hasattr(main_window, 'get_current_filtered_records'):
            try:
                records = main_window.get_current_filtered_records()
                start_date = main_window.start_date_edit.date().toString("yyyy-MM-dd")
                end_date = main_window.end_date_edit.date().toString("yyyy-MM-dd")
                print(f"[DEBUG 主窗口刷新] 方式2获取到 {len(records)} 条记录")
            except Exception as e:
                print(f"[DEBUG 主窗口刷新] 方式2失败: {e}")
                records = []
        
        # 最终更新图表
        self._final_update_chart(records, start_date, end_date)
    
    def _refresh_data_after_delay(self, method):
        """延迟后刷新数据"""
        records = []
        start_date = ""
        end_date = ""
        
        print(f"[DEBUG 延迟刷新] 开始方式{method}数据获取")
        
        if method == 1:
            try:
                records, start_date, end_date = self.parent().get_current_records_for_chart()
                print(f"[DEBUG 延迟刷新] 方式1获取到 {len(records)} 条记录")
            except Exception as e:
                print(f"[DEBUG 延迟刷新] 方式1失败: {e}")
                records = []
        
        # 如果方式1失败，尝试方式2
        if not records and hasattr(self.parent(), 'get_current_filtered_records'):
            print("[DEBUG 延迟刷新] 尝试方式2：直接获取")
            try:
                records = self.parent().get_current_filtered_records()
                start_date = self.parent().start_date_edit.date().toString("yyyy-MM-dd")
                end_date = self.parent().end_date_edit.date().toString("yyyy-MM-dd")
                print(f"[DEBUG 延迟刷新] 方式2获取到 {len(records)} 条记录")
                
                # 如果方式2也返回空数据，添加详细调试
                if not records:
                    print("[DEBUG 延迟刷新] 方式2返回空数据，检查父组件状态:")
                    print(f"[DEBUG 延迟刷新] - 父组件类型: {type(self.parent())}")
                    print(f"[DEBUG 延迟刷新] - 父组件方法存在性: {hasattr(self.parent(), 'get_current_filtered_records')}")
                    
                    # 尝试直接调用数据库获取数据
                    if hasattr(self.parent(), 'db'):
                        print("[DEBUG 延迟刷新] 尝试直接查询数据库")
                        try:
                            # 获取所有记录作为测试
                            all_records = self.parent().db.get_all_records()
                            print(f"[DEBUG 延迟刷新] 数据库总记录数: {len(all_records)}")
                            
                            # 尝试使用默认筛选条件
                            default_records = self.parent().db.get_records_by_filters()
                            print(f"[DEBUG 延迟刷新] 默认筛选记录数: {len(default_records)}")
                        except Exception as db_e:
                            print(f"[DEBUG 延迟刷新] 数据库查询失败: {db_e}")
            except Exception as e:
                print(f"[DEBUG 延迟刷新] 方式2失败: {e}")
                records = []
        
        # 最终更新图表
        self._final_update_chart(records, start_date, end_date)
    
    def _final_update_chart(self, records, start_date, end_date):
        """最终更新图表"""
        if not records:
            print("[DEBUG 最终更新] 无数据，使用空图表")
            self.show_empty_chart()
        else:
            print(f"[DEBUG 最终更新] 使用 {len(records)} 条记录更新图表")
            # 保存当前数据，供放大窗口使用
            self.current_records = records
            self.current_start_date = start_date
            self.current_end_date = end_date
            self.update_chart(records, start_date, end_date)
    
    def show_loading_chart(self):
        """显示加载中的图表"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        # 显示加载提示
        ax.text(0.5, 0.5, '加载中...', fontsize=12, ha='center', va='center', 
                transform=ax.transAxes, color='gray')
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis('off')  # 隐藏坐标轴
        
        self.canvas.draw()
    
    def show_enlarged_window(self):
        """显示放大的图表窗口（仅视觉放大，使用当前数据）"""
        print("[DEBUG 放大窗口] 开始显示放大窗口")
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"图表放大 - {self.chart_types[self.current_chart_index]}")
        dialog.resize(900, 700)
        
        layout = QVBoxLayout(dialog)
        
        # 创建放大版的图表组件（自定义版本，不包含放大按钮）
        enlarged_widget = EnlargedChartWidget(dialog, self.db)
        enlarged_widget.current_chart_index = self.current_chart_index
        enlarged_widget.title_label.setText(self.chart_types[self.current_chart_index])
        layout.addWidget(enlarged_widget)
        
        # 添加关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        # 直接使用当前图表的数据（仅视觉放大，不重新获取数据）
        print("[DEBUG 放大窗口] 使用当前图表数据进行视觉放大")
        
        # 获取当前图表的数据
        records = getattr(self, 'current_records', [])
        start_date = getattr(self, 'current_start_date', "")
        end_date = getattr(self, 'current_end_date', "")
        
        print(f"[DEBUG 放大窗口] 当前图表数据: {len(records)} 条记录")
        
        if records:
            print(f"[DEBUG 放大窗口] 使用 {len(records)} 条记录显示放大图表")
            # 直接使用当前数据更新放大图表
            enlarged_widget.current_records = records
            enlarged_widget.current_start_date = start_date
            enlarged_widget.current_end_date = end_date
            
            # 立即调用刷新方法，确保图表显示
            print("[DEBUG 放大窗口] 立即调用update_chart方法")
            enlarged_widget.update_chart(records, start_date, end_date)
            
            # 添加额外调试：检查数据是否成功传递
            print(f"[DEBUG 放大窗口] 数据传递检查 - 记录数: {len(getattr(enlarged_widget, 'current_records', []))}")
            print(f"[DEBUG 放大窗口] 数据传递检查 - 开始日期: {getattr(enlarged_widget, 'current_start_date', '无')}")
            print(f"[DEBUG 放大窗口] 数据传递检查 - 结束日期: {getattr(enlarged_widget, 'current_end_date', '无')}")
        else:
            print("[DEBUG 放大窗口] 当前无数据，显示空图表")
            enlarged_widget.show_empty_chart()
        
        dialog.exec_()


class EnlargedChartWidget(ChartWidget):
    """放大窗口专用的图表组件（不包含放大按钮）"""
    
    def __init__(self, parent=None, db=None):
        super().__init__(parent, db)
        # 标记为放大窗口，用于特殊处理
        self.is_enlarged = True
    
    def init_ui(self):
        """初始化界面（不包含放大按钮）"""
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # 顶部：切换控制区域
        control_layout = QHBoxLayout()
        
        # 左箭头按钮
        self.prev_btn = QPushButton("◀")
        self.prev_btn.setFixedSize(40, 30)
        self.prev_btn.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                font-weight: bold;
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
        """)
        self.prev_btn.clicked.connect(self.prev_chart)
        control_layout.addWidget(self.prev_btn)
        
        # 图表标题
        self.title_label = QLabel(self.chart_types[self.current_chart_index])
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50;")
        control_layout.addWidget(self.title_label, 1)
        
        # 右箭头按钮
        self.next_btn = QPushButton("▶")
        self.next_btn.setFixedSize(40, 30)
        self.next_btn.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                font-weight: bold;
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
        """)
        self.next_btn.clicked.connect(self.next_chart)
        control_layout.addWidget(self.next_btn)
        
        layout.addLayout(control_layout)
        
        # 中间：图表区域（放大尺寸）
        self.figure = Figure(figsize=(10, 8), dpi=100)  # 更大的图表尺寸
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setMinimumSize(800, 600)  # 更大的最小尺寸
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        layout.addWidget(self.canvas, 1)
        
        # 初始显示空图表
        self.show_empty_chart()
        
        # 放大窗口创建后立即尝试刷新（如果有数据）
        QTimer.singleShot(100, self._try_refresh_after_init)
    
    def _try_refresh_after_init(self):
        """初始化后尝试刷新图表"""
        print("[DEBUG _try_refresh_after_init] 开始执行初始化后刷新")
        
        # 检查当前数据状态
        print(f"[DEBUG _try_refresh_after_init] 检查数据状态:")
        print(f"[DEBUG _try_refresh_after_init] - hasattr current_records: {hasattr(self, 'current_records')}")
        print(f"[DEBUG _try_refresh_after_init] - current_records长度: {len(getattr(self, 'current_records', []))}")
        print(f"[DEBUG _try_refresh_after_init] - current_chart_index: {self.current_chart_index}")
        
        # 如果有缓存数据，立即刷新图表
        if hasattr(self, 'current_records') and self.current_records:
            print(f"[DEBUG _try_refresh_after_init] 使用缓存数据刷新: {len(self.current_records)} 条记录")
            self.update_chart(self.current_records, self.current_start_date, self.current_end_date)
        else:
            print("[DEBUG _try_refresh_after_init] 无缓存数据，保持空图表")
            # 即使没有缓存数据，也尝试强制刷新一次
            print("[DEBUG _try_refresh_after_init] 尝试强制刷新")
            self.force_refresh_chart()


class SummaryHistoryDialog(QDialog):
    """本地总结历史记录查看对话框。"""

    def __init__(self, db, parent_window=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.parent_window = parent_window
        self.selected_snapshot = None
        self.selected_history_id = None
        self.setup_ui()
        self.load_history_list()

    def setup_ui(self):
        self.setWindowTitle("本地总结历史记录")
        self.resize(1000, 700)

        layout = QHBoxLayout(self)
        left_layout = QVBoxLayout()
        right_layout = QVBoxLayout()

        self.history_list = QListWidget()
        self.history_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.history_list.currentRowChanged.connect(self.on_history_changed)
        left_layout.addWidget(self.history_list, 1)

        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setFont(QFont("Microsoft YaHei", 10))
        right_layout.addWidget(self.preview_text, 1)

        button_layout = QHBoxLayout()
        self.load_btn = QPushButton("加载到总结")
        self.export_btn = QPushButton("导出选中")
        self.delete_btn = QPushButton("删除选中")
        self.close_btn = QPushButton("关闭")
        self.load_btn.clicked.connect(self.accept_selection)
        self.export_btn.clicked.connect(self.export_selection)
        self.delete_btn.clicked.connect(self.delete_selection)
        self.close_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.load_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addWidget(self.close_btn)
        right_layout.addLayout(button_layout)

        layout.addLayout(left_layout, 2)
        layout.addLayout(right_layout, 5)

    def load_history_list(self):
        self.history_list.clear()
        for item in self.db.get_ai_summary_history_list():
            display_text = f"{item['created_at']} | {item['filter_summary']}"
            widget_item = QListWidgetItem(display_text)
            widget_item.setData(Qt.UserRole, item["id"])
            self.history_list.addItem(widget_item)

        if self.history_list.count() > 0:
            self.history_list.setCurrentRow(0)

    def on_history_changed(self, row):
        if row < 0:
            self.preview_text.clear()
            return

        item = self.history_list.item(row)
        history_id = item.data(Qt.UserRole)
        history = self.db.get_ai_summary_history(history_id)
        if not history:
            self.preview_text.setPlainText("历史记录不存在或已损坏")
            return

        self.selected_history_id = history_id
        self.selected_snapshot = history.get("snapshot", {})
        markdown_text = self.parent_window.render_summary_snapshot_markdown(self.selected_snapshot)
        self.preview_text.setMarkdown(markdown_text)

    def accept_selection(self):
        if not self.selected_snapshot:
            QMessageBox.information(self, "提示", "请先选择一条历史记录")
            return
        self.accept()

    def export_selection(self):
        if not self.selected_snapshot:
            QMessageBox.information(self, "提示", "请先选择一条历史记录")
            return
        self.parent_window.export_summary_excel(snapshot=self.selected_snapshot)

    def get_selected_history_ids(self):
        ids = []
        for item in self.history_list.selectedItems():
            history_id = item.data(Qt.UserRole)
            if history_id:
                ids.append(history_id)
        return ids

    def delete_selection(self):
        selected_ids = self.get_selected_history_ids()
        if not selected_ids:
            QMessageBox.information(self, "提示", "请先选择要删除的历史记录")
            return
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定删除选中的 {len(selected_ids)} 条本地总结历史记录吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        deleted_count = self.db.delete_ai_summary_history(selected_ids)
        self.selected_snapshot = None
        self.selected_history_id = None
        self.preview_text.clear()
        self.load_history_list()
        QMessageBox.information(self, "删除完成", f"已删除 {deleted_count} 条历史记录")


class LocalReasonCategoryDialog(QDialog):
    """本地真实退款原因分类管理窗口。"""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setup_ui()
        self.load_categories()

    def setup_ui(self):
        self.setWindowTitle("本地分类管理")
        self.resize(1100, 720)
        layout = QVBoxLayout(self)

        tip = QLabel("维护本地分类和关键词。关键词横向填写，用空格分隔；单字关键词也会参与识别。")
        layout.addWidget(tip)

        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["分类名", "关键词"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        layout.addWidget(self.table, 1)

        button_layout = QHBoxLayout()
        add_btn = QPushButton("新增分类")
        delete_btn = QPushButton("删除分类")
        save_btn = QPushButton("保存")
        close_btn = QPushButton("关闭")
        add_btn.clicked.connect(self.add_category_row)
        delete_btn.clicked.connect(self.delete_selected_categories)
        save_btn.clicked.connect(self.save_categories)
        close_btn.clicked.connect(self.reject)
        button_layout.addWidget(add_btn)
        button_layout.addWidget(delete_btn)
        button_layout.addWidget(save_btn)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)

    def load_categories(self):
        categories = [
            item for item in self.db.get_real_refund_reason_categories(active_only=False)
            if str(item.get("category_name") or "").strip() != "未明确备注"
        ]
        self.db.replace_real_refund_reason_categories(categories)
        self.table.setRowCount(0)
        for item in categories:
            self.add_category_row(item)

    def add_category_row(self, category=None):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem(str((category or {}).get("category_name", ""))))
        keywords_edit = QLineEdit()
        keywords_edit.setText(self._format_keywords_for_display((category or {}).get("keywords_text", "")))
        keywords_edit.setPlaceholderText("例如：长毛 发霉 腐烂 烂")
        self.table.setCellWidget(row, 1, keywords_edit)
        self.table.setRowHeight(row, 42)

    @staticmethod
    def _format_keywords_for_display(keywords_text):
        normalized = re.sub(r'[、,，;；\s]+', ' ', str(keywords_text or ""))
        return " ".join(item for item in normalized.split(" ") if item.strip())

    def delete_selected_categories(self):
        selected_rows = sorted({index.row() for index in self.table.selectedIndexes()}, reverse=True)
        if not selected_rows:
            QMessageBox.warning(self, "警告", "请先选择要删除的分类")
            return
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定删除选中的 {len(selected_rows)} 个分类吗？\n已归因到这些分类的历史订单不会自动清空。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        for row in selected_rows:
            self.table.removeRow(row)

    def save_categories(self):
        configs = []
        seen = set()
        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 0)
            keywords_widget = self.table.cellWidget(row, 1)

            name = str(name_item.text() if name_item else "").strip()
            if not name:
                continue
            if name in seen:
                QMessageBox.warning(self, "警告", f"分类名重复：{name}")
                return
            seen.add(name)

            configs.append({
                "category_name": name,
                "keywords_text": self._format_keywords_for_display(keywords_widget.text() if keywords_widget else ""),
                "status": "ACTIVE",
                "sort_order": row,
            })

        if not configs:
            QMessageBox.warning(self, "警告", "请至少保留一个分类")
            return

        self.db.replace_real_refund_reason_categories(configs)
        self.accept()


class AIReasonAssignmentConfirmDialog(QDialog):
    """AI归因写库前确认窗口。"""

    def __init__(self, assignments, new_categories, elapsed_seconds, parent=None):
        super().__init__(parent)
        self.assignments = assignments or []
        self.new_categories = new_categories or []
        self.elapsed_seconds = elapsed_seconds
        self.setup_ui()
        self.load_assignments()

    def setup_ui(self):
        self.setWindowTitle("确认AI归因结果")
        self.resize(1050, 680)
        layout = QVBoxLayout(self)

        summary = QLabel(
            f"AI归因耗时：{self.elapsed_seconds:.2f}秒；"
            f"待写入：{len(self.assignments)}条；"
            f"新增分类：{len(self.new_categories)}个。请确认后再写入数据库。"
        )
        layout.addWidget(summary)

        if self.new_categories:
            categories_label = QLabel("新增分类：" + "、".join(self.new_categories))
            categories_label.setWordWrap(True)
            layout.addWidget(categories_label)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["订单号", "备注", "AI归因分类", "归因说明", "是否新增分类"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        layout.addWidget(self.table, 1)

        button_layout = QHBoxLayout()
        button_layout.addStretch()
        confirm_btn = QPushButton("确认写入")
        cancel_btn = QPushButton("取消")
        confirm_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(confirm_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)

    def load_assignments(self):
        self.table.setRowCount(len(self.assignments))
        new_category_set = set(self.new_categories)
        for row, item in enumerate(self.assignments):
            record = item.get("record", {})
            values = [
                str(record.get("order_no", "")),
                str(record.get("notes", "")),
                str(item.get("category", "")),
                str(item.get("detail", "")),
                "是" if item.get("category") in new_category_set else "否",
            ]
            for col, value in enumerate(values):
                table_item = QTableWidgetItem(value)
                if col == 1:
                    table_item.setToolTip(value)
                self.table.setItem(row, col, table_item)


class ManualReasonAssignmentDialog(QDialog):
    """未归因记录批量手动归因窗口。"""

    def __init__(
        self,
        records,
        categories,
        parent=None,
        manual_assign_callback=None,
        ai_assign_callback=None,
        save_note_spec_callback=None,
        single_assign_callback=None,
    ):
        super().__init__(parent)
        self.records = records
        self.categories = categories
        self.manual_assign_callback = manual_assign_callback
        self.ai_assign_callback = ai_assign_callback
        self.save_note_spec_callback = save_note_spec_callback
        self.single_assign_callback = single_assign_callback
        self._loading = False
        self.setup_ui()
        self.load_records()

    def setup_ui(self):
        self.setWindowTitle("手动归因")
        self.resize(1100, 700)
        layout = QVBoxLayout(self)

        top_layout = QHBoxLayout()
        top_layout.addWidget(QLabel("目标分类："))
        self.category_combo = QComboBox()
        self.category_combo.addItems(self.categories)
        top_layout.addWidget(self.category_combo)
        top_layout.addStretch()
        layout.addLayout(top_layout)

        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["店铺", "订单号", "规格编码", "原始退款原因", "备注", "当前真实退款原因"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.cellClicked.connect(self.on_cell_clicked)
        layout.addWidget(self.table, 1)

        self.status_label = QLabel("提示：单击订单号可复制；规格编码和备注可编辑；当前真实退款原因可通过下拉框直接修改并保存。")
        layout.addWidget(self.status_label)

        button_layout = QHBoxLayout()
        save_note_spec_btn = QPushButton("保存备注/规格")
        assign_btn = QPushButton("确认归因")
        ai_assign_btn = QPushButton("AI归因未识别订单")
        close_btn = QPushButton("关闭")
        save_note_spec_btn.clicked.connect(self.save_note_spec_changes)
        assign_btn.clicked.connect(self.apply_manual_assignment)
        ai_assign_btn.clicked.connect(self.apply_ai_assignment)
        close_btn.clicked.connect(self.reject)
        button_layout.addWidget(save_note_spec_btn)
        button_layout.addWidget(assign_btn)
        button_layout.addWidget(ai_assign_btn)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)

    def load_records(self):
        self._loading = True
        self.table.setRowCount(len(self.records))
        for row, record in enumerate(self.records):
            columns = [
                str(record.get("store_name", "")),
                str(record.get("order_no", "")),
                str(record.get("spec_code", "") or "-"),
                str(record.get("reason", "")),
                str(record.get("notes", "")),
            ]
            for col, value in enumerate(columns):
                item = QTableWidgetItem(value)
                if col not in (2, 4):
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if col == 4:
                    item.setToolTip(value)
                if col == 0:
                    item.setData(Qt.UserRole, record.get("id"))
                self.table.setItem(row, col, item)

            combo = NoWheelComboBox()
            combo.addItem("未归因")
            combo.addItems([name for name in self.categories if name != "未归因"])
            current_reason = str(record.get("real_refund_reason") or "未归因").strip() or "未归因"
            if current_reason not in [combo.itemText(index) for index in range(combo.count())]:
                combo.addItem(current_reason)
            combo.setCurrentText(current_reason)
            combo.currentTextChanged.connect(lambda text, row=row: self.on_category_changed(row, text))
            self.table.setCellWidget(row, 5, combo)
        self._loading = False

    def on_cell_clicked(self, row, col):
        if col != 1:
            return
        item = self.table.item(row, col)
        order_no = item.text().strip() if item else ""
        if not order_no:
            return
        QApplication.clipboard().setText(order_no)
        self.status_label.setText(f"已复制订单号：{order_no}")

    def get_note_spec_changes(self):
        changes = []
        for row, record in enumerate(self.records):
            spec_item = self.table.item(row, 2)
            notes_item = self.table.item(row, 4)
            new_spec_code = str(spec_item.text() if spec_item else "").strip()
            if new_spec_code == "-":
                new_spec_code = ""
            new_notes = str(notes_item.text() if notes_item else "")
            old_spec_code = str(record.get("spec_code") or "").strip()
            old_notes = str(record.get("notes") or "")
            if new_spec_code != old_spec_code or new_notes != old_notes:
                changes.append({
                    "record": record,
                    "spec_code": new_spec_code,
                    "notes": new_notes,
                })
        return changes

    def save_note_spec_changes(self):
        if not self.save_note_spec_callback:
            return
        changes = self.get_note_spec_changes()
        if not changes:
            self.status_label.setText("没有需要保存的备注/规格修改")
            QMessageBox.information(self, "提示", "没有需要保存的修改")
            return
        result = self.save_note_spec_callback(changes) or {}
        for change in changes:
            record = change["record"]
            if str(change.get("notes") or "") != str(record.get("notes") or ""):
                record["real_refund_reason"] = ""
            record["spec_code"] = str(change.get("spec_code") or "").strip()
            record["notes"] = str(change.get("notes") or "")
        self.load_records()
        self.status_label.setText(result.get("message", "保存完成"))
        QMessageBox.information(self, "保存完成", result.get("message", "保存完成"))

    def get_selected_records(self):
        selected = []
        selected_rows = sorted({index.row() for index in self.table.selectionModel().selectedRows()})
        for row in selected_rows:
            if 0 <= row < len(self.records):
                selected.append(self.records[row])
        return selected

    def get_selected_record_ids(self):
        return [record.get("id") for record in self.get_selected_records() if record.get("id")]

    def get_selected_category(self):
        return self.category_combo.currentText().strip()

    def _remove_records_by_ids(self, record_ids):
        id_set = set(record_ids or [])
        if not id_set:
            return
        self.records = [record for record in self.records if record.get("id") not in id_set]
        self.load_records()

    def refresh_categories(self, categories):
        current = self.get_selected_category()
        self.categories = list(categories or [])
        self.category_combo.blockSignals(True)
        self.category_combo.clear()
        self.category_combo.addItems(self.categories)
        if current in self.categories:
            self.category_combo.setCurrentText(current)
        self.category_combo.blockSignals(False)
        self.load_records()

    def on_category_changed(self, row, selected_category):
        if self._loading:
            return
        if row < 0 or row >= len(self.records):
            return
        selected_category = str(selected_category or "").strip()
        if not selected_category or selected_category == "未归因":
            return
        record = self.records[row]
        if selected_category == str(record.get("real_refund_reason") or "").strip():
            return
        if not self.single_assign_callback:
            return
        result = self.single_assign_callback(record, selected_category) or {}
        if result.get("success"):
            record["real_refund_reason"] = selected_category
            record["real_refund_reason_detail"] = "当前范围归因手动修正"
            record["real_refund_reason_updated_at"] = result.get("updated_at", "")
            self.status_label.setText(f"已保存真实退款原因：{record.get('order_no', '')} -> {selected_category}")
        else:
            QMessageBox.warning(self, "保存失败", result.get("message", "保存失败"))

    def apply_manual_assignment(self):
        if not self.manual_assign_callback:
            self.accept()
            return
        selected_records = self.get_selected_records()
        selected_category = self.get_selected_category()
        if not selected_records:
            QMessageBox.warning(self, "警告", "请先选择要归因的订单")
            return
        if not selected_category:
            QMessageBox.warning(self, "警告", "请先选择目标分类")
            return
        result = self.manual_assign_callback(selected_records, selected_category) or {}
        updated_ids = result.get("assigned_ids") or [record.get("id") for record in selected_records]
        if result.get("updated_count", 0) > 0:
            self._remove_records_by_ids(updated_ids)
        QMessageBox.information(self, "手动归因", result.get("message", "归因完成"))

    def apply_ai_assignment(self):
        if not self.ai_assign_callback:
            return
        if not self.records:
            QMessageBox.information(self, "提示", "当前窗口没有未归因订单")
            return
        records_with_notes = [
            record for record in self.records
            if str(record.get("notes", "") or "").strip()
        ]
        skipped_empty_notes = len(self.records) - len(records_with_notes)
        if not records_with_notes:
            QMessageBox.information(self, "提示", "当前窗口没有带备注内容的未归因订单，无需调用AI")
            return
        skip_text = f"\n空备注订单 {skipped_empty_notes} 条不会发送。" if skipped_empty_notes else ""
        reply = QMessageBox.question(
            self,
            "确认AI归因",
            f"将把当前窗口剩余 {len(records_with_notes)} 条有备注的未归因订单发送给AI，只发送备注内容。{skip_text}\n是否继续？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        try:
            result = self.ai_assign_callback(list(records_with_notes), self) or {}
            if result.get("categories"):
                self.refresh_categories(result.get("categories"))
            assigned_ids = result.get("assigned_ids", [])
            if assigned_ids:
                self._remove_records_by_ids(assigned_ids)
            message = result.get("message", "AI归因完成")
            if skipped_empty_notes:
                message += f"\n空备注订单已跳过：{skipped_empty_notes} 条"
            QMessageBox.information(self, "AI归因", message)
        except Exception as e:
            QMessageBox.critical(self, "AI归因失败", str(e))


class CurrentRangeReasonAssignmentDialog(QDialog):
    """当前筛选范围全部订单真实退款原因查看与修正窗口。"""

    def __init__(self, records, categories, parent=None, assign_callback=None):
        super().__init__(parent)
        self.records = records or []
        self.categories = list(categories or [])
        self.assign_callback = assign_callback
        self._loading = False
        self.setup_ui()
        self.load_records()

    def setup_ui(self):
        self.setWindowTitle("当前范围归因")
        self.resize(1180, 720)
        layout = QVBoxLayout(self)

        tip = QLabel("显示当前筛选范围内全部订单。修改“当前真实退款原因”下拉框后会立即保存到数据库。")
        tip.setWordWrap(True)
        layout.addWidget(tip)

        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["店铺", "订单号", "规格编码", "原始退款原因", "备注", "当前真实退款原因"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.cellClicked.connect(self.on_cell_clicked)
        layout.addWidget(self.table, 1)

        self.status_label = QLabel("提示：单击订单号可复制；必须点开下拉框后才能选择真实退款原因。")
        layout.addWidget(self.status_label)

        button_layout = QHBoxLayout()
        button_layout.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)

    def load_records(self):
        self._loading = True
        self.table.setRowCount(len(self.records))
        for row, record in enumerate(self.records):
            columns = [
                str(record.get("store_name", "")),
                str(record.get("order_no", "")),
                str(record.get("spec_code", "") or "-"),
                str(record.get("reason", "")),
                str(record.get("notes", "")),
            ]
            for col, value in enumerate(columns):
                item = QTableWidgetItem(value)
                if col == 4:
                    item.setToolTip(value)
                if col == 0:
                    item.setData(Qt.UserRole, record.get("id"))
                self.table.setItem(row, col, item)

            combo = NoWheelComboBox()
            combo.addItem("未归因")
            combo.addItems([name for name in self.categories if name != "未归因"])
            current_reason = str(record.get("real_refund_reason") or "未归因").strip() or "未归因"
            if current_reason not in [combo.itemText(index) for index in range(combo.count())]:
                combo.addItem(current_reason)
            combo.setCurrentText(current_reason)
            combo.currentTextChanged.connect(lambda text, row=row: self.on_category_changed(row, text))
            self.table.setCellWidget(row, 5, combo)
        self._loading = False

    def on_cell_clicked(self, row, col):
        if col != 1:
            return
        item = self.table.item(row, col)
        order_no = item.text().strip() if item else ""
        if not order_no:
            return
        QApplication.clipboard().setText(order_no)
        self.status_label.setText(f"已复制订单号：{order_no}")

    def on_category_changed(self, row, selected_category):
        if self._loading:
            return
        if row < 0 or row >= len(self.records):
            return
        selected_category = str(selected_category or "").strip()
        if not selected_category or selected_category == "未归因":
            return
        record = self.records[row]
        if selected_category == str(record.get("real_refund_reason") or "").strip():
            return
        if not self.assign_callback:
            return
        result = self.assign_callback(record, selected_category) or {}
        if result.get("success"):
            record["real_refund_reason"] = selected_category
            record["real_refund_reason_detail"] = "当前范围归因手动修正"
        else:
            QMessageBox.warning(self, "保存失败", result.get("message", "保存失败"))


class NoWheelComboBox(QComboBox):
    """未展开时忽略滚轮，避免鼠标悬停误切换选项。"""

    def wheelEvent(self, event):
        if self.view().isVisible():
            super().wheelEvent(event)
        else:
            event.ignore()


# ---------------------------- 主程序入口 ---------------------------------
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = RefundManager()
    window.show()
    sys.exit(app.exec_())
